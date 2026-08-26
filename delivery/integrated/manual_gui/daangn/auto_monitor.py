"""
자동 모니터 엔진 (QThread) — 클라 자동검색 기능 전체 GUI화.

기능:
  - 다중조건(엑셀 대분류+상세: 키워드+추가/제외+최소~최대+끌올 n일전) 반복 검색
  - 전국 구단위 적응형 수집 + 프록시 로테이션(계정+프록시)
  - 신규 = 텔레그램 + 구글시트 알림 (지역/제목/가격/링크)
  - 가격변동 재알림(변동표시), 중복방지(sqlite)
  - 같은매물 다른동네 = 당근 id 상이 → 각각 인식
  - 검색 반복 전 휴식 n~n초 랜덤, 검색 전 토큰 갱신(옵션)
"""
import itertools
import time
import sqlite3
import traceback
from datetime import datetime, timedelta

from PyQt6.QtCore import QThread, pyqtSignal

from daangn_ext.search_filters import KeywordRule
from daangn_ext.adaptive import collect_region, collect_lanes, load_dong_regions
from daangn_ext import throttle
from daangn_ext.rest_scheduler import _rand_between
from daangn.notify import TelegramSender, SheetWriter


# ── 휴식 안전 범위 (GUI/외부 cfg 값이 위험해도 여기서 강제 고정) ──
CYCLE_REST_MIN = 10.0      # 사이클 사이 최소 휴식(초). 이하 = 무휴식 폴링 → 차단
CYCLE_REST_MAX = 3600.0    # 사이클 사이 최대 휴식(초)
REGION_GAP_MIN = 0.3       # 지역 간 최소 휴식(초). 이하 = 연타 → IP 스로틀
REGION_GAP_MAX = 10.0      # 지역 간 최대 휴식(초). 이상 = 전국 1바퀴 비현실적

# ── 레인 병렬 ──
# 레인은 프록시 풀을 샤딩해 쓴다. 레인당 IP 가 1개뿐이면 빈응답이 나도 교체할 곳이
# 없어 그 레인이 통째로 멈춘다(실측: 고정 IP 재시도는 성공 4/6, 중앙값 23.5회).
# 레인당 최소 이만큼은 줘야 교체가 의미를 갖는다.
MIN_IP_PER_LANE = 3
MAX_LANES = 16


def _clamp_range(lo, hi, floor, ceil, dflt_lo, dflt_hi):
    """[lo,hi] 를 [floor,ceil] 안으로 고정 + lo<=hi 보정. 값 없거나 이상하면 기본값."""
    try:
        lo = float(lo)
        hi = float(hi)
    except (TypeError, ValueError):
        return dflt_lo, dflt_hi
    lo = min(max(lo, floor), ceil)
    hi = min(max(hi, floor), ceil)
    if hi < lo:
        lo, hi = hi, lo
    return lo, hi


def _pi(v):
    try:
        return int(float(v))
    except Exception:
        return 0


class _P:
    """KeywordRule 용 dict 래퍼."""
    def __init__(self, d):
        self.name = d.get("title", "")
        self.description = d.get("content", "")


class AutoMonitor(QThread):
    log = pyqtSignal(str)
    found = pyqtSignal(dict)        # 신규/변동 매물 → 결과 테이블
    status = pyqtSignal(str)        # 현재 진행 상황 → 상태 라벨

    def __init__(self, parent, cfg: dict):
        super().__init__(parent)
        self.cfg = cfg
        self._stop = False
        self.db = sqlite3.connect(cfg.get("db_path", "./auto_seen.db"),
                                  check_same_thread=False)
        self.db.execute("CREATE TABLE IF NOT EXISTS seen("
                        "id TEXT PRIMARY KEY, price INTEGER, region TEXT, title TEXT)")
        self.db.commit()
        # 알림 송신기 — 실패는 로그로 반드시 노출, 전송은 묶어서(레이트리밋 회피)
        self._tg = TelegramSender(cfg.get("tg_token"), cfg.get("tg_chat"),
                                  log=self.log.emit,
                                  should_stop=lambda: self._stop)
        self._sheet_writer = SheetWriter(cfg.get("sheet_url"),
                                         cfg.get("sheet_cred", "./credentials.json"),
                                         log=self.log.emit)

    def stop(self):
        self._stop = True

    def _rest(self, rmin, rmax, adapt: bool = False):
        """휴식 — _stop 시 즉시 깨어남(종료 지연 방지).
        adapt=True 면 자동감속 배수를 곱한다(차단 신호가 잦으면 지역 간 간격도 벌어진다).
        사이클 휴식은 사용자가 정한 폴링 주기라 곱하지 않는다."""
        d = _rand_between(rmin, rmax)
        if adapt:
            d = throttle.scale(d)
        end = time.monotonic() + d
        while not self._stop:
            left = end - time.monotonic()
            if left <= 0:
                break
            time.sleep(min(0.2, left))
        return d

    # ── 프록시 로테이션 ──
    def _live_proxies(self) -> list:
        """현재 프록시 목록. proxy_provider 있으면 매번 새로 읽어 실행 중 추가/삭제 반영."""
        provider = self.cfg.get("proxy_provider")
        if callable(provider):
            try:
                return list(provider() or [])
            except Exception:
                pass
        return list(self.cfg.get("proxies") or [])

    def _proxy_cycle(self):
        proxies = self._live_proxies()
        if not proxies:
            return None, None
        it = itertools.cycle(proxies)
        return next(it), (lambda: next(it))

    def _plan_lanes(self, proxies) -> int:
        """이번 사이클에 쓸 레인 수.

        레인은 프록시를 **샤딩**해 쓴다(같은 IP 동시요청 = 전멸, 실측 8/8 빈응답).
        따라서 레인 수는 프록시 수를 절대 넘을 수 없고, 레인당 IP 가 너무 적으면
        빈응답 교체 여지가 사라져 오히려 느려진다 → 레인당 최소 MIN_IP_PER_LANE 개 확보.

        cfg["lanes"] 로 사용자가 지정 가능. 0/None 이면 자동(프록시 수 기준).
        """
        n_proxy = len(proxies or [])
        if n_proxy <= 1:
            return 1
        want = _pi(self.cfg.get("lanes")) or 0
        auto = max(1, n_proxy // MIN_IP_PER_LANE)
        n = want if want > 0 else auto
        return max(1, min(n, n_proxy // MIN_IP_PER_LANE or 1, MAX_LANES))

    # ── 알림 ──
    def notify(self, region, article, price, changed=None):
        title = article.get("title", ""); url = article.get("href", "")
        head = f"💱 가격변동 {changed:,}→{price:,}" if changed is not None else "🆕 신규"
        msg = f"{head}\n[{region}] {title}\n{price:,}원\n{url}"
        self.log.emit(msg)
        self._tg.enqueue(msg)
        self._sheet_writer.enqueue_row(
            [datetime.now().strftime("%Y-%m-%d %H:%M"),
             region, title, price, url,
             "가격변동" if changed is not None else "신규"])
        # 결과 테이블용
        self.found.emit({
            "region": region, "title": title, "price": price, "url": url,
            "image": article.get("thumbnail", ""), "desc": article.get("content", ""),
            "boostedAt": article.get("boostedAt", ""),
            "status": "가격변동" if changed is not None else "신규",
        })

    def _flush_notify(self, final=False):
        """대기 중인 알림 전송. 지역 1개 끝날 때마다 호출 — 유실 없이 묶어 보낸다.

        final=True: 종료 직전 마지막 전송. 중지 플래그를 무시하되 30초로 제한.
        """
        if self._tg.pending():
            if final:
                self._tg.flush(deadline=time.monotonic() + 30, ignore_stop=True)
            else:
                self._tg.flush()
        if self._sheet_writer.pending():
            ok, fail = self._sheet_writer.flush()
            if fail:
                self.log.emit(f"[시트] {fail}행 기록 실패 (텔레그램/화면 결과는 정상)")

    # ── 하위호환 래퍼 (기존 테스트/외부 호출용) ──
    def _telegram(self, text):
        ok, err = self._tg.send(text)
        if not ok and self._tg.enabled:
            self._tg._report_failure(err)
        return ok

    def _sheet_append(self, row):
        self._sheet_writer.enqueue_row(row)
        return self._sheet_writer.flush()

    # ── 대상 지역 ──
    def _regions(self):
        if self.cfg.get("scope") == "nationwide":
            # 동 단위여야 한다. 구 ID 는 당근이 대표 동으로 폴백시켜 나머지를 통째로 누락시킴.
            return [r["in"] for r in load_dong_regions(self.cfg["out_json"])]
        return self.cfg.get("regions", [])

    def _dedup_notify(self, arts, region, min_p, max_p, days):
        cutoff = datetime.now() - timedelta(days=days) if days else None
        cur = self.db
        new = changed = 0
        for a in arts:
            aid = str(a.get("id"))
            price = _pi(a.get("price"))
            if (min_p and price < min_p) or (max_p and price > max_p):
                continue
            if cutoff:
                try:
                    if datetime.fromisoformat(a["boostedAt"]).replace(tzinfo=None) < cutoff:
                        continue
                except Exception:
                    pass
            row = cur.execute("SELECT price FROM seen WHERE id=?", (aid,)).fetchone()
            if row is None:
                self.notify(region, a, price)
                cur.execute("INSERT OR REPLACE INTO seen VALUES(?,?,?,?)",
                            (aid, price, region, a.get("title", "")))
                new += 1
            elif row[0] != price:
                self.notify(region, a, price, changed=row[0])
                cur.execute("UPDATE seen SET price=? WHERE id=?", (price, aid))
                changed += 1
        cur.commit()
        return new, changed

    def run(self):
        cfg = self.cfg
        # 다중조건(엑셀) 없으면 단일조건으로
        conditions = cfg.get("conditions") or [{
            "keyword": cfg["keyword"], "extra": cfg.get("extra"),
            "exclude": cfg.get("exclude"), "min": cfg.get("min"),
            "max": cfg.get("max"), "days": cfg.get("days"),
        }]
        proxy0, next_proxy = self._proxy_cycle()
        token = cfg.get("access_token")
        rmin, rmax = _clamp_range(cfg.get("rest_min", 30), cfg.get("rest_max", 90),
                                  CYCLE_REST_MIN, CYCLE_REST_MAX, 30.0, 90.0)
        gmin, gmax = _clamp_range(cfg.get("gap_min", 0.4), cfg.get("gap_max", 1.2),
                                  REGION_GAP_MIN, REGION_GAP_MAX, 0.4, 1.2)
        cur_proxies = self._live_proxies()
        self.log.emit(f"[시작] 조건 {len(conditions)}개, 프록시 {len(cur_proxies)}개")
        self.log.emit(f"[휴식] 사이클 {rmin:.0f}~{rmax:.0f}s · 지역 간 {gmin:.1f}~{gmax:.1f}s")
        try:
            regions = self._regions()
            self.log.emit(f"[지역] {len(regions)}개 (동 단위)")
            cycle = 0
            while not self._stop:
                cycle += 1
                # 프록시 변경은 **사이클 경계에서만** 반영한다. 레인은 시작 시점에
                # 풀을 샤딩해 나눠 갖기 때문에, 도중에 목록이 바뀌면 레인끼리
                # 같은 IP 를 쥐게 될 수 있다(= 동시요청 전멸 조건).
                fresh = self._live_proxies()
                if fresh != cur_proxies:
                    self.log.emit(
                        f"[프록시] {len(cur_proxies)}개 → {len(fresh)}개 (변경 반영)")
                    cur_proxies = fresh
                self.log.emit(f"── 사이클 {cycle} ──")
                for cond in conditions:
                    if self._stop:
                        break
                    rule = KeywordRule(required=[cond["keyword"]],
                                       extra=cond.get("extra") or None, extra_mode="and",
                                       exclude=cond.get("exclude") or None)
                    done = 0
                    total_new = 0
                    missed_total = 0
                    n_lanes = self._plan_lanes(cur_proxies)

                    def on_result(reg_d, arts, cstats, _cond=cond, _rule=rule,
                                  _lanes=n_lanes):
                        """지역 하나가 끝나는 즉시 호출. 레인들이 동시에 부르지만
                        collect_lanes 가 락으로 직렬화해 준다. 중복제거·알림을 여기서
                        스트리밍 처리 — 전국이 끝날 때까지 기다리지 않는다."""
                        nonlocal done, total_new, missed_total
                        reg = reg_d["in"]
                        try:
                            if cstats.get("missed"):
                                missed_total += len(cstats["missed"])
                                self.log.emit(
                                    f"\u26a0\ufe0f [{reg}] '{_cond['keyword']}' 가격구간 "
                                    f"{len(cstats['missed'])}개 확인 실패(IP/세션 차단) — "
                                    "다음 사이클에 재시도. 프록시 부족 의심")
                            if cstats.get("expanded"):
                                self.log.emit(
                                    f"[우회] '{_cond['keyword']}' 응답 억제 → "
                                    f"'{cstats['expanded'][0]}' 로 대체 수집")
                            filtered = [a for a in arts if _rule.match(_P(a))]
                            new, chg = self._dedup_notify(
                                filtered, reg, _cond.get("min"), _cond.get("max"),
                                _cond.get("days"))
                            self._flush_notify()
                            done += 1
                            total_new += new
                            self.log.emit(
                                f"[{done}/{len(regions)}] {reg} · '{_cond['keyword']}' "
                                f"수집 {len(filtered)} · 신규 {new}"
                                + (f" · 변동 {chg}" if chg else "")
                                + f"  (누적 신규 {total_new})")
                        except Exception as e:
                            done += 1
                            self.log.emit(f"[{done}/{len(regions)}] {reg} 오류: {e}")
                        self.status.emit(
                            f"사이클 {cycle} · [{done}/{len(regions)}] "
                            f"'{_cond['keyword']}' 수집 중… (레인 {_lanes})")

                    self.log.emit(
                        f"[레인] {n_lanes}개 병렬 (프록시 {len(cur_proxies)}개"
                        + (f", 레인당 IP {len(cur_proxies)//n_lanes}개 전용)"
                           if n_lanes > 1 else " — 1레인 순차)"))
                    self.status.emit(
                        f"사이클 {cycle} · [0/{len(regions)}] "
                        f"'{cond['keyword']}' 수집 중… (레인 {n_lanes})")
                    try:
                        _, lsm = collect_lanes(
                            cond["keyword"],
                            [{"in": r} for r in regions],
                            proxies=cur_proxies or None,
                            lanes=n_lanes,
                            only_on_sale=True,
                            access_token=token,
                            should_stop=lambda: self._stop,
                            rest_range=(gmin, gmax),
                            on_result=on_result,
                        )
                        if lsm.get("skipped"):
                            self.log.emit(f"[중단] 미처리 지역 {lsm['skipped']}개")
                    except Exception as e:
                        self.log.emit(f"[수집 오류] {type(e).__name__}: {e}")

                    # 조건 1개 끝 — 커버리지 요약(누락을 눈에 보이게)
                    if missed_total:
                        self.log.emit(
                            f"[커버리지] '{cond['keyword']}' 확인 실패 구간 {missed_total}개 "
                            f"/ 지역 {len(regions)}개 — 이번 사이클 결과는 불완전할 수 있음")
                    else:
                        self.log.emit(f"[커버리지] '{cond['keyword']}' 전 구간 확인 완료")
                if self._stop:
                    break
                d = self._rest(rmin, rmax)
                self.log.emit(f"[휴식] {d:.0f}s")
        except Exception:
            self.log.emit("[치명오류]\n" + traceback.format_exc())
        finally:
            # 정지 시에도 대기 중인 알림은 마저 보낸다(유실 방지, 최대 30초)
            try:
                self._flush_notify(final=True)
            except Exception as e:
                self.log.emit(f"[알림 마무리 실패] {type(e).__name__}: {e}")
            self.log.emit(
                f"[알림 집계] 텔레그램 전송 {self._tg._sent_total}건 · "
                f"실패 {self._tg._fail_total}건")
            self.log.emit("[종료] 자동 모니터 정지")


def load_conditions_from_excel(path: str) -> list[dict]:
    """엑셀 대분류+상세조건 로드.
    열: 대분류 | 키워드 | 추가키워드 | 제외키워드 | 최소금액 | 최대금액 | 끌올일수
    (헤더명 유연 매칭, 키워드만 필수)"""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h or "").strip() for h in rows[0]]

    def col(*names):
        for i, h in enumerate(header):
            if any(n in h for n in names):
                return i
        return None
    ci = {"cat": col("대분류", "분류"), "kw": col("키워드", "keyword"),
          "extra": col("추가"), "exc": col("제외"),
          "min": col("최소"), "max": col("최대"), "days": col("끌올", "일")}
    out = []
    def num(v):
        try:
            return int(float(v))
        except Exception:
            return None
    def lst(v):
        import re
        return [x for x in re.split(r"[,\s]+", str(v or "").strip()) if x]
    for r in rows[1:]:
        kw = r[ci["kw"]] if ci["kw"] is not None else None
        if not kw:
            continue
        out.append({
            "category": r[ci["cat"]] if ci["cat"] is not None else "",
            "keyword": str(kw).strip(),
            "extra": lst(r[ci["extra"]]) if ci["extra"] is not None else [],
            "exclude": lst(r[ci["exc"]]) if ci["exc"] is not None else [],
            "min": num(r[ci["min"]]) if ci["min"] is not None else None,
            "max": num(r[ci["max"]]) if ci["max"] is not None else None,
            "days": num(r[ci["days"]]) if ci["days"] is not None else None,
        })
    return out
