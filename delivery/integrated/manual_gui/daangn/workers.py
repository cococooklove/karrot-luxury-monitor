from concurrent.futures import ThreadPoolExecutor
import io
import threading
import time
import traceback
from PyQt6.QtCore import QThread, pyqtSignal, QObject
from curl_cffi import requests

from daangn import api
from daangn.constants import MAX_RETRY_COUNT
from daangn.errors import DaangnCancelledError
from daangn.model import Product
from daangn.task import CrawlTask
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from io import BytesIO
from PIL import Image as PILImage
from openpyxl.drawing.image import Image
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from daangn.proxy_manager import ProxyManager
from daangn.utils import ReqLimitter, image_contain_resize


def get_col_width_row_height(img_width, img_height):
    col_width = img_width * 63.2 / 504.19
    row_height = img_height * 225.35 / 298.96
    return (col_width, row_height)


class ExportExcel(QThread):
    finished = pyqtSignal(bool)

    def __init__(
        self, parent: QObject | None, path: str, prds: list[Product], withImg: bool
    ):
        super().__init__(parent)
        self.setObjectName("ExportExcel")

        self.path = path
        self.prds = prds
        self.withImg = withImg
        self.cancelled = False
        self._image_size = (100, 100)
        self._image_row_height: float | None = None
        self._image_col_width: float | None = None
        self._sheets: dict[str, Worksheet] = {}
        self._sheet_rows: dict[str, int] = {}
        self._sheet_titles: set[str] = set()
        self._workbook: Workbook | None = None

    def run(self):
        try:
            wb = self._init_workbook()

            for p in self.prds:
                if self.cancelled:
                    self.finished.emit(True)
                    return

                ws, row = self._get_sheet_for_keyword(p.searched_keyword)
                self._write_row(ws, row, p)  # type: ignore

            wb.save(self.path)

        except:  # noqa: E722
            with open(f"ERROR_LOG_{int(time.time())}", "w", encoding="utf8") as f:
                traceback.print_exc(file=f)
            self.finished.emit(False)
            return

        self.finished.emit(True)

    def cancel(self):
        self.cancelled = True

    def _init_workbook(self):
        wb = Workbook()
        self._workbook = wb
        self._sheets.clear()
        self._sheet_rows.clear()
        self._sheet_titles.clear()

        if self.withImg:
            img_cell_w, img_cell_h = get_col_width_row_height(*self._image_size)
            self._image_col_width = img_cell_w
            self._image_row_height = img_cell_h
        else:
            self._image_col_width = None
            self._image_row_height = None

        return wb

    def _get_sheet_for_keyword(self, keyword: str):
        normalized = (keyword or "").strip() or "None"
        ws = self._sheets.get(normalized)
        if ws is None:
            ws = self._create_sheet_for_keyword(normalized)
            self._sheets[normalized] = ws  # type: ignore
            self._sheet_rows[normalized] = 2
        row = self._sheet_rows[normalized]
        self._sheet_rows[normalized] = row + 1
        return ws, row

    def _create_sheet_for_keyword(self, keyword: str):
        if not self._workbook:
            raise Exception("Workbook is not initialized")

        title = self._build_sheet_title(keyword)

        if not self._sheets:
            ws = self._workbook.active
            ws.title = title  # type: ignore
        else:
            ws = self._workbook.create_sheet(title)

        self._setup_sheet(ws)  # type: ignore
        return ws

    def _setup_sheet(self, ws: Worksheet):
        headers = ["지역별 순번", "이름", "설명", "가격", "주소", "링크"]
        if self.withImg:
            headers.append("이미지")
        headers.append("끌어올리기 시간")
        ws.append(headers)

        if self.withImg and self._image_col_width is not None:
            ws.column_dimensions["G"].width = self._image_col_width

    def _build_sheet_title(self, keyword: str):
        invalid_chars = "[]:*?/\\"
        sanitized = "".join(
            "_" if ch in invalid_chars else ch for ch in keyword
        ).strip()
        if not sanitized:
            sanitized = "None"

        base = sanitized
        title = base[:31]
        idx = 1
        while title in self._sheet_titles or not title:
            suffix = f"_{idx}"
            available = 31 - len(suffix)
            trimmed = base[:available] if available > 0 else base[:31]
            title = f"{trimmed}{suffix}"
            idx += 1
        self._sheet_titles.add(title)
        return title

    def _write_row(self, ws: Worksheet, row: int, product: Product):
        name, desc = self._sanitize_texts(product)
        ws.append(self._build_row_values(product, name, desc))

        self._attempt(self._set_text_cell, ws, "B", row)
        self._attempt(self._set_text_cell, ws, "C", row)
        self._attempt(self._set_price_cell, ws, row, product.price)
        self._attempt(self._write_hyperlink, ws, row, product.url)

        if self.withImg:
            self._attempt(self._insert_image, ws, row, product.image)

    def _build_row_values(self, product: Product, name: str, desc: str):
        row = [
            product.no,
            name,
            desc,
            None,
            product.area,
            None,
        ]
        if self.withImg:
            row.append(None)
        row.append(product.boostedAt.strftime("%Y-%m-%d %H:%M:%S"))
        return row

    def _sanitize_texts(self, product: Product):
        return tuple(
            ILLEGAL_CHARACTERS_RE.sub("", value)
            for value in (product.name, product.description)
        )

    def _set_text_cell(self, ws, column: str, row: int):
        ws[f"{column}{row}"].data_type = "s"

    def _set_price_cell(self, ws, row: int, price: str):
        cell = ws[f"D{row}"]
        cell.value = float(price)
        cell.number_format = "#,##0"

    def _write_hyperlink(self, ws, row: int, url: str):
        cell = ws[f"F{row}"]
        cell.value = url
        cell.hyperlink = url
        cell.style = "Hyperlink"

    def _insert_image(self, ws, row: int, url: str):
        r = requests.get(url, timeout=10, impersonate="chrome")
        r.raise_for_status()

        img_data = BytesIO(r.content)
        pilimg = PILImage.open(img_data)
        thumb = image_contain_resize(pilimg, (256, 256))

        fp = BytesIO()
        thumb.save(fp, format="png")
        fp.seek(0)

        img = Image(fp)
        img.width, img.height = self._image_size  # type: ignore

        ws.add_image(img, f"G{row}")
        if self._image_row_height is not None:
            ws.row_dimensions[row].height = self._image_row_height  # type: ignore

    def _attempt(self, fn, *args, **kwargs):
        try:
            fn(*args, **kwargs)
        except Exception:
            pass


class CancelableImageDownloader(QThread):
    finished = pyqtSignal(bytes, str)
    failed = pyqtSignal(str)

    def __init__(self, parent: QObject | None, url: str, token: str):
        super().__init__(parent)
        self.setObjectName("CrawlThread")

        self.url = url
        self.token = token
        self.running = True

    def run(self):
        try:
            r = requests.get(self.url, timeout=5, stream=True, impersonate="chrome")
            r.raise_for_status()

            content = bytearray()
            for chunk in r.iter_content(chunk_size=1024):
                if not self.running:
                    raise Exception("Stopped")

                content.extend(chunk)

            self.finished.emit(bytes(content), self.token)
        except Exception:
            self.failed.emit(self.token)

    def cancel(self):
        self.running = False


class CrawlThread(QThread):
    finished = pyqtSignal()
    error = pyqtSignal()
    message = pyqtSignal(str)
    new_products = pyqtSignal(str, list)

    def __init__(
        self,
        parent: QObject | None,
        req_min_ms: int,
        proxies: list[str],
        max_workers: int,
        tasks: list[CrawlTask],
    ) -> None:
        super().__init__(parent)
        self.setObjectName("CrawlThread")

        self.stop_event = threading.Event()
        self.proxy_manager = ProxyManager(proxies, req_min_ms)
        self.req_min_ms = req_min_ms
        # 동시요청 수는 **살아있는 IP 수를 넘을 수 없다**. 같은 IP 로 동시요청하면
        # 실측상 전멸(8/8 빈응답, adaptive 문서 참조)이라 워커가 프록시보다 많으면
        # 스스로 IP 를 태우는 꼴이 된다. 설정값은 상한일 뿐, 실제값은 여기서 깎는다.
        self.max_workers = max(1, min(max_workers, len(proxies))) if proxies else max_workers
        self.workers_clamped = self.max_workers < max_workers
        self.tasks = tasks

    def run(self) -> None:
        lock = threading.Lock()
        total_products = 0
        error_area_list = []
        completed = 0
        retry_cnt = 0
        no_proxy_limitter = ReqLimitter(self.req_min_ms)

        if self.workers_clamped:
            self.message.emit(
                f"동시 요청을 {self.max_workers}개로 낮췄습니다 "
                f"(프록시 {len(self.proxy_manager.proxies)}개 — 같은 IP 동시요청은 차단을 부릅니다)"
            )

        total_req_per_proxy = {
            proxy: 0 for proxy in ["No Proxy", *self.proxy_manager.proxies]
        }
        success_per_proxy = {
            proxy: 0 for proxy in ["No Proxy", *self.proxy_manager.proxies]
        }
        failed_per_proxy = {
            proxy: 0 for proxy in ["No Proxy", *self.proxy_manager.proxies]
        }

        try:

            def worker(task: CrawlTask):
                nonlocal completed, total_products, retry_cnt

                try:
                    last_error = None
                    for current_retry in range(MAX_RETRY_COUNT):
                        if self.stop_event.is_set():
                            break

                        if current_retry > 0:
                            with lock:
                                retry_cnt += 1

                        proxy = None
                        if self.proxy_manager.empty():
                            no_proxy_limitter.wait(self.stop_event)
                        else:
                            proxy = self.proxy_manager.acquire(self.stop_event)

                        with lock:
                            total_req_per_proxy[proxy or "No Proxy"] += 1

                        area_name, area_code = task.area
                        try:
                            # 추가기능: 키워드 포함필터(추가/제외) 규칙
                            from daangn_ext.search_filters import KeywordRule
                            rule = KeywordRule(
                                required=[task.keyword],
                                extra=task.extra_keywords or None,
                                extra_mode="and",
                                exclude=task.exclude_keywords or None,
                            )
                            fn = api.get_products_adaptive if task.adaptive else api.get_products
                            # 프록시 풀을 항상 넘긴다 — robust 가 빈응답마다 IP 를 교체할 수
                            # 있어야 한다(A/B 실측: 교체 6/6 성공 vs 고정 4/6, 중앙값 5.5회 vs 23.5회).
                            # 종전에는 적응형에만 풀을 줘서 일반 검색은 한 IP 로 30회를 두드렸다.
                            pool = self.proxy_manager.proxies or None
                            # 두 함수는 앞 두 인자 순서가 서로 반대다
                            # (get_products(area_code, area) / get_products_adaptive(area, area_code)).
                            # 위치인자로 넘기면 적응형에서 둘이 뒤바뀌어 collect_region 이
                            # regionId 대신 지역 이름("강남구 역삼동")을 받고 0건이 된다 → 이름으로 넘긴다.
                            products = fn(
                                area_code=area_code,
                                area=area_name,
                                keyword=task.keyword,
                                only_tradeable=task.only_tradeable,
                                minimum_price=task.minimum,
                                maximum_price=task.maximum,
                                proxy=proxy,
                                access_token=task.access_token,
                                rule=rule,
                                proxies=pool,
                                should_stop=self.stop_event.is_set,   # 정지 즉시 반응
                            )
                            if self.stop_event.is_set():
                                break
                            self.new_products.emit(area_name, products)

                            with lock:
                                completed += 1
                                total_products += len(products)
                                success_per_proxy[proxy or "No Proxy"] += 1
                                notify(area_name, task.keyword)

                            break

                        except Exception as e:
                            print(e)
                            with lock:
                                last_error = e
                                failed_per_proxy[proxy or "No Proxy"] += 1
                            if proxy:
                                self.proxy_manager.record_failure(proxy)
                    else:
                        with lock:
                            with io.StringIO() as sio:
                                traceback.print_exception(last_error, file=sio)
                                error_area_list.append((*task.area, sio.getvalue()))
                except DaangnCancelledError:
                    pass

            def notify(area: str, keyword: str):
                msg = f"{completed}/{len(self.tasks)} '{area}'에서 '{keyword}' 상품을 가져왔습니다  (총 {total_products}개)"
                if error_area_list:
                    msg += f" (오류 {len(error_area_list)})"
                if retry_cnt:
                    msg += f" (재시도 {retry_cnt}회)"
                self.message.emit(msg)

            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                [executor.submit(worker, task) for task in self.tasks]
                executor.shutdown(wait=True)

            # check error proxy
            error_proxy_list = []
            for proxy in self.proxy_manager.proxies:
                if proxy == "No Proxy":
                    continue

                if (
                    total_req_per_proxy.get(proxy, 0) > 0
                    and success_per_proxy.get(proxy, 0) == 0
                ):
                    error_proxy_list.append(proxy)

            if error_area_list or error_proxy_list:
                with open(f"report_{int(time.time())}.log", "w", encoding="utf8") as f:
                    f.write("상품 수=" + str(total_products) + "\n")
                    f.write("오류 수=" + str(len(error_area_list)) + "\n")

                    f.write("\n오류 프록시 목록 (모든 요청이 실패)\n")
                    for proxy in error_proxy_list:
                        f.write(
                            f"{'프록시 없는 요청' if proxy == 'No Proxy' else proxy} 요청={total_req_per_proxy.get(proxy, 0)} 성공={success_per_proxy.get(proxy, 0)} 실패={failed_per_proxy.get(proxy, 0)}\n"
                        )

                    f.write(
                        f"\n오류 지역 목록 ({MAX_RETRY_COUNT}회 재시도 동안 오류)\n"
                    )
                    for name, area, err in error_area_list:
                        f.write(f"{name} ({area})\n")
                        f.write(err)
                        f.write("\n\n")

                    f.write("\n전체 프록시 통계\n")
                    for proxy in ["No Proxy", *self.proxy_manager.proxies]:
                        f.write(
                            f"{'프록시 없는 요청' if proxy == 'No Proxy' else proxy} 요청={total_req_per_proxy.get(proxy, 0)} 성공={success_per_proxy.get(proxy, 0)} 실패={failed_per_proxy.get(proxy, 0)}\n"
                        )

        except DaangnCancelledError:
            pass

        except Exception as e:
            with open(f"ERROR_LOG_{int(time.time())}", "w", encoding="utf8") as f:
                traceback.print_exception(e, file=f)

            self.error.emit()
        finally:
            self.message.emit(
                f"작업이 정지되었습니다 (총 {total_products}개) (오류 {len(error_area_list)}개)"
            )
            self.finished.emit()

    def stop(self):
        self.stop_event.set()
