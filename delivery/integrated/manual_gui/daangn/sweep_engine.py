"""
지역 스윕 엔진 — 클라 자동검색 로직 전체. **순수 파이썬, Qt 의존 없음.**

시그널 대신 콜백(on_log/on_found/on_status)을 받는다. 그래서 이벤트 루프가 없는
헤드리스 서버 런타임이 이 클래스를 그대로 돌릴 수 있다(GUI 는 daangn.auto_monitor
의 AutoMonitor 어댑터가 콜백을 시그널로 이어 준다).

이 모듈은 PyQt 를 import 하지 않는다 — 그게 분리의 요점이다.

기능:
  - 다중조건(엑셀 대분류+상세: 키워드+추가/제외+최소~최대+끌올 n일전) 반복 검색
  - 전국 구단위 적응형 수집 + 프록시 로테이션(계정+프록시)
  - 신규 = 텔레그램 + 구글시트 알림 (지역/제목/가격/링크)
  - 가격변동 재알림(변동표시), 중복방지(sqlite)
  - 같은매물 다른동네 = 당근 id 상이 → 각각 인식
  - 검색 반복 전 휴식 n~n초 랜덤, 검색 전 토큰 갱신(옵션)
"""
import itertools
import json
import os
import time
import sqlite3
import traceback
from datetime import datetime, timedelta, timezone

from daangn_ext.app_api import SORT_RECENT
from daangn_ext.search_filters import KeywordRule
from daangn_ext.adaptive import (collect_region, collect_lanes, load_dong_regions,
                                 get_app_fallback_logger,
                                 reset_app_fallback_notices,
                                 set_app_fallback_logger)
from daangn_ext import throttle
from daangn_ext.rest_scheduler import _rand_between
from daangn.notify import TelegramSender, SheetWriter, match_line


# ── 휴식 안전 범위 (GUI/외부 cfg 값이 위험해도 여기서 강제 고정) ──
CYCLE_REST_MIN = 10.0      # 사이클 사이 최소 휴식(초). 이하 = 무휴식 폴링 → 차단
CYCLE_REST_MAX = 3600.0    # 사이클 사이 최대 휴식(초)
REGION_GAP_MIN = 0.3       # 지역 간 최소 휴식(초). 이하 = 연타 → IP 스로틀
REGION_GAP_MAX = 10.0      # 지역 간 최대 휴식(초). 이상 = 전국 1바퀴 비현실적

# ── 레인 병렬 ──
# 레인은 프록시 풀을 샤딩해 쓴다. 레인당 IP 가 1개뿐이면 빈응답이 나도 교체할 곳이
# 없어 그 레인이 통째로 멈춘다(실측: 고정 IP 재시도는 성공 4/6, 중앙값 23.5회).
# 레인당 최소 이만큼은 줘야 교체가 의미를 갖는다.
MIN_IP_PER_LANE = 3
MAX_LANES = 16


def _clamp_range(lo, hi, floor, ceil, dflt_lo, dflt_hi):
    """[lo,hi] 를 [floor,ceil] 안으로 고정 + lo<=hi 보정. 값 없거나 이상하면 기본값."""
    try:
        lo = float(lo)
        hi = float(hi)
    except (TypeError, ValueError):
        return dflt_lo, dflt_hi
    lo = min(max(lo, floor), ceil)
    hi = min(max(hi, floor), ceil)
    if hi < lo:
        lo, hi = hi, lo
    return lo, hi


def _pi(v):
    try:
        return int(float(v))
    except Exception:
        return 0


class _P:
    """KeywordRule 용 dict 래퍼."""
    def __init__(self, d):
        self.name = d.get("title", "")
        self.description = d.get("content", "")


def _noop(*_a, **_k):
    pass


# 최신순 응답 한 페이지(20건)가 덮는 시간. 2026-09-01 실측 — 역삼동 23분,
# 해운대 우동 22분, 제주 노형동 25분으로 지역 편차가 거의 없다(반경검색이
# 택배·광역 매물을 함께 물어오기 때문). 아래 수렴 판정의 기준값이다.
PAGE_SPAN_MIN = 23.0

# 처음 보는 (조건,지역)은 어디서 멈춰야 할지 모른다 — 워터마크가 없다. 그렇다고
# 끝까지 파면 첫 사이클이 통째로 과거 발굴이 되어 정작 신규를 못 따라간다.
# 그래서 첫 방문은 이 구간만 훑고 워터마크를 세운다. 그 이전 매물은 '신규'가
# 아니므로 안 봐도 된다 — 우리가 잡으려는 건 지금 올라오는 물건이다.
FIRST_VISIT_WINDOW_MIN = 120.0


def _now_iso():
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _iso_ago(minutes):
    return (datetime.now(timezone.utc)
            - timedelta(minutes=minutes)).isoformat().replace("+00:00", "Z")


def sweep_capacity(req_per_sec, page_span_min=PAGE_SPAN_MIN):
    """이 처리량으로 수렴 가능한 (지역수 x 조건수) 곱.

    최신순 + 정지 규칙에서 사이클당 요청은
        지역수 x 조건수 x (주기 / 한 페이지가 덮는 시간)
    이고, 사이클이 주기 안에 끝나야 하므로 주기가 양변에서 약분된다:
        지역수 x 조건수 = 한 페이지가 덮는 시간(초) x 초당요청수
    이 곱을 넘기면 사이클이 주기보다 길어지고, 그러면 다음 사이클은 더 깊이
    파야 해서 더 길어진다 — 발산한다. 지금까지 이 저장소의 스윕이 동 6537개를
    설정해 두고도 앞쪽 몇십 개만 돌던 이유가 정확히 이것이다.
    """
    try:
        return float(page_span_min) * 60.0 * float(req_per_sec)
    except (TypeError, ValueError):
        return 0.0


def _ckey(cond, region) -> str:
    """커서 키. 키워드만으로는 부족하다.

    엑셀 조건표는 같은 키워드를 추가어·제외어·가격대만 달리해 여러 줄로 넣을 수
    있다(load_conditions_from_excel 은 키워드로 중복제거하지 않는다). 키가
    키워드뿐이면 조건 A 가 끝낸 지역을 조건 B 가 통째로 건너뛴다 — B 는 그 지역을
    한 번도 본 적이 없는데도.
    """
    parts = [str(cond.get("keyword") or ""), str(cond.get("extra") or ""),
             str(cond.get("exclude") or ""), str(cond.get("min") or ""),
             str(cond.get("max") or ""), str(cond.get("days") or "")]
    return "\x1f".join(parts) + "\t" + str(region)


class _RegionCursor:
    """지역 순회 진행 상황을 디스크에 남겨, 사이클이 완주하지 못해도 이어가게 한다.

    왜 필요한가: 실서버 설정은 `sweep_regions` 에 동 6537개가 들어 있고 한 지역을
    최대 200페이지까지 판다. 사이클 1회는 몇 달짜리인데 앱은 배포·재부팅으로
    훨씬 자주 재시작된다. 재시작마다 목록 앞에서 다시 시작하면 앞쪽 지역만
    영원히 반복 수집되고 뒤쪽 지역은 **단 한 번도 방문되지 않는다**. 그래서
    '이번 패스에 이미 끝낸 지역'을 기억해 두고, 다음 실행은 남은 것부터 돈다.

    한 바퀴를 다 돌면 기록을 비우고 새 패스를 시작한다 — 커버리지는 느릴 수 있어도
    공평해진다. 이 파일이 없거나 깨져도 그냥 처음부터 도는 것으로 퇴화할 뿐,
    스윕을 막지 않는다(수집이 커서보다 중요하다).

    ── 워터마크 ──
    검색을 최신순(FLEA_MARKET_SORT_OPTION_RECENT)으로 돌리면 응답이 publishedAt
    내림차순으로 단조라, '지난 방문 시각까지만 페이징하고 멈춘다'가 성립한다.
    그래서 (조건,지역)마다 **마지막으로 훑은 시각**을 같이 남긴다. 다음 방문은
    그 시각까지만 파면 되므로 깊이가 볼륨에서 계산되고, 커버리지가 확률적 표본이
    아니라 보장이 된다(관련도 정렬에서는 1페이지 재현율이 15% 였다).

    워터마크는 '방문을 시작한 시각'이다. 그때 이후로 올라온 것은 다음 방문에서
    반드시 보이고, 그 이전은 이번에 이미 훑었다. 수집이 불완전했던 지역(missed)은
    워터마크를 올리지 않는다 — 올리면 그 구간을 영영 안 본다.
    """

    def __init__(self, path="./data/sweep_cursor.json", log=None, save_every=20):
        self.path = path
        self._log = log or (lambda m: None)
        self.save_every = max(1, int(save_every))
        self.passes = 1
        self._done = set()
        self._marks = {}
        self._since_write = 0
        self._load()

    def _load(self):
        try:
            with open(self.path, encoding="utf-8") as f:
                d = json.load(f)
            self.passes = int(d.get("pass") or 1)
            self._done = set(d.get("done") or [])
            self._marks = dict(d.get("marks") or {})
        except FileNotFoundError:
            pass
        except Exception as e:
            # 깨진 커서 때문에 스윕을 멈추지 않는다. 처음부터 돌 뿐이다.
            self._log(f"[커서] 읽기 실패 — 처음부터 돕니다: {str(e)[:80]}")

    def watermark(self, key):
        """이 (조건,지역)을 마지막으로 훑은 시각(ISO8601). 처음이면 None.

        None 이면 정지 규칙 없이 파야 한다 — 그 지역의 과거를 한 번도 안 봤으므로
        어디서 멈춰야 할지 알 수 없다. 호출자가 첫 방문 깊이를 따로 정한다."""
        return self._marks.get(key)

    def set_watermark(self, key, iso):
        """훑기 완료 시각 기록. 뒤로 돌리지 않는다 — 시계가 흔들려도 이미 본
        구간을 다시 보는 건 낭비일 뿐이지만, 앞당기면 그 사이가 유실된다."""
        if not key or not iso:
            return
        cur = self._marks.get(key)
        if cur and str(iso) <= cur:
            return
        self._marks[key] = str(iso)
        self._since_write += 1
        if self._since_write >= self.save_every:
            self._since_write = 0
            self._write()

    def forget_stale(self, keys):
        """설정에서 사라진 (조건,지역)의 기록을 버린다. 안 버리면 파일이
        영원히 커지고, 조건을 지웠다 되살릴 때 낡은 워터마크로 구멍이 난다."""
        live = set(keys)
        drop = [k for k in self._marks if k not in live]
        for k in drop:
            del self._marks[k]
        gone = [k for k in self._done if k not in live]
        for k in gone:
            self._done.discard(k)
        if drop or gone:
            self._write()
        return len(drop) + len(gone)

    def _write(self):
        try:
            d = os.path.dirname(os.path.abspath(self.path))
            os.makedirs(d, exist_ok=True)
            tmp = self.path + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump({"pass": self.passes, "done": sorted(self._done),
                           "marks": self._marks}, f, ensure_ascii=False)
            os.replace(tmp, self.path)
        except Exception as e:
            self._log(f"[커서] 저장 실패(진행은 계속): {str(e)[:80]}")

    def order(self, regions):
        """이번에 돌 지역 = 이 패스에서 아직 안 끝낸 것들.

        설정이 바뀌어 목록에서 사라진 지역은 기록에 남아 있어도 무시된다.
        남은 게 없으면 한 바퀴를 다 돈 것이므로 기록을 비우고 새 패스를 연다."""
        regions = list(regions or [])
        pending = [r for r in regions if r not in self._done]
        if regions and not pending:
            self.passes += 1
            self._done.clear()
            self._since_write = 0
            self._write()
            self._log(f"[커서] 전 지역 1회 순회 완료 — 패스 {self.passes} 시작")
            return regions
        if len(pending) != len(regions):
            self._log(f"[커서] 패스 {self.passes} 이어서 — 남은 지역 "
                      f"{len(pending)}/{len(regions)}개")
        return pending

    def mark(self, region):
        """지역 하나 완료. 매번 쓰면 IO 가 과해 save_every 마다 모아 쓴다."""
        if not region or region in self._done:
            return
        self._done.add(region)
        self._since_write += 1
        if self._since_write >= self.save_every:
            self._since_write = 0
            self._write()

    def flush(self):
        self._since_write = 0
        self._write()


class SweepEngine:
    """지역 스윕 엔진 — 순수 파이썬. Qt 의존 없음.

    시그널 대신 콜백을 받는다(전부 선택). GUI 는 AutoMonitor 어댑터가
    콜백을 시그널로 이어 주고, 헤드리스 서버 런타임은 콜백만으로 직접 돌린다.
      on_log(str) / on_found(dict) / on_status(str)
    """

    def __init__(self, cfg: dict, on_log=None, on_found=None, on_status=None):
        self.cfg = cfg
        self.on_log = on_log or _noop
        self.on_found = on_found or _noop
        self.on_status = on_status or _noop
        self._stop = False
        # 조건표 — 앱 알림과 같은 파일을 본다. 이게 없던 동안 스윕은 등록
        # 조건(브랜드 + 가격)만 봤는데, 브랜드는 가격 제한 없이 등록하므로
        # 사실상 그 브랜드 전 매물이 알림으로 나갔다.
        self._rules_path = cfg.get("rules_path", "./data/alert_rules.json")
        self._rules_mtime = None
        self._rules = None
        self.db = sqlite3.connect(cfg.get("db_path", "./auto_seen.db"),
                                  check_same_thread=False)
        self.db.execute("CREATE TABLE IF NOT EXISTS seen("
                        "id TEXT PRIMARY KEY, price INTEGER, region TEXT, title TEXT)")
        self.db.commit()
        # 알림 송신기 — 실패는 로그로 반드시 노출, 전송은 묶어서(레이트리밋 회피)
        self._tg = TelegramSender(cfg.get("tg_token"), cfg.get("tg_chat"),
                                  log=self._log,
                                  should_stop=lambda: self._stop)
        self._sheet_writer = SheetWriter(cfg.get("sheet_url"),
                                         cfg.get("sheet_cred", "./credentials.json"),
                                         log=self._log)

    # ── 콜백 디스패치 (기존 시그널 emit 자리) ──
    def _log(self, msg):
        self.on_log(msg)

    def _found(self, payload):
        self.on_found(payload)

    def _status(self, msg):
        self.on_status(msg)

    def stop(self):
        self._stop = True

    def _rest(self, rmin, rmax, adapt: bool = False):
        """휴식 — _stop 시 즉시 깨어남(종료 지연 방지).
        adapt=True 면 자동감속 배수를 곱한다(차단 신호가 잦으면 지역 간 간격도 벌어진다).
        사이클 휴식은 사용자가 정한 폴링 주기라 곱하지 않는다."""
        d = _rand_between(rmin, rmax)
        if adapt:
            d = throttle.scale(d)
        end = time.monotonic() + d
        while not self._stop:
            left = end - time.monotonic()
            if left <= 0:
                break
            time.sleep(min(0.2, left))
        return d

    # ── 프록시 로테이션 ──
    def _live_proxies(self) -> list:
        """현재 프록시 목록. proxy_provider 있으면 매번 새로 읽어 실행 중 추가/삭제 반영."""
        provider = self.cfg.get("proxy_provider")
        if callable(provider):
            try:
                return list(provider() or [])
            except Exception:
                pass
        return list(self.cfg.get("proxies") or [])

    def _proxy_cycle(self):
        proxies = self._live_proxies()
        if not proxies:
            return None, None
        it = itertools.cycle(proxies)
        return next(it), (lambda: next(it))

    def _plan_lanes(self, proxies) -> int:
        """이번 사이클에 쓸 레인 수.

        레인은 프록시를 **샤딩**해 쓴다(같은 IP 동시요청 = 전멸, 실측 8/8 빈응답).
        따라서 레인 수는 프록시 수를 절대 넘을 수 없고, 레인당 IP 가 너무 적으면
        빈응답 교체 여지가 사라져 오히려 느려진다 → 레인당 최소 MIN_IP_PER_LANE 개 확보.

        cfg["lanes"] 로 사용자가 지정 가능. 0/None 이면 자동(프록시 수 기준).
        """
        n_proxy = len(proxies or [])
        if n_proxy <= 1:
            return 1
        want = _pi(self.cfg.get("lanes")) or 0
        auto = max(1, n_proxy // MIN_IP_PER_LANE)
        n = want if want > 0 else auto
        return max(1, min(n, n_proxy // MIN_IP_PER_LANE or 1, MAX_LANES))

    # ── 알림 ──
    def notify(self, region, article, price, changed=None):
        title = article.get("title", ""); url = article.get("href", "")
        if changed is not None:
            msg = (f"💱 가격변동 {changed:,}→{price:,}원\n"
                   f"{title[:50]}\n📍 {region} · 지역 훑기\n{url}")
        else:
            # 앱 알림과 같은 모양으로 보낸다 — 받는 사람에게는 한 방이다.
            _, rule = self._rules_now().verdict(title, price,
                                                article.get("content") or "")
            msg = match_line(rule.label() if rule else "",
                             title, price, region, source="지역 훑기", url=url)
        self._log(msg)
        self._tg.enqueue(msg)
        self._sheet_writer.enqueue_row(
            [datetime.now().strftime("%Y-%m-%d %H:%M"),
             region, title, price, url,
             "가격변동" if changed is not None else "신규"])
        # 결과 테이블용
        self._found({
            "id": article.get("id"),
            "region": region, "title": title, "price": price, "url": url,
            "image": article.get("thumbnail", ""), "desc": article.get("content", ""),
            "boostedAt": article.get("boostedAt", ""),
            "status": "가격변동" if changed is not None else "신규",
        })

    def _flush_notify(self, final=False):
        """대기 중인 알림 전송. 지역 1개 끝날 때마다 호출 — 유실 없이 묶어 보낸다.

        final=True: 종료 직전 마지막 전송. 중지 플래그를 무시하되 30초로 제한.
        """
        if self._tg.pending():
            if final:
                self._tg.flush(deadline=time.monotonic() + 30, ignore_stop=True)
            else:
                self._tg.flush()
        if self._sheet_writer.pending():
            ok, fail = self._sheet_writer.flush()
            if fail:
                self._log(f"[시트] {fail}행 기록 실패 (텔레그램/화면 결과는 정상)")

    # ── 하위호환 래퍼 (기존 테스트/외부 호출용) ──
    def _telegram(self, text):
        ok, err = self._tg.send(text)
        if not ok and self._tg.enabled:
            self._tg._report_failure(err)
        return ok

    def _sheet_append(self, row):
        self._sheet_writer.enqueue_row(row)
        return self._sheet_writer.flush()

    # ── 대상 지역 ──
    def _regions(self):
        if self.cfg.get("scope") == "nationwide":
            # 동 단위여야 한다. 구 ID 는 당근이 대표 동으로 폴백시켜 나머지를 통째로 누락시킴.
            return [r["in"] for r in load_dong_regions(self.cfg["out_json"])]
        return self.cfg.get("regions", [])

    def _rules_now(self):
        """조건표. 파일이 바뀌면 다시 읽는다(엑셀을 새로 넣어도 재시작 불필요)."""
        try:
            mt = os.path.getmtime(self._rules_path)
        except OSError:
            mt = None
        if self._rules is None or mt != self._rules_mtime:
            from daangn_ext.alert_rules import RuleTable
            self._rules_mtime, self._rules = mt, RuleTable.load(self._rules_path)
        return self._rules

    def _dedup_notify(self, arts, region, min_p, max_p, days):
        cutoff = datetime.now() - timedelta(days=days) if days else None
        cur = self.db
        rules = self._rules_now()
        # 같은 매물을 앱 알림 경로가 이미 알렸을 수 있다. 저장소가 둘이라
        # (watch.db / auto_seen.db) 각자 처음이면 같은 매물이 두 번 나간다.
        seen_elsewhere = self.cfg.get("already_notified") or (lambda _id: False)
        new = changed = 0
        for a in arts:
            aid = str(a.get("id"))
            price = _pi(a.get("price"))
            if len(rules):
                # 조건표가 있으면 그게 진실이다 — 등록 조건 대신 이걸 쓴다.
                # 스윕에는 '상한 초과 추적'이 없다(워치리스트는 앱 경로가 만든다).
                # 조건 밖이면 seen 에도 남기지 않는다 — 나중에 값이 내려와
                # 조건 안으로 들어오면 그때 처음 알려야 한다.
                from daangn_ext.alert_rules import HIT
                if rules.verdict(a.get("title") or "", price,
                                 a.get("content") or "")[0] != HIT:
                    continue
            elif (min_p and price < min_p) or (max_p and price > max_p):
                continue
            if cutoff:
                try:
                    if datetime.fromisoformat(a["boostedAt"]).replace(tzinfo=None) < cutoff:
                        continue
                except Exception:
                    pass
            row = cur.execute("SELECT price FROM seen WHERE id=?", (aid,)).fetchone()
            if row is None:
                try:
                    dup = bool(seen_elsewhere(aid))
                except Exception:
                    dup = False
                if dup:
                    # 알리지는 않되 본 것으로 남긴다 — 다음 사이클에 또 묻지 않는다.
                    cur.execute("INSERT OR REPLACE INTO seen VALUES(?,?,?,?)",
                                (aid, price, region, a.get("title", "")))
                    continue
                self.notify(region, a, price)
                cur.execute("INSERT OR REPLACE INTO seen VALUES(?,?,?,?)",
                            (aid, price, region, a.get("title", "")))
                new += 1
            elif row[0] != price:
                self.notify(region, a, price, changed=row[0])
                cur.execute("UPDATE seen SET price=? WHERE id=?", (price, aid))
                changed += 1
        cur.commit()
        return new, changed

    def run(self):
        cfg = self.cfg
        # 다중조건(엑셀) 없으면 단일조건으로
        conditions = cfg.get("conditions") or [{
            "keyword": cfg["keyword"], "extra": cfg.get("extra"),
            "exclude": cfg.get("exclude"), "min": cfg.get("min"),
            "max": cfg.get("max"), "days": cfg.get("days"),
        }]
        proxy0, next_proxy = self._proxy_cycle()
        token = cfg.get("access_token")
        # 자동수확 연동(옵션): token_provider() 가 있으면 사이클마다 최신 access 로 갱신.
        # LDPlayer 온디바이스 수확(ld_harvest)이 accounts.json 을 신선하게 유지 → 여기서 읽음.
        token_provider = cfg.get("token_provider")
        if token_provider:
            try:
                token = token_provider() or token
                self._log("[토큰] 자동수확 연동 — 사이클마다 갱신")
            except Exception as e:
                self._log(f"[토큰] provider 초기화 실패: {e}")
        # 계정 안정화 스케줄러(옵션): 사이클마다 계정 라운드로빈 + 계정별 고정프록시 +
        # daily_cap/warmup. 활성 시 1계정-1IP 로 검색해 핑거프린트 이상·과다요청 방지.
        sched = None
        if cfg.get("stabilize"):
            try:
                from daangn_ext.account_scheduler import AccountScheduler
                sched = AccountScheduler(
                    accounts_fp=cfg.get("accounts_fp", "./accounts.json"),
                    daily_cap=int(cfg.get("daily_cap", 300)),
                    warmup_days=int(cfg.get("warmup_days", 3)),
                    log=self._log)
                self._log("[안정화] 계정 라운드로빈 + daily_cap/warmup ON")
            except Exception as e:
                self._log(f"[안정화] 스케줄러 초기화 실패(기존방식 유지): {e}")
        rmin, rmax = _clamp_range(cfg.get("rest_min", 30), cfg.get("rest_max", 90),
                                  CYCLE_REST_MIN, CYCLE_REST_MAX, 30.0, 90.0)
        gmin, gmax = _clamp_range(cfg.get("gap_min", 0.4), cfg.get("gap_max", 1.2),
                                  REGION_GAP_MIN, REGION_GAP_MAX, 0.4, 1.2)
        cur_proxies = self._live_proxies()
        # 앱API→웹크롤 폴백 경고를 운영자 로그로 끌어온다. 이걸 안 걸면 폴백이
        # stderr 로만 나가 GUI/서버 로그에서는 '매물 없음'과 구분이 안 된다.
        _prev_fallback_log = get_app_fallback_logger()
        set_app_fallback_logger(self._log)
        self._log(f"[시작] 조건 {len(conditions)}개, 프록시 {len(cur_proxies)}개")
        self._log(f"[휴식] 사이클 {rmin:.0f}~{rmax:.0f}s · 지역 간 {gmin:.1f}~{gmax:.1f}s")
        try:
            regions = self._regions()
            self._log(f"[지역] {len(regions)}개 (동 단위)")
            # 지역이 많으면 사이클은 사실상 완주하지 않는다(실서버 6537동 × 최대
            # 200페이지). 그때 재시작마다 앞에서 다시 시작하면 뒤쪽 지역은 영영
            # 방문되지 않으므로, 끝낸 지역을 기억해 다음 실행이 이어받는다.
            cursor = _RegionCursor(cfg.get("cursor_fp", "./data/sweep_cursor.json"),
                                   log=self._log)
            cycle = 0
            while not self._stop:
                cycle += 1
                reset_app_fallback_notices()   # 사이클마다 폴백 경고를 최소 1회 다시 크게
                # 커서는 (조건, 지역) 쌍 단위다. 조건마다 전 지역을 돌기 때문에
                # 지역만으로 기억하면 첫 조건이 끝낸 지역을 나머지 조건이 건너뛴다.
                cycle_started = time.monotonic()
                cycle_requests = 0
                _pairs = [_ckey(c, r) for c in conditions for r in regions]
                _pending = set(cursor.order(_pairs))
                # ── 계정 안정화: 사이클마다 계정 라운드로빈 + 그 계정 고정프록시(없으면 KR네이티브) ──
                cur_code = None
                sched_proxies = None      # sched 활성 시 이 사이클의 프록시(계정 바인딩). None=네이티브
                if sched:
                    # 전 계정 토큰 신선화(LDPlayer 수확) — provider 부작용 이용. 반환값은 무시.
                    if token_provider:
                        try:
                            token_provider()
                        except Exception as e:
                            self._log(f"[수확] 실패(계속): {str(e)[:60]}")
                    pick = sched.pick()
                    if not pick:
                        self._log("[안정화] 전 계정 캡/쿨다운 도달 — 휴식 후 재시도")
                        self._rest(rmin, rmax)
                        continue
                    cur_code = pick["code"]
                    token = pick["access"]
                    sched_proxies = [pick["proxy"]] if pick["proxy"] else None
                    self._log(
                        f"[계정] {cur_code[:6]} · 잔여 {pick['remaining']} · "
                        f"프록시 {'고정1' if pick['proxy'] else 'KR네이티브'}  ({sched.status()})")
                # 자동수확 연동(스케줄러 미사용 시): 최신 access 재조회. access 30분 만료 대응.
                elif token_provider:
                    try:
                        nt = token_provider()
                        if nt and nt != token:
                            token = nt
                            self._log("[토큰] 갱신 반영")
                    except Exception as e:
                        self._log(f"[토큰] 갱신 조회 실패: {e}")
                # 프록시 변경은 **사이클 경계에서만** 반영한다. 레인은 시작 시점에
                # 풀을 샤딩해 나눠 갖기 때문에, 도중에 목록이 바뀌면 레인끼리
                # 같은 IP 를 쥐게 될 수 있다(= 동시요청 전멸 조건).
                fresh = self._live_proxies()
                if fresh != cur_proxies:
                    self._log(
                        f"[프록시] {len(cur_proxies)}개 → {len(fresh)}개 (변경 반영)")
                    cur_proxies = fresh
                self._log(f"── 사이클 {cycle} ──")
                # 안정화 활성 시 = 이 계정 고정프록시(또는 네이티브). 아니면 = 전체 풀.
                lane_proxies = sched_proxies if sched else cur_proxies
                for cond in conditions:
                    if self._stop:
                        break
                    # 이 조건에서 아직 안 끝낸 지역만. 커서가 비어 있으면(새 패스)
                    # 전 지역이 그대로 들어온다.
                    cond_regions = [r for r in regions if _ckey(cond, r) in _pending]
                    if not cond_regions:
                        self._log(f"[커서] '{cond['keyword']}' 이 패스에서 완료 — 건너뜀")
                        continue
                    rule = KeywordRule(required=[cond["keyword"]],
                                       extra=cond.get("extra") or None, extra_mode="and",
                                       exclude=cond.get("exclude") or None)
                    done = 0
                    total_new = 0
                    missed_total = 0
                    n_lanes = self._plan_lanes(lane_proxies or [])

                    def on_result(reg_d, arts, cstats, _cond=cond, _rule=rule,
                                  _lanes=n_lanes, _regs=cond_regions):
                        """지역 하나가 끝나는 즉시 호출. 레인들이 동시에 부르지만
                        collect_lanes 가 락으로 직렬화해 준다. 중복제거·알림을 여기서
                        스트리밍 처리 — 전국이 끝날 때까지 기다리지 않는다."""
                        nonlocal done, total_new, missed_total
                        reg = reg_d["in"]
                        try:
                            if cstats.get("missed"):
                                missed_total += len(cstats["missed"])
                                self._log(
                                    f"\u26a0\ufe0f [{reg}] '{_cond['keyword']}' 가격구간 "
                                    f"{len(cstats['missed'])}개 확인 실패(IP/세션 차단) — "
                                    "다음 사이클에 재시도. 프록시 부족 의심")
                            if cstats.get("expanded"):
                                self._log(
                                    f"[우회] '{_cond['keyword']}' 응답 억제 → "
                                    f"'{cstats['expanded'][0]}' 로 대체 수집")
                            filtered = [a for a in arts if _rule.match(_P(a))]
                            new, chg = self._dedup_notify(
                                filtered, reg, _cond.get("min"), _cond.get("max"),
                                _cond.get("days"))
                            self._flush_notify()
                            done += 1
                            total_new += new
                            self._log(
                                f"[{done}/{len(_regs)}] {reg} · '{_cond['keyword']}' "
                                f"수집 {len(filtered)} · 신규 {new}"
                                + (f" · 변동 {chg}" if chg else "")
                                + f"  (누적 신규 {total_new})")
                            if cstats.get("stop_before_unapplied") is not None:
                                # 앱API 가 죽어 웹크롤로 떨어진 지역이다. 웹 결과에는
                                # 정지 규칙이 안 걸렸으므로 '이 시각 이후 전부'가
                                # 아니다. 워터마크를 올리면 그 구간이 영영 빈다.
                                self._log(
                                    f"⚠️ [{reg}] 정지 규칙 미적용(웹크롤 폴백) — "
                                    "이 지역은 다음 사이클에 같은 구간을 다시 훑는다")
                            elif not cstats.get("missed"):
                                cursor.set_watermark(_ckey(_cond, reg), visit_iso)
                            if not cstats.get("missed"):
                                # 이 지역은 이번 패스에서 **끝났다**. 사이클이 중간에
                                # 죽어도 다음 실행이 여기부터 이어간다.
                                #
                                # 확인 못 한 가격구간이 남았으면 찍지 않는다 — 바로 위에서
                                # "다음 사이클에 재시도"라고 알려 놓고 커서가 걸러버리면
                                # 그 지역은 이 패스가 끝날 때까지(6537동 스코프에서는 몇 달)
                                # 영영 불완전한 채로 남는다.
                                cursor.mark(_ckey(_cond, reg))
                        except Exception as e:
                            done += 1
                            self._log(f"[{done}/{len(_regs)}] {reg} 오류: {e}")
                        self._status(
                            f"사이클 {cycle} · [{done}/{len(_regs)}] "
                            f"'{_cond['keyword']}' 수집 중… (레인 {_lanes})")

                    self._log(
                        f"[레인] {n_lanes}개 병렬 (프록시 {len(cur_proxies)}개"
                        + (f", 레인당 IP {len(cur_proxies)//n_lanes}개 전용)"
                           if n_lanes > 1 else " — 1레인 순차)"))
                    # 이번 방문의 기준 시각. 최신순 + 정지 규칙이라 이 시각 이후에
                    # 올라온 것은 다음 방문에서 반드시 보인다. 지역이 끝나면
                    # 워터마크를 여기로 올린다(수집이 온전했을 때만).
                    visit_iso = _now_iso()
                    first_iso = _iso_ago(FIRST_VISIT_WINDOW_MIN)
                    region_payload = [
                        {"in": r,
                         "stop_before": cursor.watermark(_ckey(cond, r)) or first_iso}
                        for r in cond_regions]
                    self._status(
                        f"사이클 {cycle} · [0/{len(cond_regions)}] "
                        f"'{cond['keyword']}' 수집 중… (레인 {n_lanes})")
                    try:
                        _, lsm = collect_lanes(
                            cond["keyword"],
                            region_payload,
                            proxies=lane_proxies or None,
                            lanes=n_lanes,
                            only_on_sale=True,
                            access_token=token,
                            should_stop=lambda: self._stop,
                            rest_range=(gmin, gmax),
                            on_result=on_result,
                            # 최신순이라야 정지 규칙이 성립한다. 관련도 정렬에서는
                            # 최근 1시간 신규의 1페이지 재현율이 15% 뿐이었다.
                            sort_option=cfg.get("sort_option", SORT_RECENT),
                        )
                        cycle_requests += int(lsm.get("requests") or 0)
                        if lsm.get("skipped"):
                            self._log(f"[중단] 미처리 지역 {lsm['skipped']}개")
                        if lsm.get("app_api_fallbacks"):
                            # 웹크롤은 명품 브랜드를 억제한다('샤넬' 0건). 이 줄이 없으면
                            # 토큰만료·앱API 변경이 '매물 없음'으로만 보인다.
                            self._log(
                                f"🚨 [앱API 장애] '{cond['keyword']}' 지역 "
                                f"{lsm['app_api_fallbacks']}/{len(cond_regions)}개가 웹크롤로 폴백 "
                                f"({lsm.get('app_api_failed')}) — 명품 키워드는 0건으로 보일 수 있음. "
                                "토큰/헤더(data/config.json) 확인 필요")
                        # 안정화: 이 계정의 일일 사용량 기록(≈지역수). 차단신호면 격리.
                        if sched and cur_code:
                            sched.note(cur_code, len(cond_regions))
                            if missed_total:
                                sched.note_block(cur_code)
                    except Exception as e:
                        self._log(f"[수집 오류] {type(e).__name__}: {e}")
                    finally:
                        # 조건 하나가 끝나거나 죽어도 여기까지의 진행은 남긴다 —
                        # 안 그러면 크래시할 때마다 같은 지역을 다시 판다.
                        cursor.flush()

                    # 조건 1개 끝 — 커버리지 요약(누락을 눈에 보이게)
                    if missed_total:
                        self._log(
                            f"[커버리지] '{cond['keyword']}' 확인 실패 구간 {missed_total}개 "
                            f"/ 지역 {len(cond_regions)}개 — 이번 사이클 결과는 불완전할 수 있음")
                    else:
                        self._log(f"[커버리지] '{cond['keyword']}' 전 구간 확인 완료")
                if self._stop:
                    break
                # 사이클 1회를 실제로 돌아 봤으니 이 설정이 수렴하는지 말할 수 있다.
                # 짐작이 아니라 방금 관측한 처리량으로 판정한다.
                cyc_dt = time.monotonic() - cycle_started
                if cyc_dt > 0 and cycle_requests > 0:
                    rps = cycle_requests / cyc_dt
                    cap = sweep_capacity(rps)
                    load = len(regions) * len(conditions)
                    self._log(
                        f"[처리량] 요청 {cycle_requests}건 / {cyc_dt:.0f}s = "
                        f"{rps:.1f} req/s · 이 속도로 감당 가능한 (지역x조건) "
                        f"{cap:,.0f} · 현재 설정 {load:,}")
                    if load > cap:
                        self._log(
                            f"⚠️ [수렴 불가] 현재 설정이 감당치를 {load / cap:.1f}배 "
                            f"넘습니다. 이대로면 사이클이 갈수록 길어지고 뒤쪽 지역은 "
                            f"사실상 방문되지 않습니다. 조건 {len(conditions)}개 기준 "
                            f"지역을 {cap / max(1, len(conditions)):,.0f}개 이하로 줄이거나 "
                            f"레인·프록시를 늘리세요.")
                d = self._rest(rmin, rmax)
                self._log(f"[휴식] {d:.0f}s")
        except Exception:
            self._log("[치명오류]\n" + traceback.format_exc())
        finally:
            # 전역 로거를 원래대로. 스윕이 끝난 뒤에도 우리 _log 를 물고 있으면,
            # GUI 가 AutoMonitor 를 정리한 뒤에는 수신자 0인 시그널로 흘러들어가
            # 폴백 경고가 stderr 로도 안 남는다(고치려던 것보다 나빠진다).
            try:
                set_app_fallback_logger(_prev_fallback_log)
            except Exception:
                pass
            # 정지 시에도 대기 중인 알림은 마저 보낸다(유실 방지, 최대 30초)
            try:
                self._flush_notify(final=True)
            except Exception as e:
                self._log(f"[알림 마무리 실패] {type(e).__name__}: {e}")
            self._log(
                f"[알림 집계] 텔레그램 전송 {self._tg._sent_total}건 · "
                f"실패 {self._tg._fail_total}건")
            self._log("[종료] 자동 모니터 정지")


def load_conditions_from_excel(path: str) -> list[dict]:
    """엑셀 대분류+상세조건 로드.
    열: 대분류 | 키워드 | 추가키워드 | 제외키워드 | 최소금액 | 최대금액 | 끌올일수
    (헤더명 유연 매칭, 키워드만 필수)"""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h or "").strip() for h in rows[0]]

    def col(*names):
        for i, h in enumerate(header):
            if any(n in h for n in names):
                return i
        return None
    ci = {"cat": col("대분류", "분류"), "kw": col("키워드", "keyword"),
          "extra": col("추가"), "exc": col("제외"),
          "min": col("최소"), "max": col("최대"), "days": col("끌올", "일")}
    out = []
    def num(v):
        try:
            return int(float(v))
        except Exception:
            return None
    def lst(v):
        import re
        return [x for x in re.split(r"[,\s]+", str(v or "").strip()) if x]
    for r in rows[1:]:
        kw = r[ci["kw"]] if ci["kw"] is not None else None
        if not kw:
            continue
        out.append({
            "category": r[ci["cat"]] if ci["cat"] is not None else "",
            "keyword": str(kw).strip(),
            "extra": lst(r[ci["extra"]]) if ci["extra"] is not None else [],
            "exclude": lst(r[ci["exc"]]) if ci["exc"] is not None else [],
            "min": num(r[ci["min"]]) if ci["min"] is not None else None,
            "max": num(r[ci["max"]]) if ci["max"] is not None else None,
            "days": num(r[ci["days"]]) if ci["days"] is not None else None,
        })
    return out
