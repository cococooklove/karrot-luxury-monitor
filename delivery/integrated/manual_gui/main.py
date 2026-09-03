import json
import os
import re
from datetime import datetime, timedelta
from io import BytesIO

from typing import Any
from uuid import uuid4

from daangn.controller import MainController
from daangn.detail import render_to_html
from daangn.utils import image_contain_resize
from daangn.workers import CancelableImageDownloader, ExportExcel
from daangn.model import Product
from daangn.sweep_engine import APP_API_LANES, MIN_IP_PER_LANE, sweep_capacity
from daangn.task import CrawlTask
from daangn.ui_mainwindow import Ui_MainWindow

from PyQt6 import QtWidgets, QtCore, QtGui


WATCH_SWEEP_INTERVAL = 600          # 워치리스트 스윕 주기(초)


# ── 게스트(에뮬) 앱 트래픽 프록시 ──────────────────────────────────────────
# 토큰을 만드는 건 게스트 안의 당근 앱이고, 그 앱이 접속하는 곳이 곧 그 계정이
# 로그인한 곳이다. 서버가 미국 IP 라 검색만 KR 프록시로 돌려도 계정은 미국에서
# 붙는다. 그래서 계정에 지정한 프록시를 그 계정이 든 인스턴스의 안드로이드
# 전역 프록시로 건다(당근 앱이 그걸 존중하는 것은 실측으로 확인).
#
# 상용 프록시는 대개 자격증명이 붙는데 전역 프록시에는 그 칸이 없다. 그래서
# 호스트 루프백에 인증 없는 릴레이를 열고 업스트림에 인증을 붙여 넘긴다.
# 릴레이는 프로세스가 사는 동안 떠 있어야 하므로 여기서 하나만 들고 간다.
_GUEST_RELAY = None


def guest_proxy_sync(accounts_path="./accounts.json", log=None):
    """계정별 프록시를 게스트에 반영한다. 반환 {index: endpoint|None}.

    수확 직전에 부른다 — 프록시가 걸린 뒤에 앱이 콜드스타트해야 토큰 갱신이
    그 프록시로 나간다. 값이 안 바뀌었으면 adb 를 더 쓰지 않으므로 매 틱 불러도
    싸다. 어떤 이유로 실패해도 수확을 막지 않는다(프록시는 보조, 토큰이 본체).
    """
    global _GUEST_RELAY
    log = log or (lambda m: None)
    try:
        import ld_proxy
        from daangn_ext.proxy_relay import ProxyRelay
    except Exception as e:
        log(f"[프록시] 모듈 없음 — 게스트 프록시를 건너뜁니다: {str(e)[:60]}")
        return {}
    try:
        want = ld_proxy._account_proxies(accounts_path)
        need_relay = {px for px in want.values() if "@" in px}
        if need_relay:
            if _GUEST_RELAY is None:
                _GUEST_RELAY = ProxyRelay(bind="127.0.0.1", log=log).start()
            for px in need_relay:
                _GUEST_RELAY.add(px, px)      # 키 = URL. 같은 업스트림은 릴레이가 합친다

        def endpoint_for(px):
            if "@" not in px:
                return px.split("//", 1)[-1]
            return _GUEST_RELAY.endpoint(px) if _GUEST_RELAY else None

        return ld_proxy.apply_account_proxies(accounts_path, log=log,
                                              endpoint_for=endpoint_for)
    except Exception as e:
        log(f"[프록시] 게스트 반영 실패(계속): {type(e).__name__}: {str(e)[:80]}")
        return {}


def mirror_app_keywords_to_sweep(router, queue, log=None, enabled=False) -> int:
    """앱 알림에 배정된 키워드를 검색 스윕 대기열에도 싣는다.

    왜 필요한가: 앱 알림은 **계정 인증동네** 기반이라 그 계정이 사는 동네만 본다.
    운영 계정이 오산·평택이면 강남 명품은 원천적으로 안 보인다. 검색 스윕은
    지역을 인자로 받으므로 그 사각지대를 메우라고 있는 건데, 라우터가 키워드를
    app / sweep 중 **하나로만** 배정한다. 브랜드 20개가 앱 슬롯(상한 30)에 다
    들어가면 스윕 큐가 비고, 스윕은 아예 뜨지 않는다 — 실서버가 정확히 그 상태였다.

    그래서 '둘 다' 를 가능하게 한다. 예전에는 기본으로 꺼 두었다 — 켜면 설정된
    지역 전체의 매물이 그대로 알림이 됐기 때문이다(키워드가 조건 없이 등록돼
    있어 매칭이 곧 알림이었다). 조건표가 생기고 스윕도 그것을 보게 된 뒤로 그
    이유가 사라졌다. 이제 **조건표가 있으면 켠다** — 조건이 물량을 정한다.
    `alert_settings.json` 의 `sweep_mirror_app` 로 강제로 끌 수 있다.
    """
    _log = log or (lambda m: None)
    if not enabled or router is None or queue is None:
        return 0
    try:
        routes = router.routes()          # [{keyword, route, ...}, ...]
    except Exception as e:
        _log(f"[스윕미러] 라우트 읽기 실패: {str(e)[:80]}")
        return 0
    added = 0
    for r in routes or []:
        if (r or {}).get("route") != "app":
            continue
        kw = (r or {}).get("keyword")
        if not kw:
            continue
        try:
            if queue.add(kw):
                added += 1
        except Exception as e:
            _log(f"[스윕미러] '{kw}' 대기열 추가 실패: {str(e)[:60]}")
    if added:
        _log(f"[스윕미러] 앱 키워드 {added}개를 검색 스윕에도 실었습니다 "
             "— 계정 동네 밖 지역을 스윕이 맡습니다")
    return added


def harvest_interval() -> int:
    """수확 주기. 숫자의 소유자는 ld_autoharvest 다 — 여기 1200 을 다시 적으면
    안 된다.

    수확 간격과 토큰 신선도 임계는 한 부등식으로 묶여 있다(임계 > 간격). 두 값이
    다른 파일에 따로 적혀 있던 동안 실서버 4계정이 만료된 채 방치됐다. 그래서
    GUI 스레드도 헤드리스 루프도 여기를 거쳐 같은 값을 본다."""
    try:
        import ld_autoharvest
        return int(ld_autoharvest.HARVEST_INTERVAL)
    except Exception:
        return 1200

# 브랜드로 넓게 등록해 받고, 모델별 조건 수백 줄은 이 파일이 들고 있다가
# 알림 직전에 건다. 앱 알림 키워드 등록에는 상한이 있어(수십 개) 모델별
# 가격대를 키워드로 등록할 방법이 없기 때문이다.
ALERT_RULES_FILE = "./data/alert_rules.json"


def load_alert_rules(path=ALERT_RULES_FILE):
    """저장된 알림 룰 테이블. 없으면 빈 테이블(= 아무것도 안 거른다)."""
    from daangn_ext.alert_rules import RuleTable
    return RuleTable.load(path)


def brand_register_groups(brand_list, rules):
    """브랜드를 끌올일수별로 묶는다 → [(브랜드들, days)].

    라우터는 한 호출에 필터 하나만 받는다. 조건표는 줄마다 끌올일수가 다를
    수 있으므로 같은 값끼리 묶어 한 번에 넣는다. 브랜드 하나에 값이 여럿이면
    가장 느슨한(큰) 값을 쓴다 — 좁게 잡으면 조건에 맞는 매물을 놓친다."""
    from daangn_ext.alert_rules import brand_days
    days = brand_days(rules)
    groups: dict = {}
    for b in brand_list or []:
        b = str(b or "").strip()
        if b and b not in groups.setdefault(days.get(b), []):
            groups[days.get(b)].append(b)
    return [(ks, d) for d, ks in groups.items() if ks]


class AlertRulesCache:
    """룰 파일이 바뀌면 다시 읽는다 — 엑셀을 새로 넣어도 재시작이 필요 없다."""

    def __init__(self, path=None):
        # 경로는 만들 때 푼다 — 테스트가 모듈 상수를 바꿔 끼울 수 있게.
        self.path = path or ALERT_RULES_FILE
        self._mtime = None
        self._table = None

    def get(self):
        try:
            mt = os.path.getmtime(self.path)
        except OSError:
            mt = None
        if self._table is None or mt != self._mtime:
            self._mtime, self._table = mt, load_alert_rules(self.path)
        return self._table

    def stamp(self):
        """마지막으로 읽은 파일의 mtime — 화면이 '이미 본 파일인지' 비교하는 값."""
        self.get()
        return self._mtime


_WATCH_LABELS = {
    "entered_range": "🎯 조건 진입(가격 인하)",
    "price_down": "↓ 가격 인하",
    "price_up": "↑ 가격 인상",
    "sold": "판매완료",
    "deleted": "삭제됨",
    "republished": "끌올",
}


def mark_range_entries(events, store, router, rules=None):
    """가격이 내려와 조건 범위에 '들어온' 이벤트를 따로 표시한다.

    상한보다 비싸 알림을 보내지 않은 매물도 추적은 하고 있다(라우터가 당근에
    여유 상한으로 등록해 둔 덕에 우리 시야에는 들어와 있다). 그 값이 내려와
    조건 안으로 들어오면 그때 처음 알린다 — 처음 270만이던 매물이 250만이
    되면 알려 달라는 요구가 이것이다.

    판정 근거는 인하 전후 가격이다. 내리기 전에는 범위 밖이었고 내린 뒤
    범위 안이면 '조건 진입'이다. 원래부터 범위 안이던 매물의 인하는 지금처럼
    '가격 인하'로 남는다."""
    has_rules = rules is not None and len(rules)
    if not events or store is None or (router is None and not has_rules):
        return events
    from daangn_ext.alert_rules import HIT
    out = []
    for e in events:
        if e.get("kind") != "price_down":
            out.append(e)
            continue
        try:
            row = store.get(e.get("id")) or {}
            new, old = int(e.get("new") or 0), int(e.get("old") or 0)
            if has_rules:
                # 룰 테이블이 있으면 그게 조건의 진실이다. 제목으로 어느 룰에
                # 맞는지 다시 판정한다 — 등록 키워드(브랜드)에는 모델별
                # 가격대가 없다.
                title = e.get("title") or row.get("title") or ""

                def inside(v):
                    return rules.verdict(title, v)[0] == HIT
                ok = bool(new and old)
            else:
                cond = router.condition_for(row.get("keyword"))
                mx = cond.get("max")
                mn = cond.get("min")

                def inside(v):
                    return (mn is None or v >= int(mn)) and v <= int(mx)
                ok = bool(mx and new and old)
            if ok and inside(new) and not inside(old):
                e = dict(e, kind="entered_range")
        except Exception:
            pass
        out.append(e)
    return out


def watch_event_lines(events):
    """워치리스트 이벤트 → 알림 한 줄씩. 모르는 종류는 건너뛴다."""
    out = []
    for e in events or []:
        label = _WATCH_LABELS.get(e.get("kind"))
        if not label:
            continue
        title = e.get("title") or e.get("id") or ""
        url = e.get("url") or ""
        kind = e.get("kind")
        if kind in ("price_down", "price_up", "entered_range"):
            body = f"{int(e.get('old') or 0):,}원 → {int(e.get('new') or 0):,}원"
        elif kind == "republished":
            body = f"{e.get('old')}회 → {e.get('new')}회"
        else:
            body = ""
        out.append(" ".join(x for x in (f"[{label}]", title, body, url) if x))
    return out


def _enqueue_watch_blocks(tg, events, lines):
    """워치 이벤트를 텔레그램 매물 블록으로 적재. 이벤트가 없으면(옛 호출) 줄 요약으로."""
    from daangn.notify import watch_event_block
    n = 0
    for e in events or []:
        block = watch_event_block(e)
        if block:
            tg.enqueue_item(block)
            n += 1
    if not n and lines:
        tg.enqueue("📉 가격변동\n" + "\n".join(lines))


def watch_sweep_budget(active, interval_sec):
    """이번 스윕에서 조회할 최대 건수.

    활성 전체를 fresh 주기(4시간) 안에 한 바퀴 돈다고 보고 비례 배분한다.
    활성이 있으면 최소 1건은 본다."""
    from daangn_ext import article_watch
    active = int(active or 0)
    if active <= 0:
        return 0
    per_cycle = active * int(interval_sec) / float(article_watch.FRESH_INTERVAL)
    return max(1, int(per_cycle + 0.999))


def watch_status_text(active, next_check_at, now):
    """추적 현황 한 줄."""
    active = int(active or 0)
    if not active:
        return "추적 중 0건"
    left = int(next_check_at or 0) - int(now)
    if left <= 0:
        return f"추적 중 {active}건 · 다음 점검 대기"
    if left >= 3600:
        when = f"{left // 3600}시간 {(left % 3600) // 60}분 후"
    else:
        when = f"{max(1, left // 60)}분 후"
    return f"추적 중 {active}건 · 다음 점검 {when}"


STATE_ICONS = {
    "new": "🆕 신규",
    "tracking": "● 추적중",
    "down": "↓ 인하",
    "up": "↑ 인상",
    "paused": "⏸ 추적중단",
    "ended": "✓ 종료",
}

# 필터 버튼이 고르는 값. all 은 전부.
LISTING_FILTERS = ("all", "new", "down", "ended")


def _delta_text(first_price, price):
    """최초 감지가 대비 증감. 기준선을 모르면 '-'."""
    if not first_price or not isinstance(price, int) or not isinstance(first_price, int):
        return "-"
    d = price - first_price
    if d == 0:
        return "0원 (0.0%)"
    pct = d * 100.0 / first_price
    return f"{d:+,}원 ({pct:+.1f}%)"


def _ts_text(ts):
    import time as _t
    if not ts:
        return "-"
    try:
        return _t.strftime("%m/%d %H:%M", _t.localtime(int(ts)))
    except Exception:
        return "-"


ROUTE_NAMES = {"app": "앱 알림", "sweep": "검색 스윕"}


class RowLineDelegate(QtWidgets.QStyledItemDelegate):
    """표의 줄 구분선. 스타일시트 `::item { border-bottom }` 으로 그리면
    Qt 가 셀 배경(setBackground)을 통째로 무시해 버려서, 선은 여기서 긋는다."""
    LINE = "#F1EEE8"

    def paint(self, painter, option, index):
        super().paint(painter, option, index)
        painter.save()
        painter.setPen(QtGui.QColor(self.LINE))
        r = option.rect
        painter.drawLine(r.left(), r.bottom(), r.right(), r.bottom())
        painter.restore()


def row_lines(table):
    """모든 QTableWidget 에 붙인다 — 안 붙이면 그 표만 줄 구분선이 없다."""
    table.setItemDelegate(RowLineDelegate(table))
    return table


class RuleGrid(QtWidgets.QTableWidget):
    """조건표 — 엑셀 파일 대신 화면에서 바로 적는 표.

    엑셀이 원본이던 동안 조건 다섯 줄을 고치려면 파일을 열고, 고치고,
    저장하고, [다시 읽기]를 눌러야 했다. 파일을 옮기면 [열기]가 죽었다.
    이 표가 엑셀처럼 굴면 그 네 단계가 사라진다: 엑셀에서 복사한 것을
    Ctrl+V 로 붙이고, Delete 로 비우고, 끝 줄에 적으면 새 줄이 생긴다.
    행 번호는 엑셀과 같이 머리글이 1행이라 파서 오류의 "5행"이 표의 5다.

    값은 문자열 그대로 둔다 — 해석(가격 쉼표·브랜드 이어받기·제외 쉼표)은
    parse_rule_rows 한 곳이 엑셀과 똑같이 한다."""
    MIN_ROWS = 8
    ERR_ROW_BG = "#FCEBE8"          # 계정표 '점검필요' 와 같은 붉은 바탕
    edited = QtCore.pyqtSignal()    # 사용자가 셀을 바꿨다 — 적용 전 '수정됨'

    def __init__(self, parent=None):
        from daangn_ext.rule_grid import RULE_COLS
        super().__init__(0, len(RULE_COLS), parent)
        self.setObjectName("rulesGrid")
        row_lines(self)
        self.setHorizontalHeaderLabels(RULE_COLS)
        self.horizontalHeader().setStretchLastSection(True)
        # 행 번호 열 — 공용 헤더 QSS 의 세로 패딩(12px)이 30px 줄에서 숫자를
        # 짓눌러 깨뜨린다. 세로 헤더만 따로 폭·줄높이를 잡는다.
        vh = self.verticalHeader()
        vh.setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Fixed)
        vh.setDefaultSectionSize(30)
        vh.setFixedWidth(40)
        vh.setDefaultAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        T = QtWidgets.QAbstractItemView.EditTrigger
        self.setEditTriggers(T.DoubleClicked | T.EditKeyPressed | T.AnyKeyPressed
                             | T.SelectedClicked)
        self._errors: set[int] = set()
        self._quiet = 0                 # >0 이면 프로그램이 채우는 중 — edited 안 냄
        self.set_cells([])
        self.itemChanged.connect(self._on_item_changed)

    # ── 채우기·읽기 ──
    def _new_item(self, text=""):
        it = QtWidgets.QTableWidgetItem(str(text or ""))
        it.setFlags(it.flags() | QtCore.Qt.ItemFlag.ItemIsEditable)
        return it

    def _fill_row(self, r, vals=()):
        for c in range(self.columnCount()):
            v = vals[c] if c < len(vals) else ""
            self.setItem(r, c, self._new_item(v))

    def _ensure_rows(self, n):
        while self.rowCount() < n:
            self.insertRow(self.rowCount())
            self._fill_row(self.rowCount() - 1)

    def _relabel(self):
        from daangn_ext.rule_grid import grid_row_label
        self.setVerticalHeaderLabels([grid_row_label(i) for i in range(self.rowCount())])

    def _row_vals(self, r):
        return [(self.item(r, c).text() if self.item(r, c) else "").strip()
                for c in range(self.columnCount())]

    def _pad_tail(self):
        """끝에 빈 줄 하나는 늘 있다 — 다음 조건을 적을 자리다."""
        last = self.rowCount() - 1
        if last < 0 or any(self._row_vals(last)):
            self._ensure_rows(self.rowCount() + 1)
        self._ensure_rows(self.MIN_ROWS)
        self._relabel()

    def set_cells(self, cells):
        self._quiet += 1
        try:
            self.setRowCount(0)
            for row in cells or []:
                self.insertRow(self.rowCount())
                self._fill_row(self.rowCount() - 1, ["" if v is None else str(v) for v in row])
            self._pad_tail()
            self.mark_errors([])
        finally:
            self._quiet -= 1

    def cells(self):
        """표 → 문자열 셀. 끝의 빈 줄은 뺀다(가운데 빈 줄은 행 번호를 지키려 남긴다)."""
        rows = [self._row_vals(r) for r in range(self.rowCount())]
        while rows and not any(rows[-1]):
            rows.pop()
        return rows

    # ── 편집 ──
    def _on_item_changed(self, item):
        if self._quiet:
            return
        if item.row() == self.rowCount() - 1 and any(self._row_vals(item.row())):
            self._quiet += 1
            try:
                self._pad_tail()
            finally:
                self._quiet -= 1
        self.edited.emit()

    def paste_text(self, text):
        """현재 칸에서 시작해 탭·줄바꿈으로 나눈 값을 채운다 — 엑셀 복사 그대로."""
        from daangn_ext.rule_grid import paste_cells
        rows = paste_cells(text)
        if not rows:
            return
        r0 = max(self.currentRow(), 0)
        c0 = max(self.currentColumn(), 0)
        self._quiet += 1
        try:
            self._ensure_rows(r0 + len(rows))
            for i, vals in enumerate(rows):
                for j, v in enumerate(vals):
                    c = c0 + j
                    if c < self.columnCount():
                        self.item(r0 + i, c).setText(v)
            self._pad_tail()
        finally:
            self._quiet -= 1
        self.edited.emit()

    def copy_selected(self):
        idx = sorted((i.row(), i.column()) for i in self.selectedIndexes())
        if not idx:
            return
        lines, cur, cur_r = [], [], idx[0][0]
        for r, c in idx:
            if r != cur_r:
                lines.append("\t".join(cur)); cur, cur_r = [], r
            cur.append(self.item(r, c).text() if self.item(r, c) else "")
        lines.append("\t".join(cur))
        QtWidgets.QApplication.clipboard().setText("\n".join(lines))

    def clear_selected(self):
        idx = self.selectedIndexes()
        if not idx:
            return
        self._quiet += 1
        try:
            for i in idx:
                it = self.item(i.row(), i.column())
                if it:
                    it.setText("")
        finally:
            self._quiet -= 1
        self.edited.emit()

    def add_row(self):
        r = self.currentRow() + 1 if self.currentRow() >= 0 else self.rowCount()
        self._quiet += 1
        try:
            self.insertRow(r)
            self._fill_row(r)
            self._relabel()
        finally:
            self._quiet -= 1
        self.setCurrentCell(r, 0)

    def remove_selected_rows(self):
        rows = {i.row() for i in self.selectedIndexes()}
        if self.currentRow() >= 0:
            rows.add(self.currentRow())
        if not rows:
            return
        had = any(any(self._row_vals(r)) for r in rows)
        self._quiet += 1
        try:
            for r in sorted(rows, reverse=True):
                self.removeRow(r)
            self._pad_tail()
        finally:
            self._quiet -= 1
        if had:
            self.edited.emit()

    def keyPressEvent(self, e):
        K = QtGui.QKeySequence.StandardKey
        editing = self.state() == QtWidgets.QAbstractItemView.State.EditingState
        if e.matches(K.Paste):
            self.paste_text(QtWidgets.QApplication.clipboard().text()); return
        if e.matches(K.Copy):
            self.copy_selected(); return
        if not editing and e.key() in (QtCore.Qt.Key.Key_Delete, QtCore.Qt.Key.Key_Backspace):
            self.clear_selected(); return
        super().keyPressEvent(e)

    # ── 오류 표시 ──
    def mark_errors(self, errors):
        """파서 메시지의 "N행" 을 표의 같은 번호 줄에 붉게 칠한다."""
        rows = set()
        for msg in errors or []:
            m = re.match(r"\s*(?:\[[^\]]*\]\s*)?(\d+)행", str(msg))
            if m:
                rows.add(int(m.group(1)) - 2)
        self._quiet += 1
        try:
            for r in range(self.rowCount()):
                brush = (QtGui.QBrush(QtGui.QColor(self.ERR_ROW_BG)) if r in rows
                         else QtGui.QBrush())
                for c in range(self.columnCount()):
                    it = self.item(r, c)
                    if it:
                        it.setBackground(brush)
        finally:
            self._quiet -= 1
        self._errors = {r for r in rows if 0 <= r < self.rowCount()}

    def error_rows(self):
        return set(self._errors)


# 지역 트리 — 행에는 짧은 이름(부암동)만 보이고, 검색·표시용 전체 경로
# (서울특별시 종로구 부암동)는 이 role 에 둔다. 예전엔 모든 행이 전체 경로를
# 되풀이해 트리가 한 화면에 열 줄도 못 담았다.
AREA_FULL_ROLE = QtCore.Qt.ItemDataRole.UserRole + 1


def area_full_name(item):
    """트리 항목의 전체 경로 — 없으면 보이는 글자."""
    return item.data(0, AREA_FULL_ROLE) or item.text(0)


class TossTreeDelegate(QtWidgets.QStyledItemDelegate):
    """토스 결 지역 트리 — 행 하나를 통째로 그린다: 들여쓰기 · 셰브런 ·
    둥근 체크 · 이름. Qt 기본 트리는 브랜치 화살표와 네모 체크박스를 이미지
    파일 없이는 못 바꾸고 플랫폼마다 다르게 그린다. 델리게이트가 직접 그리면
    파일도 플랫폼 차이도 없다. 클릭도 여기서 받는다 — 부모 행은 어디를
    눌러도 접히고 펴지며, 체크는 상자를 눌러야 바뀐다. 리프는 어디를 눌러도
    체크가 바뀐다."""
    ROW_H = 40
    INDENT = 24
    BOX = 20
    BLUE = "#3182F6"
    HOVER = "#F2F4F6"
    BORDER = "#D1D6DB"
    CHEVRON = "#B0B8C1"
    TEXT = "#191F28"
    TEXT_LEAF = "#333D4B"

    def __init__(self, tree):
        super().__init__(tree)
        self.tree = tree

    @staticmethod
    def _depth(index):
        d = 0
        p = index.parent()
        while p.isValid():
            d += 1; p = p.parent()
        return d

    def _zones(self, option, index):
        """행 안의 세 구역 → (셰브런 rect, 체크 rect, 글자 x)."""
        r = option.rect
        x = r.left() + 10 + self._depth(index) * self.INDENT
        cy = r.center().y()
        chev = QtCore.QRect(x, cy - 10, 20, 20)
        x += 22
        box = QtCore.QRect(x, cy - self.BOX // 2, self.BOX, self.BOX)
        return chev, box, x + self.BOX + 10

    def sizeHint(self, option, index):
        s = super().sizeHint(option, index)
        s.setHeight(self.ROW_H)
        return s

    def paint(self, painter, option, index):
        painter.save()
        painter.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing, True)
        r = option.rect
        if option.state & QtWidgets.QStyle.StateFlag.State_MouseOver:
            painter.setPen(QtCore.Qt.PenStyle.NoPen)
            painter.setBrush(QtGui.QColor(self.HOVER))
            painter.drawRoundedRect(QtCore.QRectF(r).adjusted(4, 2, -4, -2), 10, 10)
        chev, box, tx = self._zones(option, index)
        has_kids = index.model().hasChildren(index)
        if has_kids:
            pen = QtGui.QPen(QtGui.QColor(self.CHEVRON), 1.8)
            pen.setCapStyle(QtCore.Qt.PenCapStyle.RoundCap)
            pen.setJoinStyle(QtCore.Qt.PenJoinStyle.RoundJoin)
            painter.setPen(pen); painter.setBrush(QtCore.Qt.BrushStyle.NoBrush)
            c = chev.center()
            path = QtGui.QPainterPath()
            if option.state & QtWidgets.QStyle.StateFlag.State_Open:
                path.moveTo(c.x() - 4, c.y() - 2); path.lineTo(c.x(), c.y() + 2); path.lineTo(c.x() + 4, c.y() - 2)
            else:
                path.moveTo(c.x() - 2, c.y() - 4); path.lineTo(c.x() + 2, c.y()); path.lineTo(c.x() - 2, c.y() + 4)
            painter.drawPath(path)
        state = index.data(QtCore.Qt.ItemDataRole.CheckStateRole)
        try:
            state = QtCore.Qt.CheckState(state)
        except Exception:
            state = QtCore.Qt.CheckState.Unchecked
        bf = QtCore.QRectF(box)
        if state == QtCore.Qt.CheckState.Unchecked:
            painter.setPen(QtGui.QPen(QtGui.QColor(self.BORDER), 1.5))
            painter.setBrush(QtGui.QColor("#FFFFFF"))
            painter.drawRoundedRect(bf.adjusted(0.75, 0.75, -0.75, -0.75), 6, 6)
        else:
            painter.setPen(QtCore.Qt.PenStyle.NoPen)
            painter.setBrush(QtGui.QColor(self.BLUE))
            painter.drawRoundedRect(bf, 6, 6)
            pen = QtGui.QPen(QtGui.QColor("#FFFFFF"), 2.2)
            pen.setCapStyle(QtCore.Qt.PenCapStyle.RoundCap)
            pen.setJoinStyle(QtCore.Qt.PenJoinStyle.RoundJoin)
            painter.setPen(pen); painter.setBrush(QtCore.Qt.BrushStyle.NoBrush)
            x0, y0 = bf.left(), bf.top()
            path = QtGui.QPainterPath()
            if state == QtCore.Qt.CheckState.Checked:
                path.moveTo(x0 + 5, y0 + 10.5); path.lineTo(x0 + 8.5, y0 + 14); path.lineTo(x0 + 15, y0 + 6.5)
            else:                                   # 부분 선택 = 파란 상자에 '−'
                path.moveTo(x0 + 5.5, y0 + 10); path.lineTo(x0 + 14.5, y0 + 10)
            painter.drawPath(path)
        depth = self._depth(index)
        font = QtGui.QFont(option.font)
        if has_kids:
            font.setPointSizeF(15 if depth == 0 else 14)
            font.setWeight(QtGui.QFont.Weight.Bold if depth == 0 else QtGui.QFont.Weight.DemiBold)
            painter.setPen(QtGui.QColor(self.TEXT))
        else:
            font.setPointSizeF(14); font.setWeight(QtGui.QFont.Weight.Normal)
            painter.setPen(QtGui.QColor(self.TEXT_LEAF))
        painter.setFont(font)
        painter.drawText(QtCore.QRect(tx, r.top(), r.right() - tx - 8, r.height()),
                         int(QtCore.Qt.AlignmentFlag.AlignVCenter | QtCore.Qt.AlignmentFlag.AlignLeft),
                         index.data(QtCore.Qt.ItemDataRole.DisplayRole) or "")
        painter.restore()

    def editorEvent(self, event, model, option, index):
        if event.type() == QtCore.QEvent.Type.MouseButtonDblClick:
            return True                             # 두 번 누름 = 두 번 토글. 트리 기본 펼침과 겹치지 않게 막는다
        if (event.type() != QtCore.QEvent.Type.MouseButtonRelease
                or event.button() != QtCore.Qt.MouseButton.LeftButton):
            return False
        if not (index.flags() & QtCore.Qt.ItemFlag.ItemIsUserCheckable) or not (
                index.flags() & QtCore.Qt.ItemFlag.ItemIsEnabled):
            return False
        pos = event.position().toPoint()
        chev, box, _ = self._zones(option, index)
        has_kids = model.hasChildren(index)
        hit_box = box.adjusted(-6, -6, 6, 6).contains(pos)
        if hit_box or not has_kids:
            cur = index.data(QtCore.Qt.ItemDataRole.CheckStateRole)
            new = (QtCore.Qt.CheckState.Unchecked
                   if QtCore.Qt.CheckState(cur) == QtCore.Qt.CheckState.Checked
                   else QtCore.Qt.CheckState.Checked)
            model.setData(index, new, QtCore.Qt.ItemDataRole.CheckStateRole)
            return True
        self.tree.setExpanded(index, not self.tree.isExpanded(index))
        return True


def toss_tree(tree):
    """QTreeWidget 을 토스 결로 — 델리게이트가 그리므로 Qt 의 브랜치·들여쓰기·
    선택 강조는 전부 끈다. 트리 자체는 그대로라 항목을 만드는 코드는 안 바뀐다."""
    tree.setObjectName("tossTree")
    tree.setHeaderHidden(True)
    tree.setRootIsDecorated(False)
    tree.setIndentation(0)
    tree.setExpandsOnDoubleClick(False)
    tree.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.NoSelection)
    tree.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)
    tree.setFrameShape(QtWidgets.QFrame.Shape.NoFrame)
    tree.setUniformRowHeights(True)
    tree.setMouseTracking(True)
    tree.setItemDelegate(TossTreeDelegate(tree))
    return tree


class FlowLayout(QtWidgets.QLayout):
    """줄바꿈되는 가로 배치 — 선택 칩이 창 폭에 맞춰 흐른다. Qt 에 없다."""

    def __init__(self, parent=None, hspace=6, vspace=6):
        super().__init__(parent)
        self._items = []
        self._h, self._v = hspace, vspace
        self.setContentsMargins(0, 0, 0, 0)

    def addItem(self, item): self._items.append(item)
    def count(self): return len(self._items)
    def itemAt(self, i): return self._items[i] if 0 <= i < len(self._items) else None
    def takeAt(self, i): return self._items.pop(i) if 0 <= i < len(self._items) else None
    def expandingDirections(self): return QtCore.Qt.Orientation(0)
    def hasHeightForWidth(self): return True
    def heightForWidth(self, w): return self._lay(QtCore.QRect(0, 0, w, 0), True)
    def sizeHint(self): return self.minimumSize()

    def minimumSize(self):
        s = QtCore.QSize()
        for it in self._items:
            s = s.expandedTo(it.minimumSize())
        return s

    def setGeometry(self, rect):
        super().setGeometry(rect)
        self._lay(rect, False)

    def _lay(self, rect, test):
        x, y, line_h = rect.x(), rect.y(), 0
        for it in self._items:
            s = it.sizeHint()
            if x + s.width() > rect.right() + 1 and line_h > 0:
                x = rect.x(); y += line_h + self._v; line_h = 0
            if not test:
                it.setGeometry(QtCore.QRect(QtCore.QPoint(x, y), s))
            x += s.width() + self._h
            line_h = max(line_h, s.height())
        return y + line_h - rect.y()

# 등록 표의 열. 인덱스를 손으로 세면 열이 하나 끼는 순간 조용히 어긋난다
# (실제로 삭제가 id 대신 다른 열을 읽을 뻔했다). 이름으로만 참조한다.
ALERT_COLS = ["키워드", "상태", "수집 방식", "가격범위", "제외", "추가", "끌올 기간", "id"]
ALERT_COL_KEYWORD = ALERT_COLS.index("키워드")
ALERT_COL_STATUS = ALERT_COLS.index("상태")
ALERT_COL_ROUTE = ALERT_COLS.index("수집 방식")
ALERT_COL_DAYS = ALERT_COLS.index("끌올 기간")
ALERT_COL_ID = ALERT_COLS.index("id")

# 끌올일수는 앱 알림 등록 body 에 넣을 필드가 없고 filter_by_conditions 도
# 쓰지 않는다. 검색 스윕에서만 실제로 걸린다. 표에 값만 보이고 이 사실을 안
# 적으면 앱 경로 키워드에도 걸려 있다고 믿게 된다.
DAYS_APP_TIP = ("끌올일수는 앱 알림 경로에서는 적용되지 않습니다"
                " — 검색 스윕으로 배정된 키워드에만 걸립니다")

# 등록 표의 존재 이유는 '엑셀에 넣은 조건'과 '당근 서버에 실제 걸린 키워드'의
# 차이를 보여 주는 것이다. 이 열이 없으면 표는 조건표의 복사본으로 읽히고,
# 앱 알림이 안 올 때 서버 등록이 빠졌는지 확인할 길이 없다.
REG_SERVER, REG_SWEEP, REG_MISSING, REG_UNKNOWN = "server", "sweep", "missing", "unknown"
REG_STATUS_NAMES = {REG_SERVER: "서버 등록", REG_SWEEP: "스윕 대기",
                    REG_MISSING: "미등록", REG_UNKNOWN: "확인 불가"}
REG_STATUS_TIPS = {
    REG_SERVER: "당근 앱 알림 서버에 실제 등록돼 있습니다",
    REG_SWEEP: "앱 슬롯이 차서 검색 스윕이 대신 훑습니다 (서버 등록 아님)",
    REG_MISSING: ("조건표에는 있지만 당근 서버에 아직 안 올라갔습니다."
                  " 감시를 시작하면 등록을 시도합니다 — 계속 미등록이면"
                  " 계정 토큰·앱 슬롯(30개)을 확인하세요"),
    REG_UNKNOWN: "서버 목록을 못 읽어 등록 여부를 알 수 없습니다 (토큰 확인)",
}
REG_MISSING_BG = "#FCEBE8"   # 계정표 '점검필요' 와 같은 붉은 바탕
REG_MISSING_FG = "#B3261E"
REG_UNKNOWN_FG = "#8B8474"


def alert_row_cells(keyword, route, price, exclude, uid, status=None):
    """등록 표 한 줄 → (셀 값 8개, {열 인덱스: 툴팁}). 위젯을 모르는 순수 함수.

    status 는 REG_* 중 하나(또는 None) — 이 줄이 서버에 실제 등록됐는지,
    스윕이 대신 훑는지, 조건만 있고 서버엔 없는지.

    추가키워드·끌올일수는 엑셀 조건으로만 들어오고 라우터가 `cond` 에 넣어
    둔다. 표에 안 그리면 사용자는 자기가 건 조건을 확인할 방법이 없다.

    추가키워드는 앱 경로에서 filter_by_conditions 가 건다. 끌올일수는 스윕
    경로에서만 걸리므로, 앱 경로에 값이 있으면 툴팁으로 그 사실을 붙인다 —
    엑셀 로드 때 로그로 한 번 지나가고 마는 경고라, 표에서 다시 볼 수 없으면
    걸려 있다고 믿게 된다."""
    cond = (route or {}).get("cond") or {}
    extra = ",".join(cond.get("extra") or []) or "-"
    days_n = cond.get("days")
    days = f"{days_n}일" if days_n else "-"
    vals = [str(keyword or ""), REG_STATUS_NAMES.get(status, "-"),
            ROUTE_NAMES.get((route or {}).get("route"), "-"),
            str(price or ""), str(exclude or ""), extra, days, str(uid or "")]
    tips = {}
    if status in REG_STATUS_TIPS:
        tips[ALERT_COL_STATUS] = REG_STATUS_TIPS[status]
    if route:
        tips[ALERT_COL_ROUTE] = str(route.get("reason") or "")
    if days_n and (route or {}).get("route") == "app":
        tips[ALERT_COL_DAYS] = DAYS_APP_TIP
    return vals, tips


# 어느 경로로 들어온 매물인지. 저장소에는 기록돼 있었는데 화면에 없어서
# 클라는 스윕을 켜도 무엇이 늘었는지 알 수 없었다.
SOURCE_NAMES = {"app": "앱 알림", "sweep": "지역 훑기", "feed": "동 피드"}

LISTING_COLS = ["상태", "수집", "키워드", "제목", "지역", "현재가", "Δ최초가",
                "마지막변동", "최초감지"]
LISTING_COL_TITLE = LISTING_COLS.index("제목")


def listing_display_rows(rows, now, state_filter="all"):
    """watch 행 목록 → 매물 표에 그릴 형태. 최초 감지 내림차순.

    위젯을 모르는 순수 함수로 두어 GUI 없이 검증한다."""
    from daangn_ext import article_watch as _aw
    out = []
    for r in rows or []:
        # 중복 판정만을 위한 묘비(백필이 match_seen.json 에서 옮긴 행)는 보여줄
        # 게 없다. 걸러내지 않으면 표에 빈 줄로 뜬다. 판정 기준은 '제목이 비었다'
        # 가 아니라 출처다 — 제목 없이 들어온 실제 매물이 종료되는 순간 표에서
        # 조용히 사라지면 사용자는 이유를 알 수 없다.
        if (r.get("source") or "") == _aw.SOURCE_MATCH_SEEN:
            continue
        state = _aw.state_for(r, now)
        if state_filter in ("new", "down", "ended") and state != state_filter:
            continue
        out.append({
            "id": str(r.get("id") or ""),
            "state": state,
            "icon": STATE_ICONS.get(state, state),
            "keyword": r.get("keyword") or "",
            "source": SOURCE_NAMES.get(r.get("source") or "app", "앱 알림"),
            "title": (r.get("title") or "")[:60],
            "region": r.get("region") or "",
            "price": r.get("price") or 0,
            "delta_text": _delta_text(r.get("first_price"), r.get("price")),
            "last_change_text": _ts_text(r.get("last_change")),
            "first_seen_text": _ts_text(r.get("first_seen")),
            "url": r.get("url") or "",
        })
    order = {str(r.get("id") or ""): int(r.get("first_seen") or 0)
             for r in rows or []}
    out.sort(key=lambda x: order.get(x["id"], 0), reverse=True)
    return out


SLOT_CAP_KEY = "keyword_slot_cap"


def sweep_found_to_match(payload, keyword):
    """AutoMonitor.found 페이로드 → add_from_matches 가 받는 형태.

    두 소스를 같은 문으로 들여보내야 매물 표가 하나로 유지된다."""
    aid = payload.get("id") if payload else None
    if not aid:
        return None
    from daangn_ext.article_watch import parse_iso
    return {"article_id": str(aid),
            "title": payload.get("title") or "",
            "price": payload.get("price") or 0,
            "region": payload.get("region") or "",
            "url": payload.get("url") or "",
            "time": parse_iso(payload.get("boostedAt")),
            "keyword": keyword or ""}


def sweep_keyword_for(payload, keywords):
    """스윕이 찾은 매물에 붙일 키워드 — 제목에 들어 있는 대기열 키워드 우선.

    GUI(_on_sweep_found)와 헤드리스가 같은 규칙을 써야 매물 표의 키워드 열이
    런타임마다 갈라지지 않는다."""
    kws = [k for k in (keywords or []) if k]
    title = (payload or {}).get("title") or ""
    return next((k for k in kws if k in title), kws[0] if kws else "")


def sweep_conditions(entries, extra=None, exclude=None,
                     min_price=None, max_price=None, days=None):
    """스윕 대기열 엔트리 → SweepEngine cfg["conditions"].

    가격·제외·추가키워드·끌올일수 모두 등록 당시 사용자가 넣은 값(엔트리)을
    우선한다 — 넘어온 값은 엔트리에 없을 때의 기본값이다. 기본값의 **출처**만
    런타임마다 다르고 (GUI=고급 패널 위젯, 헤드리스=alert_settings.json)
    조립 규칙은 하나다.

    extra·days 만 엔트리를 무시하고 전역값으로 덮던 시절이 있었다. 엑셀에
    행마다 적은 추가키워드·끌올일수가 스윕에서 조용히 사라져, 사용자는 자기가
    건 조건이 걸린 줄 알았다.

    '비었다'와 '안 적었다'는 구별하지 않는다 — 넷 다 비어 있으면 기본값으로
    간다. 큐가 빈 값을 아예 안 싣기 때문에(SweepQueue._norm) 구별할 방법도
    없고, exclude 가 원래 쓰던 규칙이기도 하다."""
    out = []
    for e in entries or []:
        out.append({
            "keyword": e["keyword"],
            "extra": list(e.get("extra") or []) or list(extra or []),
            "exclude": list(e.get("exclude") or []) or list(exclude or []),
            "min": e.get("min") if e.get("min") is not None else min_price,
            "max": e.get("max") if e.get("max") is not None else max_price,
            "days": e.get("days") or days,
        })
    return out


# 스윕 지역 설정 키 — GUI 위젯 값이 이 이름으로 alert_settings.json 에 저장되고
# 헤드리스가 같은 이름으로 읽는다. 이름이 갈리면 서버는 GUI 설정을 못 본다.
SWEEP_NATIONWIDE_KEY = "sweep_nationwide"


FEED_DEFAULTS = {
    "feed_enabled": True,
    "feed_categories": [31, 14],          # 여성잡화 · 남성패션/잡화
    "feed_proxies": [],                   # 비면 proxies.txt
    "feed_rps": 1.0,                      # 레인(프록시)당 초당 요청
    "feed_rest_min": 2,                   # 사이클 휴식(분)
    "sweep_app_enabled": False,           # 앱 키워드 스윕(보완층) — 버릴 계정만
    "sweep_regions_app": ["역삼동-6035"],
}


def sweep_app_enabled(settings) -> bool:
    """앱 키워드 스윕(계정 토큰으로 검색) 스위치. 기본 꺼짐.

    발굴 주경로가 계정 없는 웹 피드로 옮겨 가면서(2026-09-03 스펙) 이 경로는
    타지역 택배 매물 보완층이 됐다. 옛 키 sweep_mirror_app 은 읽어만 준다."""
    s = settings or {}
    if s.get("sweep_app_enabled") is not None:
        return bool(s["sweep_app_enabled"])
    if s.get("sweep_mirror_app") is not None:
        return bool(s["sweep_mirror_app"])
    return bool(FEED_DEFAULTS["sweep_app_enabled"])


def sweep_mirror_enabled(settings, n_rules) -> bool:
    """호환 이름 — 답은 sweep_app_enabled 하나다."""
    return sweep_app_enabled(settings)

# 지역을 아무도 고르지 않았을 때 훑을 기본 범위: 서울 + 경기(동 1,857곳).
# 명품 물량이 여기 몰려 있고, 실측 수렴 한계(지역 × 조건 ≈ 17,900) 안에서
# 브랜드 9개까지 소화한다. 그 위는 sweep_scope_for 가 구·시 단위로 낮춘다.
DEFAULT_SWEEP_SIDO = ("서울특별시", "경기도")

# 한 사이클이 주기 안에 끝나는 상한. 2026-09-01 실측: 최신순 한 페이지가
# 23분을 덮고 토큰·IP 하나로 13 req/s 가 나온다 → 1380초 × 13 ≈ 17,900.
# 넘기면 사이클이 주기보다 길어져 뒤로 밀린다(클라에게는 '알림이 늦다'로만 보인다).
# 한 사이클 예산 = 페이지폭(sweep_engine.PAGE_SPAN_MIN) × 초당 요청. 지역 × 조건이 이걸
# 넘으면 사이클이 페이지폭보다 길어져 워터마크 사이 매물을 놓치기 시작한다.
# 페이지폭 상수는 엔진의 수렴 판정(sweep_capacity)과 한 곳을 쓴다 — 재측정하면 거기만.
SWEEP_REQ_PER_LANE = 1.6        # 실측 동시 8 = 13 req/s → 레인당 1.6


def sweep_budget(lanes=None):
    """레인 수에 따른 한 사이클 예산. 0/None = 앱API 기본 레인(APP_API_LANES).

    상수로 두면 레인 1개(웹크롤 경로·수동 지정)에서도 8레인 예산으로 판정해
    동 단위를 고른 채 사이클이 8배 길어진다 — 증상은 '알림이 늦다' 뿐이다.
    호출측이 엔진과 같은 규칙으로 실제 레인 수를 넘겨야 한다(sweep_lanes_effective)."""
    n = max(1, int(lanes or 0) or APP_API_LANES)
    return int(sweep_capacity(SWEEP_REQ_PER_LANE * n))


def sweep_lanes_effective(lanes, has_token, n_proxies):
    """엔진이 실제로 돌릴 레인 수 — SweepEngine._plan_lanes 와 같은 규칙.

    토큰이 있으면 앱API 경로(IP 공유, 기본·상한 APP_API_LANES). 없으면 웹크롤 경로라
    프록시 수 ÷ MIN_IP_PER_LANE 로 묶인다. 예산은 이 수로 잡아야 엔진과 어긋나지 않는다."""
    want = int(lanes or 0)
    if has_token:
        return max(1, min(want or APP_API_LANES, APP_API_LANES))
    n_proxy = int(n_proxies or 0)
    if n_proxy <= 1:
        return 1
    auto = max(1, n_proxy // MIN_IP_PER_LANE)
    n = want if want > 0 else auto
    return max(1, min(n, n_proxy // MIN_IP_PER_LANE or 1))


SWEEP_BUDGET = sweep_budget(APP_API_LANES)


def default_sweep_regions(out_json="./OUT.json"):
    """지역 설정이 아예 없을 때 훑을 기본 동 목록.

    전국은 동 6537곳이다. 스윕 한 사이클은 (지역 × 키워드) 요청이라 키워드가
    하나뿐이어도 6537 요청 — 계정 하루 상한(daily_cap=300)의 21배다. 그래서
    '아무 설정도 없는 새 설치'가 전국을 고르면 첫 사이클에 함대 예산이 통째로
    마르고, 남은 하루는 엔진이 '전 계정 캡/쿨다운 도달' no-op 으로만 돈다.
    운영자에게 레버가 없는 상태에서 그건 기본값이 될 수 없다.

    그래서 기본값은 넓히기 쉬운 쪽으로 좁게 잡는다. 넓히는 건 지역 트리에서
    체크하거나 '전국 훑기'를 켜는 두 번의 의식적인 동작이다.

    OUT.json 이 없거나 이름이 달라져 하나도 못 고르면 빈 목록을 돌려준다 —
    그 경우에도 전국으로 떨어뜨리지 않는다. 커버리지 0 은 로그로 보이지만
    예산 고갈은 안 보인다."""
    import json as _json
    try:
        with open(out_json, encoding="utf-8") as f:
            data = _json.load(f)
    except Exception:
        return []
    out, seen = [], set()
    for block in data or []:
        if block.get("name1") not in DEFAULT_SWEEP_SIDO:
            continue
        for loc in block.get("locations") or []:
            code = f"{loc.get('name')}-{loc.get('id')}"
            if code in seen:
                continue
            seen.add(code)
            out.append(code)
    return out


def default_sweep_regions_coarse(out_json="./OUT.json"):
    """같은 범위를 구·시 단위로 줄인 목록(서울 25 + 경기 44 ≈ 69곳).

    반경 검색이 인접 동을 함께 물어오므로 구 단위여도 커버는 남는다. 조건이
    늘어 동 단위가 예산을 넘길 때 여기로 내려온다 — 조용히 뒤처지는 것보다
    성기게라도 한 바퀴를 도는 편이 낫다."""
    import json as _json
    try:
        with open(out_json, encoding="utf-8") as f:
            data = _json.load(f)
    except Exception:
        return []
    out, seen = [], set()
    for block in data or []:
        if block.get("name1") not in DEFAULT_SWEEP_SIDO:
            continue
        locs = block.get("locations") or []
        if not locs:
            continue
        loc = locs[0]                       # 그 구·시의 대표 지점 하나
        code = f"{loc.get('name')}-{loc.get('id')}"
        if code not in seen:
            seen.add(code)
            out.append(code)
    return out


def sweep_fit_budget(regions, n_conditions, coarse, log=None, budget=None):
    """지역 × 조건이 예산을 넘으면 성긴 목록으로 내린다 → (지역, 낮췄는지).

    넘긴 채로 두면 사이클이 주기보다 길어져 뒤로 밀리기만 한다. 그때 화면에
    보이는 증상은 '알림이 늦다' 뿐이라 원인을 못 찾는다.
    budget 미지정 = 기본 레인 예산(SWEEP_BUDGET)."""
    n = max(1, int(n_conditions or 1))
    budget = int(budget or SWEEP_BUDGET)
    if len(regions) * n <= budget or not coarse:
        return regions, False
    if log:
        log(f"[지역 훑기] 조건 {n}개 × {len(regions)}곳 = {len(regions) * n:,} —"
            f" 한 사이클 예산({budget:,})을 넘어 구·시 단위 {len(coarse)}곳으로"
            f" 낮춰 돕니다 ({len(coarse) * n:,}건)")
    return coarse, True


def sweep_scope_for(regions, nationwide, out_json="./OUT.json", log=None,
                    n_conditions=1, lanes=None):
    """지역 설정 → cfg 의 scope/regions. GUI(_auto_cfg_base)·헤드리스 공용.

    셋 중 하나다:
      고른 지역이 있다      → 그 지역만.
      '전국 훑기'가 켜졌다  → 전국(동 6537곳). **명시적으로 켜야만** 여기 온다.
      둘 다 아니다          → default_sweep_regions().

    옛 규칙('지역 미선택 = 전국')을 버린 이유는 default_sweep_regions 참고.
    두 런타임이 같은 규칙을 써야 GUI 에서 본 범위가 서버에서 그대로 돈다."""
    regions = [r for r in (regions or []) if r]
    if regions:
        return {"scope": "regions", "regions": regions}
    if nationwide:
        return {"scope": "nationwide"}
    dflt = default_sweep_regions(out_json)
    # 조건이 늘면 같은 지역 수라도 사이클이 길어진다. 예산을 넘으면 성긴
    # 목록으로 내린다 — 조용히 뒤처지는 것보다 한 바퀴를 도는 편이 낫다.
    dflt, lowered = sweep_fit_budget(
        dflt, n_conditions, default_sweep_regions_coarse(out_json), log=log,
        budget=sweep_budget(lanes))
    if log and not lowered:
        log(f"[지역 훑기] 기본 범위 서울·경기 {len(dflt)}곳")
    return {"scope": "regions", "regions": dflt}


# 죽은 스윕 되살리기 상한. 엔진이 뜨자마자 죽는 상황(토큰 없음·프록시 전멸)에서
# 틱마다 무한히 다시 띄우면 그때마다 실계정 요청을 태운다 — 몇 번 해보고 포기한다.
SWEEP_REVIVE_MAX = 5

# 서버 목록 씨딩 시도 상한. 첫 실행에만 필요한 조회라 실패가 이어져도 매 틱
# 요청을 새로 쓰지 않는다.
SEED_ATTEMPT_MAX = 3


def sweep_resync_action(want, have, running):
    """돌고 있는 스윕이 낡았는지 판정 — GUI·헤드리스 공용 순수 함수.

    want=대기열 키워드 집합, have=지금 도는 스윕이 떠 있는 집합(None=안 떠 있음),
    running=스레드가 살아 있는지.

    반환: "" 무동작 / "start" 시작 / "revive" 죽은 스레드 되살리기 /
          "restart" 키워드가 바뀌어 갈아끼우기(want 가 비면 정지만)."""
    want = set(want or [])
    if have is None:
        # 아직 안 떴거나 정지 요청 뒤 — 큐가 찼고 스레드가 빠졌으면 띄운다.
        return "start" if (want and not running) else ""
    if want == have and (running or not want):
        return ""
    if want == have:
        # 엔진 run() 은 루프 전체를 except 로 감싸므로 치명오류가 나면 스레드만
        # 조용히 빠진다. have 는 남아 있어 want == have 가 계속 성립하고,
        # 스윕은 세션 내내 죽은 채로 방치된다.
        return "revive"
    return "restart"


# 스윕 스레드 → 폴링 스레드 인계 큐의 상한. 폴링이 멈춰도 메모리가 무한히
# 자라지 않게 한다. 넘치면 버리고 센다(로그로 보인다).
SWEEP_FIND_QUEUE_MAX = 2000


def drain_sweep_finds(q, tracker, keywords_fn, log, limit=SWEEP_FIND_QUEUE_MAX):
    """스윕이 찾은 payload 를 **폴링 스레드에서** 워치리스트에 넣는다.

    SweepEngine 의 on_found 는 스윕 스레드에서 동기로 불린다(notify →
    _dedup_notify). 거기서 WatchStore 를 바로 만지면 폴링 루프와 같은 sqlite
    커넥션(check_same_thread=False, 락 없음)을 두 스레드가 나눠 쓰게 된다 —
    커넥션이 하나뿐이라 한쪽의 commit() 이 다른 쪽의 읽기-수정-쓰기를 중간에
    커밋하고, enforce_cap 의 active_count→oldest_active→mark 사이로 insert 가
    끼어든다.

    GUI 에는 이 위험이 없다: AutoMonitor.found 가 pyqtSignal 이라 큐 연결로
    GUI 스레드에 배달된다. 헤드리스도 같은 모양으로 맞춘 것이 이 함수다 —
    스윕 스레드는 큐에 넣기만 하고(put_nowait 는 절대 안 막힌다), sqlite 는
    폴링 스레드 하나만 만진다. 락 대신 큐를 고른 이유: 락은 '앞으로 추가될
    mutator 마다 잊지 말고 걸기'에 기대지만, 큐는 sqlite 를 단일 스레드로
    묶어 구조로 보장한다. 그리고 스윕 스레드가 폴링 루프의 긴 네트워크 스윕에
    한 순간도 붙잡히지 않는다.

    payload 정규화까지 여기서 한다 — sweep_queue 파일 읽기(keywords_fn)도
    스윕 스레드에 남기지 않기 위해서다.

    큐는 스윕·피드 두 스레드가 같이 쓴다(둘 다 sqlite 를 직접 만지면 안
    되는 이유는 위와 같다) — 하지만 라벨과 source 는 다르다. 피드는 이미
    payload["keyword"] 에 조건표가 매칭한 라벨을 실어 보내므로(FeedSweep
    만 payload 에 "verdict" 를 싣는다) 대기열 키워드로 다시 찾으면 안 되고,
    표에도 "동 피드"로 갈라 보여야 한다(GUI _on_feed_found 와 같은 규칙)."""
    import queue as _q
    # 추적기가 없으면(watch.db 열기 실패) 큐를 비우지 않는다 — 꺼내고 버리면
    # 저장소가 살아난 뒤에도 그 사이 찾은 매물이 조용히 사라진다.
    if q is None or tracker is None:
        return 0
    payloads = []
    while len(payloads) < int(limit):
        try:
            payloads.append(q.get_nowait())
        except _q.Empty:
            break
    if not payloads:
        return 0
    try:
        kws = list(keywords_fn() or [])
    except Exception:
        kws = []
    feed_norms, sweep_norms = [], []
    for p in payloads:
        if "verdict" in p:
            norm = sweep_found_to_match(p, p.get("keyword") or "")
            if norm:
                feed_norms.append(norm)
        else:
            norm = sweep_found_to_match(p, sweep_keyword_for(p, kws))
            if norm:
                sweep_norms.append(norm)
    if not feed_norms and not sweep_norms:
        return 0
    added = 0
    try:
        if feed_norms:
            added += tracker.add_from_matches(feed_norms, source="feed")
        if sweep_norms:
            added += tracker.add_from_matches(sweep_norms, source="sweep")
    except Exception as e:
        log(f"[검색스윕] 추적 등록 실패: {str(e)[:80]}")
        return 0
    if added:
        log(f"[검색스윕] 신규 매물 추적 {added}건")
    return added


def sweep_revive_step(revives, want_n):
    """되살리기 한 번을 허락할지 판정 — GUI·헤드리스 공용 순수 함수.

    sweep_resync_action 과 같은 이유로 여기 있다: 상한이 한쪽 런타임에만
    있으면 다른 쪽은 틱마다 영원히 엔진을 다시 띄운다(GUI 가 그랬다). 로그
    문구까지 같이 돌려주는 건, 세는 곳과 말하는 곳이 갈라지면 숫자와 문구가
    또 어긋나기 때문이다.

    반환 (허락?, 다음 revives, 로그 문구). 로그가 빈 문자열이면 조용히 넘어간다
    (포기 로그는 상한에 처음 닿았을 때 한 번만 나온다)."""
    n = int(revives or 0)
    if n >= SWEEP_REVIVE_MAX:
        if n == SWEEP_REVIVE_MAX:
            return False, n + 1, (
                f"[검색스윕] {SWEEP_REVIVE_MAX}회 되살렸지만 계속 죽습니다"
                " — 포기합니다(대기열이 바뀌면 다시 시도)")
        return False, n, ""
    n += 1
    return True, n, (f"[검색스윕] 스레드가 죽어 있음 — 키워드 {int(want_n)}개로"
                     f" 되살립니다 ({n}/{SWEEP_REVIVE_MAX})")


def seed_router_from_server(router, list_fn, log, state, allowed=None, prune_fn=None):
    """routes 파일이 비어 있을 때만 서버 등록 목록을 읽어 라우터에 인정시킨다.

    allowed(조건표 브랜드 집합, rule_brand_keys)를 주면 **그 안의 키워드만** 인정한다.
    빈 집합 = 조건표 없음 = 아무것도 인정하지 않는다. 나머지는 prune_fn(목록)으로
    서버에서 지운다 — 옛 일괄등록이 남긴 브랜드가 라우터로 되살아나 알림함 15칸을
    먹는 경로를 여기서 끊는다. None 이면 전부 인정(옛 동작, 테스트용).

    자동 시작(GUI 8초 지연·헤드리스 부팅)은 등록 화면을 거치지 않고 바로 폴링
    틱으로 들어간다. 그 경로가 씨딩을 안 하면 라우터는 함대가 비었다고 믿고
    이미 꽉 찬 서버 한도에 등록을 시도해 전부 실패시킨 뒤 모든 키워드를 스윕으로
    민다 — 무인 첫 실행이 통째로 앱 알림을 잃는다.

    routes 가 이미 차 있으면 목록 조회조차 하지 않으므로 첫 실행 뒤에는 공짜다.
    반대로 토큰이 없어 목록이 계속 비면 씨딩은 영원히 안 끝나므로, 시도 자체를
    state(호출자가 들고 있는 dict)로 SEED_ATTEMPT_MAX 회까지만 허용한다."""
    if router is None:
        return 0
    try:
        if router.routes():
            return 0
    except Exception:
        return 0
    n = int(state.get("n") or 0)
    if n >= SEED_ATTEMPT_MAX:
        return 0
    state["n"] = n + 1
    try:
        data = list_fn() or {}
        kws = [k.get("keyword") for k in (data.get("user_keywords") or [])]
        kws, extras = split_by_rules(kws, allowed)
        if extras:
            log(f"[라우터] 서버 등록 {len(extras)}개는 조건표에 없어 인정하지 않습니다"
                + (": " + ", ".join(extras[:6]) + (" …" if len(extras) > 6 else "")))
            if prune_fn is not None and allowed:
                try:
                    prune_fn(extras)
                except Exception as e:
                    log(f"[라우터] 서버 정리 실패: {str(e)[:80]}")
        seeded = router.seed_from_server(kws) if kws else 0
        if seeded:
            log(f"[라우터] 서버에 이미 등록된 키워드 {seeded}개를 앱 슬롯으로 인식")
        elif state["n"] >= SEED_ATTEMPT_MAX:
            log("[라우터] 서버 목록을 못 읽어 기존 등록 인식을 포기합니다"
                " — 등록 목록 새로고침 시 다시 시도됩니다")
        return seeded
    except Exception as e:
        log(f"[라우터] 기존 등록 인식 실패: {str(e)[:80]}")
        return 0


def read_token_quiet(accounts_path="./accounts.json"):
    """accounts.json 에서 가장 늦게 만료되는 access 를 읽는다 — 수확은 하지 않는다.

    수확을 소유하지 **않는** 쪽(검색 스윕)이 쓰는 provider. 스윕에 필요한 건
    '수확'이 아니라 '신선한 토큰'이고, 토큰을 신선하게 유지하는 일은 폴링 루프가
    이미 주기적으로 한다. 여기서 또 함대를 깨우면 두 스레드가 같은 LDPlayer 를
    동시에 흔들며 accounts.json 을 동시에 쓴다."""
    import json as _json
    try:
        from daangn_ext.token_manager import token_exp
        best = None
        with open(accounts_path, encoding="utf-8") as _f:
            _accs = _json.load(_f)
        for a in _accs:
            acc = a.get("access") or ""
            if acc and (best is None or token_exp(acc) > token_exp(best)):
                best = acc
        return best
    except Exception:
        return None


def harvest_token_quiet(accounts_path="./accounts.json", log=None):
    """함대 수확 후 최신 access. 수확을 **소유한** 쪽만 부른다.

    주기적 소유자는 런타임마다 하나뿐이다: GUI 는 _HarvestThread(20분),
    헤드리스는 폴링 루프(20분). 스윕은 어느 쪽에서도 수확하지 않고
    read_token_quiet 로 파일만 읽는다. 여기를 부르는 나머지 한 곳은
    사용자가 직접 누른 _multi(harvest=True) 뿐이다.

    log 를 넘기면 병합 락 경고까지 보인다 — 안 넘기면 'accounts.json 을
    다른 프로세스가 잡고 있어 수확분이 덮일 수 있다'는 경고가 삼켜진다."""
    try:
        import ld_autoharvest
        ld_autoharvest.harvest_all(accounts_path, nudge=True, log=log)
    except Exception:
        pass
    return read_token_quiet(accounts_path)


def headless_proxies(settings_path="./settings.txt",
                     accounts_path="./accounts.json"):
    """settings.txt + accounts.json 프록시 — GUI _collect_proxies 의 파일 버전.

    GUI 는 컨트롤러가 이미 읽어 둔 목록을 쓰지만 헤드리스에는 컨트롤러가 없다.
    settings.txt 형식(1줄 간격, 2줄 동시요청, 3줄~ 프록시)은 controller 와 같다."""
    out = []
    try:
        with open(settings_path, encoding="utf-8") as f:
            out += [ln.strip() for ln in f.read().splitlines()[2:]]
    except Exception:
        pass
    try:
        from daangn_ext import AccountStore
        out += AccountStore(accounts_path).proxies()
    except Exception:
        pass
    return [p for p in dict.fromkeys(out) if p]


def headless_sweep_cfg(settings, entries, notify, proxies=None,
                       proxy_provider=None, token_provider=None, log=None,
                       already_notified=None):
    """헤드리스 검색 스윕 cfg — GUI _auto_cfg_base + _sweep_cfg 와 같은 키를 만든다.

    이 한 겹만 따로 두는 이유: GUI 의 값 출처는 고급 패널 **위젯**이라 위젯이
    없는 런타임에서 그대로 부를 수 없다. 조건 조립(sweep_conditions)과 범위
    판정(sweep_scope_for)은 공유한다. 기본값은 GUI 위젯 초기값과 같게 맞췄다
    (휴식 30~90초, 지역간 0.4~1.2초, 레인 자동, 끌올 7일).

    여기서 읽는 sweep_* 키는 GUI 의 _sweep_settings_patch 가 쓴다 — 그래서
    GUI 에서 스윕을 설정하면 서버도 같은 범위로 돈다. 아무도 쓴 적이 없으면
    sweep_scope_for 의 기본 지역으로 떨어진다(전국이 아니다)."""
    s = settings or {}

    def _num(key, dflt):
        v = s.get(key)
        if v is None:
            return dflt
        try:
            return type(dflt)(v)
        except (TypeError, ValueError):
            return dflt

    cfg = {
        "rest_min": _num("sweep_rest_min", 30),
        "rest_max": _num("sweep_rest_max", 90),
        "gap_min": _num("sweep_gap_min", 0.4),
        "gap_max": _num("sweep_gap_max", 1.2),
        "lanes": _num("sweep_lanes", 0),          # 0 = 자동(프록시 수 기준)
        "tg_token": (notify or {}).get("tg_token") or None,
        "tg_chat": (notify or {}).get("tg_chat") or None,
        "sheet_url": (notify or {}).get("sheet_url") or None,
        "sheet_cred": (notify or {}).get("sheet_cred") or "./credentials.json",
        "proxies": list(proxies or []),
        "proxy_provider": proxy_provider,
        "access_token": None,
        "token_provider": token_provider,
        "stabilize": bool(token_provider),
        "accounts_fp": "./accounts.json",
        "daily_cap": 0,        # 0 = 상한 없음(account_scheduler 참고). 회전·격리만 쓴다.
        "warmup_days": 3,
        "out_json": "./OUT.json",
        "db_path": "./auto_seen.db",
    }
    if sweep_app_enabled(s):
        # 앱 키워드 스윕은 보완층이다 — 타지역 택배 매물이 목적이라 지역 1~2곳이면 된다.
        cfg.update({"scope": "regions",
                    "regions": list(s.get("sweep_regions_app") or FEED_DEFAULTS["sweep_regions_app"])})
    else:
        cfg.update(sweep_scope_for(s.get("sweep_regions"),
                                   s.get(SWEEP_NATIONWIDE_KEY),
                                   out_json=cfg["out_json"], log=log,
                                   n_conditions=len(entries or []) or 1,
                                   lanes=sweep_lanes_effective(
                                       cfg["lanes"], bool(token_provider), len(cfg["proxies"]))))
    cfg["conditions"] = sweep_conditions(
        entries,
        extra=[x for x in (s.get("sweep_extra") or []) if x],
        exclude=[x for x in (s.get("sweep_exclude") or []) if x],
        min_price=s.get("sweep_min"), max_price=s.get("sweep_max"),
        days=_num("sweep_days", 7) or None)
    if already_notified is not None:
        # 앱 알림이 이미 알린 매물은 스윕이 다시 안 알린다(저장소가 둘이다).
        cfg["already_notified"] = already_notified
    return cfg


def feed_proxies(settings, proxies_file="./proxies.txt") -> list[str]:
    """웹 프록시 목록 — 설정(조건 탭 '지역 선택' 카드의 웹 프록시 칸)이 비면 proxies.txt.

    피드 엔진과 가격추적(ProxyBudget)이 같은 풀을 쓴다. 계정용 프록시
    (settings.txt·accounts.json)와는 섞이지 않는다 — 저쪽은 계정에 묶인 IP 고
    이쪽은 계정 없는 공개 웹 경로다."""
    s = settings or {}
    v = s.get("feed_proxies")
    out = [p for p in (FEED_DEFAULTS["feed_proxies"] if v is None else v) if p]
    if out:
        return out
    try:
        with open(proxies_file, encoding="utf-8") as f:
            return [ln.strip() for ln in f if ln.strip()]
    except OSError:
        return []


def feed_cfg(settings, notify, proxies_file="./proxies.txt", out_json="./OUT.json",
             already_notified=None, log=None) -> dict:
    """웹 동 피드 엔진 cfg — GUI·헤드리스 공용. 지역은 스윕 지역 설정을 그대로 쓴다."""
    from daangn_ext.adaptive import load_dong_regions
    s = settings or {}

    def _get(key):
        v = s.get(key)
        return FEED_DEFAULTS[key] if v is None else v

    scope = sweep_scope_for(s.get("sweep_regions"), s.get(SWEEP_NATIONWIDE_KEY),
                            out_json=out_json, log=log, n_conditions=1, lanes=8)
    if scope.get("scope") == "nationwide":
        regions = [r["in"] for r in load_dong_regions(out_json)]
    else:
        regions = list(scope.get("regions") or [])
    proxies = feed_proxies(s, proxies_file)
    cfg = {
        "regions": regions,
        "categories": [int(c) for c in _get("feed_categories")],
        "proxies": proxies,
        "rps": float(_get("feed_rps")),
        "rest_min": float(_get("feed_rest_min")),
        "rules_path": "./data/alert_rules.json",
        "cursor_fp": "./data/feed_cursor.json",
        "tg_token": (notify or {}).get("tg_token") or None,
        "tg_chat": (notify or {}).get("tg_chat") or None,
    }
    if already_notified is not None:
        cfg["already_notified"] = already_notified
    return cfg


class HeadlessSweepRunner:
    """헤드리스 런타임의 검색 스윕 수명 — GUI 의 _start/_stop/_resync 를 옮긴 것.

    GUI 는 AutoMonitor(QThread)를 쓰지만 헤드리스에는 Qt 이벤트 루프가 없다.
    SweepEngine 은 순수 파이썬이므로 plain threading.Thread 로 그대로 돈다.
    '언제 갈아끼우나' 판정은 GUI 와 같은 sweep_resync_action 을 쓴다 —
    두 런타임이 다르게 판단하면 서버에서만 나는 버그가 생긴다.

    engine_factory/thread_factory 는 테스트가 진짜 스레드·네트워크 없이
    수명만 확인하려고 갈아끼우는 자리다."""

    def __init__(self, queue, cfg_builder, log, on_found,
                 engine_factory=None, thread_factory=None):
        self.queue = queue
        self.cfg_builder = cfg_builder
        self.log = log
        self.on_found = on_found
        self._engine_factory = engine_factory
        self._thread_factory = thread_factory
        self.engine = None
        self.thread = None
        self.kws = None            # 지금 도는 스윕이 떠 있는 키워드 집합
        self.revives = 0

    def _make_engine(self, cfg):
        if self._engine_factory is not None:
            return self._engine_factory(cfg, self.log, self.on_found)
        from daangn.sweep_engine import SweepEngine
        return SweepEngine(cfg, on_log=self.log, on_found=self.on_found)

    def _make_thread(self, target):
        if self._thread_factory is not None:
            return self._thread_factory(target)
        import threading
        # daemon: 되살리기 상한에 걸려 포기한 스레드가 프로세스 종료를 막으면
        # 서버 재시작이 통째로 걸린다.
        return threading.Thread(target=target, name="sweep", daemon=True)

    def running(self):
        t = self.thread
        return t is not None and bool(t.is_alive())

    def start(self):
        if self.running():
            # stop() 은 비동기다. 그 안에 다시 켜면 조용히 막혀 재시작이 통째로
            # 사라진 것처럼 보이므로 로그를 남긴다(GUI 와 같은 문구).
            self.log("[검색스윕] 아직 정지 중 — 이번 시작 요청은 건너뜁니다(다음 틱 재시도)")
            return False
        try:
            cfg = self.cfg_builder()
            if not cfg.get("conditions"):
                # conditions 가 비면 엔진이 cfg["keyword"] 로 떨어져 KeyError.
                self.log("[검색스윕] 대기열이 비어 시작하지 않습니다")
                return False
            self.engine = self._make_engine(cfg)
            self.thread = self._make_thread(self.engine.run)
            self.thread.start()
            self.kws = {c["keyword"] for c in cfg["conditions"]}
            self.log(f"[검색스윕] 시작 — 키워드 {len(self.kws)}개")
            return True
        except Exception as e:
            self.log(f"[검색스윕] 시작 실패: {str(e)[:120]}")
            return False

    def stop(self, join=0):
        """정지 요청. join 초를 주면 그만큼 기다린다(프로세스 종료 직전용)."""
        self.kws = None
        eng, t = self.engine, self.thread
        if eng is None or not self.running():
            return
        try:
            eng.stop()
        except Exception:
            pass
        self.log("[검색스윕] 정지 요청")
        if join:
            try:
                t.join(join)
            except Exception:
                pass
            if t.is_alive():
                self.log("[검색스윕] 정지 대기 초과 — 데몬 스레드로 두고 종료합니다")

    def resync(self):
        """대기열과 도는 스윕이 어긋났으면 갈아끼운다. 루프 1회에 한 번만 부른다 —
        등록이 몰아쳐도 재시작이 폴링 주기당 한 번을 넘지 않는다."""
        if self.queue is None:
            return
        try:
            want = set(self.queue.keywords())
        except Exception:
            return
        act = sweep_resync_action(want, self.kws, self.running())
        if not act:
            self.revives = 0
            return
        if act == "start":
            self.revives = 0
            self.start()
            return
        if act == "revive":
            ok, self.revives, msg = sweep_revive_step(self.revives, len(want))
            if msg:
                self.log(msg)
            if not ok:
                return
        else:
            self.revives = 0
            self.log(f"[검색스윕] 키워드 변경 {len(self.kws)}개 → {len(want)}개 — 재시작")
        self.stop()
        if want:
            self.start()


# --once 가 도는 동 수. --once 는 배포 직후 '살아 있나'를 보는 스모크다 —
# 서울·경기 1,857동 × 카테고리 2 를 다 돌면 한 시간을 넘겨 배포 스크립트가 멎는다.
FEED_ONCE_REGION_CAP = 10


def feed_once_cfg(cfg, log=None) -> dict:
    """--once 용 cfg — 앞 10동만 남긴다. 원본은 건드리지 않는다."""
    regions = list((cfg or {}).get("regions") or [])
    if len(regions) > FEED_ONCE_REGION_CAP:
        cfg = dict(cfg)
        cfg["regions"] = regions[:FEED_ONCE_REGION_CAP]
        if log:
            log(f"[피드] --once: 앞 {FEED_ONCE_REGION_CAP}동만 확인")
    return cfg


def feed_proxy_backoff_left(engine) -> float:
    """프록시 전멸로 죽은 엔진이면 남은 대기 초, 아니면 0.

    GUI(_start_feed)와 헤드리스(HeadlessFeedRunner.start)가 같은 함수를 본다 —
    따로 두면 한쪽만 죽은 프록시로 틱마다 재시작을 반복한다."""
    if engine is None or getattr(engine, "stop_reason", None) != "proxies":
        return 0.0
    import time as _t
    from daangn.feed_sweep import FeedSweep as _FS
    elapsed = _t.monotonic() - float(getattr(engine, "stopped_at", 0.0) or 0.0)
    return max(0.0, _FS.PROXY_BACKOFF_SEC - elapsed)


class HeadlessFeedRunner:
    """헤드리스 런타임의 동 피드 수명 — HeadlessSweepRunner 와 같은 모양."""

    def __init__(self, cfg_builder, log, on_found, engine_factory=None, thread_factory=None):
        self.cfg_builder = cfg_builder
        self.log = log
        self.on_found = on_found
        self._engine_factory = engine_factory
        self._thread_factory = thread_factory
        self.engine = None
        self.thread = None
        self._backoff_logged = False    # 백오프 안내는 한 번만(틱마다 아니고)

    def _make_engine(self, cfg):
        if self._engine_factory is not None:
            return self._engine_factory(cfg, self.log, self.on_found)
        from daangn.feed_sweep import FeedSweep
        return FeedSweep(cfg, on_log=self.log, on_found=self.on_found)

    def _make_thread(self, target):
        if self._thread_factory is not None:
            return self._thread_factory(target)
        import threading
        return threading.Thread(target=target, name="feed-sweep", daemon=True)

    def running(self) -> bool:
        t = self.thread
        return t is not None and t.is_alive()

    def start(self) -> bool:
        if self.running():
            self.log("[피드] 이미 돌고 있음 — 시작 요청 건너뜀")
            return False
        left = feed_proxy_backoff_left(self.engine)
        if left > 0:
            if not self._backoff_logged:
                self.log(f"[피드] 프록시 전멸 뒤 대기 — {int(left // 60) + 1}분 후 재시도")
                self._backoff_logged = True
            return False
        self._backoff_logged = False
        try:
            cfg = self.cfg_builder()
            if not cfg.get("regions"):
                self.log("[피드] 지역이 비어 시작하지 않습니다")
                return False
            self.engine = self._make_engine(cfg)
            self.thread = self._make_thread(self.engine.run)
            self.thread.start()
            self.log(f"[피드] 시작 — 동 {len(cfg['regions'])}곳 · 카테고리 {cfg.get('categories')}")
            return True
        except Exception as e:
            self.log(f"[피드] 시작 실패: {str(e)[:120]}")
            return False

    def stop(self, join=0):
        eng, t = self.engine, self.thread
        if eng is None or not self.running():
            return
        try:
            eng.stop()
        except Exception:
            pass
        self.log("[피드] 정지 요청")
        if join:
            try:
                t.join(join)
            except Exception:
                pass


def filter_by_conditions(matches, router, log=None, rules=None):
    """앱 경로 매칭에 조건표를 태운다. **조건표가 유일한 진실이다.**

    앱 알림은 당근 서버가 판정하고, 우리가 넘길 수 있는 건 최소가·최대가·
    제외어뿐이다. 추가키워드·끌올일수는 전달할 방법 자체가 없다. 그래서
    등록은 브랜드 단위로 넓게 하고 거르는 일은 여기서 조건표가 한다.

    조건표가 없으면 **아무것도 알리지 않는다.** 예전에는 '조건 없음 = 전부
    알림'이었는데, 그 상태에서 서버에 남은 브랜드 등록이 시간당 수백 건을
    쏟아냈다. 브랜드 등록 자체가 조건표에서만 나오므로(조건 탭 표 → 조건 적용),
    조건표가 없는데 매칭이 온다는 것은 서버에 낡은 등록이 남았다는 뜻이다 —
    그건 prune_to_rules 가 지운다.

    앱 매칭 payload 에는 제목·가격만 있고 본문과 끌올 시각이 없다. 그래서
    여기서는 제목과 가격까지만 본다 — 본문에만 있는 제외어와 끌올일수는
    이 경로에서 거를 수 없고 검색 스윕이 맡는다. 가격을 못 읽으면 가격
    조건은 건너뛴다(못 읽었다는 이유로 버리지 않는다).

    router 인자는 호출부 호환용으로만 남았다 — 판정에 쓰지 않는다.
    """
    items = list(matches or [])
    if not items:
        return items, [], 0
    if rules is None or not len(rules):
        if log:
            log(f"[조건표] 조건이 없어 매칭 {len(items)}건을 알리지 않습니다 — "
                "조건 탭 표에 적고 [조건 적용]을 누르세요")
        return [], [], len(items)
    from daangn_ext.alert_rules import HIT, WATCH
    from daangn_ext.search_filters import looks_wanted_ad
    kept, watch_only, cut = [], [], 0
    for m in items:
        raw_title = m.get("title") or ""
        # 삽니다/구합니다 글은 조건 유무와 무관하게 판매 매물이 아니다.
        if looks_wanted_ad(raw_title):
            cut += 1
            continue
        verdict, rule = rules.verdict(raw_title, m.get("price"))
        if verdict == HIT:
            # 어느 줄에 걸렸는지 알림에 남긴다 — 등록 키워드는 브랜드라
            # "[루이비통]" 만 뜨면 어떤 조건에 맞았는지 알 수 없다.
            kept.append(dict(m, _rule=rule.label()))
        elif verdict == WATCH:
            # 상한만 넘긴 매물은 버리지 않고 추적한다. 값이 내려오면
            # mark_range_entries 가 '조건 진입'으로 알린다.
            watch_only.append(m)
        else:
            cut += 1
    if log and (cut or watch_only):
        log(f"[매칭] 조건 불일치 {cut}건 제외"
            + (f" · 상한 초과 {len(watch_only)}건은 인하 대기로 추적" if watch_only else ""))
    return kept, watch_only, cut


def rule_brand_keys(table) -> set:
    """조건표의 브랜드 집합(정규화). 조건표가 없으면 빈 집합 = 어떤 키워드도 인정하지 않는다.

    서버·라우터에 있는 키워드가 '조건표 것인가'를 판정하는 유일한 기준이다."""
    from daangn_ext.alert_rules import brands
    from daangn_ext.search_filters import normalize_text
    if table is None or not len(table):
        return set()
    return {normalize_text(b) for b in brands(table.rules) if b}


def split_by_rules(keywords, allowed):
    """키워드 목록을 (조건표에 있는 것, 없는 것)으로 가른다. allowed=None 이면 전부 인정."""
    from daangn_ext.search_filters import normalize_text
    ok, extra = [], []
    for kw in keywords or []:
        kw = str(kw or "").strip()
        if not kw:
            continue
        (ok if allowed is None or normalize_text(kw) in allowed else extra).append(kw)
    return ok, extra


def prune_to_rules(router, fleet, allowed, log, core_only=False) -> int:
    """조건표에 없는 키워드를 라우터·대기열·서버 등록에서 지운다 → 지운 수.

    조건표가 비었으면 **아무것도 지우지 않는다.** 파일을 못 읽은 한 번의 오류로
    함대 등록을 전부 날리면 안 된다 — 그때는 filter_by_conditions 가 알림만 멈춘다.
    비우고 싶으면 [전체 삭제]가 따로 있다.

    남은 등록을 지워야 하는 이유는 낭비만이 아니다. 알림함은 15건에서 잘리고
    페이징이 없다 — 조건표에 없는 브랜드의 매칭이 그 15칸을 먹으면 진짜 매칭이
    밀려 나간다."""
    log = log or (lambda m: None)
    if not allowed:
        return 0
    n = 0
    if router is not None:
        try:
            _, extras = split_by_rules([r.get("keyword") for r in router.routes()], allowed)
            for kw in extras:
                router.remove(kw)
                n += 1
                log(f"[조건표] '{kw}' 는 조건표에 없어 등록에서 뺍니다")
        except Exception as e:
            log(f"[조건표] 라우터 정리 실패: {str(e)[:80]}")
    if fleet is not None:
        try:
            n += int(fleet.delete_not_in(allowed, log=log, core_only=core_only) or 0)
        except Exception as e:
            log(f"[조건표] 서버 등록 정리 실패: {str(e)[:80]}")
    return n

def dedupe_new_matches(matches, watch_store, fallback):
    """폴링 결과에서 아직 안 본 매치만 고른다 → (fresh, dropped).

    watch 테이블이 '본 매물'의 진실이다 — dead 행을 지우지 않으므로 판매완료·
    삭제된 매물이 다시 떠도 재알림하지 않는다.

    다만 watch 는 article_id 로만 키를 잡는다(add_from_matches). 그래서
    article_id 가 없는 매치(광고 등 알림 인박스 id 만 있는 payload)는 저장소에
    물어봐야 영원히 None 이 돌아온다 — 폴링마다 재알림이다. 그 키는 같은 DB 의
    seen_key 테이블이 받는다(watch 행이 아니다 — 매물이 아니므로 표에도 안 뜨고
    조회 대상도 아니다). 재시작해도 남는다.

    프로세스 안의 fallback 집합은 저장소를 아예 못 연 경우의 마지막 방어선으로만
    남는다(그때는 어차피 남길 곳이 없다). GUI·헤드리스가 같은 문을 쓰게 한 곳에
    둔다(따로 두면 한쪽만 고쳐진다).

    fallback 은 이 함수가 갱신한다. dropped 는 키가 아예 없어 버린 건수다.
    """
    fresh, dropped = [], 0
    durable = getattr(watch_store, "seen_key_add", None) if watch_store else None
    for m in matches or []:
        art = str((m or {}).get("article_id") or "")
        key = art or str((m or {}).get("id") or "")
        if not key:
            dropped += 1
            continue
        if art and watch_store is not None:
            if watch_store.get(art) is not None:
                continue
        elif durable is not None:
            if not durable(key):        # 이미 기록된 키 — 재알림 금지
                continue
        elif key in fallback:
            continue
        else:
            fallback.add(key)
        fresh.append(m)
    return fresh, dropped


def headless_watch_due(last_sweep, now, interval):
    """헤드리스 루프에서 이번 회에 스윕할 차례인지.

    아직 한 번도 안 돌았으면(0/None) 바로 돈다 — --once 스모크가 성립하도록."""
    if not last_sweep:
        return True
    return int(now) - int(last_sweep) >= int(interval)


class _HarvestThread(QtCore.QThread):
    """백그라운드 자동 수확 — 앱 실행 중 주기적으로 LDPlayer/폰서 토큰 갱신.
    accounts.json 을 항상 신선하게 유지 → 수동 수확 불필요. access 30분 만료 전 갱신."""
    tick = QtCore.pyqtSignal(str)

    def __init__(self, interval=None, accounts="./accounts.json"):
        super().__init__()
        self.interval = harvest_interval() if interval is None else interval
        self.accounts = accounts
        # 첫 틱은 아직 측정값이 없으니 설정 간격을 그대로 기준으로 쓴다.
        self._period = float(self.interval)
        self._stop = False

    def run(self):
        import time as _t
        while not self._stop:
            started = _t.monotonic()
            hstats = {}
            # 프록시를 먼저 걸어야 이어지는 콜드스타트가 그 프록시로 나간다.
            guest_proxy_sync(self.accounts, log=self.tick.emit)
            try:
                import ld_autoharvest
                u, i, t, h = ld_autoharvest.harvest_all(
                    self.accounts, nudge=True, log=self.tick.emit, stats=hstats)
                self.tick.emit(f"[자동수확] {h}계정 갱신 · 총 {t}계정" if h
                               else "[자동수확] 대상 없음(LDPlayer/폰 확인)")
            except Exception as e:
                self.tick.emit(f"[자동수확] 실패: {str(e)[:60]}")
            # 다음 수확은 '가장 먼저 죽는 토큰'의 만료 직후다. 앱이 만료 전에는
            # 갱신을 거절하므로(ld_autoharvest 상수 주석의 실측 로그), 고정 주기로
            # 돌면 최악의 경우 그 주기만큼 만료된 채 방치된다 — 실제로 20분을
            # 그렇게 보내고 폴링이 "전계정(0)" 으로 헛돌았다.
            import ld_autoharvest as _LA
            spent = _t.monotonic() - started
            delay = _LA.next_harvest_delay(hstats.get("min_remaining"),
                                           ceil=self.interval)
            for _ in range(max(1, int(delay - spent))):
                if self._stop:
                    return
                _t.sleep(1)

    def stop(self):
        self._stop = True


class _AddInstanceThread(QtCore.QThread):
    """`.ldbk` 복원 워커. 복원은 몇 분 걸리므로 GUI 스레드에서 돌리면 안 된다."""
    line = QtCore.pyqtSignal(str)
    done = QtCore.pyqtSignal(object, object)     # (결과 dict|None, 오류문자열|None)

    def __init__(self, ldbk, name):
        super().__init__()
        self.ldbk = ldbk
        self.name = name

    def run(self):
        try:
            import ld_instance
            res = ld_instance.add_from_ldbk(
                self.ldbk, self.name, app_dir=".", log=self.line.emit)
        except Exception as e:
            self.done.emit(None, f"{type(e).__name__}: {str(e)[:200]}")
            return
        self.done.emit(res, None)


class _ThumbThread(QtCore.QThread):
    """매칭 썸네일 비동기 다운로드 — GUI 프리징 없이 사진 로드.
    jobs: [(QTableWidgetItem, url)]. 다운로드 후 (item, bytes) emit → 메인서 setIcon."""
    loaded = QtCore.pyqtSignal(object, bytes)

    def __init__(self, jobs):
        super().__init__()
        self.jobs = jobs

    def run(self):
        try:
            import httpx
            c = httpx.Client(timeout=8, follow_redirects=True)
        except Exception:
            return
        for item, url in self.jobs:
            if not url:
                continue
            try:
                r = c.get(url)
                if r.status_code == 200 and r.content:
                    self.loaded.emit(item, r.content)
            except Exception:
                pass
        try:
            c.close()
        except Exception:
            pass


class _NotifyThread(QtCore.QThread):
    """매칭 알림 백그라운드 — 텔레그램(실시간) + 구글시트(검색가능 히스토리). GUI 안 멈춤."""
    log = QtCore.pyqtSignal(str)

    def __init__(self, notify, items):
        super().__init__()
        self.notify = notify or {}
        self.items = items

    def run(self):
        import time as _t
        emit = lambda m: self.log.emit(m)
        tok, chat = self.notify.get("tg_token"), self.notify.get("tg_chat")
        if tok and chat:
            try:
                from daangn.notify import TelegramSender, item_block
                tg = TelegramSender(tok, chat, log=emit)
                for m in self.items:
                    tg.enqueue_item(item_block(
                        "신규 매물", m.get("region"), m.get("title"), m.get("price"),
                        m.get("url"), stamp=m.get("time"), stamp_label="등록"))
                tg.flush()
                emit(f"[텔레그램] {len(self.items)}건 전송")
            except Exception as e:
                emit(f"[텔레그램] 실패: {str(e)[:50]}")
        # 구글시트 히스토리(선택)
        if self.notify.get("sheet_url"):
            try:
                from daangn.notify import SheetWriter
                sw = SheetWriter(self.notify.get("sheet_url"),
                                 self.notify.get("sheet_cred") or "./credentials.json", log=emit)
                for m in self.items:
                    ts = _t.strftime("%Y-%m-%d %H:%M", _t.localtime(int(m.get("time") or 0))) \
                        if m.get("time") else ""
                    sw.enqueue_row([ts, m.get("keyword") or "", m.get("title") or "",
                                    m.get("price") or "", m.get("region") or "",
                                    m.get("_account") or "", m.get("url") or ""])
                wrote, failed = sw.flush()
                if wrote:
                    emit(f"[구글시트] {wrote}행 기록")
            except Exception as e:
                emit(f"[구글시트] 실패: {str(e)[:50]}")


class _WatchNotifyThread(QtCore.QThread):
    """워치리스트 변동 알림 — 텔레그램 + 구글시트. GUI 안 멈춤."""
    log = QtCore.pyqtSignal(str)

    def __init__(self, notify, lines, events=None):
        super().__init__()
        self.notify = notify or {}
        self.lines = list(lines)          # 로그·구글시트용 한 줄 요약
        self.events = list(events or [])  # 텔레그램 블록용 원본 이벤트

    def run(self):
        import time as _t
        emit = lambda m: self.log.emit(m)
        if not self.lines:
            return
        tok, chat = self.notify.get("tg_token"), self.notify.get("tg_chat")
        if tok and chat:
            try:
                from daangn.notify import TelegramSender
                tg = TelegramSender(tok, chat, log=emit)
                _enqueue_watch_blocks(tg, self.events, self.lines)
                tg.flush()
                emit(f"[텔레그램] 변동 {len(self.lines)}건 전송")
            except Exception as e:
                emit(f"[텔레그램] 실패: {str(e)[:50]}")
        if self.notify.get("sheet_url"):
            try:
                from daangn.notify import SheetWriter
                sw = SheetWriter(self.notify.get("sheet_url"),
                                 self.notify.get("sheet_cred") or "./credentials.json",
                                 log=emit)
                ts = _t.strftime("%Y-%m-%d %H:%M")
                for ln in self.lines:
                    sw.enqueue_row([ts, "가격변동", ln])
                wrote, failed = sw.flush()
                if wrote:
                    emit(f"[구글시트] {wrote}행 기록")
            except Exception as e:
                emit(f"[구글시트] 실패: {str(e)[:50]}")


class _WatchSweepThread(QtCore.QThread):
    """워치리스트 재조회 — 네트워크가 GUI 를 막지 않게 백그라운드에서 돈다.

    sqlite 커넥션은 GUI 스레드(add_from_matches)와 공유한다. 모든 store 메서드가
    단문 + 즉시 commit 이라 커넥션 뮤텍스로 직렬화된다 — 여러 문장을 한 트랜잭션에
    묶게 되면 그때는 잠금이 필요하다."""
    done = QtCore.pyqtSignal(list, int, int)      # events, active_count, next_due_at
    log = QtCore.pyqtSignal(str)

    def __init__(self, tracker, store, budget, interval):
        super().__init__()
        self._tracker, self._store = tracker, store
        self._budget, self._interval = budget, interval
        self._stopped = False

    def stop(self):
        """다음 항목부터 멈춘다. sweep 은 provider 가 None 이면 그 자리서 끝낸다."""
        self._stopped = True

    def _provider(self):
        return None if self._stopped else self._budget.next()

    def run(self):
        events = []
        try:
            # 첫 스윕 전(최대 10분) 패널이 '추적 중 0건'으로 보이지 않게 먼저 현황만
            self.done.emit([], self._store.active_count(), self._store.next_due_at())
        except Exception:
            pass
        try:
            dropped = self._tracker.enforce_cap()
            if dropped:
                self.log.emit(f"[가격추적] 상한 초과 {dropped}건 추적 중단")
            n = watch_sweep_budget(self._store.active_count(), self._interval)
            if n:
                self._budget.reload()
                events = self._tracker.sweep(self._provider, n)
                if getattr(self._tracker, "last_sweep_exhausted", False):
                    self.log.emit("[가격추적] 계정 예산 소진 — 남은 대상은 다음 회차로")
        except Exception as e:
            self.log.emit(f"[가격추적] 스윕 실패: {str(e)[:120]}")
        try:
            self.done.emit(events, self._store.active_count(),
                           self._store.next_due_at())
        except Exception:
            self.done.emit(events, 0, 0)


class _AlertWorker(QtCore.QThread):
    """알림 API 호출을 백그라운드로(GUI 프리징 방지). fn(log_emit) 실행 후 결과 emit."""
    done = QtCore.pyqtSignal(object)
    log = QtCore.pyqtSignal(str)

    def __init__(self, fn):
        super().__init__()
        self._fn = fn

    def run(self):
        try:
            self.done.emit(self._fn(self.log.emit))
        except Exception as e:
            self.log.emit(f"[오류] {type(e).__name__}: {e}")
            self.done.emit(None)


def emul_reconcile(rows, attached, is_window):
    """스캔 행 + 부착 중인 창 → (live, detach).

    attached: ld index -> 탭에 부착된 child hwnd.
    부착된 창은 WS_CHILD 라 EnumWindows(top-level 열거)에 안 잡힌다. 그래서
    스캔 결과에 없거나 다른 창(같은 pid 의 툴바 등)으로 잡혀도 '사라졌다'가
    아니다 — 부착 창의 생사는 IsWindow 로만 본다. 죽었으면 detach 에 넣고,
    그 인스턴스가 새 창으로 다시 떴으면 스캔이 준 새 창을 live 에 남긴다.
    """
    live = {r["index"]: r for r in rows
            if r["running"] and r["top_hwnd"] and is_window(r["top_hwnd"])}
    by_index = {r["index"]: r for r in rows}
    detach = []
    for idx, hwnd in attached.items():
        if not is_window(hwnd):
            detach.append(idx)
            continue
        r = dict(live.get(idx) or by_index.get(idx)
                 or {"index": idx, "name": "", "pid": 0})
        r["top_hwnd"] = hwnd
        r["running"] = True
        if not r.get("title"):
            r["title"] = r.get("name") or f"인스턴스 {idx}"
        live[idx] = r
    return live, detach


class _LdListThread(QtCore.QThread):
    """ldconsole list2 조회 — LDPlayer 가 hang 해도 GUI 가 멈추지 않게 별도 스레드."""
    rows = QtCore.pyqtSignal(object)

    def __init__(self, console):
        super().__init__()
        self._console = console

    def run(self):
        import ldwin
        self.rows.emit(ldwin.list_instances(self._console))


class _EmbedHost(QtWidgets.QWidget):
    """LDPlayer 창 하나를 담는 네이티브 호스트 위젯. 크기 변하면 자식 창도 맞춘다."""

    def __init__(self, embedder, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_NativeWindow, True)
        self.setMinimumSize(360, 480)
        self._embedder = embedder
        self.child_hwnd = 0

    def host_hwnd(self):
        return int(self.winId())

    def resizeEvent(self, ev):
        super().resizeEvent(ev)
        if self.child_hwnd:
            self._embedder.fit(self.child_hwnd, self.width(), self.height())

    def showEvent(self, ev):
        super().showEvent(ev)
        if self.child_hwnd:
            self._embedder.fit(self.child_hwnd, self.width(), self.height())

    def mousePressEvent(self, ev):
        super().mousePressEvent(ev)
        if self.child_hwnd:
            self._embedder.focus(self.child_hwnd)


class _ThumbCaptureThread(QtCore.QThread):
    """인스턴스 창 썸네일 캡처. PrintWindow 는 대상이 hang 하면 같이 멈추므로
    반드시 워커에서 돌린다(GUI 는 결과가 늦게 올 뿐 얼지 않는다)."""
    shot = QtCore.pyqtSignal(object)     # (ld index, (w, h, bytes) | None)

    def __init__(self, targets):
        super().__init__()
        self._targets = list(targets)    # [(index, hwnd), ...]

    def run(self):
        import ldwin
        for idx, hwnd in self._targets:
            try:
                self.shot.emit((idx, ldwin.capture(hwnd)))
            except Exception:
                self.shot.emit((idx, None))


class _InstanceCard(QtWidgets.QFrame):
    """인스턴스 하나를 나타내는 카드 — 썸네일 + 이름 + 상태. 클릭하면 탭으로 연다."""
    clicked = QtCore.pyqtSignal(int)

    THUMB_W, THUMB_H = 150, 252

    def __init__(self, index, name, parent=None):
        super().__init__(parent)
        self.index = index
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setStyleSheet(
            "QFrame { background:#FBFAF7; border:1px solid #DDD6C9; border-radius:10px; }"
            "QFrame:hover { border:1px solid #8A6D1F; }")
        lay = QtWidgets.QVBoxLayout(self)
        lay.setContentsMargins(8, 8, 8, 8); lay.setSpacing(5)

        self.thumb = QtWidgets.QLabel(self)
        self.thumb.setFixedSize(self.THUMB_W, self.THUMB_H)
        self.thumb.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.thumb.setWordWrap(True)
        self.thumb.setStyleSheet(
            "background:#F2F1EE; border:none; border-radius:6px;"
            "color:#78705F; font-size:12px;")
        self.thumb.setText("화면 불러오는 중…")

        self.title = QtWidgets.QLabel(name, self)
        self.title.setStyleSheet("border:none; color:#1F1B16; font-weight:700; font-size:13px;")
        self.title.setToolTip(f"{name} (index {index})")
        self.state = QtWidgets.QLabel("", self)
        self.state.setStyleSheet("border:none; color:#78705F; font-size:11px;")

        lay.addWidget(self.thumb); lay.addWidget(self.title); lay.addWidget(self.state)

    def set_frame(self, pixmap):
        if pixmap is None:
            self.thumb.setText("화면 캡처 불가 — 클릭하면 탭에서 직접 볼 수 있음")
            return
        self.thumb.setPixmap(pixmap.scaled(
            self.THUMB_W, self.THUMB_H,
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation))

    def set_state(self, text, attached):
        self.state.setText(text)
        self.state.setStyleSheet(
            "border:none; font-size:11px; color:%s;" % ("#8A6D1F" if attached else "#78705F"))
        self.setStyleSheet(
            "QFrame { background:#FBFAF7; border-radius:10px; border:1px solid %s; }"
            "QFrame:hover { border:1px solid #8A6D1F; }"
            % ("#8A6D1F" if attached else "#DDD6C9"))

    def mouseReleaseEvent(self, ev):
        super().mouseReleaseEvent(ev)
        if ev.button() == Qt.MouseButton.LeftButton and self.rect().contains(ev.pos()):
            self.clicked.emit(self.index)


from PyQt6.QtGui import (
    QCloseEvent,
    QFontDatabase,
    QRegularExpressionValidator,
    QPixmap,
    QDesktopServices,
)
from PyQt6.QtCore import Qt, QRegularExpression, QItemSelection, QUrl
from PyQt6.QtWidgets import QApplication, QMainWindow
from PyQt6.QtWidgets import QProgressDialog
from PIL import Image as PILImage
from openpyxl import load_workbook


# ── 명품 프리미엄 테마 (라이트 + 골드) ──
# base #FFFFFF · surface #F7F5F1 · card #FBFAF7 · border #DDD6C9
# gold #9A7B2E (deep #8A6D1F / fill #C9A751) · ink #1F1B16 · muted #6B6355
# 2026-09-02 다크에서 전환. 클라가 RDP 로 서버 화면을 보는데 색 압축을 거치면
# 어두운 배경의 대비가 무너져 팝업 글자가 안 보였다. 색만 바꾸고 배치는 그대로다.
APP_QSS = """
/* 라이트 팔레트 (2026-09-02). 클라가 RDP 로 서버 화면을 보는데, 어두운 배경은
   RDP 색 압축을 거치며 대비가 무너져 팝업 글자가 안 보였다. 흰 바탕에 진한
   글자로 바꾸고, 브랜드 금색은 흰 배경에서도 읽히도록 어둡게 내렸다(#8A6D1F).

   본문 대비 14:1, 보조 텍스트 7:1 이상 — 저사양 원격 화면에서도 읽히는 값이다.
   팝업(QMessageBox/QInputDialog/QFileDialog)은 **명시적으로** 지정한다. 예전에는
   지정이 없어, 부모 없이 뜬 경고창이 스타일을 못 받고 제각각으로 보였다. */
* { font-family: 'Pretendard', '.AppleSystemUIFont', 'Apple SD Gothic Neo', 'Malgun Gothic', Helvetica; font-size: 14px; font-weight: 500; color: #1F1B16; }
QMainWindow, QWidget { background: #FFFFFF; }
QToolTip { background:#2A251E; color:#F7F5F1; border:1px solid #2A251E; padding:7px 10px; border-radius:8px; font-size:14px; }

#brandBar { background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #FBFAF7, stop:1 #F2EFE9); border-bottom: 1px solid #DDD6C9; }

QTabWidget::pane { border: none; background: transparent; }
QTabBar { qproperty-drawBase: 0; left: 18px; }
QTabBar::tab { background: transparent; color: #6B6355; padding: 11px 20px; margin-right: 4px; margin-top: 8px; border: none; border-bottom: 2px solid transparent; font-size: 16px; font-weight: 700; }
QTabBar::tab:selected { color: #8A6D1F; border-bottom: 2px solid #9A7B2E; }
QTabBar::tab:hover:!selected { color: #1F1B16; }

QLabel { color: #3A342B; background: transparent; }

QLineEdit, QSpinBox, QDoubleSpinBox, QComboBox { background: #FFFFFF; border: 1.5px solid #C9C1B2; border-radius: 11px; padding: 8px 13px; color: #1F1B16; font-size: 15px; min-height: 24px; selection-background-color: #E7D3A6; selection-color: #1F1B16; }
QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus, QComboBox:focus { background: #FFFDF8; border: 1.5px solid #9A7B2E; }
QLineEdit:disabled, QSpinBox:disabled, QComboBox:disabled { background: #F2F1EE; color: #8B8474; }
QLineEdit::placeholder { color: #8B8474; }
QComboBox QAbstractItemView { background: #FFFFFF; color: #1F1B16; border: 1px solid #C9C1B2; border-radius: 10px; font-size: 15px; selection-background-color: #F0E6CC; selection-color: #1F1B16; outline: none; padding: 4px; }
QComboBox::drop-down { border: none; width: 22px; }
QSpinBox::up-button, QSpinBox::down-button,
QDoubleSpinBox::up-button, QDoubleSpinBox::down-button { width: 16px; border: none; background: transparent; }

QPushButton { background: #F7F5F1; color: #2A251E; border: 1px solid #C9C1B2; border-radius: 11px; padding: 9px 15px; font-weight: 700; font-size: 14px; min-height: 22px; }
QPushButton:hover { background: #FFFFFF; border-color: #9A7B2E; color: #1F1B16; }
QPushButton:pressed { background: #EDE9E0; }
QPushButton:disabled { color: #A69E8D; border-color: #E4DFD5; background: #F7F5F1; }
QPushButton#startBtn { border: none; color: #241C08; padding: 10px 24px; font-size: 15px; font-weight: 800;
  background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #E3C57F, stop:1 #C9A751); }
QPushButton#startBtn:hover { background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #EDD292, stop:1 #D6B45E); }
QPushButton#startBtn:pressed { background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #C9A751, stop:1 #B08F3E); }

QCheckBox { color: #3A342B; spacing: 8px; font-size: 15px; min-height: 24px; background: transparent; }
QCheckBox::indicator { width: 19px; height: 19px; border: 1.5px solid #B5AC9A; border-radius: 6px; background: #FFFFFF; }
QCheckBox::indicator:checked { background: #9A7B2E; border-color: #9A7B2E; image: none; }
QCheckBox::indicator:hover { border-color: #9A7B2E; }

QGroupBox { background: #FBFAF7; border: 1px solid #DDD6C9; border-radius: 16px; margin-top: 18px; padding: 22px 18px 16px 18px; font-size: 16px; font-weight: 800; color: #8A6D1F; }
QGroupBox::title { subcontrol-origin: margin; left: 10px; top: 2px; padding: 0 6px; letter-spacing: 1px; }

/* 접이식 섹션 — 카드가 아니라 목록의 한 줄이다. 접었을 때 빈 상자가 남으면
   화면에 유령 카드가 생긴다(실제로 그렇게 보였다). 얇은 구분선 하나만 둔다. */
QGroupBox#sectionBox { background: transparent; border: none; border-top: 1px solid #E8E3D9;
  border-radius: 0; margin-top: 6px; padding: 14px 2px 0 2px;
  font-size: 15px; font-weight: 700; color: #6B6357; letter-spacing: 0; }
QGroupBox#sectionBox::title { subcontrol-origin: margin; left: 0px; top: 4px; padding: 0 2px; letter-spacing: 0; }
QGroupBox#sectionBox:checked { color: #1F1B16; }
QGroupBox#sectionBox:hover { color: #9A7B2E; }
QGroupBox#sectionBox::indicator { width: 0px; height: 0px; margin: 0; border: none; }
/* 카드 첫 섹션 — 카드 머리글 바로 아래라 구분선이 필요 없다. */
QGroupBox#sectionBox[firstInCard="true"] { border-top: none; margin-top: 2px; }
QGroupBox#sectionBox[firstInCard="bare"] { border-top: none; margin-top: 0; padding-top: 0; }

/* 조건 탭 — 단계 카드. 회색 바탕 위에 흰 카드 두 장(① 조건, ② 지역)이
   순서대로 놓인다. 접이식 줄 세 개가 한 면에 늘어서 있던 동안 어디가 시작이고
   무엇이 필수인지 보이지 않았다. 카드가 순서를, 배지가 단계를 말한다. */
QWidget#rulesPage { background: #F5F3EE; }
QFrame#stepCard { background: #FFFFFF; border: 1px solid #E8E3D9; border-radius: 20px; }
QLabel#stepBadge { background: #9A7B2E; color: #FFFFFF; border-radius: 14px; min-width: 28px; max-width: 28px;
  min-height: 28px; max-height: 28px; font-size: 14px; font-weight: 800; qproperty-alignment: AlignCenter; }
QLabel#stepTitle { color: #1F1B16; font-size: 19px; font-weight: 800; }
QLabel#stepSub { color: #8B857A; font-size: 13px; font-weight: 500; }
QLabel#mutedNote { color: #8B857A; font-size: 13px; }

/* 설정 탭 — 토스웹 문법. 회색 바탕 위 흰 카드, 카드 안은 목록 행(라벨 왼쪽,
   입력 오른쪽, 행 사이 얇은 선). 주 동작(저장) 하나만 금색이고 나머지는 글자
   버튼이다. 예전엔 베이지 QGroupBox 세 개가 금색 제목을 띄우고 있어 조건 탭의
   단계 카드와 다른 앱처럼 보였다. */
QWidget#settingsPage { background: #F5F3EE; }
QWidget#settingsCol { background: transparent; }
/* 에뮬레이터 탭 — 같은 회색 바탕 위 카드 둘(인스턴스 목록 · 부착 화면).
   제목 줄 하나에 상태와 버튼을 모은다. 카드 안 탭·스크롤은 테두리를 뺀다. */
QWidget#emulPage { background: #F5F3EE; }
QWidget#emulPage QSplitter::handle { background: transparent; width: 12px; }
QFrame#stepCard QTabWidget::pane { border: none; }
QFrame#stepCard QScrollArea { background: transparent; }
QFrame#settingRow { background: transparent; border: none; border-bottom: 1px solid #F1EEE8; }
QFrame#settingRow[last="true"] { border-bottom: none; }
QLabel#rowLabel { color: #3A342B; font-size: 14px; font-weight: 700; }
QPushButton#dangerBtn { background: transparent; border: none; color: #B4342A; padding: 9px 10px; font-weight: 700; }
QPushButton#dangerBtn:hover { background: #FCEBE8; border-radius: 10px; }
QPushButton#dangerBtn:disabled { color: #D8A9A4; background: transparent; }
/* 보조 동작은 글자 버튼 — 주 동작(금색) 옆에서 같은 무게로 보이면 안 된다. */
QPushButton#linkBtn { background: transparent; border: none; color: #8A6D1F; padding: 9px 10px; font-weight: 700; }
QPushButton#linkBtn:hover { background: #FBF6EA; border-radius: 10px; }
QPushButton#linkBtn:disabled { color: #B5AC9A; background: transparent; }
QPushButton#ghostBtn { background: #F2F0EC; border: none; border-radius: 10px; color: #3A342B; padding: 8px 14px; font-size: 13px; }
QPushButton#ghostBtn:hover { background: #E8E4DC; }

/* 지역 트리 — 토스 결. 테두리 없는 흰 면, 행은 델리게이트가 그린다(둥근
   파란 체크·셰브런·hover 회색). 검색은 회색 알약, 보조 동작은 글자 버튼,
   선택 수는 파란 칩. 결과 탭과 같은 파랑(#3182F6) 하나만 쓴다. */
QTreeWidget#tossTree { background: #FFFFFF; border: none; border-radius: 0; padding: 0; outline: none; }
QTreeWidget#tossTree::item { padding: 0; border: none; background: transparent; }
QTreeWidget#tossTree::item:selected, QTreeWidget#tossTree::item:hover { background: transparent; color: #191F28; }
QTreeWidget#tossTree QScrollBar:vertical { background: transparent; width: 8px; margin: 2px 2px; }
QTreeWidget#tossTree QScrollBar::handle:vertical { background: #E5E8EB; border-radius: 4px; min-height: 36px; }
QTreeWidget#tossTree QScrollBar::handle:vertical:hover { background: #D1D6DB; }
QLineEdit#tossSearch { background: #F2F4F6; border: 1.5px solid transparent; border-radius: 12px;
  padding: 9px 14px; font-size: 15px; color: #191F28; min-height: 22px; }
QLineEdit#tossSearch:focus { background: #FFFFFF; border: 1.5px solid #3182F6; }
QPushButton#tossTextBtn { background: transparent; border: none; border-radius: 10px; color: #4E5968;
  padding: 9px 10px; font-size: 13px; font-weight: 700; }
QPushButton#tossTextBtn:hover { background: #F2F4F6; color: #191F28; }
QLabel#tossCountChip { background: #F2F4F6; color: #8B95A1; border-radius: 13px; padding: 5px 12px;
  font-size: 13px; font-weight: 800; }
QLabel#tossCountChip[some="true"] { background: #E8F3FF; color: #3182F6; }
QWidget#tossChipRow { background: transparent; }
QPushButton#tossChip { background: #E8F3FF; color: #1B64DA; border: none; border-radius: 14px;
  padding: 6px 12px; font-size: 13px; font-weight: 700; }
QPushButton#tossChip:hover { background: #D6E9FF; }
QLabel#tossChipMore { color: #8B95A1; font-size: 13px; font-weight: 700; padding: 6px 4px; }

/* 목록 필터 — 고르는 값이지 누르는 명령이 아니다. 버튼처럼 도드라지면
   [감시 시작] 과 같은 무게로 보인다. 선택된 하나만 진하게. */
QPushButton#filterChip { background: #F2F0EC; color: #6B6357; border: none;
  border-radius: 14px; padding: 7px 16px; font-size: 13px; font-weight: 700; }
QPushButton#filterChip:hover { background: #E8E4DC; color: #3A342B; }
QPushButton#filterChip:checked { background: #2A251E; color: #FFFFFF; }

/* ── 결과 탭 — 토스 웹 결. 테두리 대신 여백과 회색 단계가 구조를 말하고,
   파랑(#3182F6) 하나만 누르는 것에 쓴다. 다른 탭은 건드리지 않는다. */
QWidget#resultsPage { background: #F9FAFB; }
QWidget#resultsPage QLabel { color: #4E5968; }
QWidget#resultsPage QPushButton#startBtn { background: #3182F6; color: #FFFFFF; border: none; border-radius: 12px;
  padding: 12px 22px; font-size: 15px; font-weight: 700; }
QWidget#resultsPage QPushButton#startBtn:hover { background: #1B64DA; }
QWidget#resultsPage QPushButton#startBtn:pressed { background: #1957C2; }
QWidget#resultsPage QPushButton#startBtn:checked { background: #F2F4F6; color: #191F28; }
QWidget#resultsPage QPushButton#startBtn:checked:hover { background: #E5E8EB; }
QWidget#resultsPage QPushButton#startBtn:disabled { background: #E5E8EB; color: #B0B8C1; }
QWidget#resultsPage QLabel#statusLine { font-size: 14px; color: #6B7684; }
QWidget#resultsPage QLabel#watchCount { font-size: 14px; color: #8B95A1; }
QFrame#resultCard { background: #FFFFFF; border: none; border-radius: 20px; }
QFrame#resultCard QPushButton#filterChip { background: #F2F4F6; color: #4E5968; border: none; border-radius: 17px;
  padding: 8px 16px; font-size: 14px; font-weight: 600; }
QFrame#resultCard QPushButton#filterChip:hover { background: #E5E8EB; color: #191F28; }
QFrame#resultCard QPushButton#filterChip:checked { background: #E8F3FF; color: #3182F6; }
QFrame#resultCard QTableWidget { background: #FFFFFF; border: none; border-radius: 0; padding: 0; font-size: 15px;
  color: #191F28; selection-background-color: #F2F4F6; }
QFrame#resultCard QTableWidget::item { padding: 12px 10px; }
QFrame#resultCard QTableWidget::item:selected { background: #F2F4F6; color: #191F28; }
QFrame#resultCard QHeaderView::section { background: #FFFFFF; color: #8B95A1; padding: 10px 10px 12px 10px; border: none;
  border-bottom: 1px solid #F2F4F6; font-size: 13px; font-weight: 600; letter-spacing: 0; }
QFrame#resultCard QTableCornerButton::section { background: #FFFFFF; border: none; }
QWidget#resultsPage QGroupBox#sectionBox { border: none; margin-top: 0; padding: 12px 4px 0 4px;
  color: #8B95A1; font-size: 14px; font-weight: 600; }
QWidget#resultsPage QGroupBox#sectionBox:checked { color: #4E5968; }
QWidget#resultsPage QGroupBox#sectionBox:hover { color: #3182F6; }
QWidget#resultsPage QTextEdit { background: #FFFFFF; border: none; border-radius: 16px; padding: 12px 14px;
  font-size: 13px; color: #6B7684; }
QWidget#resultsPage QProgressBar { background: #F2F4F6; border: none; border-radius: 3px; min-height: 6px; max-height: 6px; }
QWidget#resultsPage QProgressBar::chunk { border-radius: 3px; background: #3182F6; }
QWidget#resultsPage QScrollBar::handle:vertical { background: #E5E8EB; }
QWidget#resultsPage QScrollBar::handle:vertical:hover { background: #D1D6DB; }

QTreeWidget, QListWidget, QTableWidget, QTextEdit, QTextBrowser { background: #FFFFFF; border: 1px solid #DDD6C9; border-radius: 13px; padding: 4px; font-size: 14px; color: #1F1B16; }
QTreeWidget::item, QListWidget::item { padding: 6px 4px; border-radius: 8px; }
QTreeWidget::item:selected, QListWidget::item:selected { background: #F0E6CC; color: #1F1B16; }
QTreeWidget::item:hover, QListWidget::item:hover { background: #F7F5F1; }

QTableWidget { gridline-color: transparent; }
/* ::item 에 color 를 주면 setForeground 가, border 를 주면 setBackground 가
   전부 무시된다(QStyleSheetStyle). 실제로 미등록 빨강·계정표 점검필요 바탕이
   그래서 안 보였다. 글자색은 위 QTableWidget 규칙이 물려주고, 줄 구분선은
   RowLineDelegate 가 그린다. 여기엔 padding 만 둔다. */
QTableWidget::item { padding: 11px 6px; }
QTableWidget::item:selected { background: #F5EFDD; color: #1F1B16; }
QTableWidget { selection-background-color: #F5EFDD; }
QHeaderView::section { background: #FFFFFF; color: #8B857A; padding: 12px 8px; border: none; border-bottom: 1px solid #EAE6DE; font-weight: 700; font-size: 12px; letter-spacing: 0.3px; }
/* 조건 표의 행 번호 — 엑셀처럼 옆에 붙는 숫자. 세로 패딩이 있으면 30px 줄에서 깨진다. */
QTableWidget#rulesGrid QHeaderView::section:vertical { padding: 0 4px; border-bottom: 1px solid #F1EEE8; border-right: 1px solid #EAE6DE; font-weight: 500; color: #B0A99C; }
QTableCornerButton::section { background: #F7F5F1; border: none; }

QProgressBar { background: #EDE9E0; border: 1px solid #DDD6C9; border-radius: 7px; min-height: 24px; max-height: 24px; color: #1F1B16; font-weight: 800; font-size: 13px; letter-spacing: 0.5px; text-align: center; }
QProgressBar::chunk { border-radius: 6px; background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #C9A751, stop:1 #E3C57F); }

QScrollArea { border: none; background: transparent; }
QSplitter::handle { background: transparent; }
QStatusBar { background: #F7F5F1; color: #5C5449; border-top: 1px solid #DDD6C9; font-size: 13px; }
QStatusBar::item { border: none; }

/* 팝업 — 부모 없이 뜨는 것까지 확실히 잡는다. */
QDialog, QMessageBox, QInputDialog, QFileDialog, QColorDialog, QFontDialog { background: #FFFFFF; }
QMessageBox QLabel, QInputDialog QLabel { color: #1F1B16; font-size: 14px; }
QMessageBox QPushButton, QInputDialog QPushButton, QFileDialog QPushButton { min-width: 88px; padding: 8px 16px; }
QFileDialog QListView, QFileDialog QTreeView { background: #FFFFFF; color: #1F1B16; }

QScrollBar:vertical { background: transparent; width: 10px; margin: 4px 2px; }
QScrollBar::handle:vertical { background: #C9C1B2; border-radius: 5px; min-height: 40px; }
QScrollBar::handle:vertical:hover { background: #A69E8D; }
QScrollBar:horizontal { background: transparent; height: 10px; margin: 2px 4px; }
QScrollBar::handle:horizontal { background: #C9C1B2; border-radius: 5px; min-width: 40px; }
QScrollBar::handle:horizontal:hover { background: #A69E8D; }
QScrollBar::add-line, QScrollBar::sub-line { width: 0; height: 0; }
QScrollBar::add-page, QScrollBar::sub-page { background: transparent; }
"""


class NotifyTestThread(QtCore.QThread):
    """알림 설정 테스트 — 네트워크 호출이 GUI 를 멈추지 않게 별도 스레드."""
    result = QtCore.pyqtSignal(dict)

    def __init__(self, parent, cfg):
        super().__init__(parent)
        self.cfg = cfg

    def run(self):
        from daangn.notify import run_test
        try:
            res = run_test(self.cfg.get("tg_token"), self.cfg.get("tg_chat"),
                           self.cfg.get("sheet_url"), self.cfg.get("sheet_cred"))
        except Exception as e:
            res = {"tg_ok": False, "tg_msg": f"{type(e).__name__}: {e}",
                   "sheet_ok": None, "sheet_msg": "테스트 실패"}
        self.result.emit(res)


class HealthCheckThread(QtCore.QThread):
    """프록시 풀 진단 — IP 당 1요청. 네트워크가 GUI 를 멈추지 않게 별도 스레드."""
    result = QtCore.pyqtSignal(dict)
    progress = QtCore.pyqtSignal(int, int)

    def __init__(self, parent, proxies):
        super().__init__(parent)
        self.proxies = list(proxies)

    def run(self):
        from daangn_ext.health import health_check
        try:
            res = health_check(self.proxies,
                               on_progress=lambda d, t: self.progress.emit(d, t))
        except Exception as e:
            res = {"error": f"{type(e).__name__}: {e}"}
        self.result.emit(res)


class _TokenRefreshThread(QtCore.QThread):
    """검색 직전 access 토큰 확보 — **GUI 스레드에서 돌리면 안 된다**.

    안에서 부르는 ld_autoharvest.harvest_all(nudge=True) 는 LDPlayer 함대를
    깨워 정품 앱이 갱신한 토큰을 읽어오는 일이다. 전국 검색 규모에서는 수십
    초가 걸리고, 그동안 GUI 스레드에 있으면 이벤트 루프가 멈춘다 → 창이 얼어
    리페인트조차 안 된다. 클라가 본 '검색을 눌러도 한참 아무 변화 없음'의
    실제 정체가 이것이다. 진행 문구를 아무리 찍어도 그려질 틈이 없으므로,
    이 스레드는 진행 표시의 장식이 아니라 전제다.

    워커 스레드에서 위젯을 만지면 Qt 는 조용히 죽는다 — 여기서는 시그널만
    쏘고, 화면 갱신·모달은 전부 GUI 쪽 슬롯이 한다.
    """
    log = QtCore.pyqtSignal(str)
    done = QtCore.pyqtSignal(object, str)   # (access_token | None, 오류문구)

    def __init__(self, parent, accounts="./accounts.json"):
        super().__init__(parent)
        self.setObjectName("TokenRefreshThread")
        self.accounts = accounts

    def run(self):
        import json as _json
        # 1) LDPlayer 온디바이스 수확 — 정품 앱이 갱신한 access 를 su 로 직접 읽어
        #    accounts.json 에 병합. HTTP refresh(api.kr.karrotmarket.com)는 WAF 403
        #    이라 이 경로가 실질 갱신책이다.
        try:
            import ld_autoharvest
            ld_autoharvest.harvest_all(
                self.accounts, nudge=True,
                log=lambda m: self.log.emit(str(m)))
        except Exception as e:
            self.log.emit(f"[수확 건너뜀] {str(e)[:60]}")
        # 2) accounts.json 에서 남은 수명이 가장 긴 access 반환
        try:
            from daangn_ext.token_manager import token_exp
            best = None
            with open(self.accounts, encoding="utf-8") as _f:
                _accs = _json.load(_f)
            for a in _accs:
                acc = a.get("access") or ""
                if acc and (best is None or token_exp(acc) > token_exp(best)):
                    best = acc
            if best:
                self.done.emit(best, "")
                return
        except Exception:
            pass
        # 3) 폴백: 기존 HTTP refresh (LDPlayer 없음/수확실패 시. WAF면 실패할 수 있음)
        try:
            from daangn_ext import TokenManager, AccountStore, bind_to_token_manager
            store = AccountStore(self.accounts)
            if not store.rows:
                self.done.emit(
                    None,
                    "계정 없음. LDPlayer 를 켜거나 '계정+프록시 추가'로 등록하세요.")
                return
            tm = TokenManager()
            bind_to_token_manager(store, tm)
            tm.refresh_all()
            accs = list(tm.accounts.values())
            self.done.emit(tm.ensure_safe(accs[0]) if accs else None, "")
        except Exception as e:
            self.done.emit(None, f"토큰 갱신 오류: {e}")


class MyProgressDialog(QProgressDialog):
    def closeEvent(self, a0: QCloseEvent):  # type: ignore
        if a0.spontaneous():
            a0.ignore()


class SortUserRoleItem(QtWidgets.QTableWidgetItem):
    def __lt__(self, other):
        return self.data(Qt.ItemDataRole.UserRole) < other.data(
            Qt.ItemDataRole.UserRole
        )


class MainWindow(QMainWindow):
    EXCEL_HEADER_ALIASES = {
        "지역": "area",
        "키워드": "keyword",
        "최소가격": "minimum",
        "최대가격": "maximum",
    }

    # 클라 요구(2026-09-01): 수동 검색과 '매물 감시+에뮬레이터'를 **별도 프로그램**으로
    # 쓰고 싶다. 코드를 둘로 복제하면 두 벌을 따로 고쳐야 하고, 실제로 갈라야 하는
    # 것은 화면과 백그라운드 기동뿐이다. 그래서 한 코드베이스에 실행 모드를 둔다.
    #
    # 상태는 파일로 공유된다. 수동 검색은 앱API 토큰이 있어야 하는데 그 토큰을
    # 만드는 건 감시 프로그램의 수확기다 — 둘 다 accounts.json 을 본다. 동시에
    # 쓰는 건 이미 프로세스 간 파일락이 막고 있다(ld_autoharvest._file_lock).
    # 3탭 합본(옛 'all') 모드는 **없다.** 합본 창은 수확·폴링·라우터를 같이
    # 소유해, 따로 띄운 매물 감시 창과 같은 keyword_routes.json 을 놓고 다툰다 —
    # 실서버 2026-09-02 에 엑셀 조건이 통째로 사라진 경로가 이것이다. 모드
    # 인자가 없거나 모르는 값이면 매물 감시로 뜬다(백그라운드 소유자는 하나).
    MODES = {
        "manual": {"tabs": ("manual",), "background": False,
                   "title": "수동 검색"},
        "watch":  {"tabs": ("rules", "results", "emul", "settings"),
                   "background": True, "title": "매물 감시"},
    }

    def __init__(self, mode="watch"):
        super().__init__()
        self.mode = mode if mode in self.MODES else "watch"
        self._mode_cfg = self.MODES[self.mode]
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self._init_state()
        self._setup_ui()

        self.controller = MainController(self)

        if not self.controller.is_ready():
            self.alert(f"DB open failed: {self.controller.db_error}")
            return

        self._setup_model()
        self._init_tree()
        self._connect_signals()

        self._load_proxy()
        self._setup_tabs()
        self.sb.showMessage("프로그램이 시작되었습니다.")

        if self._mode_cfg["title"]:
            self.setWindowTitle(self.windowTitle() + " — " + self._mode_cfg["title"])
        # 수동 검색 전용 모드는 수확기를 돌리지 않는다. 두 프로그램이 동시에
        # 함대를 깨우면 같은 인스턴스에 force-stop 이 겹치고, 수동 쪽은 토큰을
        # 소비만 하면 되지 만들 필요가 없다(감시 프로그램이 만든 것을 파일로 읽는다).
        if not self._mode_cfg["background"]:
            self._harvest_thread = None
            return
        # 백그라운드 자동 수확 — 토큰 항상 신선 유지(수동 불필요).
        self._harvest_thread = _HarvestThread(accounts="./accounts.json")
        self._harvest_thread.tick.connect(self._on_harvest_tick)
        self._harvest_thread.start()

    def _alog(self, msg):
        """운영 로그 한 줄 — 화면(매물 감시 탭)과 karrot_monitor.log 양쪽에 남긴다.

        무인 서버에서는 위젯을 볼 사람이 없다. 파일에 안 남기면 앱이 살아 있는지
        확인할 방법이 data/*.json mtime 뿐이 된다(실제로 그래서 토큰 만료를 늦게 봤다).
        시각 형식은 헤드리스 런타임 log() 와 같게 맞춘다 — 두 로그를 나란히 읽는다.
        """
        import time as _t
        # 위젯은 아직 안 만들어졌을 수 있다(_setup_ui 이전 초기화 경로).
        if hasattr(self, "alertLog"):
            try:
                self.alertLog.append(msg)
            except Exception:
                pass
        try:
            print(f"[{_t.strftime('%H:%M:%S')}] {msg}", flush=True)
        except Exception:
            pass

    def _on_harvest_tick(self, msg):
        import time as _t
        self._last_harvest_ts = _t.time()
        # 상태바는 3탭 공용이다 — 여기 찍으면 수동 검색 화면에도 자동수확 로그가
        # 따라다닌다. 이 메시지의 자리는 '매물 감시' 탭 로그다.
        self._alog(msg)
        # 상태 한 줄 즉시 갱신
        try:
            self._refresh_alert_health()
        except Exception:
            pass

    def _setup_tabs(self):
        """기존 수동 UI 를 탭으로 감싸고 자동 탭 추가. 스크롤로 어떤 창크기든 다 보이게."""
        self.takeCentralWidget()                    # 기존 중앙위젯 버림(위젯은 재사용)
        self.tabs = QtWidgets.QTabWidget(self)
        # 위젯은 **모드와 무관하게 전부 만든다**. 화면에 안 띄울 뿐이다 —
        # 여기저기서 서로의 위젯을 참조하고 있어 안 만들면 AttributeError 로
        # 흩어진다. 에뮬 탭은 생성 시 창을 붙이지 않으므로(탭 활성화 때 붙는다)
        # 안 보이는 채로 만들어 두는 비용이 없다.
        show = self._mode_cfg["tabs"]
        manual_w = self._scroll(self._build_manual_tab())
        # 검색 스윕 스레드 핸들 — 탭은 없어졌지만 엔진은 라우터가 부린다.
        self.auto_monitor = None
        self.feed_monitor = None
        # 마지막 피드 엔진 — 모니터를 버려도 '왜 멈췄나'는 남아야 백오프가 선다.
        self._feed_last_engine = None
        self._app_sweep_off_logged = False
        # 매물 감시는 조건·결과·설정 세 탭으로 나뉜다. 한 탭에 접이식 네 개로
        # 쌓았던 동안 설정은 고급 패널·알림 창·계정 창 세 곳에 흩어졌고,
        # 클라는 어디를 펴야 하는지 몰랐다. 위젯은 한 함수가 다 만든다.
        rules_w, results_w, settings_w = (
            self._scroll(p) for p in self._build_alert_tab())
        # 에뮬레이터는 실제 창을 끼워넣으므로 스크롤로 감싸지 않는다(크기 = 탭 영역).
        emul_w = self._build_emul_tab()
        for key, page, title in (("manual", manual_w, "수동 검색"),
                                 ("rules", rules_w, "조건"),
                                 ("results", results_w, "결과"),
                                 ("emul", emul_w, "에뮬레이터"),
                                 ("settings", settings_w, "설정")):
            if key in show:
                self.tabs.addTab(page, title)
        # 하루에 수십 번 보는 것은 결과다 — 조건은 첫 탭이지만 첫 화면은 아니다.
        if "results" in show:
            self.tabs.setCurrentWidget(results_w)
        self.tabs.currentChanged.connect(self._on_tab_changed)
        # 명품 브랜드 헤더(골드 워드마크) + 탭
        central = QtWidgets.QWidget()
        cl = QtWidgets.QVBoxLayout(central); cl.setContentsMargins(0, 0, 0, 0); cl.setSpacing(0)
        header = QtWidgets.QWidget(); header.setObjectName("brandBar")
        hl = QtWidgets.QHBoxLayout(header); hl.setContentsMargins(26, 15, 26, 13); hl.setSpacing(14)
        brand = QtWidgets.QLabel("❖  L U X E")
        brand.setStyleSheet("color:#8A6D1F; font-size:22px; font-weight:800; letter-spacing:5px;")
        sub = QtWidgets.QLabel("명품 실시간 모니터")
        sub.setStyleSheet("color:#6B6355; font-size:13px; letter-spacing:3px; padding-top:6px;")
        hl.addWidget(brand); hl.addWidget(sub); hl.addStretch(1)
        # 우측 상단: 지금 도는 판과 설치 시각. 서버에 붙은 사람이 로그를 뒤지지
        # 않고 어느 판인지 보게 — 배포 각인(data/deployed.json)을 읽는다.
        self.versionLabel = QtWidgets.QLabel()
        self.versionLabel.setObjectName("versionLabel")
        self.versionLabel.setAlignment(QtCore.Qt.AlignmentFlag.AlignRight
                                       | QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.versionLabel.setTextInteractionFlags(
            QtCore.Qt.TextInteractionFlag.TextSelectableByMouse)
        hl.addWidget(self.versionLabel)
        self._init_version_label()
        cl.addWidget(header); cl.addWidget(self.tabs, 1)
        self.setCentralWidget(central)
        self.resize(1200, 840)                      # 하단 위젯 안 잘리게 넉넉히
        self.setStyleSheet(APP_QSS)                  # 전역 스타일(앱 단위로도 건다)
        self._apply_card_shadows()

    # ── 판(버전) 표시 ──
    def _init_version_label(self):
        """배포 각인의 판·설치 시각만 적는다. '최신' 판정은 하지 않는다 —
        틀린 최신 표시가 없는 표시보다 나쁘다. 최신인지는 커밋 기록과 대조."""
        from daangn_ext.version import local_version, version_label
        v = local_version(".")
        self.versionLabel.setStyleSheet(
            "color:#8C8578; font-size:12px; letter-spacing:1px; padding-top:6px;")
        self.versionLabel.setText(version_label(v))
        self.versionLabel.setToolTip(
            f"실행 중인 판: {v.get('sha') or '미상'}\n"
            f"설치: {v.get('installed') or '-'} ({v.get('source') or '기록 없음'})")

    def on_emul_add_clicked(self):
        """계정 추가 — .ldbk 를 골라 새 인스턴스로 복원한다.

        복원은 파일이 GB 단위라 몇 분 걸린다. GUI 스레드에서 돌리면 창이 얼어
        사용자가 죽은 줄 알고 강제 종료한다 — 강제 종료는 인스턴스 설정을 잘라
        영구 고장을 만든 전례가 있다. 그래서 워커 스레드에서 돌리고 진행 상황을
        그대로 보여준다."""
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("계정 추가")
        dlg.resize(680, 460)
        v = QtWidgets.QVBoxLayout(dlg); v.setSpacing(10)
        v.addWidget(QtWidgets.QLabel(
            "계정 백업(.ldbk)을 새 에뮬레이터 인스턴스로 복원합니다.\n"
            "복원한 인스턴스가 곧 계정입니다 — 프로그램에 토큰을 직접 넣는 방법은"
            " 당근이 막아 두었습니다.", parent=dlg))

        form = QtWidgets.QFormLayout()
        pathEdit = QtWidgets.QLineEdit(dlg)
        pathEdit.setPlaceholderText("계정 백업 파일 (.ldbk)")
        pickBtn = QtWidgets.QPushButton("파일 선택", dlg)
        row = QtWidgets.QHBoxLayout(); row.addWidget(pathEdit, 1); row.addWidget(pickBtn)
        rowW = QtWidgets.QWidget(); rowW.setLayout(row)
        nameEdit = QtWidgets.QLineEdit(dlg)
        nameEdit.setPlaceholderText("이 계정을 알아볼 이름 (예: 강남1)")
        form.addRow("백업 파일", rowW)
        form.addRow("이름", nameEdit)
        v.addLayout(form)

        logView = QtWidgets.QPlainTextEdit(dlg)
        logView.setReadOnly(True)
        v.addWidget(logView, 1)

        btns = QtWidgets.QHBoxLayout()
        runBtn = QtWidgets.QPushButton("복원 시작", dlg); runBtn.setObjectName("startBtn")
        closeBtn = QtWidgets.QPushButton("닫기", dlg)
        btns.addStretch(1); btns.addWidget(closeBtn); btns.addWidget(runBtn)
        v.addLayout(btns)

        def pick():
            p, _ = QtWidgets.QFileDialog.getOpenFileName(
                dlg, "계정 백업 선택", "", "LDPlayer 백업 (*.ldbk)")
            if p:
                pathEdit.setText(p)
                if not nameEdit.text().strip():
                    import os as _os
                    nameEdit.setText(_os.path.splitext(_os.path.basename(p))[0])
        pickBtn.clicked.connect(pick)

        state = {"thread": None}

        def done(res, err):
            runBtn.setEnabled(True)
            if err:
                logView.appendPlainText("실패: " + err)
                QtWidgets.QMessageBox.warning(dlg, "실패", err)
                return
            msg = (f"인덱스 {res['index']} 로 복원했습니다.\n\n"
                   "다음 순서:\n"
                   "1) 원격 접속(RDP)한 상태에서 그 인스턴스를 켜세요.\n"
                   "   에뮬레이터는 화면이 붙어 있어야 켜집니다.\n"
                   "2) 앱이 로그인돼 있으면 잠시 후 계정 목록에 자동으로 나타납니다.\n"
                   "3) 매물 감시 → 계정+프록시 에서 프록시를 지정하세요.")
            if res.get("warnings"):
                msg += "\n\n확인 필요:\n· " + "\n· ".join(res["warnings"])
            QtWidgets.QMessageBox.information(dlg, "복원 완료", msg)
            self._emul_scan()

        def run():
            ldbk = pathEdit.text().strip()
            name = nameEdit.text().strip()
            if not ldbk or not name:
                QtWidgets.QMessageBox.warning(dlg, "확인", "백업 파일과 이름을 정하세요.")
                return
            runBtn.setEnabled(False)
            logView.appendPlainText("시작합니다. 파일이 커서 몇 분 걸립니다…")
            th = _AddInstanceThread(ldbk, name)
            state["thread"] = th                 # GC 로 스레드가 사라지지 않게 붙든다
            th.line.connect(logView.appendPlainText)
            th.done.connect(done)
            th.start()
        runBtn.clicked.connect(run)
        closeBtn.clicked.connect(dlg.accept)
        dlg.exec()

    def _apply_card_shadows(self):
        """그룹박스 카드에 소프트 그림자(토스 입체감). QSS box-shadow 대체."""
        from PyQt6.QtWidgets import QGraphicsDropShadowEffect
        from PyQt6.QtGui import QColor
        for gb in self.findChildren(QtWidgets.QGroupBox):
            if gb.objectName() == "sectionBox":
                continue          # 섹션은 카드가 아니다 — 그림자가 붙으면 유령 상자가 된다
            eff = QGraphicsDropShadowEffect(gb)
            eff.setBlurRadius(28); eff.setOffset(0, 6)
            eff.setColor(QColor(0, 0, 0, 38))
            gb.setGraphicsEffect(eff)

    def _scroll(self, inner):
        sa = QtWidgets.QScrollArea()
        sa.setWidgetResizable(True)
        sa.setWidget(inner)
        return sa

    # ── 에뮬레이터 탭 (LDPlayer 인스턴스를 한 화면에서) ────────────────
    # LDPlayer 는 인스턴스마다 독립 창을 띄우고 탭 UI 가 없다. 계정이 늘면
    # 창이 화면을 뒤덮는다. 그래서 LDPlayer 창은 앱 바깥에 절대 두지 않는다:
    #   1) 스캔이 찾은 실행 중 인스턴스는 전부 오른쪽 탭으로 부착한다
    #   2) 부착 못 한 창(실패·틈새)은 화면 밖으로 치우고 다음 스캔에서 재시도
    #   3) 왼쪽 카드는 탭으로 가는 목차
    # 대가: 앱이 비정상 종료하면 자식이 된 인스턴스 창이 같이 파괴된다.
    # 정상 종료는 _emul_shutdown 이 창을 전부 되돌린다.
    EMUL_THUMB_BATCH = 6         # 한 틱에 캡처할 인스턴스 수(라운드로빈)

    def _build_emul_tab(self):
        import ldwin
        self._ldwin = ldwin
        self._emul = ldwin.Embedder()
        self._emul_hosts = {}            # ld index -> _EmbedHost (부착됨)
        self._emul_cards = {}            # ld index -> _InstanceCard
        self._emul_live = {}             # ld index -> 마지막 스캔 행
        self._emul_order = []            # 부착 순서(초과 시 가장 오래된 것부터 뗌)
        self._emul_thread = None
        self._emul_thumb_thread = None
        self._emul_thumb_cursor = 0
        self._emul_rescued = False
        self._emul_closing = False
        self._emul_seen = set()          # 감시가 이미 치운 창(hwnd)
        self._emul_rescan = False        # 스캔 도중 새 창이 떴다 → 끝나면 바로 한 번 더
        self._emul_console = ldwin.find_console()

        # 회색 바탕 위 제목 줄 + 카드 둘. 맨 위 줄이 제목·상태·버튼을 한 번에
        # 말한다 — 예전엔 흰 바탕에 라벨과 버튼이 흩어져 어디가 시작인지 없었다.
        w = QtWidgets.QWidget()
        w.setObjectName("emulPage")
        v = QtWidgets.QVBoxLayout(w)
        v.setContentsMargins(20, 18, 20, 18); v.setSpacing(12)

        bar = QtWidgets.QHBoxLayout(); bar.setSpacing(10)
        _tt = QtWidgets.QVBoxLayout(); _tt.setSpacing(2)
        _t = QtWidgets.QLabel("에뮬레이터"); _t.setObjectName("stepTitle")
        self.emulStatus = QtWidgets.QLabel("확인 중…")
        self.emulStatus.setObjectName("stepSub")
        _tt.addWidget(_t); _tt.addWidget(self.emulStatus)
        self.emulRefreshBtn = QtWidgets.QPushButton("새로고침", w)
        self.emulRefreshBtn.setObjectName("ghostBtn")
        self.emulRefreshBtn.clicked.connect(lambda: self._emul_scan())
        # 계정을 늘리는 유일한 경로다. '계정+프록시' 의 refresh 토큰 추가는 당근
        # WAF 이전 경로라 지금은 동작하지 않는다 — 토큰은 에뮬 안 앱에서만 나온다.
        self.emulAddBtn = QtWidgets.QPushButton("계정 추가(.ldbk 복원)", w)
        self.emulAddBtn.setObjectName("startBtn")
        self.emulAddBtn.setToolTip(
            "계정 백업(.ldbk)을 새 에뮬레이터 인스턴스로 복원합니다.\n"
            "복원한 인스턴스가 곧 계정입니다.")
        self.emulAddBtn.clicked.connect(self.on_emul_add_clicked)
        for b in (self.emulRefreshBtn, self.emulAddBtn):
            b.setSizePolicy(QtWidgets.QSizePolicy.Policy.Fixed,
                            QtWidgets.QSizePolicy.Policy.Fixed)
        bar.addLayout(_tt, 1)
        bar.addWidget(self.emulRefreshBtn, 0, Qt.AlignmentFlag.AlignVCenter)
        bar.addWidget(self.emulAddBtn, 0, Qt.AlignmentFlag.AlignVCenter)
        v.addLayout(bar)

        split = QtWidgets.QSplitter(Qt.Orientation.Horizontal, w)

        # 좌: 인스턴스 썸네일 그리드 (카드)
        left, lv = self._step_card(None, "인스턴스", "클릭하면 그 탭으로 이동합니다.")
        gridHost = QtWidgets.QWidget()
        gridHost.setStyleSheet("background: transparent;")
        self.emulGrid = QtWidgets.QGridLayout(gridHost)
        self.emulGrid.setContentsMargins(0, 0, 0, 0); self.emulGrid.setSpacing(10)
        self.emulGrid.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
        self.emulEmpty = QtWidgets.QLabel(
            "실행 중인 LDPlayer 인스턴스가 없습니다.\n"
            "자동 수확/모니터가 인스턴스를 부팅하면 여기 나타납니다.")
        self.emulEmpty.setWordWrap(True)
        self.emulEmpty.setObjectName("mutedNote")
        self.emulGrid.addWidget(self.emulEmpty, 0, 0, 1, 2)
        sa = QtWidgets.QScrollArea(); sa.setWidgetResizable(True); sa.setWidget(gridHost)
        sa.setFrameShape(QtWidgets.QFrame.Shape.NoFrame)
        lv.addWidget(sa, 1)
        split.addWidget(left)

        # 우: 부착된 인스턴스 탭 (카드)
        right, rv = self._step_card(None, "화면", "")
        self.emulTabs = QtWidgets.QTabWidget(right)
        self.emulTabs.setDocumentMode(True)
        self.emulTabs.setTabsClosable(False)   # 닫으면 창이 바탕화면으로 나간다
        self.emulHint = QtWidgets.QLabel(
            "실행 중인 인스턴스 화면이 여기 탭으로 자동으로 붙습니다.")
        self.emulHint.setWordWrap(True)
        self.emulHint.setObjectName("mutedNote")
        rv.addWidget(self.emulHint)
        rv.addWidget(self.emulTabs, 1)
        split.addWidget(right)

        split.setStretchFactor(0, 0); split.setStretchFactor(1, 1)
        split.setSizes([360, 820])
        v.addWidget(split, 1)

        # 목록 스캔은 다른 탭에 있어도 계속 — 새로 뜬 창을 바로 치워야 하니까.
        self._emul_timer = QtCore.QTimer(self)
        self._emul_timer.setInterval(8000)
        self._emul_timer.timeout.connect(lambda: self._emul_scan())
        # 새 창 감시 — list2 스캔(8s)만으로는 launch 뒤 창이 몇 초씩 바탕화면에
        # 보인다. EnumWindows 는 싸므로 자주 돌려 뜨자마자 치우고 스캔을 앞당긴다.
        self._emul_watch_timer = QtCore.QTimer(self)
        self._emul_watch_timer.setInterval(150)
        self._emul_watch_timer.timeout.connect(self._emul_watch_tick)
        # 썸네일은 이 탭을 보고 있을 때만.
        self._emul_thumb_timer = QtCore.QTimer(self)
        self._emul_thumb_timer.setInterval(3000)
        self._emul_thumb_timer.timeout.connect(self._emul_thumb_tick)

        if not ldwin.IS_WINDOWS:
            self.emulStatus.setText("Windows 전용 기능입니다.")
            self.emulRefreshBtn.setEnabled(False)
        elif not self._emul_console:
            self.emulStatus.setText("ldconsole.exe 를 못 찾음 — LDPlayer 설치경로 확인")
        else:
            self._emul_timer.start()
            self._emul_watch_timer.start()
            QtCore.QTimer.singleShot(0, lambda: self._emul_scan())
        return w

    def _on_tab_changed(self, i):
        """에뮬레이터 탭을 보고 있을 때만 썸네일을 캡처한다(불필요한 부하 방지)."""
        if not hasattr(self, "emulTabs"):
            return
        if self.tabs.tabText(i) == "에뮬레이터":
            self._emul_scan()
            self._emul_thumb_tick()
            self._emul_thumb_timer.start()
        else:
            self._emul_thumb_timer.stop()

    # -- 스캔 ---------------------------------------------------------------
    def _emul_scan(self):
        if not getattr(self, "_emul_console", None) or not self._ldwin.IS_WINDOWS:
            return
        if self._emul_thread is not None and self._emul_thread.isRunning():
            return
        t = _LdListThread(self._emul_console)
        t.rows.connect(self._emul_apply)
        self._emul_thread = t
        t.start()

    def _emul_apply(self, rows):
        """스캔 반영 — 카드 동기화, 창 치우기 정책 적용, 죽은 인스턴스 정리."""
        if getattr(self, "_emul_closing", True):   # 종료 중 도착한 신호는 무시
            return
        ld = self._ldwin
        attached = {idx: host.child_hwnd for idx, host in self._emul_hosts.items()}
        live, detach = emul_reconcile(rows, attached, ld.is_window)

        if not self._emul_rescued:
            # 지난 세션이 크래시로 죽었으면 화면 밖에 창이 남아 있다 — 되살린다.
            self._emul_rescued = True
            n = sum(1 for r in live.values() if ld.rescue(r["top_hwnd"]))
            if n:
                print(f"[에뮬레이터] 이전 세션이 남긴 창 {n}개 복구")
        self._emul_live = live

        changed = False
        for idx in detach:                   # 부착 창이 정말 죽은 경우만
            self._emul_detach(idx)
        for idx in list(self._emul_cards):
            if idx not in live:
                card = self._emul_cards.pop(idx)
                card.setParent(None); card.deleteLater()
                changed = True
        for idx in sorted(live):
            if idx in self._emul_cards:
                continue
            r = live[idx]
            card = _InstanceCard(idx, r.get("title") or r["name"] or f"인스턴스 {idx}")
            card.clicked.connect(self._emul_card_clicked)
            self._emul_cards[idx] = card
            changed = True
        if changed:
            self._emul_relayout()

        # 실행 중 인스턴스는 전부 탭으로. 못 붙인 창은 바탕화면 대신 화면 밖으로.
        for idx in sorted(live):
            if idx not in self._emul_hosts:
                self._emul_attach(idx)
        for idx, r in live.items():
            if idx not in self._emul_hosts:
                self._emul.stow(r["top_hwnd"])
        self._emul.prune()
        self._emul_seen = {h for h in self._emul_seen if ld.is_window(h)}
        self._emul_sync_states()
        if self._emul_rescan:            # 스캔 도중 뜬 창 — 8초 기다리지 않는다
            self._emul_rescan = False
            QtCore.QTimer.singleShot(0, lambda: self._emul_scan())

    def _emul_watch_tick(self):
        """새 LDPlayer 창 감시 — 뜨자마자 화면 밖으로 치우고 스캔을 앞당긴다.

        부팅 중인 창은 list2 가 아직 index 를 못 주므로 hwnd 만으로 치운다.
        LDPlayer 가 부팅 중 창 위치를 제자리로 되돌리는 경우도 다시 치운다.
        첫 스캔의 잔재 복구(rescue)보다 먼저 돌면 우리가 치운 창을 되살려
        버리므로 복구가 끝난 뒤에만 일한다.
        """
        if self._emul_closing or not self._emul_rescued:
            return
        attached = {h.child_hwnd for h in self._emul_hosts.values()}
        fresh = False
        for win in self._ldwin.player_windows():
            hw = win["hwnd"]
            if hw in attached:
                continue
            if hw not in self._emul_seen:
                fresh = True
                self._emul_seen.add(hw)
            self._emul.stow(hw)
        if not fresh:
            return
        if self._emul_thread is not None and self._emul_thread.isRunning():
            self._emul_rescan = True
        else:
            self._emul_scan()

    def _emul_relayout(self):
        """카드 그리드 재배치 — 2열 고정(카드 폭이 고정이라 열 수도 고정)."""
        while self.emulGrid.count():
            self.emulGrid.takeAt(0)
        if not self._emul_cards:
            self.emulEmpty.setVisible(True)
            self.emulGrid.addWidget(self.emulEmpty, 0, 0, 1, 2)
            return
        self.emulEmpty.setVisible(False)
        for n, idx in enumerate(sorted(self._emul_cards)):
            card = self._emul_cards[idx]
            self.emulGrid.addWidget(card, n // 2, n % 2)
            card.setVisible(True)

    def _emul_sync_states(self):
        for idx, card in self._emul_cards.items():
            if idx in self._emul_hosts:
                card.set_state("탭에 붙음", True)
            else:
                card.set_state("부착 대기 · 창 치움", False)
        self.emulHint.setVisible(not self._emul_hosts)
        self.emulStatus.setText(
            f"인스턴스 {len(self._emul_live)}개 · 탭 {len(self._emul_hosts)}개")

    # -- 부착/분리 -----------------------------------------------------------
    def _emul_card_clicked(self, idx):
        host = self._emul_hosts.get(idx)
        if host is not None:
            self.emulTabs.setCurrentWidget(host)
            return
        self._emul_attach(idx)               # 아직 못 붙은 창이면 지금 재시도

    def _emul_attach(self, idx):
        r = self._emul_live.get(idx)
        if r is None or idx in self._emul_hosts:
            return

        host = _EmbedHost(self._emul, self.emulTabs)
        pos = self.emulTabs.addTab(host, r.get("title") or r["name"]
                                   or f"인스턴스 {idx}")
        host.child_hwnd = r["top_hwnd"]
        ok = self._emul.embed(r["top_hwnd"], host.host_hwnd(),
                              max(host.width(), 1), max(host.height(), 1))
        if not ok:
            host.child_hwnd = 0
            self.emulTabs.removeTab(pos)
            host.deleteLater()
            self._emul.stow(r["top_hwnd"])   # 바탕화면에 두지 않는다 — 다음 스캔에서 재시도
            self.emulStatus.setText(f"인스턴스 {idx} 화면 부착 실패 — 다음 스캔에서 재시도")
            return
        self._emul_hosts[idx] = host
        self._emul_order.append(idx)
        if self.emulTabs.count() == 1:
            self.emulTabs.setCurrentIndex(pos)   # 첫 탭만 — 보던 탭을 뺏지 않는다
        self._emul_sync_states()

    def _emul_detach(self, idx):
        """탭을 걷고 창을 top-level 로 되돌린 뒤 바로 화면 밖으로 치운다.
        (살아 있는 창이 바탕화면에 나타나는 순간이 없어야 한다)"""
        host = self._emul_hosts.pop(idx, None)
        if idx in self._emul_order:
            self._emul_order.remove(idx)
        if host is None:
            return
        hwnd, host.child_hwnd = host.child_hwnd, 0
        self._emul.release(hwnd)
        pos = self.emulTabs.indexOf(host)
        if pos >= 0:
            self.emulTabs.removeTab(pos)
        host.deleteLater()
        self._emul.stow(hwnd)
        self._emul_sync_states()

    # -- 썸네일 --------------------------------------------------------------
    def _emul_thumb_tick(self):
        """부착 안 된 인스턴스를 라운드로빈으로 캡처. 100개여도 한 틱 부하는 고정."""
        if self._emul_thumb_thread is not None and self._emul_thumb_thread.isRunning():
            return
        pending = [i for i in sorted(self._emul_cards) if i not in self._emul_hosts]
        if not pending:
            return
        n = len(pending)
        start = self._emul_thumb_cursor % n
        take = [pending[(start + k) % n] for k in range(min(self.EMUL_THUMB_BATCH, n))]
        self._emul_thumb_cursor = (start + len(take)) % n

        targets = [(i, self._emul_live[i]["top_hwnd"]) for i in take
                   if i in self._emul_live]
        if not targets:
            return
        t = _ThumbCaptureThread(targets)
        t.shot.connect(self._emul_thumb_ready)
        self._emul_thumb_thread = t
        t.start()

    def _emul_thumb_ready(self, payload):
        if self._emul_closing:
            return
        idx, shot = payload
        card = self._emul_cards.get(idx)
        if card is None:
            return
        if not shot:
            card.set_frame(None)
            return
        w, h, raw = shot
        img = QtGui.QImage(raw, w, h, w * 4,
                           QtGui.QImage.Format.Format_RGB32).copy()
        card.set_frame(QtGui.QPixmap.fromImage(img))

    # -- 종료 ---------------------------------------------------------------
    def _emul_shutdown(self):
        """붙여둔 창은 원래대로, 치워둔 창은 화면으로. 안 하면 부모와 함께
        파괴되거나 화면 밖에 남아 사용자가 찾을 수 없게 된다."""
        self._emul_closing = True        # 늦게 도착하는 워커 신호를 전부 무시
        for name in ("_emul_timer", "_emul_watch_timer", "_emul_thumb_timer"):
            t = getattr(self, name, None)
            if t is not None:
                t.stop()
        # 워커부터 세운 뒤 상태를 비운다(정리된 위젯을 늦은 신호가 건드리지 않게).
        for name in ("_emul_thread", "_emul_thumb_thread"):
            t = getattr(self, name, None)
            if t is not None and t.isRunning() and not t.wait(2000):
                t.terminate(); t.wait(500)
        self._emul_hosts = {}
        self._emul_order = []
        self._emul_cards = {}
        self._emul_live = {}
        emb = getattr(self, "_emul", None)
        if emb is not None:
            try:
                emb.release_all()
                emb.unstow_all()
            except Exception as e:
                print(f"[에뮬레이터] 창 복구 실패: {type(e).__name__}: {e}")

    # ── 키워드 알림 탭 ─────────────────────────────────────────────
    # ── 상태칩 → 대응 항목 ──
    # 칩은 값을 보여줄 뿐 조절은 조건·설정 탭이 한다. 사용자가 "커버리지 78%" 를
    # 보고 그 설정으로 가는 길이 없으면 칩은 죽은 글자다(설계 §1: 칩 클릭은 해당
    # 고급 패널 항목으로 스크롤한다). 대응 항목이 없는 칩은 여기 넣지 않는다 —
    # 없는 목적지를 지어내면 사용자가 엉뚱한 곳으로 간다.
    # 상태 한 줄의 조각 순서. 값은 _set_status 가 채운다.
    STATUS_ORDER = ("token", "accounts", "coverage", "poll", "rules", "feed", "notify")

    TAB_HELP = ("키워드 알림을 계정에 등록 → 매물 뜨면 토큰폴링으로 실시간 수신.\n"
                "1계정 = 인증동네 + 인접 지역 커버. 여러 계정(다른 동네) = 전국.")

    def _refresh_rules_view(self):
        """조건표 요약과 표를 파일 상태에 맞춘다.

        파일이 밖에서 바뀌어도(서버에서 교체, 전체 삭제) 캐시가 mtime 으로
        잡는다. 화면은 그 뒤에 따라와야 하므로 폴링 틱마다 한 번 부른다.
        표는 파일이 바뀐 뒤 한 번만 다시 채우고, 사용자가 고치는 중이면
        절대 덮지 않는다 — 적고 있는 줄이 사라지는 것이 최악이다."""
        from daangn_ext.alert_rules import brands
        table = self._alert_rules.get()
        rules = table.rules
        stamp = self._alert_rules.stamp()
        if not self._rules_dirty and stamp != self._grid_seen_stamp:
            from daangn_ext.rule_grid import rules_to_grid
            self.rulesGrid.set_cells(rules_to_grid(rules))
            self._grid_seen_stamp = stamp
        # 섹션 제목이 '조건 없음'을 이미 말한다 — 본문은 다음 할 일을 말한다.
        text = (table.detail() if rules
                else "아직 조건이 없습니다. 표에 적고 [조건 적용]을 누르세요.")
        if self._rules_dirty:
            text += "    ·    표를 고쳤습니다 — 미적용, [조건 적용]을 누르세요"
        self.rulesSummary.setText(text)
        self.rulesApplyBtn.setText("조건 적용 (수정됨)" if self._rules_dirty else "조건 적용")
        self._set_status("rules",
                       f"조건 {len(rules)}" if rules else "조건 없음",
                       "ok" if rules else "off")
        # 접혀 있어도 무엇이 걸려 있는지는 보여야 한다 — 섹션 제목이 그 자리다.
        box = getattr(self, "condBox", None)
        if box is not None:
            n_b = len(brands(rules)) if rules else 0
            box._baseTitle = (f"조건 {len(rules)}개 · 브랜드 {n_b}개"
                              if rules else "조건 없음 — 엑셀을 넣기 전까지 알리지 않습니다")
            self._sync_box_visible(box, box.isChecked())

    def _sync_box_visible(self, box, on):
        """체크형 QGroupBox 는 체크를 풀어도 자식을 '비활성'으로만 만든다.
        접으려면 직접 숨겨야 한다. qt_ 접두 내부 위젯은 건드리지 않는다."""
        # qt_ 로 시작하는 이름은 Qt 내부 위젯(콤보 팝업 스크롤 컨테이너 등)이라
        # 건드리지 않는다 — 표시 여부를 Qt 가 스스로 관리한다.
        on = bool(on)

        def _kept_hidden(wdg):
            """조건표로 옮겨간 옛 입력들 — 살려는 두되 화면에는 안 낸다."""
            while wdg is not None and wdg is not box:
                if getattr(wdg, "_keepHidden", False):
                    return True
                wdg = wdg.parentWidget()
            return False

        for c in box.findChildren(QtWidgets.QWidget):
            if c.objectName().startswith("qt_") or _kept_hidden(c):
                continue
            # 표·트리의 머리글은 뷰가 스스로 관리한다(setHeaderHidden). 여기서
            # 켜 버리면 숨긴 머리글이 되살아난다.
            if isinstance(c, QtWidgets.QHeaderView):
                continue
            c.setVisible(on)
        # 체크박스 대신 화살표로 접힘을 알린다 — 설정을 켜고 끄는 것이 아니라
        # 펼치고 접는 것이므로 체크 표시는 뜻이 어긋난다.
        if box.objectName() == "sectionBox":
            base = getattr(box, "_baseTitle", None)
            if base is None:
                base = box.title().lstrip("▸▾ ").strip()
                box._baseTitle = base
            box.setTitle((("▾  " if on else "▸  ") + base) if base else "")
        # 자식을 숨겨도 레이아웃 여백은 남는다 — 접은 그룹이 빈 상자로 화면을
        # 먹는다. 접히면 제목 줄 높이까지 줄인다.
        lay = box.layout()
        if lay is not None:
            if on:
                lay.setContentsMargins(*getattr(box, "_openMargins", (9, 9, 9, 9)))
            else:
                if not hasattr(box, "_openMargins"):
                    m = lay.contentsMargins()
                    box._openMargins = (m.left(), m.top(), m.right(), m.bottom())
                lay.setContentsMargins(0, 0, 0, 0)
        box.setMaximumHeight(16777215 if on
                             else box.fontMetrics().height() + 24)

    def _step_card(self, n, title, sub):
        """단계 카드 한 장 — 번호 배지 + 제목 + 한 줄 설명. 안에 넣을 레이아웃을
        같이 돌려준다. 카드는 순서를 말하고, 사용자는 위에서 아래로 읽는다.
        n=None 이면 배지 없는 일반 카드(설정 탭처럼 순서가 없는 곳)."""
        card = QtWidgets.QFrame()
        card.setObjectName("stepCard")
        cl = QtWidgets.QVBoxLayout(card)
        cl.setContentsMargins(22, 20, 22, 20); cl.setSpacing(12)
        head = QtWidgets.QHBoxLayout(); head.setSpacing(12)
        if n is not None:
            badge = QtWidgets.QLabel(str(n)); badge.setObjectName("stepBadge")
            head.addWidget(badge, 0, QtCore.Qt.AlignmentFlag.AlignTop)
        tv = QtWidgets.QVBoxLayout(); tv.setSpacing(2)
        t = QtWidgets.QLabel(title); t.setObjectName("stepTitle")
        tv.addWidget(t)
        if sub:
            d = QtWidgets.QLabel(sub); d.setObjectName("stepSub"); d.setWordWrap(True)
            tv.addWidget(d)
        head.addLayout(tv, 1)
        cl.addLayout(head)
        return card, cl

    def _setting_row(self, label, widget, last=False):
        """설정 카드의 목록 행 — 왼쪽 라벨(고정폭), 오른쪽 입력. 행 아래 얇은 선.
        QFormLayout 은 라벨을 입력 높이에 맞춰 띄우고 선이 없어 어디까지가 한
        행인지 안 보였다."""
        row = QtWidgets.QFrame()
        row.setObjectName("settingRow")
        if last:
            row.setProperty("last", "true")
        h = QtWidgets.QHBoxLayout(row)
        h.setContentsMargins(0, 6, 0, 6); h.setSpacing(12)
        lab = QtWidgets.QLabel(label, row); lab.setObjectName("rowLabel")
        lab.setFixedWidth(96)
        h.addWidget(lab, 0, QtCore.Qt.AlignmentFlag.AlignVCenter)
        h.addWidget(widget, 1)
        return row

    def _collapsible(self, title, inner, checked=False, first=False):
        """접이식 그룹 한 덩어리. 자주 안 보는 것을 화면에서 치우되 지우지는
        않는다 — 안의 위젯은 그대로 살아 있어 갱신 코드가 안 바뀐다.
        first=True 는 카드 머리글 바로 아래 오는 섹션 — 위 구분선을 뺀다."""
        box = QtWidgets.QGroupBox(title)
        box.setObjectName("sectionBox")
        if first:
            box.setProperty("firstInCard", "true" if title else "bare")
        box.setCheckable(True)
        box.setChecked(bool(checked))
        box.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        lay = QtWidgets.QVBoxLayout(box)
        if isinstance(inner, QtWidgets.QLayout):
            lay.addLayout(inner)
        else:
            lay.addWidget(inner)
        box.toggled.connect(lambda on, b=box: self._sync_box_visible(b, on))
        self._sync_box_visible(box, box.isChecked())
        return box

    def _build_alert_tab(self):
        """매물 감시 위젯을 전부 만들고 (조건, 결과, 설정) 세 페이지로 돌려준다.

        위젯을 만드는 순서와 보이는 순서를 갈라 둔 이유가 있다. 여기저기서
        서로의 위젯을 참조하므로 생성은 한 함수에 두고, 어느 탭에 놓을지는
        맨 아래 조립부가 한 번에 정한다. 한 탭에 접이식 네 개로 쌓았던
        동안 클라는 하루에 수십 번 보는 매물 표 아래 접힌 것을 못 찾았다."""
        w = QtWidgets.QWidget()               # 결과 페이지 — 매물 표의 집
        w.setObjectName("resultsPage")        # 토스 결(QSS) 은 이 이름 아래서만 산다
        v = QtWidgets.QVBoxLayout(w); v.setContentsMargins(24, 20, 24, 20); v.setSpacing(16)
        cond_v = QtWidgets.QVBoxLayout()      # '감시 조건' 에 들어갈 것
        listing_v = QtWidgets.QVBoxLayout()   # 매물 — 화면의 주인공

        # 전국 커버 %. 감시 시작이 계산해 넣고 상태 한 줄이 보여 준다.
        self._cover_pct = 0

        self._watch_label = QtWidgets.QLabel("추적 중 0건")
        self._watch_label.setObjectName("watchCount")

        # ── 컨트롤 바: 토글 하나 + 상태 한 줄 ──
        # 결과 탭에서 누를 것은 감시 시작뿐이다. 상태칩 다섯 개가 버튼처럼
        # 서 있던 동안 클라는 눌러 봤고, 둘은 아무 데도 안 데려갔다. 현황
        # 상자는 같은 값을 한 번 더 보였다. 상태는 읽는 것이다 — 라벨 한 줄.
        top = QtWidgets.QHBoxLayout(); top.setSpacing(10)
        self.watchToggleBtn = QtWidgets.QPushButton("▶ 감시 시작")
        self.watchToggleBtn.setObjectName("startBtn")
        self.watchToggleBtn.setCheckable(True)
        self.watchToggleBtn.setToolTip(self.TAB_HELP)
        self.watchToggleBtn.clicked.connect(self.on_watch_toggle)
        top.addWidget(self.watchToggleBtn)
        self._status = {k: ("", "off") for k in self.STATUS_ORDER}
        self.statusLine = QtWidgets.QLabel("")
        self.statusLine.setObjectName("statusLine")
        self.statusLine.setTextFormat(QtCore.Qt.TextFormat.RichText)
        top.addWidget(self.statusLine)
        top.addSpacing(8)
        top.addWidget(self._watch_label, 1)

        # ── 감시 조건: 표에 바로 적는다 ──
        # 등록 경로는 이 표 하나다. 예전에는 '명품20 전계정등록' 버튼과 수동
        # 키워드 폼이 따로 있어 브랜드만 등록해 놓고 조건표를 안 넣는 상태가
        # 만들어졌다 — 그러면 모델·가격대와 무관하게 브랜드 전 매물이 알림으로
        # 쏟아진다. 조건표와 어긋나는 뒷문은 두지 않는다.
        # 그 다음엔 엑셀 파일이 원본이었다. 다섯 줄 고치는 데 파일 열기·고치기·
        # 저장·[다시 읽기] 네 단계였고, 파일을 옮기면 [열기]가 죽었다. 표를
        # 여기 두고 엑셀처럼 굴게 한다 — 엑셀은 표를 채우는 보조 경로로 남는다.
        self.rulesHint = QtWidgets.QLabel(
            "브랜드만 필수. 제품명을 비우면 그 브랜드 전체, 띄어쓰기는 무시합니다"
            " — '오버 더 문'은 '오버더문'도 잡습니다. 제외는 쉼표로 구분."
            " 엑셀에서 복사해 Ctrl+V 로 붙여 넣어도 됩니다.")
        self.rulesHint.setObjectName("mutedNote")
        self.rulesHint.setWordWrap(True)
        cond_v.addWidget(self.rulesHint)
        self.rulesGrid = RuleGrid(w)
        self.rulesGrid.setMinimumHeight(240)
        cond_v.addWidget(self.rulesGrid)
        self.rulesApplyBtn = QtWidgets.QPushButton("조건 적용")
        self.rulesApplyBtn.setObjectName("startBtn")
        self.rulesApplyBtn.setToolTip(
            "표의 조건을 저장하고 브랜드를 등록합니다 — 표에 없는 브랜드는 해제됩니다")
        self.rulesAddRowBtn = QtWidgets.QPushButton("줄 추가")
        self.rulesAddRowBtn.setObjectName("linkBtn")
        self.rulesAddRowBtn.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.rulesDelRowBtn = QtWidgets.QPushButton("선택 줄 삭제")
        self.rulesDelRowBtn.setObjectName("linkBtn")
        self.rulesDelRowBtn.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.rulesImportBtn = QtWidgets.QPushButton("엑셀 불러오기")
        self.rulesImportBtn.setObjectName("linkBtn")
        self.rulesImportBtn.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.rulesImportBtn.setToolTip(
            "엑셀 파일(브랜드·제품명·최소가격·최대가격·제외)을 읽어 표를 채웁니다."
            " 적용은 [조건 적용]을 눌러야 됩니다")
        _rb = QtWidgets.QHBoxLayout(); _rb.setSpacing(4)
        _rb.addWidget(self.rulesApplyBtn)
        _rb.addSpacing(6)
        _rb.addWidget(self.rulesAddRowBtn)
        _rb.addWidget(self.rulesDelRowBtn)
        _rb.addStretch(1)
        _rb.addWidget(self.rulesImportBtn)
        cond_v.addLayout(_rb)
        self.rulesSummary = QtWidgets.QLabel("조건 없음")
        self.rulesSummary.setObjectName("mutedNote")
        self.rulesSummary.setWordWrap(True)
        cond_v.addWidget(self.rulesSummary)

        # 진행 표시 — 등록은 '계정 수 × 키워드 수' 만큼 요청이라 20초 넘게 걸린다.
        # 그동안 화면이 아무 말도 안 하면 사용자는 실패한 줄 알고 다시 누르거나
        # '전체 삭제'로 손을 뻗는다(실서버 2026-09-02 에 등록 21건이 그렇게 날아갔다).
        self.alertBusyRow = QtWidgets.QWidget(w)
        _bh = QtWidgets.QHBoxLayout(self.alertBusyRow)
        _bh.setContentsMargins(0, 0, 0, 0)
        self.alertBusyBar = QtWidgets.QProgressBar(self.alertBusyRow)
        self.alertBusyBar.setRange(0, 0)          # 진행률을 모른다 — 무한 막대
        self.alertBusyBar.setMaximumWidth(160)
        self.alertBusyBar.setTextVisible(False)
        self.alertBusyLabel = QtWidgets.QLabel("", self.alertBusyRow)
        _bh.addWidget(self.alertBusyBar)
        _bh.addWidget(self.alertBusyLabel, 1)
        self.alertBusyRow.setVisible(False)

        # 커버 동네 정보
        self.alertSubLabel = QtWidgets.QLabel("동네 정보: (감시를 시작하면 채워집니다)")

        # 등록 목록
        self.alertTable = row_lines(QtWidgets.QTableWidget(0, len(ALERT_COLS), w))
        self.alertTable.setHorizontalHeaderLabels(ALERT_COLS)
        self.alertTable.verticalHeader().setVisible(False)
        self.alertTable.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.alertTable.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        # 열이 7개다. 나머지를 기본 폭(~100px)으로 두면 끌올·id 가 가로
        # 스크롤 뒤로 밀려, 방금 추가한 열이 안 보인다.
        _hh = self.alertTable.horizontalHeader()
        for _c in range(len(ALERT_COLS)):
            _hh.setSectionResizeMode(
                _c, QtWidgets.QHeaderView.ResizeMode.Stretch if _c == ALERT_COL_KEYWORD
                else QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.alertTable.setShowGrid(False)
        self.alertTable.setMinimumHeight(220)
        # id 는 서버 내부 식별자다 — 사람이 읽을 값이 아니다. 열은 남긴다
        # (인덱스 상수와 셀 생성 함수가 쓴다). 화면에서만 감춘다.
        self.alertTable.setColumnHidden(ALERT_COL_ID, True)

        # 삭제 경로는 '전체 삭제' 하나다. 낱개 삭제는 조건표와 등록을 어긋나게
        # 만들고, 표 바로 아래 있던 동안 실서버에서 등록 21건이 날아갔다
        # (2026-09-02). 설정 탭의 고급(접힘)에 둔다.
        self.alertDelAllBtn = QtWidgets.QPushButton("전체 삭제")

        # ── 화면에서 뺀 튜닝값(저장·복원이 계속 쓴다) ──
        # 폴링·커버 계산은 감시 시작이 스스로 한다. 계정 현황·프록시·진단은
        # 계정+프록시 창 안으로 옮겼다 — 버튼 여덟 개가 나란히 있던 동안
        # 클라는 어느 것을 눌러야 하는지 몰랐다.
        self.alertPollInterval = QtWidgets.QSpinBox(); self.alertPollInterval.setRange(30, 3600)
        self.alertPollInterval.setValue(120); self.alertPollInterval.setSuffix("초")
        self.alertCoverMode = QtWidgets.QComboBox()
        self.alertCoverMode.addItem("전국 풀커버", False)
        self.alertCoverMode.addItem("핵심지역 집중", True)
        self.alertCoverMode.setToolTip(
            "전국=모든 유효계정 사용 / 핵심지역=명품 밀집동네(강남·분당·해운대 등) 계정만 "
            "→ 적은 계정으로 거래량 대부분 커버. 등록·폴링·집계 전부 이 모드 따름.")
        # 무인 운영 스위치(실행 시 자동 폴링·크래시 자동복구·야간 감속)는
        # 체크박스가 아니라 기본값이다 — 서버 상주로 도는 앱에서 끌 이유가
        # 없고, 꺼 두면 감시가 조용히 멈춘다. 옛 alert_settings.json 값은 존중한다.
        # 부팅 자동실행은 설치본(install.ps1)이 LDPlayer 순차 기동 뒤에 앱을
        # 띄우는 경로로 등록한다 — 앱이 따로 등록하면 로그온 때 둘이 동시에
        # 인스턴스를 띄워 함대가 hang 한다.
        sweep = self._build_sweep_settings()

        # ── 조건 페이지: 조건표 → 훑을 지역 → 등록 상태 ──
        # 등록 상태는 클라가 넣은 조건이 아니라 시스템이 당근에 올린 결과다.
        # 조건과 같은 상자에 두면 둘을 같은 것으로 읽는다 — 상자를 가른다.
        # 두 단계 카드다: ① 조건(조건 표 → 등록된 조건 표) ② 지역. 접이식
        # 줄 세 개가 한 면에 늘어서 있던 동안 무엇이 먼저고 무엇이 필수인지
        # 안 보였다. 카드가 순서를, 배지가 단계를 말한다.
        rules_w = QtWidgets.QWidget()
        rules_w.setObjectName("rulesPage")
        rv = QtWidgets.QVBoxLayout(rules_w); rv.setContentsMargins(20, 18, 20, 18); rv.setSpacing(12)
        card1, c1 = self._step_card(
            1, "감시 조건",
            "표에 브랜드·제품명·가격대를 적고 [조건 적용]을 누르세요. 엑셀에서 복사해 붙여 넣어도 됩니다.")
        self.condBox = self._collapsible("감시 조건", cond_v, checked=True, first=True)
        c1.addWidget(self.condBox)
        reg_v = QtWidgets.QVBoxLayout()
        # 이 표는 조건표의 복사본이 아니라 '서버에 실제 걸렸나'를 보는 창이다.
        # 이름·설명이 그걸 말하지 않으면 엑셀이 있는데 왜 또 있냐는 질문이 온다.
        _rh = QtWidgets.QLabel(
            "조건표 ≠ 서버 등록. 빨간 <b>미등록</b>은 조건은 있지만 당근에 아직"
            " 안 올라간 키워드 — 앱 알림이 안 오는 첫 번째 이유입니다.")
        _rh.setObjectName("mutedNote")
        _rh.setWordWrap(True)
        reg_v.addWidget(_rh)
        self.alertSubLabel.setObjectName("mutedNote")
        reg_v.addWidget(self.alertSubLabel)
        reg_v.addWidget(self.alertTable, 1)
        self.regBox = self._collapsible("등록된 감시 조건 — 당근 서버 등록 상태",
                                        reg_v, checked=True)
        c1.addWidget(self.regBox, 1)
        rv.addWidget(card1)

        card2, c2 = self._step_card(
            2, "지역 선택",
            "고른 지역만 훑습니다. 아무것도 안 고르면 서울·경기를 훑습니다.")
        # 카드 제목이 곧 섹션 제목이다 — 같은 말을 한 줄 아래 또 쓰지 않는다.
        self.areaBox = self._collapsible("", sweep, checked=True, first=True)
        c2.addWidget(self.areaBox)
        rv.addWidget(card2)
        rv.addStretch(1)

        # ── 설정 페이지: 알림(인라인) → 계정·프록시(창) → 고급(접힘) ──
        # 조건 탭과 같은 회색 바탕 위 흰 카드 세 장. 카드 머리 = 제목 + 한 줄
        # 설명, 안은 목록 행. 배지는 없다 — 설정엔 순서가 없다.
        # 내용은 입력 네 칸과 버튼 셋뿐이다 — 카드를 창 폭에 늘리면 빈 흰 면만
        # 커진다. 본문 열을 720px 로 묶어 가운데 두고 나머지는 회색 여백.
        settings_w = QtWidgets.QWidget()
        settings_w.setObjectName("settingsPage")
        _outer = QtWidgets.QHBoxLayout(settings_w); _outer.setContentsMargins(20, 18, 20, 18)
        _col = QtWidgets.QWidget(settings_w); _col.setObjectName("settingsCol")
        _col.setMaximumWidth(720)
        # 좌우 여백 1 : 열 6 — 열은 720 에서 잘리고, 창이 좁으면 비율대로 준다.
        _outer.addStretch(1); _outer.addWidget(_col, 6); _outer.addStretch(1)
        sv = QtWidgets.QVBoxLayout(_col); sv.setContentsMargins(0, 0, 0, 0); sv.setSpacing(12)
        self.notifyBox, nl = self._step_card(
            None, "알림", "새 매물과 가격 변동을 텔레그램·구글시트로 보냅니다.")
        self._build_notify_form(self.notifyBox, nl)
        sv.addWidget(self.notifyBox)
        # 계정 현황·프록시 목록·프록시 진단은 계정+프록시 창 안에 있다 —
        # 전부 같은 accounts.json 을 보는 화면이라 한 입구로 모았다. fleet
        # 상태·진단 서브창이 딸려 있어 이것만은 인라인으로 펴지 않는다.
        acct, al_v = self._step_card(
            None, "계정 · 프록시", "계정 현황과 프록시 목록·진단은 한 창에 모여 있습니다.")
        al = QtWidgets.QHBoxLayout(); al.setSpacing(10)
        self.healthLabel.setObjectName("statusLine")
        al.addWidget(self.healthLabel, 1)
        self.autoAccountsBtn = QtWidgets.QPushButton("계정+프록시 열기  ›", acct)
        self.autoAccountsBtn.setObjectName("ghostBtn")
        self.autoAccountsBtn.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.autoAccountsBtn.clicked.connect(self.on_accounts_btn_clicked)
        self.autoAccountsBtn.setSizePolicy(QtWidgets.QSizePolicy.Policy.Fixed,
                                          QtWidgets.QSizePolicy.Policy.Fixed)
        al.addWidget(self.autoAccountsBtn, 0, QtCore.Qt.AlignmentFlag.AlignVCenter)
        al_v.addLayout(al)
        sv.addWidget(acct)
        # 고급 = 되돌리기. 평소엔 쓸 일이 없어 카드 안에서 접어 둔다. 삭제는
        # 빨간 글자 버튼 — 금색 주 동작과 같은 무게로 보이면 안 된다.
        adv, adv_l = self._step_card(
            None, "고급", "되돌리기 동작입니다. 평소엔 쓸 일이 없습니다.")
        adv_v = QtWidgets.QVBoxLayout()
        _dr = QtWidgets.QHBoxLayout(); _dr.setSpacing(10)
        _dl = QtWidgets.QLabel("당근에 등록된 키워드를 전부 지웁니다. "
                               "조건을 다시 적용하면 다시 등록됩니다.")
        _dl.setObjectName("mutedNote"); _dl.setWordWrap(True)
        _dr.addWidget(_dl, 1)
        self.alertDelAllBtn.setObjectName("dangerBtn")
        self.alertDelAllBtn.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        _dr.addWidget(self.alertDelAllBtn, 0, QtCore.Qt.AlignmentFlag.AlignVCenter)
        adv_v.addLayout(_dr)
        self.advancedBox = self._collapsible("되돌리기 보기", adv_v, first=True)
        adv_l.addWidget(self.advancedBox)
        sv.addWidget(adv)
        sv.addStretch(1)

        # ── 매물 (신규·추적을 한 표에) ──
        listing_v.setContentsMargins(20, 16, 20, 8); listing_v.setSpacing(12)
        fbar = QtWidgets.QHBoxLayout(); fbar.setSpacing(8)
        self.listingFilter = QtWidgets.QButtonGroup(w)
        for i, (key, label) in enumerate(
                (("all", "전체"), ("new", "🆕 신규"),
                 ("down", "↓ 인하"), ("ended", "✓ 종료"))):
            b = QtWidgets.QPushButton(label); b.setCheckable(True)
            b.setObjectName("filterChip")
            b.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
            b.setChecked(key == "all")
            b.setProperty("filterKey", key)
            self.listingFilter.addButton(b, i)
            fbar.addWidget(b)
        fbar.addStretch(1)
        self.listingFilter.setExclusive(True)
        self.listingFilter.buttonClicked.connect(
            lambda _b: self._refresh_listing_table())
        listing_v.addLayout(fbar)

        self.listingTable = row_lines(QtWidgets.QTableWidget(0, len(LISTING_COLS), w))
        self.listingTable.setHorizontalHeaderLabels(
            LISTING_COLS)
        self.listingTable.verticalHeader().setVisible(False)
        self.listingTable.setEditTriggers(
            QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.listingTable.setSelectionBehavior(
            QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        # 제목만 늘어나고 나머지는 내용 폭이다 — 기본 폭에선 가격·날짜가 "…" 로 잘렸다.
        _lh = self.listingTable.horizontalHeader()
        for _c in range(len(LISTING_COLS)):
            _lh.setSectionResizeMode(
                _c, QtWidgets.QHeaderView.ResizeMode.Stretch if _c == LISTING_COL_TITLE
                else QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.listingTable.setShowGrid(False)
        self.listingTable.setAlternatingRowColors(False)
        self.listingTable.verticalHeader().setDefaultSectionSize(48)
        self.listingTable.itemDelegate().LINE = "#F2F4F6"
        self.listingTable.setFrameShape(QtWidgets.QFrame.Shape.NoFrame)
        self.listingTable.setMinimumHeight(440)
        self.listingTable.setToolTip(
            "앱 알림과 검색 스윕이 같은 표로 들어옵니다. 두 번 누르면 매물이 열립니다.")
        self.listingTable.setSortingEnabled(True)
        self.listingTable.itemDoubleClicked.connect(self.on_listing_open)
        listing_v.addWidget(self.listingTable, 1)

        self.alertLog = QtWidgets.QTextEdit(); self.alertLog.setReadOnly(True); self.alertLog.setMaximumHeight(110)

        # ── 결과 페이지 조립: 자주 보는 것이 위 ──
        # ① 지금 상태 한 줄  ② 매물  ③ 로그
        v.addLayout(top)
        v.addWidget(self.alertBusyRow)
        card = QtWidgets.QFrame(w); card.setObjectName("resultCard")
        card.setLayout(listing_v)
        v.addWidget(card, 1)
        self.logBox = self._collapsible("로그", self.alertLog, checked=True)
        v.addWidget(self.logBox)

        self.alertDelAllBtn.clicked.connect(self.on_alert_delete_all)
        self.rulesApplyBtn.clicked.connect(self.on_rules_apply)
        self.rulesImportBtn.clicked.connect(self.on_rules_import_excel)
        self.rulesAddRowBtn.clicked.connect(self.rulesGrid.add_row)
        self.rulesDelRowBtn.clicked.connect(self.rulesGrid.remove_selected_rows)
        self.rulesGrid.edited.connect(self._on_rules_edited)
        self._rules_dirty = False        # 표를 고쳤는데 아직 적용 안 함
        self._grid_seen_stamp = object() # 표에 마지막으로 채운 파일의 mtime
        self._alert_worker = None
        self._alert_rules = AlertRulesCache()
        self._match_links = {}
        self._last_harvest_ts = 0
        self._last_poll_ts = 0
        self._last_new = 0
        self._alert_poll_timer = QtCore.QTimer(self)
        self._alert_poll_timer.timeout.connect(self._auto_poll_tick)  # 자동폴링=전국(전계정)
        # ── 워치리스트(가격변동 추적) ──
        # 중복 판정은 watch(article_id) + seen_key(인박스 id) 테이블이 한다.
        # 이 집합은 저장소를 아예 못 열었을 때만 쓰는 마지막 방어선이다.
        self._match_seen_fallback = set()
        self._watch_store = None
        self._watch_tracker = None
        self._watch_budget = None
        self._watch_thread = None
        self._watch_threads = []
        self._watch_timer = None
        self._watch_active = 0
        self._watch_next_due = 0
        self._watch_init_error = ""
        try:
            from daangn_ext import article_watch as _aw
            self._watch_store = _aw.WatchStore("./data/watch.db")
            self._watch_tracker = _aw.WatchTracker(self._watch_store)
            # 프록시 목록은 회차마다 다시 읽는다(reload) — 설정에서 웹 프록시를
            # 고쳐도 앱을 껐다 켜지 않고 다음 회차부터 반영된다.
            self._watch_budget = _aw.ProxyBudget(
                feed_proxies(self._load_alert_settings()),
                provider=lambda: feed_proxies(self._load_alert_settings()))
            self._watch_timer = QtCore.QTimer(self)
            self._watch_timer.timeout.connect(self._watch_sweep_tick)
        except Exception as e:
            self._watch_init_error = str(e)[:200]
            print("워치리스트 초기화 실패:", self._watch_init_error)
        # ── 키워드 라우터 · 감시 컨트롤러 ──
        self._router = None
        self._supervisor = None
        self._sweep_queue = None
        self._sweep_kws = None          # 지금 도는 스윕이 떠 있는 키워드 집합
        try:
            from daangn_ext.sweep_queue import SweepQueue
            from daangn_ext.keyword_router import KeywordRouter, DEFAULT_SLOT_CAP
            from daangn_ext.supervisor import SupervisorController, SupervisorPolicy
            self._sweep_queue = SweepQueue("./data/sweep_queue.json")
            cap = int(self._load_alert_settings().get(SLOT_CAP_KEY)
                      or DEFAULT_SLOT_CAP)
            self._router = KeywordRouter(self._alert_fleet(), self._sweep_queue,
                                         slot_cap=cap)
            # 워치 타이머가 없으면(워치리스트 초기화 실패) 컨트롤러를 세우지 않는다 —
            # start() 가 None.start() 로 죽는다. 감시 자체를 막고 이유를 남긴다.
            if self._watch_timer is None:
                raise RuntimeError(
                    f"워치 타이머 없음 — {self._watch_init_error or '워치리스트 초기화 실패'}")
            policy = SupervisorPolicy(lambda: int(self.alertPollInterval.value()),
                                      self._night_factor,
                                      sweep_interval=WATCH_SWEEP_INTERVAL)
            self._supervisor = SupervisorController(
                policy, self._alert_poll_timer, self._watch_timer,
                self._sweep_queue,
                start_search_sweep=self._start_search_sweep,
                stop_search_sweep=self._stop_search_sweep,
                start_feed=self._start_feed, stop_feed=self._stop_feed)
        except Exception as e:
            self._alog(f"[감시] 초기화 실패: {str(e)[:120]}")
        if self._supervisor is None:
            self.watchToggleBtn.setEnabled(False)
            self.watchToggleBtn.setToolTip(
                "감시 컨트롤러 초기화 실패 — 로그 확인 후 재실행하세요")
        self.alertPollInterval.valueChanged.connect(
            lambda _v: self._supervisor.retune() if self._supervisor else None)
        # 실시간 헬스줄(토큰/수확/폴링) — 무인 신뢰 위해 5초마다 갱신
        self._alert_health_timer = QtCore.QTimer(self)
        self._alert_health_timer.timeout.connect(self._refresh_alert_health)
        self._alert_health_timer.start(5000)
        QtCore.QTimer.singleShot(500, self._refresh_alert_health)
        # 저장된 매물을 켜자마자 그린다. watch.db 는 재시작에도 남지만(watch
        # 테이블엔 DELETE 가 없다) 표를 채우는 트리거가 전부 이벤트뿐이라 —
        # 필터 클릭·신규 매칭·스윕 완료 — 켜면 빈 표가 뜨고, 감시를 시작해
        # 최대 30분(야간)을 기다려야 채워졌다. 사용자에겐 매물이 사라진
        # 것으로 보인다. 생성자가 끝난 뒤 그리도록 미룬다.
        QtCore.QTimer.singleShot(0, self._refresh_listing_table)
        # 커버 모드 복원 + 저장
        try:
            if self._load_alert_settings().get("core_only"):
                self.alertCoverMode.setCurrentIndex(1)
        except Exception:
            pass
        self.alertCoverMode.currentIndexChanged.connect(
            lambda _i: self._save_alert_settings({"core_only": bool(self._core_only())}))
        # 실행 시 자동 폴링(무인) — 지연 시작(토큰 수확 대기)
        if self._alert_setting("autostart"):
            if getattr(self, "_mode_cfg", {}).get("background", True):
                QtCore.QTimer.singleShot(8000, self._autostart_poll)
        self._refresh_rules_view()
        return rules_w, w, settings_w

    _STATUS_COLORS = {"ok": "#4E5968", "warn": "#F27A00",
                      "bad": "#F04452", "off": "#B0B8C1"}

    def _set_status(self, key, text, level="ok"):
        """상태 한 줄의 조각 하나를 바꾸고 줄을 다시 그린다. 나쁜 값만 색이 난다."""
        import html as _html
        st = getattr(self, "_status", None)
        line = getattr(self, "statusLine", None)
        if st is None or line is None:
            return
        st[key] = (text, level)
        parts = []
        for k in self.STATUS_ORDER:
            t, lv = st.get(k, ("", "off"))
            if not t:
                continue
            c = self._STATUS_COLORS.get(lv, self._STATUS_COLORS["ok"])
            w = 800 if lv in ("warn", "bad") else 600
            parts.append(f'<span style="color:{c}; font-weight:{w};">{_html.escape(t)}</span>')
        line.setText("&nbsp;&nbsp;·&nbsp;&nbsp;".join(parts))

    def _sync_advanced_visible(self, on):
        """고급 패널 접기 — 접이식 공통 규칙(_sync_box_visible)을 그대로 쓴다."""
        self._sync_box_visible(self.advancedBox, on)

    _ALERT_SETTINGS_FILE = "./data/alert_settings.json"

    def _load_alert_settings(self):
        import json as _json
        try:
            with open(self._ALERT_SETTINGS_FILE, encoding="utf-8") as _f:
                return _json.load(_f)
        except Exception:
            return {}

    # 무인 운영 스위치 기본값. 예전엔 체크박스였고 기본이 꺼짐이라, 설치 뒤
    # 누가 켜 주기 전까지 감시는 앱을 켜도 시작하지 않았다.
    _ALERT_DEFAULTS = {"autostart": True, "night": True, "crash_recover": True}

    def _alert_setting(self, key):
        """저장값이 있으면 그것, 없으면 무인 운영 기본값."""
        v = self._load_alert_settings().get(key)
        return self._ALERT_DEFAULTS.get(key, False) if v is None else bool(v)

    def _save_alert_settings(self, patch):
        import json as _json, os as _os
        try:
            cur = self._load_alert_settings(); cur.update(patch)
            _os.makedirs(_os.path.dirname(self._ALERT_SETTINGS_FILE), exist_ok=True)
            with open(self._ALERT_SETTINGS_FILE, "w", encoding="utf-8") as _f:
                _json.dump(cur, _f)
        except Exception:
            pass

    _BOOT_KEY = r"Software\Microsoft\Windows\CurrentVersion\Run"
    _BOOT_NAME = "KarrotLuxeMonitor"
    # install.ps1 이 서버에 심는 부팅 진입점. 이쪽은 ldboot.ps1 을 거쳐 LDPlayer 를
    # 한 대씩 띄운 뒤에야 앱을 부른다. 우리 이름으로 하나 더 등록하면 로그온 때
    # 앱과 ldboot 이 동시에 인스턴스를 launch 해 게스트 커널이 hang 한다.
    _BOOT_NAME_INSTALLER = "LDPlayerBoot"

    _MODE_FLAG = {"manual": " --manual", "watch": " --watch"}

    def _boot_command(self):
        """부팅 시 실행할 커맨드 — frozen exe면 exe, 개발이면 pythonw + main.py.
        크래시 자동복구 켜져 있으면 --watchdog(감시자 모드)로 실행.

        **지금 이 창의 모드를 반드시 실어 보낸다.** 예전에는 --watchdog 만 붙여서,
        매물감시(--watch) 창에서 부팅 자동실행을 켜면 다음 부팅부터 3탭 합본이
        떴다. 합본은 이제 모드 자체가 없지만, 커맨드에는 모드를 명시해 둔다 —
        모르는 모드는 매물 감시로 보낸다."""
        import sys as _sys
        wd = " --watchdog" if self._alert_setting("crash_recover") else ""
        mode = self._MODE_FLAG.get(getattr(self, "mode", "watch"), " --watch")
        if getattr(_sys, "frozen", False):
            return f'"{_sys.executable}"{wd}{mode}'
        script = os.path.abspath(__file__)
        exe = _sys.executable
        pyw = os.path.join(os.path.dirname(exe), "pythonw.exe")
        launcher = pyw if os.path.exists(pyw) else exe
        return f'"{launcher}" "{script}"{wd}{mode}'

    def _installer_boot_entry(self) -> bool:
        """install.ps1 이 심은 부팅 진입점이 있는지."""
        import sys as _sys
        if _sys.platform != "win32":
            return False
        try:
            import winreg
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, self._BOOT_KEY) as k:
                winreg.QueryValueEx(k, self._BOOT_NAME_INSTALLER)
            return True
        except Exception:
            return False

    def _boot_autostart_enabled(self):
        import sys as _sys
        if _sys.platform != "win32":
            return False
        if self._installer_boot_entry():
            return True          # 설치본 경로로 이미 켜져 있다
        try:
            import winreg
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, self._BOOT_KEY) as k:
                winreg.QueryValueEx(k, self._BOOT_NAME)
            return True
        except Exception:
            return False

    def _set_boot_autostart(self, enable):
        import sys as _sys
        if _sys.platform != "win32":
            self._alog("[부팅 자동실행] Windows 전용 — Mac/Linux 미지원")
            return
        if enable and self._installer_boot_entry():
            # 설치본이 이미 ldboot 경로로 등록해 뒀다. 여기서 하나 더 넣으면
            # 로그온 때 앱과 ldboot 이 동시에 인스턴스를 띄워 함대가 hang 한다.
            self._alog(
                "[부팅 자동실행] 이미 켜져 있습니다 — 설치본이 LDPlayer 순차 기동 뒤에 "
                "앱을 띄우는 경로로 등록해 뒀습니다. 따로 등록하지 않습니다.")
            return
        try:
            import winreg
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, self._BOOT_KEY, 0,
                                winreg.KEY_SET_VALUE) as k:
                if enable:
                    winreg.SetValueEx(k, self._BOOT_NAME, 0, winreg.REG_SZ,
                                      self._boot_command())
                else:
                    try:
                        winreg.DeleteValue(k, self._BOOT_NAME)
                    except FileNotFoundError:
                        pass
            self._alog(f"[부팅 자동실행] {'등록' if enable else '해제'}됨")
            if not enable and self._installer_boot_entry():
                self._alog(
                    "[부팅 자동실행] 다만 설치본 경로(LDPlayer 순차 기동)는 그대로입니다 — "
                    "부팅 시 앱은 계속 뜹니다.")
        except Exception as e:
            self._alog(f"[부팅 자동실행] 실패: {str(e)[:60]}")

    def _autostart_poll(self):
        """실행 시 자동감시 — 이미 돌고 있으면 건드리지 않는다(중복 방지)."""
        if self._supervisor and not self._supervisor.is_running():
            self._alog("[감시] 실행 시 자동 시작")
            self.on_watch_toggle()

    def _refresh_alert_health(self):
        """토큰/자동수확/자동폴링 실시간 헬스 한 줄. 무인 운영 신뢰의 핵심 지표."""
        import json as _json, time as _t
        now = _t.time()
        # 토큰 상태
        alive = expired = 0
        soonest = None
        try:
            from daangn_ext.keyword_alert_api import token_remaining
            with open("./accounts.json", encoding="utf-8") as _f:
                _accs = _json.load(_f)
            for a in _accs:
                acc = a.get("access") or ""
                if not acc:
                    expired += 1; continue
                rem = token_remaining(acc)
                if rem > 60:
                    alive += 1
                    soonest = rem if soonest is None else min(soonest, rem)
                else:
                    expired += 1
        except Exception:
            pass
        tok_ok = alive > 0
        soon = f", 임박 {soonest // 60}분" if soonest is not None else ""
        tok = f"{'🟢' if tok_ok else '🔴'} 토큰 {alive}개 유효{soon}" + (f" · 만료 {expired}" if expired else "")
        # 자동수확
        try:
            hv_on = self._harvest_thread.isRunning()
            iv = getattr(self._harvest_thread, "interval", harvest_interval())
        except Exception:
            hv_on, iv = False, 1200
        if hv_on and self._last_harvest_ts:
            nxt = max(0, int(iv - (now - self._last_harvest_ts)))
            hv = f"🟢 자동수확 ON(다음 {nxt // 60}분 {nxt % 60}초)"
        elif hv_on:
            hv = "🟡 자동수확 ON(첫 수확 대기)"
        else:
            hv = "🔴 자동수확 OFF"
        # 자동폴링
        if self._alert_poll_timer.isActive():
            last = _t.strftime("%H:%M:%S", _t.localtime(self._last_poll_ts)) if self._last_poll_ts else "-"
            nf = self._night_factor()
            eff = self.alertPollInterval.value() * nf
            night = f"·야간×{nf}" if nf > 1 else ""
            pl = (f"🟢 자동폴링 ON({eff}초{night} · 마지막 {last}"
                  f" · 직전신규 {self._last_new})")
        else:
            pl = "⚪ 자동폴링 OFF"
        # 텔레그램 배선 여부(무인 알림 도달 확인)
        nt = getattr(self, "_notify", {}) or {}
        tg = "🟢 텔레그램 연결" if (nt.get("tg_token") and nt.get("tg_chat")) else "⚪ 텔레그램 미설정"
        if nt.get("sheet_url"):
            tg += " · 🟢 시트"
        # 상태 한 줄 — 같은 값을 짧게. 자세한 문장은 툴팁에.
        self._set_status("token", f"토큰 {alive}", "ok" if tok_ok else "bad")
        n_acc = alive + expired
        self._set_status("accounts", f"계정 {n_acc}",
                       "ok" if n_acc else "warn")
        pct = int(getattr(self, "_cover_pct", 0) or 0)
        self._set_status("coverage", f"커버리지 {pct}%" if pct else "커버리지 -",
                       "ok" if pct else "warn")
        if self._alert_poll_timer.isActive():
            rem = max(0, self._alert_poll_timer.remainingTime()) // 1000
            self._set_status("poll", f"다음폴링 {rem // 60}:{rem % 60:02d}", "ok")
        else:
            self._set_status("poll", "폴링 OFF", "off")
        tg_ok = bool(nt.get("tg_token") and nt.get("tg_chat"))
        self._set_status("notify", "텔레그램 연결" if tg_ok else "텔레그램 미설정",
                         "ok" if tg_ok else "warn")
        self.statusLine.setToolTip("\n".join([tok, hv, pl, tg]))
        self._refresh_watch_panel()

    def _refresh_watch_panel(self):
        """가격 추적 현황 한 줄 갱신 — 스윕 스레드가 남긴 캐시만 읽는다(sqlite 동시접근 방지)."""
        try:
            import time as _t
            err = getattr(self, "_watch_init_error", "")
            if err:
                self._watch_label.setText(f"추적 꺼짐 — 초기화 실패: {err[:80]}")
                return
            self._watch_label.setText(watch_status_text(
                getattr(self, "_watch_active", 0),
                getattr(self, "_watch_next_due", 0), int(_t.time())))
        except Exception:
            pass

    def on_alert_fleet(self):
        """계정 팜 현황 다이얼로그 — 계정별 상태 색상코딩, 열려있는 동안 5초 자동갱신."""
        from PyQt6.QtGui import QColor, QBrush
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("계정 팜 현황")
        dlg.resize(760, 560)
        lay = QtWidgets.QVBoxLayout(dlg)
        summ = QtWidgets.QLabel("")
        summ.setStyleSheet("font-weight:700; font-size:15px; padding:6px 4px; color:#8A6D1F;")
        lay.addWidget(summ)
        cols = ["계정", "동네", "만료(분)", "핵심", "실패", "상태"]
        tbl = row_lines(QtWidgets.QTableWidget(0, len(cols), dlg))
        tbl.setHorizontalHeaderLabels(cols)
        tbl.verticalHeader().setVisible(False)
        tbl.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        tbl.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.Stretch)
        lay.addWidget(tbl, 1)

        def refresh():
            try:
                rows = self._multi().fleet_status()
            except Exception as e:
                summ.setText(f"조회 실패: {str(e)[:60]}"); return
            alive = sum(1 for r in rows if r["alive"])
            core = sum(1 for r in rows if r["core"] and r["alive"])
            banned = sum(1 for r in rows if r["banned"])
            summ.setText(f"총 {len(rows)}계정 · 유효 {alive} · 핵심(유효) {core}"
                         + (f" · ⚠️점검필요 {banned}" if banned else ""))
            tbl.setRowCount(0)
            for r in rows:
                i = tbl.rowCount(); tbl.insertRow(i)
                exp = r["exp_min"]
                exp_txt = "만료" if exp < 0 or (not r["alive"]) else f"{exp}"
                vals = [r["code"], r["region"], exp_txt,
                        "★" if r["core"] else "", str(r["fail"]),
                        "점검필요" if r["banned"] else ("정상" if r["alive"] else "만료")]
                for c, val in enumerate(vals):
                    it = QtWidgets.QTableWidgetItem(val)
                    if r["banned"]:
                        it.setBackground(QBrush(QColor("#FCEBE8")))
                    elif not r["alive"]:
                        it.setForeground(QBrush(QColor("#8B8474")))
                    elif r["core"]:
                        it.setForeground(QBrush(QColor("#8A6D1F")))
                    tbl.setItem(i, c, it)

        btnRow = QtWidgets.QHBoxLayout()
        editBtn = QtWidgets.QPushButton("핵심지역 편집")
        resetBtn = QtWidgets.QPushButton("상태 초기화")
        resetBtn.setToolTip("실패/점검필요 플래그 초기화 — 토큰만료로 인한 오탐 해소")
        refreshBtn = QtWidgets.QPushButton("새로고침")
        btnRow.addWidget(editBtn); btnRow.addWidget(resetBtn)
        btnRow.addWidget(refreshBtn); btnRow.addStretch(1)
        lay.addLayout(btnRow)
        editBtn.clicked.connect(lambda: self._edit_core_regions(refresh))
        refreshBtn.clicked.connect(lambda: refresh())

        def do_reset():
            try:
                self._multi().reset_state(); self._alog("[팜] 계정 상태 초기화")
            except Exception:
                pass
            refresh()
        resetBtn.clicked.connect(do_reset)

        refresh()
        t = QtCore.QTimer(dlg); t.timeout.connect(refresh); t.start(5000)
        dlg.finished.connect(lambda _=0: t.stop())
        dlg.show()

    def _edit_core_regions(self, on_saved=None):
        """핵심지역 키워드 편집 다이얼로그 — 동네명 부분일치. 저장 시 core 판정 즉시 반영."""
        m = self._multi()
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("핵심지역 편집")
        dlg.resize(440, 460)
        lay = QtWidgets.QVBoxLayout(dlg)
        lay.addWidget(QtWidgets.QLabel(
            "명품 밀집 지역 키워드(한 줄에 하나). 계정 인증동네명에 포함되면 '핵심'.\n"
            "예: 강남, 청담, 분당, 해운대"))
        edit = QtWidgets.QPlainTextEdit(dlg)
        edit.setPlainText("\n".join(m.core_keywords()))
        lay.addWidget(edit, 1)
        row = QtWidgets.QHBoxLayout()
        saveBtn = QtWidgets.QPushButton("저장"); saveBtn.setObjectName("startBtn")
        resetBtn = QtWidgets.QPushButton("기본값 복원")
        cancelBtn = QtWidgets.QPushButton("취소")
        row.addWidget(saveBtn); row.addWidget(resetBtn); row.addStretch(1); row.addWidget(cancelBtn)
        lay.addLayout(row)

        def do_save():
            kws = [ln.strip() for ln in edit.toPlainText().splitlines() if ln.strip()]
            if m.save_core_keywords(kws):
                self._alog(f"[핵심지역] {len(kws)}개 저장")
                if on_saved:
                    on_saved()
                dlg.accept()
        def do_reset():
            edit.setPlainText("\n".join(m.CORE_REGION_KEYWORDS))
        saveBtn.clicked.connect(do_save)
        resetBtn.clicked.connect(do_reset)
        cancelBtn.clicked.connect(dlg.reject)
        dlg.exec()

    def _alert_api(self):
        """스레드 내에서 호출 — thread-safe 토큰 수확 후 KeywordAlertAPI 반환."""
        from daangn_ext.keyword_alert_api import KeywordAlertAPI
        token = self._harvest_token_quiet()
        if not token:
            raise RuntimeError("유효 토큰 없음 — LDPlayer/폰 수확 확인")
        return KeywordAlertAPI(token)

    def _quiet_keyword_list(self, core_only=False):
        """서버에 이미 등록된 키워드 목록 — 수확 없이 accounts.json 토큰만 쓴다.

        씨딩 전용. _alert_api() 를 쓰면 안 된다: 그쪽은 조회 전에 LDPlayer 부팅 +
        함대 전체 수확을 먼저 돌리므로, 첫 실행 씨딩 한 번에 에뮬레이터 팜이 뜬다.
        헤드리스 _server_keyword_list 와 같은 문이다 — 등록은 전 계정에 같은
        키워드를 쓰므로 유효한 계정 하나만 봐도 함대 상태를 안다."""
        from daangn_ext.keyword_alert_api import KeywordAlertAPI
        valid = self._multi()._valid(core_only)
        if not valid:
            return {}
        _code, access, proxy = valid[0]
        api = KeywordAlertAPI(access, "./data/config.json", proxy=proxy)
        try:
            return api.list()
        finally:
            try:
                api.close()
            except Exception:
                pass

    def _alert_run(self, fn, on_done=None, queue=False, label="처리 중") -> bool:
        """작업 하나를 워커로 돌린다. **시작했거나 대기에 넣었으면 True.**

        예전에는 바쁠 때 아무 값도 안 돌려줬고, 부르는 쪽은 그걸 안 봤다. 엑셀
        불러오기는 작업이 버려진 뒤에도 "배정 중입니다"를 띄웠다 — 사용자는
        성공으로 읽고, 표가 비어 있으니 '전체 삭제'를 눌러 등록을 통째로 날렸다
        (실서버 2026-09-02 00:14 로그).

        queue=True 는 거절 대신 **한 자리 대기열**에 넣는다. 무인 서버는
        자동수확·자동폴링이 수시로 돌아서 사용자가 누른 순간이 그 창에 겹치는 게
        일상이다. 사용자가 방금 시킨 일은 미뤄서라도 해야지 버리면 안 된다.
        자리가 하나뿐인 이유는 같은 버튼을 연타했을 때 같은 일을 여러 번 하지
        않기 위해서다 — 마지막 것이 이긴다."""
        if self._alert_worker and self._alert_worker.isRunning():
            if not queue:
                self.alert("이전 작업 진행 중 — 잠시 후")
                return False
            self._alert_pending = (fn, on_done, label)
            self._alog("[대기] 이전 작업이 끝나면 이어서 실행합니다")
            self._alert_busy(f"{label} — 앞 작업이 끝나면 시작합니다")
            return True
        self._alog("── 작업 시작 ──")
        self._alert_busy(f"{label}… 잠시만 기다리세요")
        self._alert_worker = _AlertWorker(fn)
        self._alert_worker.log.connect(lambda m: self._alog(m))
        if on_done:
            self._alert_worker.done.connect(on_done)
        self._alert_worker.finished.connect(self._alert_drain)
        self._alert_worker.start()
        return True

    def _alert_busy(self, text):
        """진행 표시. text 가 비면 감춘다.

        등록·삭제는 '계정 수 × 키워드 수' 만큼 요청이라 20초를 넘긴다. 그동안
        화면이 조용하면 사용자는 실패한 줄 알고 다시 누르거나 '전체 삭제'로
        손을 뻗는다 — 실서버에서 등록 21건이 그렇게 사라졌다."""
        row = getattr(self, "alertBusyRow", None)
        if row is None:                  # 아직 탭이 안 만들어진 초기화 경로
            return
        if text:
            self.alertBusyLabel.setText(str(text))
            row.setVisible(True)
        else:
            row.setVisible(False)

    def _alert_drain(self):
        """앞 작업이 끝났다 — 대기시켜 둔 작업이 있으면 이어서 돌린다.

        `finished` 는 run() 이 막 반환한 시점이라 isRunning() 이 아직 True 일 수
        있다. 그때 바로 시작하면 스스로를 다시 대기열에 넣어 영원히 안 돈다.
        그래서 아직 돌고 있으면 조금 뒤에 다시 본다."""
        pend = getattr(self, "_alert_pending", None)
        if not pend:
            self._alert_busy("")         # 더 할 일이 없다 — 진행 표시를 내린다
            return
        if self._alert_worker and self._alert_worker.isRunning():
            QtCore.QTimer.singleShot(200, self._alert_drain)
            return
        self._alert_pending = None
        self._alert_run(pend[0], pend[1], label=pend[2], queue=True)

    def _pi(self, s):
        s = (s or "").strip().replace(",", "")
        return int(s) if s.isdigit() else None

    _ROUTE_NAMES = ROUTE_NAMES

    def _log_route(self, res):
        """라우터 결과 한 줄. route 는 None 일 수 있다(빈 키워드)."""
        name = self._ROUTE_NAMES.get(res.get("route"), "배정 안 됨")
        self._alog(
            f"[키워드] {res.get('keyword')} → {name} ({res.get('reason')})")

    def _safe_alert_list(self, log):
        """현재 계정의 등록 목록. 토큰이 없으면 None — 등록 자체는 이미 끝났다."""
        try:
            return self._alert_api().list()
        except Exception as e:
            log(f"[목록] 조회 실패: {str(e)[:80]}")
            return None

    def _alert_route_done(self, payload):
        if not payload:
            return
        if payload.get("route"):
            self._log_route(payload["route"])
        self._alert_populate(payload.get("list"))

    def _alert_routes_done(self, payload):
        """라우터 여러 건 결과 → 경로 로그 + 목록 갱신."""
        if not payload:
            return
        for r in payload.get("routes") or []:
            self._log_route(r)
        self._alert_populate(payload.get("list"))

    def on_alert_delete_all(self):
        # 삭제 경로는 하나여야 한다. 조건표와 등록을 따로 지우게 두면 조건만
        # 지운 반쪽 상태가 생기고, 그때는 브랜드 전 매물이 알림으로 쏟아진다.
        n_rules = len(self._alert_rules.get())
        n_reg = len(self._router.routes()) if self._router else 0
        if not ask_yes_no(
                self, "전체 삭제", "감시를 처음 상태로 되돌릴까요?",
                f"· 조건 {n_rules}줄이 지워집니다\n"
                f"· 당근에 등록된 키워드 {n_reg}개가 해제됩니다\n"
                "· 되돌릴 수 없습니다 — 조건을 다시 적어야 합니다",
                danger=True):
            return
        try:
            from daangn_ext.alert_rules import RuleTable as _RT
            os.makedirs(os.path.dirname(ALERT_RULES_FILE), exist_ok=True)
            _RT().save(ALERT_RULES_FILE)
            self._refresh_rules_view()
            self._alog(f"[조건표] 조건 {n_rules}줄을 지웠습니다")
        except Exception as e:
            self._alog(f"[조건표] 조건 지우기 실패: {str(e)[:80]}")
        # 워커가 바쁘면 아래 _alert_run 이 거절한다. 그런데 라우터 비우기는 그
        # 앞에서 이미 끝나 있어, 배정만 사라지고 앱 등록은 남는 반쪽 상태가 됐다.
        # 지울 수 없으면 아무것도 건드리지 않는다.
        if self._alert_worker and self._alert_worker.isRunning():
            self.alert("이전 작업 진행 중 — 잠시 후")
            return
        # 라우터 배정도 함께 비운다 — 남겨두면 슬롯이 찬 채로 전부 스윕행이 된다.
        # 관측 상한도 되돌린다. '처음 상태'에 낮아진 상한이 남아 있으면 다시 넣은
        # 엑셀이 슬롯에 못 들어가고 전부 스윕행이 된다 — GUI 에서 상한을
        # 되돌리는 길은 이것 하나다(서버는 --reset-cap).
        if self._router:
            try:
                for r in list(self._router.routes()):
                    self._router.remove(r["keyword"])
                if self._router.reset_observed_cap():
                    self._alog("[라우터] 앱 슬롯 상한 관측치 초기화 —"
                               " 다음 등록부터 설정 상한을 다시 씁니다")
            except Exception as e:
                self._alog(
                    f"[라우터] 배정 비우기 실패 — 앱 슬롯이 남습니다: {str(e)[:80]}")
        def job(log):
            api = self._alert_api()
            n = api.delete_all(log=log); log(f"총 {n}건 삭제")
            return api.list()
        self._alert_run(job, self._alert_populate, label="전체 삭제 중")

    def _routes_map(self):
        try:
            return {r["keyword"]: r for r in (self._router.routes() if self._router
                                              else [])}
        except Exception:
            return {}

    def _alert_row(self, r, keyword, route, price, exclude, uid, status=None):
        """표 한 줄. 값 구성은 alert_row_cells(순수 함수)가 한다.
        미등록 줄은 붉은 바탕 + 붉고 굵은 키워드·상태 — 표에서 눈에 띄어야 할
        유일한 줄이다."""
        vals, tips = alert_row_cells(keyword, route, price, exclude, uid, status)
        for c, val in enumerate(vals):
            cell = QtWidgets.QTableWidgetItem(val)
            if c in tips:
                cell.setToolTip(tips[c])
            if status == REG_MISSING:
                cell.setBackground(QtGui.QBrush(QtGui.QColor(REG_MISSING_BG)))
                if c in (ALERT_COL_KEYWORD, ALERT_COL_STATUS):
                    cell.setForeground(QtGui.QBrush(QtGui.QColor(REG_MISSING_FG)))
                    f = cell.font(); f.setBold(True); cell.setFont(f)
            elif status == REG_UNKNOWN and c == ALERT_COL_STATUS:
                cell.setForeground(QtGui.QBrush(QtGui.QColor(REG_UNKNOWN_FG)))
            self.alertTable.setItem(r, c, cell)

    def _queue_entries(self):
        """스윕 대기열 엔트리. SweepQueue 는 __len__ 이 있어 비면 falsy 다 —
        존재 여부는 is not None 으로 본다."""
        try:
            return (self._sweep_queue.entries()
                    if self._sweep_queue is not None else [])
        except Exception:
            return []

    def _alert_populate(self, data):
        """등록 목록 그리기.

        토큰이 없어 앱 목록을 못 읽어도(data 가 None) 대기열은 그린다 — 방금 스윕으로
        밀린 키워드가 안 보이면 사용자는 등록이 삼켜진 줄 안다. 대신 앱 목록을 못
        읽었다는 사실을 로그에 남겨, 짧아진 표를 '키워드가 사라졌다'로 읽지 않게 한다."""
        entries = self._queue_entries()
        routes = self._routes_map()
        # 라우터가 **항상 있는 진실**이다. 엑셀로 넣은 조건은 등록이 끝나는 즉시
        # 여기 들어오지만, 서버 목록 조회는 토큰이 없으면 실패한다. 예전에는 그때
        # 표를 통째로 비워 두고 로그 한 줄만 남겼는데, 사용자에게는 '엑셀을
        # 올렸는데 아무것도 안 뜬다'로 보였다. 셋을 합쳐 그린다.
        if not data and not entries and not routes:
            return
        if not data:
            self._alog(
                "[목록] 앱 등록 목록을 못 읽었습니다 — 저장된 조건으로 표시합니다"
                " (삭제하려면 목록을 읽을 수 있어야 합니다)")
        kws = (data or {}).get("user_keywords") or []
        # 이 브랜치 이전에 일괄등록된 키워드는 routes 파일에 없다. 첫 실행에서
        # 그걸 인정하지 않으면 라우터는 함대가 비었다고 믿고 이미 꽉 찬 서버
        # 한도에 계속 등록을 시도한다.
        if self._router is not None and kws:
            try:
                _ok_kws, _ = split_by_rules([k.get("keyword") for k in kws],
                                            rule_brand_keys(self._alert_rules.get()))
                seeded = self._router.seed_from_server(_ok_kws)
                if seeded:
                    self._alog(
                        f"[라우터] 서버에 이미 등록된 키워드 {seeded}개를 앱 슬롯으로 인식")
            except Exception as e:
                self._alog(f"[라우터] 기존 등록 인식 실패: {str(e)[:80]}")
            # 씨딩이 방금 route 를 만들었다 — 위에서 읽어 둔 맵은 그걸 모른다.
            # 낡은 맵으로 그리면 경로 열이 전부 '-' 로 뜬다.
            routes = self._routes_map()
        self.alertTable.setRowCount(0)
        shown = set()
        for k in kws:
            r = self.alertTable.rowCount(); self.alertTable.insertRow(r)
            price = ""
            if k.get("min_price") or k.get("max_price"):
                price = f"{k.get('min_price') or ''}~{k.get('max_price') or ''}"
            kw = k.get("keyword", "")
            shown.add(kw)
            self._alert_row(r, kw, routes.get(kw), price,
                            ",".join(k.get("exclude_keywords") or []),
                            str(k.get("id", "")), REG_SERVER)
        # 스윕으로 밀린 키워드는 앱 목록에 없다 — 여기 안 보이면 사용자는
        # 등록이 삼켜진 줄 안다. 대기열에서 끌어와 함께 그린다.
        for e in entries:
            if e["keyword"] in shown:
                continue
            r = self.alertTable.rowCount(); self.alertTable.insertRow(r)
            price = ""
            if e.get("min") or e.get("max"):
                price = f"{e.get('min') or ''}~{e.get('max') or ''}"
            self._alert_row(r, e["keyword"], routes.get(e["keyword"]), price,
                            ",".join(e.get("exclude") or []), "", REG_SWEEP)
            shown.add(e["keyword"])
        # 서버 목록에도 대기열에도 없지만 라우터가 아는 키워드 — 엑셀로 방금 넣은
        # 것들이 여기 해당한다. 조건은 라우터의 cond 가 갖고 있다.
        # 앱 경로인데 서버 목록에 없으면 '미등록'(빨강). 단 서버 목록 자체를 못
        # 읽었으면(data 없음) 등록됐는지 모르는 것이지 안 된 게 아니다 — 그때
        # 빨갛게 칠하면 멀쩡한 등록을 지우러 가게 만든다. '확인 불가'로 둔다.
        for kw, route in (routes or {}).items():
            if kw in shown:
                continue
            cond = (route or {}).get("cond") or {}
            r = self.alertTable.rowCount(); self.alertTable.insertRow(r)
            price = ""
            if cond.get("min") or cond.get("max"):
                price = f"{cond.get('min') or ''}~{cond.get('max') or ''}"
            if (route or {}).get("route") == "sweep":
                status = REG_SWEEP
            elif not data:
                status = REG_UNKNOWN
            else:
                status = REG_MISSING
            self._alert_row(r, kw, route, price,
                            ",".join(cond.get("exclude") or []), "", status)
            shown.add(kw)
        subs = (data or {}).get("subscription_infos") or []
        if subs:
            txt = " · ".join(f"{s.get('name')}({s.get('ranged_regions_count')}지역"
                             + (",알림ON" if s.get('enable_notification') else ",알림OFF") + ")"
                             for s in subs)
            self.alertSubLabel.setText(f"커버 동네: {txt}")

    # ── 신규 매칭 폴링 ──
    def _set_sleep_block(self, block):
        """자동폴링 중 PC 절전 차단(Windows). 절전 걸리면 감시 멈추는 함정 방지.
        block=True 상시 각성 유지, False 해제. Mac/Linux no-op."""
        import sys as _sys
        if _sys.platform != "win32":
            return
        try:
            import ctypes
            ES_CONTINUOUS = 0x80000000
            ES_SYSTEM_REQUIRED = 0x00000001
            flags = (ES_CONTINUOUS | ES_SYSTEM_REQUIRED) if block else ES_CONTINUOUS
            ctypes.windll.kernel32.SetThreadExecutionState(flags)
        except Exception:
            pass

    def on_watch_toggle(self):
        """감시 토글 — 폴링·워치 스윕·검색 스윕이 한 수명을 공유한다."""
        if not self._supervisor:
            self._alog("[감시] 컨트롤러가 없습니다")
            self.watchToggleBtn.setChecked(False)
            return
        if self._supervisor.is_running():
            self._supervisor.stop()
            self._set_sleep_block(False)
            self.watchToggleBtn.setText("▶ 감시 시작")
            self.watchToggleBtn.setChecked(False)
            self._alog("[감시] 정지")
        else:
            self._supervisor.start()
            self._set_sleep_block(True)
            self.watchToggleBtn.setText("■ 감시 정지")
            self.watchToggleBtn.setChecked(True)
            nf = self._night_factor()
            night = f" · 야간감속 ×{nf}" if nf > 1 else ""
            self._alog(
                f"[감시] 시작 · 폴링 {self.alertPollInterval.value()}초{night} · 절전차단")
            self._watch_kickoff()          # 첫 회차는 기다리지 않는다

    def _match_populate(self, matches):
        import time as _t
        self._last_poll_ts = _t.time()
        if matches is None:
            self._last_new = 0
            return
        matches, watch_only, _ = filter_by_conditions(
            matches, getattr(self, "_router", None), self._alog,
            rules=self._alert_rules.get())
        new_items, dropped = dedupe_new_matches(
            matches, self._watch_store, self._match_seen_fallback)
        new = len(new_items)
        self._last_new = new
        if dropped:
            self._alog(f"[매칭] id 없는 payload {dropped}건 건너뜀")
        # 추적 등록은 알림 여부와 무관하게 한다 — 상한을 넘겨 알리지 않은
        # 매물도 값이 내려오는지 지켜봐야 '조건 진입'을 잡을 수 있다.
        try:
            added = self._watch_tracker.add_from_matches(new_items + watch_only) \
                if self._watch_tracker else 0
            if added:
                self._alog(f"[가격추적] {added}건 추적 시작")
        except Exception as e:
            self._alog(f"[가격추적] 등록 실패: {str(e)[:80]}")
        if new:
            self._alog(f"[매칭] 신규 {new}건 추가")
            self._notify_matches(new_items)
        if new or watch_only:
            self._refresh_listing_table()
        try:
            self._refresh_alert_health()
        except Exception:
            pass

    def _notify_matches(self, items):
        """신규 매칭 → 텔레그램(실시간) + 구글시트(히스토리). 백그라운드 발송(GUI 안 멈춤)."""
        nt = getattr(self, "_notify", {}) or {}
        has_tg = nt.get("tg_token") and nt.get("tg_chat")
        has_sheet = bool(nt.get("sheet_url"))
        if not (items and (has_tg or has_sheet)):
            return
        if not hasattr(self, "_notify_threads"):
            self._notify_threads = []
        th = _NotifyThread(nt, list(items))
        th.log.connect(self._alog)
        th.finished.connect(lambda t=th: self._notify_threads.remove(t)
                            if t in self._notify_threads else None)
        self._notify_threads.append(th)
        th.start()

    def _watch_sweep_tick(self):
        """10분마다 워치리스트를 예산만큼 재조회한다(백그라운드)."""
        if not self._watch_tracker:
            return
        th = getattr(self, "_watch_thread", None)
        if th is not None and th.isRunning():
            return                      # 이전 스윕이 아직 돌면 건너뛴다
        th = _WatchSweepThread(self._watch_tracker, self._watch_store,
                               self._watch_budget, WATCH_SWEEP_INTERVAL)
        th.log.connect(self._alog)
        th.done.connect(self._on_watch_swept)
        self._watch_thread = th
        th.start()

    def _on_watch_swept(self, events, active, next_due):
        self._watch_active = int(active)
        self._watch_next_due = int(next_due)
        if events:
            self._notify_watch_events(events)
        self._refresh_listing_table()

    def _notify_watch_events(self, events):
        events = mark_range_entries(events, self._watch_store,
                                    getattr(self, "_router", None),
                                    rules=self._alert_rules.get())
        lines = watch_event_lines(events)
        if not lines:
            return
        for line in lines:
            self._alog(f"[가격추적] {line}")
        self._refresh_listing_table()
        nt = getattr(self, "_notify", {}) or {}
        if not ((nt.get("tg_token") and nt.get("tg_chat")) or nt.get("sheet_url")):
            return
        th = _WatchNotifyThread(getattr(self, "_notify", {}) or {}, lines, events)
        th.log.connect(self._alog)
        th.finished.connect(lambda t=th: self._watch_threads.remove(t)
                            if t in self._watch_threads else None)
        self._watch_threads.append(th)
        th.start()

    def _listing_filter_key(self):
        b = self.listingFilter.checkedButton()
        return b.property("filterKey") if b else "all"

    def _refresh_listing_table(self):
        """저장소에서 읽어 표를 다시 그린다. 행 수가 수백 단위라 전면 갱신으로 족하다."""
        if not getattr(self, "_watch_store", None):
            return
        import time as _t
        try:
            rows = listing_display_rows(self._watch_store.listing_rows(),
                                        int(_t.time()), self._listing_filter_key())
        except Exception as e:
            self._alog(f"[매물표] 갱신 실패: {str(e)[:80]}")
            return
        self.listingTable.setSortingEnabled(False)
        self.listingTable.setRowCount(0)
        for r in rows:
            i = self.listingTable.rowCount()
            self.listingTable.insertRow(i)
            vals = [r["icon"], r["source"], r["keyword"], r["title"], r["region"],
                    f"{r['price']:,}" if r["price"] else "-",
                    r["delta_text"], r["last_change_text"], r["first_seen_text"]]
            for c, val in enumerate(vals):
                cell = QtWidgets.QTableWidgetItem(val)
                if c == 0:
                    cell.setData(QtCore.Qt.ItemDataRole.UserRole, r["url"])
                    cell.setData(QtCore.Qt.ItemDataRole.UserRole + 1, r["id"])
                self.listingTable.setItem(i, c, cell)
        self.listingTable.setSortingEnabled(True)

    def on_listing_open(self, item):
        """더블클릭 → 가격 이력 + 매물 링크."""
        cell0 = self.listingTable.item(item.row(), 0)
        if cell0 is None:
            return
        url = cell0.data(QtCore.Qt.ItemDataRole.UserRole) or ""
        aid = cell0.data(QtCore.Qt.ItemDataRole.UserRole + 1) or ""
        hist = []
        try:
            hist = self._watch_store.price_history(str(aid))
        except Exception:
            pass
        import time as _t
        lines = [f"{_t.strftime('%m/%d %H:%M', _t.localtime(h['ts']))}  "
                 f"{h['price']:,}원" for h in hist] or ["가격 이력 없음"]
        dlg = QtWidgets.QMessageBox(self)
        dlg.setWindowTitle("가격 이력")
        dlg.setText("\n".join(lines))
        if url:
            open_btn = dlg.addButton("매물 열기",
                                     QtWidgets.QMessageBox.ButtonRole.ActionRole)
        else:
            open_btn = None
        dlg.addButton(QtWidgets.QMessageBox.StandardButton.Close)
        dlg.exec()
        if open_btn is not None and dlg.clickedButton() is open_btn:
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))

    def _set_thumb(self, item, data):
        """다운로드된 썸네일 바이트 → 아이콘(메인스레드서 안전하게 setIcon)."""
        try:
            from PyQt6.QtGui import QPixmap, QIcon
            pm = QPixmap()
            if pm.loadFromData(data):
                item.setIcon(QIcon(pm.scaled(
                    56, 56, QtCore.Qt.AspectRatioMode.KeepAspectRatio,
                    QtCore.Qt.TransformationMode.SmoothTransformation)))
        except Exception:
            pass

    # ── 전국(전 계정) 멀티계정 ──
    def _multi(self, harvest=False):
        """MultiAccountAlerts 반환. harvest=True 면 LDPlayer 수확 먼저(느림·프로덕션용).
        기본은 accounts.json 기존 토큰 사용(빠름 — coverage/poll 즉시)."""
        from daangn_ext.keyword_alert_api import MultiAccountAlerts
        if harvest:
            try:
                import ld_autoharvest
                ld_autoharvest.harvest_all(
                    "./accounts.json", nudge=True,
                    log=lambda m: self.sb.showMessage(m, 4000))
            except Exception:
                pass
        return MultiAccountAlerts("./accounts.json", "./data/config.json")

    def _alert_fleet(self):
        """전계정 알림 함대. 라우터와 폴링이 같은 생성 경로를 쓰게 한 곳으로 모은다.

        MultiAccountAlerts 는 accounts.json 을 호출 때마다 다시 읽는다 —
        인스턴스를 들고 있어도 수확으로 갱신된 토큰이 그대로 반영된다."""
        return self._multi()

    def _core_only(self):
        """커버 모드 = 핵심지역 집중이면 True. (콤보 currentData)"""
        try:
            return bool(self.alertCoverMode.currentData())
        except Exception:
            return False

    def _night_factor(self):
        """야간 감속 배수 — 새벽0~7시 ×3, 늦밤22~24·이른7~9시 ×2, 그외 ×1."""
        if not self._alert_setting("night"):
            return 1
        import time as _t
        h = _t.localtime().tm_hour
        if 0 <= h < 7:
            return 3
        if 22 <= h < 24 or 7 <= h < 9:
            return 2
        return 1

    def _auto_poll_tick(self):
        """자동폴링 타이머 틱 — 간격 재조정은 컨트롤러가, 승격은 라우터가 맡는다.
        이전 폴링이 아직 진행 중이면 조용히 스킵(무인: 모달 팝업 금지)."""
        if self._supervisor:
            self._supervisor.retune()
        # 조건 파일이 밖에서 바뀌었을 수 있다 — 캐시가 mtime 으로 잡고,
        # 화면 숫자는 여기서 따라간다.
        try:
            self._refresh_rules_view()
        except Exception:
            pass
        # 스윕 재동기화만 GUI 스레드에 남는다 — QThread 를 만들고 세우는 일이라
        # 워커로 못 옮긴다. 대신 네트워크를 타지 않는다.
        self._resync_search_sweep()
        self._resync_feed()
        if self._alert_worker and self._alert_worker.isRunning():
            self._alog("[자동폴링] 이전 폴링 진행 중 — 이번 틱 스킵")
            return
        co = self._core_only()
        state = self.__dict__.setdefault("_seed_state", {})

        def job(log):
            # 씨딩·승격은 **워커 안에서** 돈다. 씨딩은 20초 타임아웃짜리 HTTP
            # 조회라 타이머 콜백(=GUI 스레드)에서 하면 창이 그대로 멈춘다.
            # 자동 시작은 등록 화면(_alert_populate)을 거치지 않으므로 여기서
            # 씨딩하지 않으면 무인 첫 실행이 함대가 빈 줄 알고 꽉 찬 서버 한도에
            # 등록을 시도해 전부 스윕으로 민다. routes 가 차 있으면 조회조차 안 한다.
            # 승격은 씨딩 결과에 기대므로 반드시 뒤에 온다.
            _allowed = rule_brand_keys(self._alert_rules.get())
            seed_router_from_server(
                self._router, lambda: self._quiet_keyword_list(co), log, state,
                allowed=_allowed,
                prune_fn=lambda extras: self._alert_fleet().delete_keywords(
                    extras, log=log, core_only=co))
            if self._router:
                try:
                    for p in self._router.rebalance(core_only=co, log=log):
                        log(f"[라우터] {p['keyword']} → 앱 알림 승격")
                except Exception as e:
                    log(f"[라우터] 승격 실패: {str(e)[:80]}")
                mirror_app_keywords_to_sweep(
                    self._router, self._sweep_queue, log=log,
                    enabled=sweep_mirror_enabled(
                        self._load_alert_settings(),
                        len(self._alert_rules.get())))
            return self._alert_fleet().poll_all(log=log, core_only=co)

        self._alert_run(job, self._match_populate, label="자동 폴링 중")

    def _watch_kickoff(self):
        """감시 시작 직후 한 번: 전 계정 폴링 → 커버 동네 집계 → 등록 목록.

        셋 다 같은 계정 토큰을 쓰는 네트워크 작업이라 워커 하나에 이어 붙인다.
        _alert_run 의 대기열은 한 자리라 셋을 따로 넣으면 마지막 것만 남는다."""
        co = self._core_only()

        def job(log):
            out = {"matches": None, "cov": None, "list": None}
            # 셋 중 하나가 죽어도 나머지는 한다 — 첫 폴링이 실패했다고 커버
            # 현황과 등록 목록까지 비워 두면 화면이 '아무것도 없음'으로 읽힌다.
            try:
                out["matches"] = self._alert_fleet().poll_all(log=log, core_only=co)
            except Exception as e:
                log(f"[폴링] 첫 확인 실패: {str(e)[:80]}")
            try:
                cov = self._multi().coverage(log=log, core_only=co)
                log(f"커버 동네 {len(cov)}개 · 합산 {sum(int(c[2] or 0) for c in cov)}지역")
                out["cov"] = cov
            except Exception as e:
                log(f"[커버] 집계 실패: {str(e)[:80]}")
            out["list"] = self._safe_alert_list(log)
            return out

        def done(out):
            out = out or {}
            self._match_populate(out.get("matches"))
            self._update_coverage(out.get("cov"))
            self._alert_populate(out.get("list"))

        self._alert_run(job, done, queue=True, label="첫 확인 중")

    def on_alert_poll_all(self):
        co = self._core_only()
        def job(log):
            return self._alert_fleet().poll_all(log=log, core_only=co)
        self._alert_run(job, self._match_populate, label="전 계정 알림함 확인 중")

    def on_alert_coverage(self):
        co = self._core_only()
        def job(log):
            cov = self._multi().coverage(log=log, core_only=co)
            total = sum(int(c[2] or 0) for c in cov)
            log(f"커버 동네 {len(cov)}개 · 합산 {total}지역")
            for code, name, cnt in cov:
                log(f"  {code}: {name} ({cnt}지역)")
            return cov
        self._alert_run(job, self._update_coverage, label="커버 동네 조회 중")

    def _update_coverage(self, cov):
        """커버 동네 조회 결과 → 전국 커버 %(상태 한 줄) + 증설 안내 한 줄(로그)."""
        if cov is None:
            return
        KOREA_DONG = 3500            # 전국 행정동 대략
        AVG_RANGE = 78              # 계정당 커버 지역(39지역 x 인증동네 2곳)
        codes = {c[0] for c in cov}
        n_acc = len(codes)
        dongs = {(c[0], c[1]) for c in cov}         # 계정×동네
        total_regions = sum(int(c[2] or 0) for c in cov)
        # 겹침 감안 실효 커버(대략 70%)
        eff = int(total_regions * 0.7)
        pct = min(100, int(eff / KOREA_DONG * 100))
        need_full = max(0, round(KOREA_DONG / (AVG_RANGE * 0.7)) - n_acc)  # 전국까지 추가계정(대략)
        self._cover_pct = pct
        self._alog(f"[커버] 계정 {n_acc}개 · 동네 {len(dongs)}곳 · 실효 ~{eff}지역"
                   f" / 전국 {KOREA_DONG}동 ({pct}%) · 전국까지 약 +{need_full}계정"
                   " — 계정은 서로 다른 동네에 두어야 커버가 넓어집니다")
        self._refresh_alert_health()

    def _build_auto_area_tree(self, parent):
        """자동용 지역 트리 — 수동과 동일한 시도>구>동 3단계. 미선택 시 전국."""
        import json as _json
        tree = toss_tree(QtWidgets.QTreeWidget(parent))
        tree.setMinimumWidth(240)
        self.auto_area_leaves = []
        CK = Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsAutoTristate
        try:
            with open("./OUT.json", encoding="utf-8") as _f:
                data = _json.load(_f)
        except Exception:
            return tree
        # 수동 _init_tree 와 동일: 블록(구) 단위 순회 → 같은 동 코드 보장
        sido = {}          # name1(시도) -> [블록(구), ...]
        for block in data:
            sido.setdefault(block["name1"], []).append(block)
        # 시도가 최상위다. '지역' 루트 행은 [전체 선택] 버튼이 대신한다.
        # 행에는 그 단계 이름만(종로구 · 부암동), 전체 경로는 AREA_FULL_ROLE.
        for s in sido:                      # 원본 순서(수동과 동일). 정렬 안 함
            top = QtWidgets.QTreeWidgetItem(tree, [s])
            top.setData(0, AREA_FULL_ROLE, s)
            top.setFlags(top.flags() | CK); top.setCheckState(0, Qt.CheckState.Unchecked)
            for block in sido[s]:
                gu_txt = f"{block['name1']} {block['name2']}".strip()
                guit = QtWidgets.QTreeWidgetItem(top, [(block["name2"] or "").strip() or gu_txt])
                guit.setData(0, AREA_FULL_ROLE, gu_txt)
                guit.setFlags(guit.flags() | CK); guit.setCheckState(0, Qt.CheckState.Unchecked)
                for loc in block["locations"]:
                    leaf = QtWidgets.QTreeWidgetItem(guit, [loc["name"]])
                    leaf.setData(0, AREA_FULL_ROLE, f"{gu_txt} {loc['name']}".strip())
                    leaf.setFlags(leaf.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                    leaf.setCheckState(0, Qt.CheckState.Unchecked)
                    leaf.setData(0, Qt.ItemDataRole.UserRole, f"{loc['name']}-{loc['id']}")
                    self.auto_area_leaves.append(leaf)
        return tree

    def _tree_panel(self, tree, leaves):
        """트리 + 지역검색 + 전체선택/해제 + 선택 칩 (토스 결).

        검색은 전체 경로(AREA_FULL_ROLE) 부분일치 — 행에는 '부암동'만 보여도
        '종로'로 찾힌다. 검색 중엔 맞는 동이 없는 시도·구 행을 숨긴다.
        고른 동은 검색줄 아래 칩으로 보인다(× 로 해제) — 트리가 접혀 있어도
        무엇을 골랐는지 한눈에 본다."""
        panel = QtWidgets.QWidget(self)
        panel.setObjectName("tossTreePanel")
        v = QtWidgets.QVBoxLayout(panel)
        v.setContentsMargins(0, 0, 0, 0); v.setSpacing(8)
        search = QtWidgets.QLineEdit(panel)
        search.setObjectName("tossSearch")
        search.setPlaceholderText("지역 검색  (예: 강남, 분당)")
        search.setClearButtonEnabled(True)
        search.setMinimumWidth(160)                  # 좁은 수동 탭 패널에서 "…" 로 접히지 않게

        def _parents():
            seen, out = set(), []
            for leaf in leaves:
                p = leaf.parent()
                while p is not None and id(p) not in seen:
                    seen.add(id(p)); out.append(p); p = p.parent()
            return out
        parents = _parents()

        def do_filter(text):
            text = text.strip()
            for leaf in leaves:
                match = (text == "" or text in area_full_name(leaf))
                leaf.setHidden(not match)
            # 리프를 정한 뒤 부모 — 보이는 자식이 하나도 없으면 숨긴다.
            # 깊은 쪽(구)부터 봐야 시도가 구의 결과를 읽는다.
            for p in sorted(parents, key=lambda it: -TossTreeDelegate._depth(
                    tree.indexFromItem(it))):
                if not text:
                    p.setHidden(False); continue
                any_vis = any(not p.child(i).isHidden() for i in range(p.childCount()))
                p.setHidden(not any_vis)
                p.setExpanded(any_vis)
        search.textChanged.connect(do_filter)

        # 검색·전체 선택/해제·선택 수를 한 줄에 — 트리 위 두 줄이 세로로
        # 쌓이면 지역 목록이 화면 아래로 밀린다.
        hb = QtWidgets.QHBoxLayout(); hb.setSpacing(4)
        hb.addWidget(search, 1)
        ball = QtWidgets.QPushButton("전체 선택", panel); ball.setObjectName("tossTextBtn")
        bclr = QtWidgets.QPushButton("전체 해제", panel); bclr.setObjectName("tossTextBtn")
        for b in (ball, bclr):
            b.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        cnt = QtWidgets.QLabel("0곳 선택", panel)
        cnt.setObjectName("tossCountChip")
        ball.clicked.connect(lambda: [l.setCheckState(0, Qt.CheckState.Checked)
                                      for l in leaves if not l.isHidden()])
        bclr.clicked.connect(lambda: [l.setCheckState(0, Qt.CheckState.Unchecked)
                                      for l in leaves])

        # 선택 칩 — 고른 동을 '종로구 부암동' 칩으로. 8개 넘으면 '외 N곳'.
        chips = QtWidgets.QWidget(panel)
        chips.setObjectName("tossChipRow")
        chips_l = FlowLayout(chips)
        chips.hide()
        CHIP_MAX = 8

        def _clear_chips():
            while chips_l.count():
                it = chips_l.takeAt(0)
                if it.widget():
                    it.widget().deleteLater()

        def _rebuild_chips(picked):
            _clear_chips()
            for leaf in picked[:CHIP_MAX]:
                gu = leaf.parent().text(0) if leaf.parent() is not None else ""
                b = QtWidgets.QPushButton(f"{gu} {leaf.text(0)}".strip() + "  ✕", chips)
                b.setObjectName("tossChip")
                b.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
                b.setToolTip("누르면 선택에서 뺍니다")
                b.clicked.connect(lambda _=False, l=leaf: l.setCheckState(0, Qt.CheckState.Unchecked))
                chips_l.addWidget(b)
            if len(picked) > CHIP_MAX:
                more = QtWidgets.QLabel(f"외 {len(picked) - CHIP_MAX:,}곳", chips)
                more.setObjectName("tossChipMore")
                chips_l.addWidget(more)
            chips.setVisible(bool(picked))
            chips.updateGeometry()

        # 디바운스: 시도/구 체크 시 itemChanged 수천 발생 → 매번 전체스캔하면 O(N²) 폭발.
        # 타이머로 변경을 모아 한 번만 카운트.
        timer = QtCore.QTimer(panel); timer.setSingleShot(True); timer.setInterval(150)

        def do_count():
            picked = [l for l in leaves if l.checkState(0) == Qt.CheckState.Checked]
            cnt.setText(f"{len(picked):,}곳 선택")
            cnt.setProperty("some", bool(picked))
            cnt.style().unpolish(cnt); cnt.style().polish(cnt)
            _rebuild_chips(picked)
        timer.timeout.connect(do_count)
        tree.itemChanged.connect(lambda *_: timer.start())
        hb.addSpacing(4); hb.addWidget(ball); hb.addWidget(bclr); hb.addSpacing(4); hb.addWidget(cnt)
        v.addLayout(hb)
        v.addWidget(chips)
        v.addWidget(tree, 1)
        return panel

    def _selected_auto_regions(self):
        return [it.data(0, Qt.ItemDataRole.UserRole)
                for it in getattr(self, "auto_area_leaves", [])
                if it.checkState(0) == Qt.CheckState.Checked]

    def _build_sweep_settings(self):
        """훑을 지역 — 고급 패널에 그대로 들어가는 위젯(테두리 없음).

        예전엔 '지역 훑기' 상자에 휴식·지역 간 간격·레인 스핀박스가 같이
        있었다. 클라가 만질 값이 아니고, 나란히 있으면 지역을 고르다 그것도
        건드린다. 화면에는 지역 트리와 전국 체크 하나만 남긴다 — 나머지는
        살려 두되 숨긴다(저장·복원과 옛 설정 파일이 계속 그 값을 쓴다)."""
        box = QtWidgets.QWidget(self)
        gv = QtWidgets.QVBoxLayout(box)
        gv.setContentsMargins(0, 0, 0, 0); gv.setSpacing(6)

        # 지역 — 미선택이면 기본 지역(서울·경기). 전국은 아래 체크박스로만.
        self.autoAreaTree = self._build_auto_area_tree(box)
        area = self._tree_panel(self.autoAreaTree, self.auto_area_leaves)
        area.setMaximumHeight(520)                   # 행 40px — 검색줄·칩 빼고 여덟 줄쯤
        gv.addWidget(area)

        # 전국은 의식적으로 켜야 한다 — 동 6537곳 × 키워드가 한 사이클이라
        # 계정 하루 상한(300)의 수십 배다. 미선택이 곧 전국이던 옛 규칙은
        # 서버에서 '아무 설정 없음 = 전국'이 되어 첫 사이클에 예산을 말렸다.
        self.autoNationwide = QtWidgets.QCheckBox("전국 훑기", box)
        self.autoNationwide.setChecked(False)
        _dflt_n = len(default_sweep_regions("./OUT.json"))
        self.autoNationwide.setToolTip(
            "체크하면 전국 동 단위(약 6537곳)를 훑는다. 한 사이클 요청 =\n"
            "지역 수 × 조건 수라 계정 하루 상한(300)을 금방 넘긴다 —\n"
            f"끄면 위에서 고른 지역만, 아무것도 안 고르면 기본 {_dflt_n}동"
            "(서울·경기)만 훑는다.")
        _nw = QtWidgets.QHBoxLayout(); _nw.setSpacing(8)
        _nw.addWidget(self.autoNationwide)
        _nl = QtWidgets.QLabel(f"아무것도 안 고르면 서울·경기 {_dflt_n}동을 훑습니다")
        _nl.setObjectName("mutedNote")
        _nw.addWidget(_nl)
        _nw.addStretch(1)
        gv.addLayout(_nw)

        # ── 웹 동 피드(계정 없이 발굴) ──
        self.feedEnabledChk = QtWidgets.QCheckBox("동 피드 발굴 (계정 없음)", box); self.feedEnabledChk.setChecked(True)
        self.feedCat31 = QtWidgets.QCheckBox("여성잡화", box); self.feedCat31.setChecked(True)
        self.feedCat14 = QtWidgets.QCheckBox("남성패션/잡화", box); self.feedCat14.setChecked(True)
        self.feedCat5 = QtWidgets.QCheckBox("여성의류", box); self.feedCat5.setChecked(False)
        _fc = QtWidgets.QHBoxLayout(); _fc.setSpacing(10)
        for w in (self.feedCat31, self.feedCat14, self.feedCat5): _fc.addWidget(w)
        _fc.addStretch(1)
        self.feedProxies = QtWidgets.QPlainTextEdit(box); self.feedProxies.setPlaceholderText("웹 프록시 한 줄에 하나 (비우면 proxies.txt)")
        self.feedProxies.setMaximumHeight(72)
        self.feedRps = QtWidgets.QDoubleSpinBox(box); self.feedRps.setRange(0.1, 5.0); self.feedRps.setDecimals(1); self.feedRps.setSingleStep(0.1); self.feedRps.setValue(1.0); self.feedRps.setFixedWidth(72)
        self.feedRestMin = QtWidgets.QSpinBox(box); self.feedRestMin.setRange(0, 120); self.feedRestMin.setValue(2); self.feedRestMin.setFixedWidth(72)
        self.sweepAppChk = QtWidgets.QCheckBox("앱 키워드 스윕(보완층, 스윕 계정만)", box); self.sweepAppChk.setChecked(False)
        gv.addWidget(self._setting_row("동 피드", self.feedEnabledChk))
        _fcw = QtWidgets.QWidget(box); _fcw.setLayout(_fc)
        gv.addWidget(self._setting_row("카테고리", _fcw))
        gv.addWidget(self._setting_row("웹 프록시", self.feedProxies))
        _frps = QtWidgets.QHBoxLayout(); _frps.setContentsMargins(0, 0, 0, 0)
        _frps.addWidget(self.feedRps); _frps.addStretch(1)
        _frpsw = QtWidgets.QWidget(box); _frpsw.setLayout(_frps)
        gv.addWidget(self._setting_row("초당 요청/레인", _frpsw))
        _frest = QtWidgets.QHBoxLayout(); _frest.setContentsMargins(0, 0, 0, 0)
        _frest.addWidget(self.feedRestMin); _frest.addStretch(1)
        _frestw = QtWidgets.QWidget(box); _frestw.setLayout(_frest)
        gv.addWidget(self._setting_row("사이클 휴식(분)", _frestw))
        gv.addWidget(self._setting_row("앱 스윕", self.sweepAppChk))
        for w in (self.feedEnabledChk, self.feedCat31, self.feedCat14, self.feedCat5, self.sweepAppChk):
            w.toggled.connect(lambda *_: self._save_alert_settings(self._feed_settings_patch()))
        for w in (self.feedRps, self.feedRestMin):
            w.valueChanged.connect(lambda *_: self._save_alert_settings(self._feed_settings_patch()))
        self.feedProxies.textChanged.connect(lambda: self._save_alert_settings(self._feed_settings_patch()))

        self.autoExtra = QtWidgets.QLineEdit(box); self.autoExtra.setPlaceholderText("추가 키워드")
        self.autoExclude = QtWidgets.QLineEdit(box); self.autoExclude.setPlaceholderText("제외 키워드")
        self.autoMin = QtWidgets.QLineEdit(box); self.autoMin.setPlaceholderText("최소가"); self.autoMin.setFixedWidth(96)
        self.autoMax = QtWidgets.QLineEdit(box); self.autoMax.setPlaceholderText("최대가"); self.autoMax.setFixedWidth(96)
        self.autoDays = QtWidgets.QSpinBox(box); self.autoDays.setRange(0, 365); self.autoDays.setValue(7); self.autoDays.setFixedWidth(72)
        # 사이클 휴식(초) — 하한 10s: 그 아래는 무휴식 폴링 = 봇 패턴 → 차단
        self.autoRestMin = QtWidgets.QSpinBox(box); self.autoRestMin.setRange(10, 3600); self.autoRestMin.setValue(30); self.autoRestMin.setFixedWidth(72)
        self.autoRestMax = QtWidgets.QSpinBox(box); self.autoRestMax.setRange(10, 3600); self.autoRestMax.setValue(90); self.autoRestMax.setFixedWidth(72)
        # 지역 간 휴식(초) — 전국 구단위 수백 요청 사이 간격. 0.3s 미만 연타 = IP 스로틀
        self.autoGapMin = QtWidgets.QDoubleSpinBox(box); self.autoGapMin.setRange(0.3, 10.0)
        self.autoGapMin.setDecimals(1); self.autoGapMin.setSingleStep(0.1)
        self.autoGapMin.setValue(0.4); self.autoGapMin.setFixedWidth(72)
        self.autoGapMax = QtWidgets.QDoubleSpinBox(box); self.autoGapMax.setRange(0.3, 10.0)
        self.autoGapMax.setDecimals(1); self.autoGapMax.setSingleStep(0.1)
        self.autoGapMax.setValue(1.2); self.autoGapMax.setFixedWidth(72)
        # min<=max 강제 — 뒤집힌 범위 입력 자체를 막는다
        for lo, hi in ((self.autoRestMin, self.autoRestMax), (self.autoGapMin, self.autoGapMax)):
            hi.setMinimum(lo.value()); lo.setMaximum(hi.value())
            lo.valueChanged.connect(hi.setMinimum)
            hi.valueChanged.connect(lo.setMaximum)
        self.autoRestMin.setToolTip("사이클(전국 1바퀴) 사이 랜덤 휴식. 10~3600초")
        self.autoRestMax.setToolTip("사이클(전국 1바퀴) 사이 랜덤 휴식. 10~3600초")
        self.autoGapMin.setToolTip("지역(구) 요청 사이 랜덤 휴식. 0.3~10.0초")
        self.autoGapMax.setToolTip("지역(구) 요청 사이 랜덤 휴식. 0.3~10.0초")
        # 레인 = 동시에 도는 수집 갈래. 앱API(토큰 있음)는 IP 를 공유해 동시요청하므로
        # 0 = APP_API_LANES(8, 실측 13 req/s 무스로틀). 웹크롤 폴백 경로만 프록시를
        # 샤딩해 프록시 수 ÷ 3 으로 묶인다(같은 IP 동시요청 = 빈응답).
        self.autoLanes = QtWidgets.QSpinBox(box)
        self.autoLanes.setRange(0, 16); self.autoLanes.setValue(0)
        self.autoLanes.setSpecialValueText("자동"); self.autoLanes.setFixedWidth(72)
        self.autoLanes.setToolTip(
            f"동시 수집 갈래(레인) 수. 0=자동(앱API {APP_API_LANES}개, 실측 무스로틀 상한).\n"
            "앱API 는 레인이 IP 를 공유한다 — 프록시 수와 무관.\n"
            "웹크롤 폴백은 프록시 수 ÷ 3 으로 자동 조정된다(같은 IP 동시요청 = 빈응답).")
        # 토큰 갱신 체크박스는 없다 — LDPlayer 앱이 갱신한 토큰을 수확하는 것이
        # 유일한 경로라(PC 직접 갱신은 WAF 가 막는다) 끄는 순간 토큰 0 이 된다.
        self._notify = self._load_notify()

        # 가격·추가·제외·끌올은 조건표가, 휴식·지역 간 간격·레인은 기본값이
        # 정한다. 화면에 남겨 두면 적어도 아무 일이 안 일어나서 "적었는데 왜
        # 안 걸리지"만 남는다. 위젯 자체는 살려 둔다 — 저장·복원과 옛 설정
        # 파일이 이 값을 계속 쓴다.
        self._sweepLegacy = QtWidgets.QWidget(box)
        self._sweepLegacy._keepHidden = True
        _lg = QtWidgets.QHBoxLayout(self._sweepLegacy)
        _lg.setContentsMargins(0, 0, 0, 0)
        for _w in (self.autoMin, self.autoMax, self.autoExtra,
                   self.autoExclude, self.autoDays,
                   self.autoRestMin, self.autoRestMax,
                   self.autoGapMin, self.autoGapMax, self.autoLanes,
                   self.alertPollInterval, self.alertCoverMode):
            _lg.addWidget(_w)
        self._sweepLegacy.setVisible(False)
        gv.addWidget(self._sweepLegacy)

        # 시작 버튼은 없다 — 스윕은 감시 토글이 켜고 끈다. 알림·계정 설정은
        # 설정 탭에 있다. 조건은 조건 탭의 표 하나로 모았다.
        # 여기에도 엑셀 버튼이 있던 동안, 등록용과 알림용이 따로 있는 줄
        # 알고 클라가 같은 시트를 양쪽에 넣었다.

        # 복원이 먼저, 배선이 나중이다 — 순서가 바뀌면 복원값이 저장을 유발해
        # 첫 실행에서 '기본값을 사용자가 고른 값'으로 굳혀 버린다.
        self._restore_sweep_settings()
        self._wire_sweep_settings(box)
        return box

    # ── 스윕 설정 영속화 ──
    # 이 값들은 GUI 위젯에만 있었고 alert_settings.json 에는 한 번도 쓰이지
    # 않았다. 그래서 서버(headless_sweep_cfg)는 늘 빈 설정을 읽어 전국으로
    # 떨어졌고, 운영자에게는 좁힐 레버가 아예 없었다. 여기서 위젯 값을
    # headless_sweep_cfg 가 읽는 바로 그 키로 저장한다.
    def _sweep_settings_patch(self):
        return {
            "sweep_regions": self._selected_auto_regions(),
            SWEEP_NATIONWIDE_KEY: bool(self.autoNationwide.isChecked()),
            "sweep_extra": self._splt(self.autoExtra.text()),
            "sweep_exclude": self._splt(self.autoExclude.text()),
            "sweep_min": self._num(self.autoMin.text()),
            "sweep_max": self._num(self.autoMax.text()),
            "sweep_days": int(self.autoDays.value()),
            "sweep_rest_min": int(self.autoRestMin.value()),
            "sweep_rest_max": int(self.autoRestMax.value()),
            "sweep_gap_min": float(self.autoGapMin.value()),
            "sweep_gap_max": float(self.autoGapMax.value()),
            "sweep_lanes": int(self.autoLanes.value()),
        }

    def _restore_sweep_settings(self):
        """저장된 값을 위젯에 되돌린다. 없으면 위젯 초기값 그대로 둔다."""
        s = self._load_alert_settings()
        if not isinstance(s, dict):
            return

        def _txt(widget, key):
            v = s.get(key)
            if isinstance(v, (list, tuple)):
                v = ", ".join(str(x) for x in v)
            if v is not None:
                widget.setText(str(v))

        def _spin(widget, key):
            v = s.get(key)
            if v is None:
                return
            try:
                widget.setValue(type(widget.value())(v))
            except (TypeError, ValueError):
                pass

        def _spin_range(lo, hi, lo_key, hi_key):
            """min<=max 커플링을 잠시 풀고 두 값을 되돌린 뒤 다시 건다.

            커플링(lo.valueChanged→hi.setMinimum, hi.valueChanged→lo.setMaximum)이
            걸린 채로 복원하면 지금 위젯에 들어 있는 값이 저장값을 잘라낸다.
            어느 쪽을 먼저 넣어도 마찬가지라 순서로는 못 피한다 — 현재 하한이
            200 이면 상한 140 은 200 으로 잘리고, 그 반대도 같다.
            커플링이 건드리지 않는 lo.minimum()/hi.maximum() 이 원래 한계다."""
            floor, ceil = lo.minimum(), hi.maximum()
            lo.setMaximum(ceil)
            hi.setMinimum(floor)
            _spin(lo, lo_key)
            _spin(hi, hi_key)
            hi.setMinimum(lo.value())
            lo.setMaximum(hi.value())
        try:
            _txt(self.autoExtra, "sweep_extra")
            _txt(self.autoExclude, "sweep_exclude")
            _txt(self.autoMin, "sweep_min")
            _txt(self.autoMax, "sweep_max")
            _spin(self.autoDays, "sweep_days")
            _spin_range(self.autoRestMin, self.autoRestMax,
                        "sweep_rest_min", "sweep_rest_max")
            _spin_range(self.autoGapMin, self.autoGapMax,
                        "sweep_gap_min", "sweep_gap_max")
            _spin(self.autoLanes, "sweep_lanes")
            self.autoNationwide.setChecked(bool(s.get(SWEEP_NATIONWIDE_KEY)))
            want = {str(r) for r in (s.get("sweep_regions") or []) if r}
            if want:
                for leaf in getattr(self, "auto_area_leaves", []):
                    if leaf.data(0, Qt.ItemDataRole.UserRole) in want:
                        leaf.setCheckState(0, Qt.CheckState.Checked)
            if s.get("feed_enabled") is not None: self.feedEnabledChk.setChecked(bool(s["feed_enabled"]))
            cats = s.get("feed_categories")
            if isinstance(cats, list):
                self.feedCat31.setChecked(31 in cats); self.feedCat14.setChecked(14 in cats); self.feedCat5.setChecked(5 in cats)
            if isinstance(s.get("feed_proxies"), list): self.feedProxies.setPlainText("\n".join(s["feed_proxies"]))
            if s.get("feed_rps") is not None: self.feedRps.setValue(float(s["feed_rps"]))
            if s.get("feed_rest_min") is not None: self.feedRestMin.setValue(int(s["feed_rest_min"]))
            self.sweepAppChk.setChecked(sweep_app_enabled(s))
        except Exception as e:
            print(f"[스윕설정 복원 실패] {type(e).__name__}: {e}")

    def _wire_sweep_settings(self, parent):
        """위젯 변경 → alert_settings.json 저장. 디바운스 필수 —
        시도/구를 한 번 체크하면 itemChanged 가 수백 발 나온다."""
        self._sweep_save_timer = QtCore.QTimer(parent)
        self._sweep_save_timer.setSingleShot(True)
        self._sweep_save_timer.setInterval(300)
        self._sweep_save_timer.timeout.connect(self._save_sweep_settings)
        kick = self._sweep_save_timer.start
        self.autoAreaTree.itemChanged.connect(lambda *_: kick())
        self.autoNationwide.toggled.connect(lambda *_: kick())
        for w in (self.autoExtra, self.autoExclude, self.autoMin, self.autoMax):
            w.textChanged.connect(lambda *_: kick())
        for w in (self.autoDays, self.autoRestMin, self.autoRestMax,
                  self.autoGapMin, self.autoGapMax, self.autoLanes):
            w.valueChanged.connect(lambda *_: kick())

    def _save_sweep_settings(self):
        self._save_alert_settings(self._sweep_settings_patch())

    # ── 알림 설정 저장/불러오기 ──
    NOTIFY_FILE = "./notify.json"
    NOTIFY_DEFAULT = {"tg_token": "", "tg_chat": "", "sheet_url": "",
                      "sheet_cred": "./credentials.json"}

    def _load_notify(self):
        """notify.json 에서 알림 설정 복원. 없거나 깨졌으면 기본값."""
        cfg = dict(self.NOTIFY_DEFAULT)
        try:
            with open(self.NOTIFY_FILE, encoding="utf-8") as f:
                saved = json.load(f)
            if isinstance(saved, dict):
                for k in cfg:
                    if isinstance(saved.get(k), str):
                        cfg[k] = saved[k]
        except FileNotFoundError:
            pass
        except Exception as e:
            print(f"[알림설정 로드 실패] {type(e).__name__}: {e}")
        return cfg

    def _save_notify(self):
        """알림 설정 저장. 봇 토큰이 들어가므로 파일 권한 0600."""
        try:
            with open(self.NOTIFY_FILE, "w", encoding="utf-8") as f:
                json.dump(self._notify, f, ensure_ascii=False, indent=2)
            try:
                os.chmod(self.NOTIFY_FILE, 0o600)
            except OSError:
                pass
            return True, ""
        except Exception as e:
            return False, f"{type(e).__name__}: {e}"

    def _build_notify_form(self, box, lay=None):
        """알림 설정 — 설정 탭에 바로 펼친 폼(텔레그램·구글시트).

        예전엔 '알림 설정' 버튼 → 다이얼로그였다. 고급 패널을 펴고, 버튼을
        찾고, 창을 띄우는 세 걸음 뒤에야 토큰 칸이 나왔다. 설정 탭이 생긴
        뒤로는 폼이 그 자리에 있다. 값은 self._notify 가 원본이고 저장은
        notify.json 이다 — 다이얼로그 때와 같다.
        lay 를 주면 그 레이아웃(카드 본문)에 이어 붙인다."""
        n = self._notify
        v = lay if lay is not None else QtWidgets.QVBoxLayout(box)
        v.setSpacing(6)
        self.notifyToken = QtWidgets.QLineEdit(n["tg_token"], box)
        self.notifyToken.setPlaceholderText("텔레그램 봇 토큰 (예: 123456:AA...)")
        self.notifyChat = QtWidgets.QLineEdit(n["tg_chat"], box)
        self.notifyChat.setPlaceholderText("chat_id / 방 (예: -1001234567890)")
        self.notifySheet = QtWidgets.QLineEdit(n["sheet_url"], box)
        self.notifySheet.setPlaceholderText("구글시트 주소(선택)")
        self.notifyCred = QtWidgets.QLineEdit(n["sheet_cred"], box)
        self.notifyCred.setPlaceholderText("구글 서비스계정 JSON 키 경로(시트 쓸 때만 필요)")
        self.notifyCredBtn = QtWidgets.QPushButton("찾기", box)
        self.notifyCredBtn.setFixedWidth(60)
        credRow = QtWidgets.QWidget(box); credLay = QtWidgets.QHBoxLayout(credRow)
        credLay.setContentsMargins(0, 0, 0, 0); credLay.setSpacing(6)
        credLay.addWidget(self.notifyCred, 1); credLay.addWidget(self.notifyCredBtn)

        def pick_cred():
            path, _ = QtWidgets.QFileDialog.getOpenFileName(
                self, "서비스계정 JSON 키 선택", "", "JSON (*.json)")
            if path:
                self.notifyCred.setText(path)
        self.notifyCredBtn.clicked.connect(pick_cred)

        rows = QtWidgets.QVBoxLayout(); rows.setSpacing(0)
        fields = (("텔레그램 토큰", self.notifyToken), ("텔레그램 방", self.notifyChat),
                  ("구글시트", self.notifySheet), ("시트 인증파일", credRow))
        for i, (lab, wdg) in enumerate(fields):
            rows.addWidget(self._setting_row(lab, wdg, last=(i == len(fields) - 1)))
        v.addLayout(rows)
        _hint = QtWidgets.QLabel("[저장]을 눌러야 적용되고, notify.json 에 남습니다.", box)
        _hint.setObjectName("mutedNote"); _hint.setWordWrap(True)
        v.addWidget(_hint)
        self.notifyResult = QtWidgets.QLabel("", box); self.notifyResult.setWordWrap(True)
        self.notifyResult.setStyleSheet("color:#5C5449; font-size:14px;")
        v.addWidget(self.notifyResult)

        bb = QtWidgets.QHBoxLayout()
        self.notifyTestBtn = QtWidgets.QPushButton("테스트 발송", box)
        self.notifyTestBtn.setObjectName("linkBtn")
        self.notifyTestBtn.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.notifySaveBtn = QtWidgets.QPushButton("저장", box)
        self.notifySaveBtn.setObjectName("startBtn")
        bb.addWidget(self.notifyTestBtn); bb.addStretch(1); bb.addWidget(self.notifySaveBtn)
        v.addLayout(bb)
        self.notifyTestBtn.clicked.connect(self.on_notify_test)
        self.notifySaveBtn.clicked.connect(self.on_notify_save)

    def _collect_notify(self):
        return {"tg_token": self.notifyToken.text().strip(),
                "tg_chat": self.notifyChat.text().strip(),
                "sheet_url": self.notifySheet.text().strip(),
                "sheet_cred": self.notifyCred.text().strip() or "./credentials.json"}

    def _refresh_notify_form(self):
        """self._notify → 폼. 파일에서 다시 읽었을 때 화면을 맞춘다."""
        n = self._notify
        self.notifyToken.setText(n["tg_token"]); self.notifyChat.setText(n["tg_chat"])
        self.notifySheet.setText(n["sheet_url"]); self.notifyCred.setText(n["sheet_cred"])

    def on_notify_test(self):
        """폼에 적힌 값으로 시험 발송 — 저장 전 값이다. 결과는 폼 아래 줄에."""
        cur = self._collect_notify()
        result, test = self.notifyResult, self.notifyTestBtn
        if not (cur["tg_token"] and cur["tg_chat"]) and not cur["sheet_url"]:
            result.setStyleSheet("color:#5C5449; font-size:14px;")
            result.setText("⚠️ 텔레그램(토큰+방) 또는 구글시트 주소를 먼저 입력하세요.")
            return
        test.setEnabled(False); test.setText("보내는 중…")
        result.setStyleSheet("color:#5C5449; font-size:14px;")
        result.setText("전송 시도 중…")
        # 부모는 MainWindow — 창이 먼저 닫혀도 실행 중 스레드가 삭제되지 않게
        self._notify_test = NotifyTestThread(self, cur)

        def done(res):
            try:
                render(res)
            except RuntimeError:
                pass            # 결과 도착 전 창이 닫힌 경우

        def render(res):
            lines = []
            if cur["tg_token"] or cur["tg_chat"]:
                lines.append(("✅ 텔레그램: " if res["tg_ok"] else "❌ 텔레그램: ")
                             + res["tg_msg"])
            if res["sheet_ok"] is None:
                if cur["sheet_url"]:
                    lines.append("❌ 구글시트: " + res["sheet_msg"])
            else:
                lines.append(("✅ 구글시트: " if res["sheet_ok"] else "❌ 구글시트: ")
                             + res["sheet_msg"])
            bad = (cur["tg_token"] and not res["tg_ok"]) or res["sheet_ok"] is False
            result.setStyleSheet("color:#B4342A;" if bad else "color:#2E7D32;")
            result.setText("\n".join(lines) or "테스트할 항목 없음")
            test.setEnabled(True); test.setText("테스트 발송")
        self._notify_test.result.connect(done)
        self._notify_test.start()

    def on_notify_save(self):
        """폼 → self._notify → notify.json. 인증파일 공란은 기본 경로로 보정."""
        self._notify.update(self._collect_notify())
        self._refresh_notify_form()             # 보정된 값(기본 경로)을 되비춘다
        saved, err = self._save_notify()
        if not saved:
            self.alert(f"알림 설정 저장 실패 — {err}\n(이번 실행 동안만 적용됩니다)")
            self.notifyResult.setStyleSheet("color:#B4342A;")
            self.notifyResult.setText("❌ 저장 실패 — " + err)
            return False
        self.notifyResult.setStyleSheet("color:#2E7D32;")
        self.notifyResult.setText("✅ 저장했습니다")
        return True

    @staticmethod
    def _mask_proxy(p: str) -> str:
        """비밀번호 마스킹 표시."""
        if "@" in p and "://" in p:
            scheme, rest = p.split("://", 1)
            if "@" in rest:
                cred, host = rest.split("@", 1)
                user = cred.split(":", 1)[0]
                return f"{scheme}://{user}:****@{host}"
        return p

    def _proxy_sources(self):
        """[(proxy, source)] — source 는 'settings' | 'account'."""
        rows = []
        seen = set()
        for p in self.controller.proxies:
            if p and p not in seen:
                seen.add(p)
                rows.append((p, "settings"))
        try:
            from daangn_ext import AccountStore
            for p in AccountStore("./accounts.json").proxies():
                if p and p not in seen:
                    seen.add(p)
                    rows.append((p, "account"))
        except Exception:
            pass
        return rows

    def on_proxy_view_clicked(self):
        dlg = QtWidgets.QDialog(self)
        dlg.resize(620, 440)
        v = QtWidgets.QVBoxLayout(dlg)

        info = QtWidgets.QLabel(parent=dlg)
        info.setWordWrap(True)
        v.addWidget(info)

        listw = QtWidgets.QListWidget(dlg)
        listw.setSelectionMode(
            QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        v.addWidget(listw, 1)

        def reload_list():
            rows = self._proxy_sources()
            listw.clear()
            for proxy, src in rows:
                tag = "settings.txt" if src == "settings" else "계정저장소"
                item = QtWidgets.QListWidgetItem(
                    f"[{tag}] {self._mask_proxy(proxy)}")
                item.setData(Qt.ItemDataRole.UserRole, (proxy, src))
                listw.addItem(item)
            dlg.setWindowTitle(f"적용 프록시 {len(rows)}개")
            if rows:
                info.setText(
                    f"현재 적용 중인 프록시 {len(rows)}개 (settings.txt + 계정저장소). "
                    "여러 개 선택 후 삭제 가능.")
            else:
                info.setText(
                    "프록시 없음 — 아래 '추가' 버튼 또는 '계정+프록시 추가/관리' 사용")

        def on_add():
            text, ok = QtWidgets.QInputDialog.getText(
                dlg, "프록시 추가", "프록시 주소 (여러 개는 줄바꿈/쉼표로 구분)",
                QtWidgets.QLineEdit.EchoMode.Normal,
                "http://user:pass@host:port")
            if not ok:
                return
            cands = [c.strip() for c in text.replace(",", "\n").splitlines()]
            cands = [c for c in cands if c]
            added, errs = 0, []
            for c in cands:
                if "://" not in c:
                    c = "http://" + c
                err = self.controller.add_proxy(c)
                if err:
                    errs.append(f"{self._mask_proxy(c)}: {err}")
                else:
                    added += 1
            reload_list()
            if errs:
                QtWidgets.QMessageBox.warning(
                    dlg, "일부 추가 실패",
                    f"{added}개 추가됨.\n\n" + "\n".join(errs))

        def on_del():
            items = listw.selectedItems()
            if not items:
                QtWidgets.QMessageBox.information(
                    dlg, "삭제", "삭제할 프록시를 선택하세요.")
                return
            picked = [it.data(Qt.ItemDataRole.UserRole) for it in items]
            names = "\n".join(self._mask_proxy(p) for p, _ in picked)
            if not ask_yes_no(
                    dlg, "프록시 삭제", f"프록시 {len(picked)}개를 삭제할까요?",
                    f"{names}\n\n계정저장소 항목은 계정은 남고 프록시 연결만"
                    " 해제됩니다.", danger=True):
                return

            errs = []
            acct_targets = [p for p, src in picked if src == "account"]
            for proxy, src in picked:
                if src == "settings":
                    err = self.controller.remove_proxy(proxy)
                    if err:
                        errs.append(f"{self._mask_proxy(proxy)}: {err}")
            if acct_targets:
                try:
                    from daangn_ext import AccountStore
                    store = AccountStore("./accounts.json")
                    for r in store.rows:
                        if r.get("proxy") in acct_targets:
                            r["proxy"] = None
                    store.save()
                except Exception as e:
                    errs.append(f"계정저장소 저장 실패: {e}")

            reload_list()
            if errs:
                QtWidgets.QMessageBox.warning(
                    dlg, "일부 삭제 실패", "\n".join(errs))

        btns = QtWidgets.QHBoxLayout()
        addBtn = QtWidgets.QPushButton("추가", dlg)
        addBtn.clicked.connect(on_add)
        delBtn = QtWidgets.QPushButton("삭제", dlg)
        delBtn.clicked.connect(on_del)
        closeBtn = QtWidgets.QPushButton("닫기", dlg)
        closeBtn.clicked.connect(dlg.accept)
        btns.addWidget(addBtn); btns.addWidget(delBtn)
        btns.addStretch(1); btns.addWidget(closeBtn)
        v.addLayout(btns)

        reload_list()
        dlg.exec()

    def _wrap(self, layout):
        c = QtWidgets.QWidget(self); c.setLayout(layout); return c

    def _on_rules_edited(self):
        """표가 바뀌었다 — 적용 전까지 '수정됨' 으로 보인다."""
        self._rules_dirty = True
        self._refresh_rules_view()

    def on_rules_apply(self):
        """표의 조건을 읽어 확인받고 조건표로 저장하고 브랜드를 등록한다.

        오류 줄은 붉게 칠하고 건너뛴다 — 나머지는 들어간다. 빈 표는 조건
        삭제가 아니라 거절이다. 지우는 길은 [전체 삭제] 하나여야 등록 해제까지
        같이 간다."""
        from daangn_ext.alert_rules import parse_rule_rows
        from daangn_ext.rule_grid import grid_to_rows
        cells = self.rulesGrid.cells()
        if not cells:
            QtWidgets.QMessageBox.information(
                self, "빈 조건표",
                "적은 조건이 없습니다. 표에 브랜드부터 적으세요.\n"
                "조건을 모두 지우려면 [전체 삭제]를 쓰세요.")
            return
        rules, errors = parse_rule_rows(grid_to_rows(cells))
        self.rulesGrid.mark_errors(errors)
        self._apply_rules(rules, errors, self)

    def on_rules_import_excel(self):
        """엑셀 파일을 읽어 표를 채운다 — 적용은 [조건 적용]이 한다.

        수백 줄을 옮길 때는 붙여넣기보다 파일이 안전하다. 표를 통째로 바꾸고
        '수정됨' 으로 두어, 무엇이 들어왔는지 보고 나서 적용하게 한다."""
        from daangn_ext.alert_rules import load_rules_from_excel
        from daangn_ext.rule_grid import rules_to_grid
        p, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "조건 엑셀 선택", "", "Excel (*.xlsx *.xlsm)")
        if not p:
            return
        try:
            rules, errors = load_rules_from_excel(p)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "오류", f"엑셀 로드 오류:\n{e}")
            return
        if not rules:
            QtWidgets.QMessageBox.warning(
                self, "실패", "읽은 조건이 없습니다.\n" + "\n".join(errors[:3]))
            return
        self.rulesGrid.set_cells(rules_to_grid(rules))
        for msg in errors[:8]:
            self._alog(f"[조건표] {msg}")
        self._alog(f"[조건표] 엑셀 {os.path.basename(p)} → 표에 {len(rules)}줄"
                   " 불러옴 — [조건 적용]을 누르면 반영됩니다")
        self._on_rules_edited()

    def _apply_rules(self, rules, errors, parent) -> bool:
        """읽은 룰을 확인받고 조건표로 저장하고 브랜드를 등록한다."""
        from daangn_ext.alert_rules import RuleTable, brands
        if not rules:
            QtWidgets.QMessageBox.warning(
                parent, "실패", "읽은 조건이 없습니다.\n" + "\n".join(errors[:3]))
            return False
        bs = brands(rules)
        if not self._confirm_rules(parent, rules, bs, errors):
            return False
        for msg in errors[:8]:
            self._alog(f"[조건표] {msg}")
        os.makedirs(os.path.dirname(ALERT_RULES_FILE), exist_ok=True)
        RuleTable(rules).save(ALERT_RULES_FILE)
        # 방금 저장한 것은 표에서 나온 것이다 — 다시 채워 오류 줄을 지우지 않는다.
        self._grid_seen_stamp = self._alert_rules.stamp()
        self._rules_dirty = False
        self._refresh_rules_view()
        self._alog(f"[조건표] {len(rules)}줄 적용 · 브랜드 {len(bs)}개 등록"
                   f" ({', '.join(bs[:8])}{' …' if len(bs) > 8 else ''})")
        self._register_rule_brands(parent, bs, rules)
        QtWidgets.QMessageBox.information(
            parent, "적용됨",
            f"조건 {len(rules)}줄을 적용했습니다.\n"
            f"브랜드 {len(bs)}개를 등록 중입니다 — 수집은 넓게,"
            " 알림은 이 조건표로 거릅니다.")
        return True

    def _confirm_rules(self, parent, rules, bs, errors) -> bool:
        """적용 전에 무엇이 들어가는지 먼저 보여준다 — 엉뚱한 파일을
        고른 것을 되돌릴 수 있는 지점은 여기뿐이다."""
        try:
            free = int((self._router.capacity() or {}).get("free") or 0)
            cap_line = f"슬롯 여유: {free}칸 (브랜드 {len(bs)}개 필요)"
            if len(bs) > free:
                cap_line += f" — {len(bs) - free}개는 검색 스윕으로 돌립니다"
        except Exception:
            cap_line = "슬롯 여유: 확인 불가"
        box = QtWidgets.QMessageBox(parent)
        box.setWindowTitle("이대로 적용할까요?")
        box.setText(f"조건 {len(rules)}줄 읽음\n"
                    f"등록할 브랜드 {len(bs)}개: {' · '.join(bs[:8])}"
                    + (" …" if len(bs) > 8 else "") + f"\n{cap_line}"
                    + (f"\n확인할 내용 {len(errors)}건" if errors else ""))
        if errors:
            box.setDetailedText("\n".join(errors))
        box.setStandardButtons(QtWidgets.QMessageBox.StandardButton.Ok
                               | QtWidgets.QMessageBox.StandardButton.Cancel)
        box.button(QtWidgets.QMessageBox.StandardButton.Ok).setText("적용")
        box.button(QtWidgets.QMessageBox.StandardButton.Cancel).setText("취소")
        return box.exec() == QtWidgets.QMessageBox.StandardButton.Ok

    def _register_rule_brands(self, parent, bs, rules):
        """브랜드만 등록한다. 가격·제외는 넘기지 않는다 — 당근 서버가
        먼저 거르면 조건표가 볼 매물 자체가 없어진다."""
        if not self._router:
            self._alog("[조건표] 라우터가 없어 브랜드 등록을 건너뜁니다")
            return
        groups = brand_register_groups(bs, rules)
        co = self._core_only()

        def job(log):
            res = []
            for ks, d in groups:
                res.extend(self._router.add_many(
                    ks, None, None, None, core_only=co, log=log,
                    days=d, replace_cond=True) or [])
            # 조건표에서 빠진 브랜드는 라우터·서버에서 지운다 — 엑셀이 진실이다.
            prune_to_rules(self._router, self._alert_fleet(),
                           rule_brand_keys(self._alert_rules.get()), log, co)
            return {"routes": res, "list": self._safe_alert_list(log)}
        # queue=True — 자동수확·자동폴링이 도는 중이라도 버리지 않는다.
        if not self._alert_run(job, self._alert_routes_done, queue=True,
                               label="브랜드 등록 중"):
            QtWidgets.QMessageBox.warning(
                parent, "등록 보류",
                "조건표는 적용했지만 지금은 브랜드 등록을 시작할 수 없습니다.\n"
                "로그를 확인하고 '목록 새로고침' 후 다시 시도하세요.")

    def _account_proxies(self):
        """accounts.json 프록시 — mtime 캐시 (자동 모니터가 구마다 호출하므로)."""
        path = "./accounts.json"
        try:
            st = os.stat(path)
            stamp = (st.st_mtime_ns, st.st_size)
        except OSError:
            self._acct_proxy_cache = (None, [])
            return []
        cached_stamp, cached = getattr(self, "_acct_proxy_cache", (None, []))
        if cached_stamp == stamp:
            return cached
        try:
            from daangn_ext import AccountStore
            proxies = AccountStore(path).proxies()
        except Exception:
            proxies = []
        self._acct_proxy_cache = (stamp, proxies)
        return proxies

    def _collect_proxies(self):
        """settings.txt 프록시 + 계정저장소 프록시 합침 → 자동 로테이션용."""
        proxies = list(self.controller.proxies) + self._account_proxies()
        return [p for p in dict.fromkeys(proxies) if p]

    def _auto_cfg_base(self):
        """검색 스윕 공통 cfg — 옛 on_auto_start_clicked 에서 그대로 떼어냈다.

        keyword/extra 는 넣지 않는다. 대기열 키워드를 _sweep_cfg 가 conditions 로
        채우기 때문이다. 나머지 값(속도·프록시·계정·알림)은 의미가 그대로다."""
        cfg = {
            "rest_min": self.autoRestMin.value(),
            "rest_max": self.autoRestMax.value(),
            "gap_min": self.autoGapMin.value(),
            "gap_max": self.autoGapMax.value(),
            "lanes": self.autoLanes.value(),        # 0 = 자동(프록시 수 기준)
            "tg_token": self._notify["tg_token"] or None,
            "tg_chat": self._notify["tg_chat"] or None,
            "sheet_url": self._notify["sheet_url"] or None,
            "sheet_cred": self._notify.get("sheet_cred") or "./credentials.json",
            "proxies": self._collect_proxies(),
            # 실행 중 프록시 추가/삭제 반영용 (조건 루프마다 재조회)
            "proxy_provider": self._collect_proxies,
            # 초기 토큰도 워커스레드의 token_provider 가 파일에서 읽는다(메인스레드
            # 접근 없음). 파일을 신선하게 유지하는 건 _HarvestThread 다.
            # 같은 매물을 앱 알림이 이미 알렸으면 스윕은 다시 안 알린다.
            # 저장소가 둘이라(watch.db / auto_seen.db) 이걸 안 이으면 두 번 간다.
            "already_notified": self._already_notified,
            "access_token": None,
            # 사이클마다 최신 access 재조회 — accounts.json 을 **읽기만** 한다.
            # GUI 의 수확 소유자는 _HarvestThread(20분 주기, __init__ 에서 기동)
            # 하나뿐이다. 여기서 harvest_all 을 부르면 휴식주기(30~90초)마다 함대를
            # 통째로 깨워 _HarvestThread 와 동시에 adb 를 두드리고 accounts.json 을
            # 같이 쓴다 — LDPlayer 순차기동 규칙도 그때 깨진다.
            "token_provider": self._read_token_quiet,
            # 계정 안정화(밴회피): 사이클마다 계정 라운드로빈 + 계정별 고정프록시(없으면 KR네이티브)
            # + daily_cap/warmup. 수확 갱신 켜졌을 때 함께 활성(다계정 전제).
            "stabilize": True,
            "accounts_fp": "./accounts.json",
            "daily_cap": 0,        # 0 = 상한 없음(account_scheduler 참고). 회전·격리만 쓴다.
            "warmup_days": 3,
            "out_json": "./OUT.json",
            "db_path": "./auto_seen.db",
        }
        # 범위 판정은 헤드리스(headless_sweep_cfg)와 같은 함수를 쓴다 —
        # 여기서만 '미선택=전국'이면 GUI 에서 본 범위와 서버가 도는 범위가 갈린다.
        s = self._load_alert_settings()
        if sweep_app_enabled(s):
            # 앱 키워드 스윕은 보완층이다 — 타지역 택배 매물이 목적이라 지역 1~2곳이면 된다.
            cfg.update({"scope": "regions",
                        "regions": list(s.get("sweep_regions_app") or FEED_DEFAULTS["sweep_regions_app"])})
        else:
            cfg.update(sweep_scope_for(self._selected_auto_regions(),
                                       self.autoNationwide.isChecked(),
                                       out_json=cfg["out_json"],
                                       log=self._alog,
                                       n_conditions=len(self._queue_entries()) or 1,
                                       # GUI 는 항상 수확 토큰 경로(앱API)다.
                                       lanes=sweep_lanes_effective(cfg["lanes"], True, 0)))
        return cfg

    @staticmethod
    def _splt(text):
        """쉼표·공백 구분 → 리스트. 옛 on_auto_start_clicked 의 splt 와 같다."""
        import re
        return [x for x in re.split(r"[,\s]+", (text or "").strip()) if x]

    @staticmethod
    def _num(text):
        text = (text or "").strip().replace(",", "")
        return int(text) if text.isdigit() else None

    def _sweep_cfg(self):
        """스윕 큐의 키워드로 AutoMonitor cfg 를 만든다. 지역·속도는 고급 패널 값.

        가격·제외는 등록 당시 사용자가 넣은 값(큐 엔트리)을 우선한다 — 고급 패널
        값은 엔트리에 없을 때의 기본값이다."""
        panel_min = self._num(self.autoMin.text())
        panel_max = self._num(self.autoMax.text())
        panel_excl = self._splt(self.autoExclude.text())
        extra = self._splt(self.autoExtra.text())
        days = self.autoDays.value() or None
        conditions = sweep_conditions(self._queue_entries(), extra=extra,
                                      exclude=panel_excl, min_price=panel_min,
                                      max_price=panel_max, days=days)
        cfg = dict(self._auto_cfg_base())
        cfg["conditions"] = conditions
        return cfg

    def _already_notified(self, article_id) -> bool:
        """이미 본 매물인가. 워치리스트가 그 사실의 주인이다.

        키가 둘이다 — 앱은 숫자 id, 피드는 슬러그 href. href 행은 첫 상세
        조회에서 숫자 id 로 옮겨 가지만(WatchTracker.check_one 재키잉) url 은
        href 그대로 남으므로, id 로 못 찾으면 url 로 한 번 더 본다."""
        store = getattr(self, "_watch_store", None)
        if store is None:
            return False
        try:
            key = str(article_id)
            return bool(store.get(key) or store.get_by_url(key))
        except Exception:
            return False

    def _dispose_auto_monitor(self):
        """다 쓴 AutoMonitor 를 놓아준다 — 시그널을 끊고 Qt 에 반납한다.

        돌고 있는 스레드는 건드리지 않는다(호출자가 이미 정지시킨 뒤에 부른다).
        반환값은 실제로 버렸는지 — 테스트가 누수 여부를 이걸로 본다."""
        am = self.auto_monitor
        self.auto_monitor = None
        if am is None:
            return False
        for sig in ("log", "found"):
            try:
                getattr(am, sig).disconnect()
            except (TypeError, RuntimeError, AttributeError):
                pass                    # 연결이 없으면 disconnect 가 TypeError
        try:
            am.deleteLater()
        except (RuntimeError, AttributeError):
            pass
        return True

    def _app_sweep_gate(self, settings=None) -> bool:
        """앱 키워드 스윕이 켜져 있는가. 꺼져 있으면 도는 스윕을 세우고 False.

        헤드리스 폴링 루프의 같은 자리와 같은 규칙이다 — 안내는 켜짐→꺼짐
        전이마다 한 번만 남긴다. 틱마다 찍으면(폴링 30~90초) 하루치 로그가
        같은 줄로 덮인다."""
        s = self._load_alert_settings() if settings is None else settings
        if sweep_app_enabled(s):
            self._app_sweep_off_logged = False
            return True
        if not getattr(self, "_app_sweep_off_logged", False):
            self._alog("[검색스윕] 앱 스윕 꺼짐(설정) — 피드가 발굴합니다")
            self._app_sweep_off_logged = True
        am = self.auto_monitor
        if am is not None and am.isRunning():
            self._stop_search_sweep()
        return False

    def _start_search_sweep(self):
        if not self._app_sweep_gate():
            return
        if self.auto_monitor is not None and self.auto_monitor.isRunning():
            # stop() 은 비동기다(최대 8초). 그 안에 다시 켜면 여기서 조용히 막혔다 —
            # 재시작이 통째로 사라지는 것처럼 보이므로 로그를 남긴다.
            self._alog(
                "[검색스윕] 아직 정지 중 — 이번 시작 요청은 건너뜁니다(다음 틱 재시도)")
            return
        try:
            cfg = self._sweep_cfg()
            if not cfg.get("conditions"):
                # conditions 가 비면 AutoMonitor 가 cfg["keyword"] 로 떨어져 KeyError.
                # 컨트롤러의 큐 검사에 기대지 않고 여기서 직접 막는다.
                self._alog("[검색스윕] 대기열이 비어 시작하지 않습니다")
                return
            from daangn.auto_monitor import AutoMonitor
            # 죽은 모니터를 버리고 간다. 안 버리면 되살릴 때마다 MainWindow 에
            # 매달린 QThread 가 한 개씩 쌓이고(수명이 창과 같다), 끊지 않은
            # log/found 시그널은 죽은 객체에서도 계속 슬롯을 때린다.
            self._dispose_auto_monitor()
            self.auto_monitor = AutoMonitor(self, cfg)
            self.auto_monitor.log.connect(self._alog)
            self.auto_monitor.found.connect(self._on_sweep_found)
            self.auto_monitor.start()
            # 이 스윕이 어떤 키워드 집합으로 떠 있는지 기억한다 — cfg 는 스냅샷이다.
            self._sweep_kws = {c["keyword"] for c in cfg["conditions"]}
            self._alog(
                f"[검색스윕] 시작 — 키워드 {len(self._sweep_kws)}개")
        except Exception as e:
            self._alog(f"[검색스윕] 시작 실패: {str(e)[:120]}")

    def _stop_search_sweep(self):
        am = self.auto_monitor
        self._sweep_kws = None
        if am is not None and am.isRunning():
            am.stop()
            self._alog("[검색스윕] 정지 요청")

    def _feed_settings_patch(self):
        cats = [c for c, w in ((31, self.feedCat31), (14, self.feedCat14), (5, self.feedCat5)) if w.isChecked()]
        return {
            "feed_enabled": bool(self.feedEnabledChk.isChecked()),
            "feed_categories": cats,
            "feed_proxies": [ln.strip() for ln in self.feedProxies.toPlainText().splitlines() if ln.strip()],
            "feed_rps": float(self.feedRps.value()),
            "feed_rest_min": int(self.feedRestMin.value()),
            "sweep_app_enabled": bool(self.sweepAppChk.isChecked()),
        }

    def _dispose_feed_monitor(self):
        """다 쓴 FeedMonitor 를 놓아준다 — 시그널을 끊고 Qt 에 반납한다.

        _dispose_auto_monitor 와 같은 모양이다: 돌고 있는 스레드는 건드리지
        않는다(호출자가 이미 정지시켰거나, 끝나서 죽은 것만 부른다)."""
        fm = self.feed_monitor
        self.feed_monitor = None
        if fm is None:
            return False
        for sig in ("log", "found", "status"):
            try:
                getattr(fm, sig).disconnect()
            except (TypeError, RuntimeError, AttributeError):
                pass                    # 연결이 없으면 disconnect 가 TypeError
        try:
            fm.deleteLater()
        except (RuntimeError, AttributeError):
            pass
        return True

    def _feed_log_once(self, key, msg):
        """같은 사유로는 한 번만 말한다. 시작에 성공하면 래치가 풀린다.

        _start_feed 는 폴링 틱마다 불린다(_resync_feed) — 안 걸어 두면 '설정에서
        꺼져 있음' 한 줄이 하루에 수백 번 쌓여 로그가 못 쓰게 된다."""
        seen = self.__dict__.setdefault("_feed_logged", set())
        if key in seen:
            return
        seen.add(key)
        self._alog(msg)

    def _on_feed_status(self, text):
        """피드 상태 한 줄. 프록시 사망만 경고색으로 — 나머지는 진행 표시다."""
        from daangn.feed_sweep import PROXY_DEAD_STATUS
        self._set_status("feed", text,
                         "warn" if text == PROXY_DEAD_STATUS else "ok")

    def _start_feed(self):
        s = self._load_alert_settings()
        if not s.get("feed_enabled", FEED_DEFAULTS["feed_enabled"]):
            self._feed_log_once("off", "[피드] 설정에서 꺼져 있음"); return
        if len(self._alert_rules.get().rules) == 0:
            self._feed_log_once("rules", "[피드] 조건표가 비어 있어 시작하지 않습니다"); return
        fm = self.feed_monitor
        if fm is not None and fm.isRunning():
            return
        # 프록시가 전멸해 죽은 엔진이면 물러선다 — 헤드리스와 같은 함수를 본다.
        # 죽은 프록시로는 다시 띄워도 첫 요청에서 같은 자리에 눕는다.
        left = feed_proxy_backoff_left(getattr(self, "_feed_last_engine", None))
        if left > 0:
            self._feed_log_once(
                "backoff", f"[피드] 프록시 전멸 뒤 대기 — {int(left // 60) + 1}분 후 재시도")
            return
        try:
            self._dispose_feed_monitor()
            cfg = feed_cfg(s, self._notify, already_notified=self._already_notified, log=self._alog)
            from daangn.feed_monitor import FeedMonitor
            self.feed_monitor = FeedMonitor(self, cfg)
            self.feed_monitor.log.connect(self._alog)
            self.feed_monitor.found.connect(self._on_feed_found)
            self.feed_monitor.status.connect(self._on_feed_status)
            self.feed_monitor.start()
            # 엔진은 모니터를 버린 뒤에도 정지 사유를 들고 있어야 한다 — 다음 틱의
            # 백오프 판정이 이걸 본다.
            self._feed_last_engine = self.feed_monitor.engine
            self.__dict__["_feed_logged"] = set()
            self._set_status("feed", f"피드 {len(cfg['regions'])}동 · 레인 {max(1, len(cfg['proxies']))}", "ok")
        except Exception as e:
            self._alog(f"[피드] 시작 실패: {str(e)[:120]}")

    def _stop_feed(self):
        fm = self.feed_monitor
        if fm is not None and fm.isRunning():
            fm.stop()
        self._set_status("feed", "", "off")

    def _resync_feed(self):
        """감시 중인데 피드가 죽어 있으면 다시 띄운다 — 헤드리스 폴링 루프와 같은 자리.

        엔진은 프록시 전멸·예외로 조용히 죽는다. 되살리는 손이 없으면 발굴
        주경로가 통째로 멈춘 채 앱 알림만 남는다(클라 눈에는 '알림이 준다')."""
        if not (self._supervisor and self._supervisor.is_running()):
            return
        fm = self.feed_monitor
        if fm is not None and fm.isRunning():
            return
        self._start_feed()

    def _on_feed_found(self, payload):
        """피드가 찾은 매물 → 워치리스트(추적) + 결과 표. GUI 스레드에서 불린다.

        _on_sweep_found 은 대기열 키워드에서 라벨을 다시 찾지만, 피드는 이미
        payload["keyword"] 에 매칭된 조건표 라벨(브랜드)을 실어 보낸다 —
        그대로 쓴다."""
        try:
            norm = sweep_found_to_match(payload, payload.get("keyword") or "")
            if norm and self._watch_tracker:
                if self._watch_tracker.add_from_matches([norm], source="feed"):
                    self._refresh_listing_table()
        except Exception as e:
            self._alog(f"[피드] 추적 등록 실패: {str(e)[:80]}")

    def _resync_search_sweep(self):
        """돌고 있는 스윕이 낡은 키워드 집합인지 보고, 다르면 갈아끼운다.

        cfg 는 AutoMonitor 를 만들 때 한 번 찍은 스냅샷이다. 감시 중에 키워드가
        큐로 밀려오면 아무도 그걸 훑지 않고, 반대로 rebalance 가 앱으로 승격시키면
        스윕이 이미 앱이 보는 키워드를 계속 훑어 요청을 두 번 쓴다.

        폴링 틱에서만 부른다 — 등록이 몰아쳐도 재시작은 폴링 주기당 한 번을 넘지 않는다."""
        if not (self._supervisor and self._supervisor.is_running()):
            return
        # 스위치 검사가 먼저다 — 뒤에 두면 꺼져 있어도 'start' 판정까지 가서
        # _start_search_sweep 이 틱마다 거절 로그를 찍는다.
        if not self._app_sweep_gate():
            return
        if self._sweep_queue is None:
            return
        try:
            want = set(self._sweep_queue.keywords())
        except Exception:
            return
        have = getattr(self, "_sweep_kws", None)
        am = self.auto_monitor
        running = am is not None and am.isRunning()
        # 판정은 헤드리스(HeadlessSweepRunner.resync)와 같은 함수를 쓴다 —
        # 따로 두면 서버에서만 나는 버그가 생긴다.
        act = sweep_resync_action(want, have, running)
        if not act:
            self._sweep_revives = 0
            return
        if act == "start":
            self._sweep_revives = 0
            self._start_search_sweep()
            return
        if act == "revive":
            # 상한 계산도 헤드리스와 같은 함수를 쓴다. 예전엔 여기에 상한이
            # 아예 없어서, 뜨자마자 죽는 엔진을 틱마다 영원히 다시 띄우고
            # 죽은 AutoMonitor 를 창에 쌓았다.
            ok, self._sweep_revives, msg = sweep_revive_step(
                getattr(self, "_sweep_revives", 0), len(want))
            if msg:
                self._alog(msg)
            if not ok:
                return
        else:
            self._sweep_revives = 0
            self._alog(
                f"[검색스윕] 키워드 변경 {len(have)}개 → {len(want)}개 — 재시작")
        self._stop_search_sweep()          # _sweep_kws 를 None 으로 되돌린다
        if want:
            self._start_search_sweep()     # 아직 정지 중이면 다음 틱이 이어받는다

    def _on_sweep_found(self, payload):
        """검색 스윕이 찾은 매물도 앱 알림과 같은 문으로 워치리스트에 들어간다."""
        kw = ""
        try:
            kw = sweep_keyword_for(payload, self._sweep_queue.keywords())
        except Exception:
            pass
        norm = sweep_found_to_match(payload, kw)
        if not norm or not self._watch_tracker:
            return
        try:
            if self._watch_tracker.add_from_matches([norm], source="sweep"):
                self._refresh_listing_table()
        except Exception as e:
            self._alog(f"[검색스윕] 추적 등록 실패: {str(e)[:80]}")


    def _init_state(self):
        # 검색 수명 상태 — 토큰 스레드 유무가 '준비 중'의 유일한 근거다.
        self._token_thread = None
        self._search_cancelled = False
        self._crawl_builder = None
        self.worker_thread: CancelableImageDownloader | None = None
        self.current_loaded_url: str | None = None
        self.current_download_task_token: str | None = None
        self.all_last_child: list[QtWidgets.QTreeWidgetItem] = []
        self.area_lookup: dict[str, list[tuple[str, str]]] = {}
        self.preview_image_size = (256, 256)

    def _setup_ui(self):
        self.ui.detailView.setReadOnly(True)
        # 링크 클릭 = 내부 이동(빈칸) 금지, 시스템 브라우저로 열기
        self.ui.detailView.setOpenLinks(False)
        self.ui.detailView.setOpenExternalLinks(False)
        self.ui.detailView.anchorClicked.connect(self._open_detail_link)
        self.sb: QtWidgets.QStatusBar = self.statusBar()  # type: ignore
        # 제목은 브랜드명 고정. 프록시 수·동시요청·대기시간은 자동/수동 공용
        # 크롤러 설정이라 수동 검색 화면 제목에 얹을 이유가 없다 — 지금 상태는
        # '매물 감시' 탭의 IP/간격 표시가 보여준다.
        self.setWindowTitle("LUXE — 명품 실시간 모니터")
        self._setup_health_indicator()   # 위젯만 만든다. 배치는 _build_alert_tab.

        for edit in (self.ui.minimumEdit, self.ui.maximumEdit):
            edit.setValidator(
                QRegularExpressionValidator(QRegularExpression("[0-9]*"), edit)
            )

        self.ui.prdImg.setScaledContents(True)
        self.ui.prdImg.setFixedSize(*self.preview_image_size)

        self._setup_extra_ui()

    def _setup_health_indicator(self):
        """'매물 감시' 탭에 상시 표시 — 쓸 수 있는 IP 수 + 현재 요청간격(자동감속 반영).
        차단 대응은 전부 자동이라, 사용자에게 필요한 건 설정이 아니라 **지금 상태**다.

        예전엔 상태바에 있었다. 상태바는 세 탭 공용이라 수동 검색 화면에도
        자동 인프라 표시가 따라다녔다. 쓰는 곳에만 둔다."""
        self.healthLabel = QtWidgets.QLabel("")
        self.healthLabel.setStyleSheet("color:#5C5449; font-size:13px;")
        self._health_thread = None
        self._health_timer = QtCore.QTimer(self)
        self._health_timer.timeout.connect(self._refresh_health_indicator)
        self._health_timer.start(2000)

    def _refresh_health_indicator(self):
        from daangn_ext import proxy_budget, throttle
        if not getattr(self, "controller", None):
            return
        try:
            pool = self._collect_proxies()
        except Exception:
            return
        st = proxy_budget.pool_status(pool)
        parts = [f"IP {st['alive']}/{st['total']}"]
        if st["cooling"]:
            parts[0] += f" (쿨다운 {st['cooling']}, {st['next_free_in']:.0f}초 후 해제)"
        parts.append(throttle.describe(self.controller.req_min_ms))
        self.healthLabel.setText("  ·  ".join(parts))

    def on_health_check_clicked(self):
        if self._health_thread and self._health_thread.isRunning():
            return
        pool = self._collect_proxies()
        if not pool and not self.ask("등록된 프록시가 없습니다. 직결 IP 만 진단할까요?"):
            return
        self._health_thread = HealthCheckThread(self, pool)
        self._health_thread.progress.connect(
            lambda d, t: self.sb.showMessage(f"진단 중… {d}/{t}"))
        self._health_thread.result.connect(self._show_health_report)
        self._health_thread.start()

    def _show_health_report(self, res: dict):
        self._health_thread = None
        if res.get("error"):
            self.alert(f"진단 실패: {res['error']}")
            return
        from daangn_ext.health import report_text
        self.sb.showMessage(res["verdict"])
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("프록시 진단")
        dlg.resize(720, 520)
        v = QtWidgets.QVBoxLayout(dlg)
        head = QtWidgets.QLabel(f"판정: {res['verdict']}")
        head.setStyleSheet("font-weight:800; font-size:16px; color:#8A6D1F; letter-spacing:1px;")
        act = QtWidgets.QLabel(f"대응: {res['action']}")
        act.setWordWrap(True)
        v.addWidget(head); v.addWidget(act)
        box = QtWidgets.QPlainTextEdit(report_text(res))
        box.setReadOnly(True)
        box.setFont(QFontDatabase.systemFont(QFontDatabase.SystemFont.FixedFont))
        v.addWidget(box, 1)
        btn = QtWidgets.QPushButton("닫기"); btn.clicked.connect(dlg.accept)
        v.addWidget(btn, 0, Qt.AlignmentFlag.AlignRight)
        dlg.exec()

    def _open_detail_link(self, url: QUrl) -> None:
        """상세뷰 링크 = 시스템 브라우저로. 상대경로면 당근 도메인 부착."""
        if url.isRelative() or not url.scheme():
            url = QUrl("https://www.daangn.com" + url.toString())
        QDesktopServices.openUrl(url)

    def _setup_extra_ui(self):
        """추가기능 위젯 생성(배치는 _build_manual_tab 에서)."""
        self.extraEdit = QtWidgets.QLineEdit()
        self.extraEdit.setPlaceholderText("추가 키워드 (쉼표/공백 구분, 모두 포함)")
        self.excludeEdit = QtWidgets.QLineEdit()
        self.excludeEdit.setPlaceholderText("제외 키워드 (쉼표/공백 구분)")

    def _build_manual_tab(self):
        """수동 탭 — 자동 탭과 통일된 그룹박스 레이아웃. 클라 원본 위젯 재사용."""
        w = QtWidgets.QWidget(self)
        outer = QtWidgets.QHBoxLayout(w)
        outer.setContentsMargins(4, 4, 4, 4); outer.setSpacing(0)
        split = QtWidgets.QSplitter(Qt.Orientation.Horizontal, w)
        split.setChildrenCollapsible(False); split.setHandleWidth(6)
        outer.addWidget(split)

        # 좌: 지역 선택 (검색+전체선택 패널)
        self.ui.areaTree.setMinimumWidth(200)
        toss_tree(self.ui.areaTree)
        split.addWidget(self._tree_panel(self.ui.areaTree, self.all_last_child))

        # 중앙: 슬림 필터바(2줄) + 대형 결과 테이블 + 하단 보조바
        center = QtWidgets.QWidget(w); cl = QtWidgets.QVBoxLayout(center)
        cl.setContentsMargins(8, 0, 8, 0); cl.setSpacing(10)

        self.ui.keywordEdit.setPlaceholderText("검색 키워드")
        self.ui.minimumEdit.setPlaceholderText("최소가"); self.ui.minimumEdit.setFixedWidth(96)
        self.ui.maximumEdit.setPlaceholderText("최대가"); self.ui.maximumEdit.setFixedWidth(96)
        self.ui.startBtn.setText("검색")

        fc = QtWidgets.QGroupBox(center); fc.setTitle("")
        fv = QtWidgets.QVBoxLayout(fc); fv.setContentsMargins(14, 12, 14, 12); fv.setSpacing(8)
        r0 = QtWidgets.QHBoxLayout(); r0.setSpacing(8)
        r0.addWidget(self.ui.keywordEdit, 3)
        r0.addWidget(self.ui.minimumEdit); r0.addWidget(QtWidgets.QLabel("~"))
        r0.addWidget(self.ui.maximumEdit); r0.addWidget(self.ui.startBtn)
        r1 = QtWidgets.QHBoxLayout(); r1.setSpacing(8)
        r1.addWidget(self.extraEdit, 1); r1.addWidget(self.excludeEdit, 1)
        r1.addSpacing(10)
        r1.addWidget(self.ui.onlyTradeableCheck)
        # 진행 표시 — 전국 검색은 지역이 200곳을 넘고 몇 분이 걸린다. 누른 뒤
        # 화면이 그대로면 사용자는 눌리지 않은 줄 안다(클라 실제 반응). 상태바
        # 한 줄로는 부족해서 필터바 바로 아래, 결과표 위에 둔다.
        self.searchProgress = QtWidgets.QProgressBar()
        self.searchProgress.setRange(0, 0)          # 준비 단계는 진행률을 모른다
        self.searchProgress.setTextVisible(True)
        self.searchProgress.setVisible(False)
        self.searchProgressLabel = QtWidgets.QLabel("")
        self.searchProgressLabel.setStyleSheet(
            "color:#6B6355; font-size:12px;")
        self.searchProgressLabel.setVisible(False)
        fv.addLayout(r0); fv.addLayout(r1)
        fv.addWidget(self.searchProgress); fv.addWidget(self.searchProgressLabel)
        cl.addWidget(fc)

        hdr = QtWidgets.QHBoxLayout()
        rl0 = QtWidgets.QLabel("검색 결과"); rl0.setStyleSheet("font-weight:800; color:#8A6D1F; font-size:16px; letter-spacing:1px;")
        hdr.addWidget(rl0); hdr.addStretch(1)
        cl.addLayout(hdr)
        self.ui.itemListView.setMinimumHeight(360)
        cl.addWidget(self.ui.itemListView, 1)

        # 하단 보조 바
        self.ui.saveToExcelBtn.setText("엑셀 저장")
        self.ui.crawlFromExcelBtn.setText("엑셀 크롤링")
        sb = QtWidgets.QHBoxLayout(); sb.setSpacing(8)
        for b in (self.ui.saveToExcelBtn, self.ui.crawlFromExcelBtn):
            b.setSizePolicy(QtWidgets.QSizePolicy.Policy.Fixed,
                            QtWidgets.QSizePolicy.Policy.Fixed)
        sb.addWidget(self.ui.saveToExcelBtn); sb.addWidget(self.ui.crawlFromExcelBtn)
        sb.addStretch(1)
        cl.addLayout(sb)
        split.addWidget(center)

        # 우: 상세
        right = QtWidgets.QWidget(w); rl = QtWidgets.QVBoxLayout(right)
        rl.setContentsMargins(8, 0, 0, 0)
        rl.addWidget(QtWidgets.QLabel("상세"))
        rl.addWidget(self.ui.prdImg, 0, Qt.AlignmentFlag.AlignHCenter)
        rl.addWidget(self.ui.detailView, 1)
        split.addWidget(right)
        split.setStretchFactor(0, 0); split.setStretchFactor(1, 1); split.setStretchFactor(2, 0)
        split.setSizes([240, 640, 300])
        return w

    def _setup_model(self):
        self.products_model = self.controller.create_model(self)
        header = self.ui.itemListView.horizontalHeader()
        header.setSortIndicator(0, Qt.SortOrder.AscendingOrder)  # type: ignore
        self.ui.itemListView.setSortingEnabled(True)
        self.ui.itemListView.setModel(self.products_model)

    def _connect_signals(self):
        self.ui.startBtn.clicked.connect(self.on_start_btn_clicked)
        self.ui.crawlFromExcelBtn.clicked.connect(self.crawl_from_excel)
        self.ui.saveToExcelBtn.clicked.connect(self.save_to_excel)

        selection_model = self.ui.itemListView.selectionModel()
        assert selection_model
        selection_model.selectionChanged.connect(self.on_select_prd)

        self.controller.task_error.connect(self._handle_task_error)
        self.controller.task_finished.connect(self._handle_task_finished)
        self.controller.task_message.connect(self._handle_task_message)
        self.controller.task_progress.connect(self._handle_task_progress)

    def _load_proxy(self):
        err = self.controller.load_proxy_settings()
        if err:
            self.alert(err)

    def _init_tree(self):
        self.all_last_child.clear()
        self.area_lookup = {}
        with open("./OUT.json", "r", encoding="utf8") as f:
            AREA_DATA = json.loads(f.read())

        AREA_ROOT: dict[str, list[Any]] = {}
        for area in AREA_DATA:
            lis = AREA_ROOT.setdefault(area["name1"], [])
            lis.append(area)

        # 시도가 최상위 — '지역' 루트 행은 [전체 선택] 버튼이 대신한다. 행에는
        # 그 단계 이름만 보이고 전체 경로는 AREA_FULL_ROLE 에 둔다(자동 트리와 같다).
        all_locations: list[tuple[str, str]] = []

        for sido in AREA_ROOT:
            parent = QtWidgets.QTreeWidgetItem(self.ui.areaTree)
            parent.setText(0, f"{sido}")
            parent.setData(0, AREA_FULL_ROLE, f"{sido}")
            parent.setFlags(parent.flags() | Qt.ItemFlag.ItemIsAutoTristate | Qt.ItemFlag.ItemIsUserCheckable)
            parent.setCheckState(0, Qt.CheckState.Unchecked)

            sido_locations: list[tuple[str, str]] = []

            for area in AREA_ROOT[sido]:
                name1 = area["name1"] or ""
                name2 = area["name2"] or ""
                child1_txt = f"{name1} {name2}".strip()

                child1 = QtWidgets.QTreeWidgetItem(parent)
                child1.setText(0, name2.strip() or child1_txt)
                child1.setData(0, AREA_FULL_ROLE, child1_txt)
                child1.setFlags(child1.flags() | Qt.ItemFlag.ItemIsAutoTristate | Qt.ItemFlag.ItemIsUserCheckable)
                child1.setCheckState(0, Qt.CheckState.Unchecked)

                area_locations: list[tuple[str, str]] = []

                for loc in area["locations"]:
                    child2 = QtWidgets.QTreeWidgetItem(child1)
                    child2.setFlags(child2.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                    child2.setText(0, loc["name"])
                    child2.setData(0, AREA_FULL_ROLE, f"{child1_txt} {loc['name']}".strip())
                    child2.setData(
                        0, Qt.ItemDataRole.UserRole, f"{loc['name']}-{loc['id']}"
                    )
                    child2.setCheckState(0, Qt.CheckState.Unchecked)
                    self.all_last_child.append(child2)

                    location_entry = (
                        area_full_name(child2),
                        child2.data(0, Qt.ItemDataRole.UserRole),
                    )
                    area_locations.append(location_entry)
                    sido_locations.append(location_entry)
                    all_locations.append(location_entry)

                    self._register_area_mapping(area_full_name(child2), [location_entry])
                    self._register_area_mapping(str(loc["id"]), [location_entry])

                self._register_area_mapping(child1_txt, area_locations)

            self._register_area_mapping(sido, sido_locations)

        self._register_area_mapping("전국", all_locations)

    def _register_area_mapping(
        self, key: str, locations: list[tuple[str, str]]
    ) -> None:
        if not key or not locations:
            return

        existing = self.area_lookup.setdefault(key, [])
        existing_codes = {code for _, code in existing}
        for name, code in locations:
            if code not in existing_codes:
                existing.append((name, code))
                existing_codes.add(code)

    def _enter_task(self):
        self.ui.startBtn.setText("정지")
        self.ui.crawlFromExcelBtn.setEnabled(False)
        self.ui.saveToExcelBtn.setEnabled(False)

    def _leave_task(self):
        self.ui.startBtn.setText("시작")
        self.ui.crawlFromExcelBtn.setEnabled(True)
        self.ui.saveToExcelBtn.setEnabled(True)

    def alert(self, text: str):
        message_box = QtWidgets.QMessageBox(self)
        message_box.setWindowIcon(app_icon())
        message_box.setWindowTitle("알림")
        message_box.setText(text)
        message_box.setIcon(QtWidgets.QMessageBox.Icon.Information)
        message_box.exec()

    def ask(self, text: str):
        reply = QtWidgets.QMessageBox(self)
        reply.setWindowIcon(app_icon())
        reply.setWindowTitle("알림")
        reply.setText(text)
        reply.setStandardButtons(
            QtWidgets.QMessageBox.StandardButton.Yes
            | QtWidgets.QMessageBox.StandardButton.No
        )

        return reply.exec() == QtWidgets.QMessageBox.StandardButton.Yes

    def on_start_btn_clicked(self):
        # 토큰 준비 구간에는 controller.task 가 아직 없다. is_task_running() 만
        # 보면 여기서 두 번째 검색이 그대로 시작된다(같은 함대를 두 번 깨운다).
        if getattr(self, "_token_thread", None) is not None:
            if not self.ask("토큰을 준비하고 있습니다. 검색을 취소할까요?"):
                return
            # ask() 는 중첩 이벤트루프를 돈다. 그 사이에 토큰이 도착해
            # _on_token_ready 가 이미 검색을 시작했을 수 있다 — 그때 이 플래그는
            # 아무것도 멈추지 못한다(이미 지나간 분기다). 실제로 도는 작업을 세운다.
            if self._token_thread is None:
                if self.controller.is_task_running():
                    self.controller.stop_task()
                    self._set_search_progress("정지 중…")
                return
            self._search_cancelled = True
            self._set_search_progress("취소 중… 토큰 준비가 끝나면 멈춥니다")
            return

        # if stop btn
        if self.controller.is_task_running():
            if self.controller.is_task_stopping():
                self.alert("이미 정지하고 있습니다")
                return
            if not self.ask("정지 하시겠습니까?"):
                return
            self.controller.stop_task()
            return

        onlyTradeable = self.ui.onlyTradeableCheck.isChecked()
        keyword = self.ui.keywordEdit.text()
        minimum = int(val) if (val := self.ui.minimumEdit.text().strip()) else None
        maximum = int(val) if (val := self.ui.maximumEdit.text().strip()) else None

        if not keyword:
            self.alert("키워드가 비어있습니다")
            return

        if minimum is not None and maximum is not None:
            if int(minimum) > int(maximum):
                self.alert("최대 가격이 더 큽니다")
                return

        area_id_list = []
        for ch in self.all_last_child:
            if ch.checkState(0) == Qt.CheckState.Checked:
                area_id_list.append((area_full_name(ch), ch.data(0, Qt.ItemDataRole.UserRole)))

        if not area_id_list:
            self.alert("선택된 지역이 없습니다")
            return

        ok = self.ask(
            f"{len(area_id_list)}개 지역에서 검색을 시작합니다.\n키워드: {keyword}"
        )
        if not ok:
            return

        # 추가/제외 키워드 파싱
        extra = self._split_keywords(self.extraEdit.text())
        exclude = self._split_keywords(self.excludeEdit.text())
        # 앱API 통일: 수동도 항상 app-API(search-bff) 경로. 웹크롤(robust_fetch) 폐지.
        # (체크박스와 무관하게 True — 자동/수동 동일 데이터소스)
        adaptive = True

        # 토큰 갱신은 실패 시 모달을 띄운다. 모달이 중첩 이벤트루프를 도는 동안
        # 버튼이 살아 있으면 두 번째 작업이 재진입으로 시작될 수 있다 → 먼저 잠근다.
        # 토큰 확보(수십 초)는 워커 스레드가 한다 — GUI 스레드에서 돌리면 창이
        # 얼어 진행 표시가 그려지지 않는다. 작업 목록은 토큰이 온 뒤에 짓는다.
        self._start_crawl(lambda access_token: [
            CrawlTask(
                area=area,
                keyword=keyword,
                only_tradeable=onlyTradeable,
                minimum=minimum,
                maximum=maximum,
                extra_keywords=extra,
                exclude_keywords=exclude,
                adaptive=adaptive,
                access_token=access_token,
            )
            for area in area_id_list
        ])

    def _read_token_quiet(self) -> str | None:
        """스윕 스레드의 토큰 provider — accounts.json 을 **읽기만** 한다.

        GUI 에서 수확을 소유하는 건 _HarvestThread(20분 주기) 하나뿐이다.
        스윕이 필요한 건 '수확'이 아니라 '신선한 토큰'이고, 그 파일을 신선하게
        유지하는 일은 이미 그 스레드가 한다. GUI 미접근(showMessage/alert 금지)."""
        return read_token_quiet("./accounts.json")

    def _harvest_token_quiet(self) -> str | None:
        """수확 후 최신 access. 사용자가 부른 작업(_alert_api)에서만 쓴다 —
        주기적 수확은 _HarvestThread 소유다. GUI 미접근. 본체는 헤드리스와 공유."""
        return harvest_token_quiet("./accounts.json")

    # ── 검색 수명: 토큰 확보(스레드) → 작업 시작 → 진행 표시 ──────────────
    #
    # 예전에는 여기서 토큰 갱신을 동기로 부르고 바로 start_task 를 했다. 갱신이
    # 수십 초 걸리는 동안 GUI 스레드가 막혀 창이 얼었고, 표는 이미 비워진 뒤라
    # 사용자에게는 '눌렀는데 아무 일도 안 일어남'으로 보였다.

    def _start_crawl(self, build_tasks):
        """토큰을 워커 스레드로 확보한 뒤 build_tasks(access_token) 으로 검색 시작.

        build_tasks 는 토큰이 도착한 다음에 불린다 — CrawlTask 가 토큰을 품기
        때문에 그 전에 지을 수 없다."""
        self._enter_task()          # 준비 중 재진입 방지(정지 버튼으로 바뀐다)
        self.clearItemList()
        self._crawl_builder = build_tasks
        self._search_cancelled = False
        self._show_search_progress("토큰을 준비하고 있습니다… (LDPlayer 수확)")
        th = _TokenRefreshThread(self)
        th.log.connect(self._set_search_progress)
        th.done.connect(self._on_token_ready)
        th.finished.connect(th.deleteLater)   # 검색마다 스레드 객체가 쌓이지 않게
        self._token_thread = th
        th.start()

    def _on_token_ready(self, access_token, err):
        self._token_thread = None
        if self._search_cancelled:
            self._abort_search("검색을 취소했습니다")
            return
        if err:
            self._abort_search("")
            self.alert(err)
            return
        try:
            tasks = self._crawl_builder(access_token)
            self.searchProgress.setRange(0, max(1, len(tasks)))
            self.searchProgress.setValue(0)
            self._set_search_progress(f"0/{len(tasks)} 지역 — 검색을 시작합니다")
            self.controller.start_task(tasks)
        except Exception as e:
            self._abort_search("")
            self.alert(str(e))

    def _abort_search(self, msg):
        self._hide_search_progress()
        self._leave_task()
        if msg:
            self.sb.showMessage(msg, 5000)

    def _show_search_progress(self, text):
        self.searchProgress.setRange(0, 0)      # 남은 시간을 모르는 준비 단계
        self.searchProgress.setVisible(True)
        self.searchProgressLabel.setVisible(True)
        self._set_search_progress(text)

    def _set_search_progress(self, text):
        self.searchProgressLabel.setText(str(text)[:160])

    def _hide_search_progress(self):
        self.searchProgress.setVisible(False)
        self.searchProgressLabel.setVisible(False)

    def _handle_task_progress(self, done, total):
        self.searchProgress.setRange(0, max(1, total))
        self.searchProgress.setValue(done)
        self.searchProgress.setFormat(f"{done}/{total} 지역  (%p%)")

    def on_accounts_btn_clicked(self):
        """계정+프록시 추가/관리 다이얼로그."""
        from daangn_ext import AccountStore
        from daangn_ext.account_store import account_role
        store = AccountStore("./accounts.json")
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("계정 + 프록시 관리")
        dlg.resize(520, 360)
        v = QtWidgets.QVBoxLayout(dlg)

        listw = QtWidgets.QListWidget(dlg)

        def _name(r):
            # 수확으로 들어온 계정은 label 이 비어 있고 code 만 있다. 화면에
            # '(무명)' 만 뜨면 어느 계정에 프록시를 붙이는지 알 수 없다.
            return (r.get("label") or r.get("code") or "(무명)")

        def reload_list():
            keep = listw.currentRow()
            listw.clear()
            for r in store.rows:
                listw.addItem(
                    f"{_name(r)}  |  {'스윕' if account_role(r) == 'sweep' else '알림'}"
                    f"  |  {r.get('proxy') or '프록시없음'}")
            if 0 <= keep < listw.count():
                listw.setCurrentRow(keep)
        reload_list()
        v.addWidget(QtWidgets.QLabel("등록된 계정", parent=dlg))
        v.addWidget(listw)

        form = QtWidgets.QFormLayout()
        labelEdit = QtWidgets.QLineEdit(dlg); labelEdit.setPlaceholderText("별칭/전화 (선택)")
        proxyEdit = QtWidgets.QLineEdit(dlg); proxyEdit.setPlaceholderText("http://user:pass@host:port")
        form.addRow("별칭", labelEdit)
        form.addRow("프록시", proxyEdit)
        roleBox = QtWidgets.QComboBox(dlg)
        roleBox.addItem("알림 계정 (앱 알림·등록만, 검색 안 함)", "alert")
        roleBox.addItem("스윕 계정 (앱 키워드 검색 전용 — 버려도 되는 계정)", "sweep")
        form.addRow("역할", roleBox)
        v.addLayout(form)

        v.addWidget(QtWidgets.QLabel(
            "계정을 고르면 그 계정의 프록시를 고쳐 저장할 수 있습니다. 비우고 저장하면 직결.\n"
            "· 계정을 늘리는 건 [에뮬레이터] 탭에서 LDPlayer 에 .ldbk 를 복원하는 것입니다\n"
            "  (PC 직접 갱신은 당근 WAF 가 막아 토큰은 에뮬레이터 앱에서만 나옵니다).",
            parent=dlg))

        # 같은 파일을 보는 화면들의 입구 — 계정 팜 현황·프록시 목록·프록시 진단.
        # 매물 감시 탭에 나란히 있던 것을 여기로 모았다.
        sub = QtWidgets.QHBoxLayout()
        fleetBtn = QtWidgets.QPushButton("계정 현황", dlg)
        fleetBtn.setToolTip("계정별 동네·토큰만료·핵심여부·폴링실패·밴격리 — 팜 운영 한 눈에")
        fleetBtn.clicked.connect(self.on_alert_fleet)
        proxyViewBtn = QtWidgets.QPushButton(
            f"프록시 목록 ({len(self._collect_proxies())})", dlg)
        proxyViewBtn.clicked.connect(self.on_proxy_view_clicked)
        diagBtn = QtWidgets.QPushButton("프록시 진단", dlg)
        diagBtn.setToolTip(
            "프록시를 IP 당 1회씩 찔러 '지금 막힌 건지, 막혔으면 어디가 문제인지' 판정")
        diagBtn.clicked.connect(self.on_health_check_clicked)
        sub.addWidget(fleetBtn); sub.addWidget(proxyViewBtn); sub.addWidget(diagBtn)
        sub.addStretch(1)
        v.addLayout(sub)

        btns = QtWidgets.QHBoxLayout()
        saveProxyBtn = QtWidgets.QPushButton("선택 계정 저장", dlg)
        saveProxyBtn.setObjectName("startBtn")
        delBtn = QtWidgets.QPushButton("선택 삭제", dlg)
        closeBtn = QtWidgets.QPushButton("닫기", dlg)
        btns.addWidget(saveProxyBtn); btns.addWidget(delBtn)
        btns.addStretch(1); btns.addWidget(closeBtn)
        v.addLayout(btns)

        def on_pick(i):
            """선택한 계정의 현재 프록시를 입력칸에 올린다 — 고치려면 보여야 한다."""
            if 0 <= i < len(store.rows):
                r = store.rows[i]
                proxyEdit.setText(r.get("proxy") or "")
                labelEdit.setText(r.get("label") or "")
                roleBox.setCurrentIndex(0 if account_role(r) == "alert" else 1)
        listw.currentRowChanged.connect(on_pick)

        def do_save_proxy():
            i = listw.currentRow()
            if not (0 <= i < len(store.rows)):
                QtWidgets.QMessageBox.warning(dlg, "확인", "먼저 계정을 고르세요.")
                return
            r = store.rows[i]
            key = r.get("code") or r.get("label") or r.get("refresh")
            val = proxyEdit.text().strip()
            # 역할과 프록시는 다른 필드다 — 프록시 저장이 실패했다고 역할까지
            # 버리면, 클라가 역할을 고치고 [저장]을 눌러도 조용히 사라진다.
            ok = store.set_proxy(key, val)
            if ok:
                self._alog(f"[프록시] {_name(r)} → {val or '직결'}")
            if store.set_role(key, roleBox.currentData()):
                self._alog(f"[계정] {_name(r)} 역할 → {roleBox.currentText()}")
            if ok:
                reload_list()
                QtWidgets.QMessageBox.information(
                    dlg, "저장됨",
                    f"{_name(r)} 의 프록시를 {'지웠습니다(직결)' if not val else val} 로 바꿨습니다.\n"
                    "다음 검색·폴링부터 적용됩니다.")
            else:
                QtWidgets.QMessageBox.warning(
                    dlg, "실패",
                    "저장하지 못했습니다. accounts.json 을 다른 프로그램이 쓰고 있는지"
                    " 확인하세요(로그에 사유가 남습니다).")
        saveProxyBtn.clicked.connect(do_save_proxy)

        def do_del():
            i = listw.currentRow()
            if not (0 <= i < len(store.rows)):
                QtWidgets.QMessageBox.warning(dlg, "확인", "먼저 계정을 고르세요.")
                return
            r = store.rows[i]
            # 되돌릴 수 없는 삭제다. refresh 토큰은 이 PC 에서 재발급할 수 없다
            # (당근 WAF 가 PC 갱신을 막는다) — 다시 넣으려면 LDPlayer 에 .ldbk 를
            # 복원해야 한다. 확인 없이 지우면 감시 계정이 조용히 하나 줄어든다.
            if not ask_yes_no(
                    dlg, "계정 삭제", f"'{_name(r)}' 계정을 목록에서 지울까요?",
                    "이 PC 에서는 로그인 토큰을 다시 만들 수 없습니다."
                    " 되돌리려면 LDPlayer 에 .ldbk 를 복원해야 합니다.\n\n"
                    "지운 내용은 accounts.json.deleted 에 남습니다.", danger=True):
                return
            key = r.get("code") or r.get("label") or r.get("refresh")
            if store.remove(key):
                self._alog(f"[계정] {_name(r)} 삭제 — accounts.json.deleted 에 보관")
                reload_list()
            else:
                QtWidgets.QMessageBox.warning(
                    dlg, "실패",
                    "지우지 못했습니다. accounts.json 을 다른 프로그램이 쓰고 있는지"
                    " 확인하세요.")
        delBtn.clicked.connect(do_del)
        closeBtn.clicked.connect(dlg.accept)
        dlg.exec()

    @staticmethod
    def _split_keywords(text: str) -> list[str]:
        """추가/제외 키워드 입력을 쉼표·공백으로 쪼갠다. 수동 검색과 엑셀 크롤링이 공유한다."""
        import re
        return [k for k in re.split(r"[,\s]+", (text or "").strip()) if k]

    def crawl_from_excel(self):
        if self.controller.is_task_running():
            self.alert("작업이 진행 중입니다")
            return

        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, caption="엑셀 파일 선택", filter="*.xlsx"
        )
        if not path:
            return

        try:
            tasks, errors = self._read_excel_tasks(path)
        except Exception as e:
            self.alert(f"엑셀 파일을 읽는 중 오류가 발생했습니다\n{e}")
            return

        if errors:
            preview = "\n".join(errors[:5])
            if len(errors) > 5:
                preview += f"\n... 총 {len(errors)}개 행에서 오류가 발생했습니다"
            self.alert(preview)
            return

        if not tasks:
            self.alert("엑셀에 유효한 검색 조건이 없습니다")
            return

        # 화면에 입력한 추가/제외 키워드와 토큰 갱신 설정을 엑셀 조건에도 그대로 태운다.
        # (예전엔 여기서 빠져 있어 같은 조건인데 수동 검색과 결과가 달랐고,
        #  adaptive 가 False 라 구단위 분할이 없어 상한 290 건에서 잘렸다.)
        # 엑셀은 '전국' 한 칸이 전국 동단위로 펼쳐진다. 게다가 adaptive 경로는 지역마다
        # 가격분할을 돌리므로 요청 수가 곱으로 늘어난다 → 수동 검색과 같이 규모를 먼저 알린다.
        total_areas = sum(len(t["areas"]) for t in tasks)
        if not self.ask(
            f"{len(tasks)}개 조건 × 총 {total_areas}개 지역에서 검색을 시작합니다.\n"
            "지역이 많으면 계정 하루 요청 상한을 넘길 수 있습니다."
        ):
            return

        extra = self._split_keywords(self.extraEdit.text())
        exclude = self._split_keywords(self.excludeEdit.text())
        tradeable = self.ui.onlyTradeableCheck.isChecked()

        def build(access_token):
            requests = []
            for task in tasks:
                for area in task["areas"]:
                    requests.append(
                        CrawlTask(
                            area=area,
                            keyword=task["keyword"],
                            only_tradeable=tradeable,
                            minimum=task["minimum"],
                            maximum=task["maximum"],
                            extra_keywords=extra,
                            exclude_keywords=exclude,
                            adaptive=True,
                            access_token=access_token,
                        )
                    )
            return requests

        self._start_crawl(build)

    def _read_excel_tasks(self, path: str) -> tuple[list[dict[str, Any]], list[str]]:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))  # type: ignore
        finally:
            wb.close()

        if not rows:
            return [], ["엑셀에 데이터가 없습니다"]

        header_map: dict[str, int] = {}
        for idx, value in enumerate(rows[0]):
            mapped = self.EXCEL_HEADER_ALIASES.get(str(value))
            if mapped and mapped not in header_map:
                header_map[mapped] = idx

        required = {"area", "keyword"}
        missing = required - header_map.keys()
        if missing:
            # 사용자가 엑셀에 실제로 적어야 하는 건 한글 머리글이다 → 내부 키 대신 한글로 안내.
            ko = {v: k for k, v in self.EXCEL_HEADER_ALIASES.items()}
            names = ", ".join(ko.get(m, m) for m in sorted(missing))
            return [], [f"엑셀 첫 행에 다음 열이 필요합니다: {names}"]

        area_idx = header_map["area"]
        keyword_idx = header_map["keyword"]
        minimum_idx = header_map.get("minimum")
        maximum_idx = header_map.get("maximum")

        errors: list[str] = []
        tasks: list[dict[str, Any]] = []

        def row_value(row: tuple[Any, ...], idx: int | None):
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row_idx, row in enumerate(rows[1:], start=2):
            area_raw = row_value(row, area_idx)
            keyword_raw = row_value(row, keyword_idx)

            # skip blank rows
            if not any(
                (
                    area_raw,
                    keyword_raw,
                    row_value(row, minimum_idx),
                    row_value(row, maximum_idx),
                )
            ):
                continue

            area_text = str(area_raw).strip() if area_raw is not None else ""
            keyword = str(keyword_raw).strip() if keyword_raw is not None else ""

            if not area_text:
                errors.append(f"{row_idx}행: 지역이 비어있습니다")
                continue

            if not keyword:
                errors.append(f"{row_idx}행: 키워드가 비어있습니다")
                continue

            area_tokens = [
                token.strip() for token in area_text.split(",") if token.strip()
            ]
            combined_areas: list[tuple[str, str]] = []
            missing_tokens: list[str] = []
            seen_codes: set[str] = set()

            for token in area_tokens:
                resolved = self._resolve_area_entry(token)
                if not resolved:
                    missing_tokens.append(token)
                    continue

                for entry in resolved:
                    if entry[1] not in seen_codes:
                        combined_areas.append(entry)
                        seen_codes.add(entry[1])

            if missing_tokens:
                errors.append(
                    f"{row_idx}행: 다음 지역을 찾을 수 없습니다 - {', '.join(missing_tokens)}"
                )
                continue

            if not combined_areas:
                errors.append(f"{row_idx}행: '{area_text}' 지역을 찾을 수 없습니다")
                continue

            try:
                minimum = (
                    self._parse_price_value(row_value(row, minimum_idx))
                    if minimum_idx is not None
                    else None
                )
            except ValueError:
                errors.append(f"{row_idx}행: 최소가격을 숫자로 변환할 수 없습니다")
                continue

            try:
                maximum = (
                    self._parse_price_value(row_value(row, maximum_idx))
                    if maximum_idx is not None
                    else None
                )
            except ValueError:
                errors.append(f"{row_idx}행: 최대가격을 숫자로 변환할 수 없습니다")
                continue

            if minimum is not None and maximum is not None and minimum > maximum:
                errors.append(f"{row_idx}행: 최소가격이 최대가격보다 큽니다")
                continue

            tasks.append(
                {
                    "areas": combined_areas,
                    "keyword": keyword,
                    "minimum": minimum,
                    "maximum": maximum,
                }
            )

        return tasks, errors

    def _resolve_area_entry(self, area_text: str) -> list[tuple[str, str]]:
        return self.area_lookup.get(area_text, [])

    def _parse_price_value(self, raw: Any) -> int | None:
        if raw is None:
            return None
        if isinstance(raw, (int, float)):
            return int(raw)
        text = str(raw).strip()
        if not text:
            return None
        text = text.replace(",", "")
        return int(float(text))

    def _prompt_recent_days_filter(self) -> tuple[bool, int | None]:
        while True:
            days_text, ok = QtWidgets.QInputDialog.getText(
                self,
                "필터링",
                "최근 며칠 이내의 상품만 저장할까요?\n(비워두면 전체 저장)",
            )
            if not ok:
                return False, None

            days_text = days_text.strip()
            if not days_text:
                return True, None

            try:
                days = int(days_text)
                if days <= 0:
                    raise ValueError
                return True, days
            except ValueError:
                self.alert("1 이상의 숫자를 입력해주세요.")

    def _filter_recent_products(
        self, products: list[Product], days: int | None
    ) -> list[Product]:
        if days is None:
            return products

        cutoff = datetime.now() - timedelta(days=days)
        return [prd for prd in products if prd.boostedAt >= cutoff]

    def on_select_prd(self, selected: QItemSelection, deselected: QItemSelection):
        prev_row = None
        prev_indexes = deselected.indexes()
        if prev_indexes:
            prev_row = prev_indexes[0].row()

        indexes = selected.indexes()
        if not indexes:
            return

        curr_row = indexes[0].row()

        if prev_row == curr_row:
            return

        curr_id = self.products_model.data(self.products_model.index(curr_row, 0))

        prd = self.controller.get_product_by_id(curr_id)
        if prd:
            self.load_detail(prd)
        else:
            self.alert("상품 정보를 불러오지 못했습니다")

    def load_detail(self, prd: Product):
        if self.current_loaded_url == prd.url:
            return
        self.current_loaded_url = prd.url

        document = self.ui.detailView.document()
        if document:
            self.ui.detailView.setHtml(render_to_html(prd))

        self.current_download_task_token = str(uuid4())

        if self.worker_thread is not None:
            self.worker_thread.cancel()

        if not prd.image:
            self.ui.prdImg.clear()
            return

        self.worker_thread = CancelableImageDownloader(
            self, prd.image, self.current_download_task_token
        )

        def on_download_finished(data: bytes, token: str):
            if token != self.current_download_task_token:
                return

            try:
                pil_img = PILImage.open(BytesIO(data))
                normalized = image_contain_resize(pil_img, self.preview_image_size)

                buffer = BytesIO()
                normalized.save(buffer, format="PNG")
                buffer.seek(0)

                qp = QPixmap()
                qp.loadFromData(buffer.getvalue())
                self.ui.prdImg.setPixmap(qp)
            except Exception:
                self.alert("이미지 로딩 오류")

        def on_download_failed(token: str):
            if token != self.current_download_task_token:
                return

            self.alert("이미지 다운로드 오류")

        self.worker_thread.finished.connect(on_download_finished)
        self.worker_thread.failed.connect(on_download_failed)
        self.worker_thread.start()

    def save_to_excel(self):
        if self.controller.is_task_running():
            self.alert("작업이 진행 중입니다")
            return

        products = list(self.controller.get_current_data())

        if not products:
            self.alert("크롤링된 상품이 없습니다")
            return

        proceed, recent_days = self._prompt_recent_days_filter()
        if not proceed:
            return

        filtered_products = self._filter_recent_products(products, recent_days)
        if not filtered_products:
            self.alert("조건에 맞는 상품이 없습니다")
            return

        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, filter="*.xlsx")
        if not path:
            return

        yesno = self.ask("이미지도 저장을 하시겠습니까?")
        export = ExportExcel(self, path, filtered_products, yesno)
        dlg = MyProgressDialog("저장 중", "취소", 0, 0, self)
        dlg.setWindowModality(Qt.WindowModality.ApplicationModal)

        dlg.canceled.disconnect()
        dlg.canceled.connect(lambda: export.cancel())

        def on_finished(success: bool):
            dlg.close()

            if export.cancelled:
                return

            if success:
                self.alert("저장되었습니다")
                try:
                    os.startfile(os.path.dirname(path))  # type: ignore
                except Exception:
                    pass
            else:
                self.alert("저장 중 오류")

        export.finished.connect(on_finished)
        export.start()

        dlg.show()

    def closeEvent(self, ev: QCloseEvent) -> None:  # type: ignore
        running = self.controller.is_task_running()
        prompt = (
            "작업이 진행 중입니다.\n정지하고 종료하시겠습니까?"
            if running
            else "종료 하시겠습니까?"
        )
        if not self.ask(prompt):
            ev.ignore()
            return

        # 임베드한 LDPlayer 창부터 떼어낸다(이후 정리에서 지연/예외가 나도 안전하게).
        self._emul_shutdown()

        if running:
            if not self.controller.is_task_stopping():
                self.controller.stop_task()
            task = self.controller.task
            if task is not None and task.isRunning():
                if not task.wait(8000):
                    task.terminate()
                    task.wait(2000)

        # 종료 전 실행 중인 QThread 정리 (미정리 시 SIGABRT)
        # 토큰 스레드는 controller.task 가 아직 없는 구간에 돈다 — 위 running
        # 분기가 못 잡는다. harvest_all 은 수십 초라 기다림이 길다.
        tok = getattr(self, "_token_thread", None)
        if tok is not None and tok.isRunning():
            self._search_cancelled = True
            if not tok.wait(8000):
                tok.terminate()
                tok.wait(2000)
        if self.auto_monitor is not None and self.auto_monitor.isRunning():
            self.auto_monitor.stop()
            if not self.auto_monitor.wait(3000):
                self.auto_monitor.terminate()
                self.auto_monitor.wait(2000)
        # 피드도 QThread 다 — 안 세우면 종료 때 SIGABRT 로 떨어진다. 사이클 안의
        # 요청 하나(최대 25초 타임아웃)를 기다려야 해서 auto_monitor 보다 넉넉하다.
        if self.feed_monitor is not None and self.feed_monitor.isRunning():
            self.feed_monitor.stop()
            if not self.feed_monitor.wait(8000):
                self.feed_monitor.terminate()
                self.feed_monitor.wait(2000)
        if self.worker_thread is not None and self.worker_thread.isRunning():
            self.worker_thread.cancel()
            self.worker_thread.wait(3000)
        if getattr(self, "_watch_timer", None):
            self._watch_timer.stop()
        th = getattr(self, "_watch_thread", None)
        if th is not None and th.isRunning():
            th.stop()                      # 다음 항목부터 중단
            if not th.wait(5000):          # 진행 중인 요청 하나는 기다린다
                th.terminate()
                th.wait(1000)
        for t in list(getattr(self, "_watch_threads", []) or []):
            if t.isRunning() and not t.wait(3000):
                t.terminate()
                t.wait(1000)
        alive = th is not None and th.isRunning()
        if getattr(self, "_watch_store", None) and not alive:
            self._watch_store.close()

    def clearItemList(self):
        self.controller.clear_products()

    def _handle_task_error(self):
        self._hide_search_progress()
        self._leave_task()
        self.alert("작업 중 오류\nERROR_LOG 파일에 오류가 저장되었습니다")

    def _handle_task_finished(self):
        self._hide_search_progress()
        self.alert("작업이 정지되었습니다")
        self._leave_task()

    def _handle_task_message(self, msg: str):
        self.sb.showMessage(msg)


ASSET_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")
FONT_DIR = os.path.join(ASSET_DIR, "fonts")
APP_ICON_PATH = os.path.join(ASSET_DIR, "icon.png")


def app_icon() -> QtGui.QIcon:
    """명품 앱 아이콘. 없으면 빈 QIcon(기본 로켓 대신 Qt 기본)."""
    return QtGui.QIcon(APP_ICON_PATH) if os.path.isfile(APP_ICON_PATH) else QtGui.QIcon()


def load_bundled_fonts() -> list[str]:
    """번들 Pretendard 등록. 미설치 환경에서도 QSS font-family 가 먹도록."""
    families: list[str] = []
    if not os.path.isdir(FONT_DIR):
        return families
    for name in sorted(os.listdir(FONT_DIR)):
        if not name.lower().endswith((".otf", ".ttf")):
            continue
        fid = QFontDatabase.addApplicationFont(os.path.join(FONT_DIR, name))
        if fid != -1:
            families.extend(QFontDatabase.applicationFontFamilies(fid))
    return families


def _setup_logging():
    """exe 옆 karrot_monitor.log 에 stdout/stderr + 모든 미처리 예외 기록(원격 디버깅용)."""
    import sys as _sys, os as _os, traceback as _tb, datetime as _dt
    base = _os.path.dirname(_sys.executable) if getattr(_sys, "frozen", False) \
        else _os.path.dirname(_os.path.abspath(__file__))
    try:
        logf = open(_os.path.join(base, "karrot_monitor.log"), "a",
                    encoding="utf-8", buffering=1)
    except Exception:
        return
    def stamp():
        return _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    logf.write(f"\n===== 시작 {stamp()} =====\n")

    class _Tee:
        def __init__(self, *s): self.s = [x for x in s if x]
        def write(self, d):
            for x in self.s:
                try: x.write(d)
                except Exception: pass
        def flush(self):
            for x in self.s:
                try: x.flush()
                except Exception: pass
    _sys.stdout = _Tee(_sys.__stdout__, logf)
    _sys.stderr = _Tee(_sys.__stderr__, logf)

    def _hook(et, ev, tb):
        logf.write(f"[{stamp()}] 미처리 예외:\n"
                   + "".join(_tb.format_exception(et, ev, tb)) + "\n")
        logf.flush()
    _sys.excepthook = _hook
    try:
        import threading as _th
        def _thook(a):
            logf.write(f"[{stamp()}] 스레드 예외:\n"
                       + "".join(_tb.format_exception(a.exc_type, a.exc_value, a.exc_traceback)) + "\n")
            logf.flush()
        _th.excepthook = _thook
    except Exception:
        pass
    print(f"[로그] {logf.name}")


def _chdir_app_dir():
    """앱 폴더로 작업디렉터리 고정 — 상대경로(accounts.json/OUT.json/data/) 견고.
    Windows 부팅 자동실행(cwd=system32)·바탕화면 실행 등 어디서 켜도 데이터 정상."""
    import sys as _sys
    try:
        base = os.path.dirname(_sys.executable) if getattr(_sys, "frozen", False) \
            else os.path.dirname(os.path.abspath(__file__))
        if base:
            os.chdir(base)
    except Exception as e:
        # chdir 실패 = 상대경로(data/·notify.json 등) 전부 깨짐 → 조용히 넘기면 무인서 오작동. 반드시 표면화.
        print(f"[치명] 작업디렉터리 이동 실패 → 상대경로 오작동 가능: {e}", file=_sys.stderr, flush=True)


def _child_cmd():
    """자식(실제 앱) 실행 커맨드 — frozen exe면 exe --child, 개발이면 python main.py --child.

    실행 모드(--manual / --watch)는 **그대로 물려준다**. 안 물려주면 워치독이
    한 번 재시작하는 순간 수동 전용으로 띄운 프로그램이 다른 모드로 되살아나고,
    수확기까지 같이 도는 프로그램이 두 개가 된다(함대를 동시에 깨운다).
    모드가 없으면 매물 감시다 — 합본은 없다."""
    import sys as _sys
    mode = [a for a in _sys.argv if a in ("--manual", "--watch")][:1] or ["--watch"]
    if getattr(_sys, "frozen", False):
        return [_sys.executable, "--child"] + mode
    return [_sys.executable, os.path.abspath(__file__), "--child"] + mode


def _run_watchdog():
    """크래시 자동복구 감시자 — 앱을 자식으로 실행, 비정상 종료 시 재시작.
    정상 종료(코드 0=유저가 창 닫음)면 감시 종료. 재시작 백오프 5·10·…·60초."""
    import subprocess as _sp, time as _t, sys as _sys
    _chdir_app_dir()
    print(f"[워치독] 감시 시작 — 크래시 시 자동 재시작")
    fails = 0
    while True:
        try:
            p = _sp.Popen(_child_cmd())
            code = p.wait()
        except Exception as e:
            print(f"[워치독] 실행 실패: {e}"); code = 1
        if code == 0:
            print("[워치독] 앱 정상 종료 — 감시 종료"); break
        fails += 1
        wait = min(60, 5 * fails)
        print(f"[워치독] 앱 비정상 종료(코드 {code}) — {wait}초 후 재시작 (누적 {fails}회)")
        _t.sleep(wait)


def ask_yes_no(parent, title, text, detail="", danger=False) -> bool:
    """예/아니오 확인창. 예를 골랐으면 True.

    QMessageBox.question 을 그대로 쓰면 버튼이 'Yes/No' 영문으로 나온다 —
    Qt 한국어 번역이 배포에 안 들어 있어서다. 클라가 보는 화면이라 버튼 글자를
    직접 준다. 기본 선택은 언제나 '아니오'다: 이 창이 뜨는 자리는 되돌리기
    어려운 삭제뿐이라, 엔터를 잘못 눌러 지워지는 일이 없어야 한다."""
    box = QtWidgets.QMessageBox(parent)
    box.setIcon(QtWidgets.QMessageBox.Icon.Warning if danger
                else QtWidgets.QMessageBox.Icon.Question)
    box.setWindowTitle(title)
    box.setText(text)
    if detail:
        box.setInformativeText(detail)
    yes = box.addButton("예", QtWidgets.QMessageBox.ButtonRole.YesRole)
    no = box.addButton("아니오", QtWidgets.QMessageBox.ButtonRole.NoRole)
    box.setDefaultButton(no)
    box.exec()
    return box.clickedButton() is yes


def _raise_window(w):
    """다른 실행이 아이콘을 눌렀다 — 이 창을 앞으로 올린다.

    최소화 상태면 편다. 창을 그냥 raise_() 만 하면 최소화된 창은 안 뜬다."""
    try:
        w.setWindowState(
            (w.windowState() & ~QtCore.Qt.WindowState.WindowMinimized)
            | QtCore.Qt.WindowState.WindowActive)
        w.show()
        w.raise_()
        w.activateWindow()
    except Exception:
        pass


def _run_app():
    _chdir_app_dir()
    _setup_logging()
    try:
        # 고배율(125·150%) 디스플레이에서 글자가 뭉개지지 않게 — 반올림 없이 실제 배율로 렌더
        QApplication.setHighDpiScaleFactorRoundingPolicy(
            Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)
        app = QApplication([])
        app.setWindowIcon(app_icon())   # Dock/작업표시줄/팝업 전부 명품 아이콘
        load_bundled_fonts()
        _f = app.font()                 # 힌팅 최대 — 원격 화면에서 획이 흐려지는 것 방지
        _f.setHintingPreference(QtGui.QFont.HintingPreference.PreferFullHinting)
        _f.setStyleStrategy(QtGui.QFont.StyleStrategy.PreferAntialias)
        app.setFont(_f)
        # 스타일은 **앱 전체**에 건다. MainWindow 에만 걸면 부모 없이 뜨는 팝업
        # (QMessageBox.warning(None, ...) 등)이 스타일을 못 받아 시스템 기본
        # 모습으로 나온다 — 본창과 톤이 달라 클라가 '창이 깨졌다'고 읽었다.
        app.setStyleSheet(APP_QSS)
        # 클라가 두 프로그램으로 쓴다: --manual(수동 검색) / --watch(매물 감시+에뮬).
        # 인자가 없으면 매물 감시다. 3탭 합본은 없다 — 옛 바로가기·작업이 인자
        # 없이 불러도 합본이 뜨지 않는다.
        import sys as _sysarg
        _mode = "watch"
        if "--manual" in _sysarg.argv:
            _mode = "manual"
        elif "--watch" in _sysarg.argv:
            _mode = "watch"
        # 수확·폴링·라우터를 도는 창은 하나여야 한다. 둘이면 각자의
        # KeywordRouter 가 같은 keyword_routes.json 을 서로 덮어 엑셀 조건이
        # 사라진다(실서버 2026-09-02: --watch 아이콘 + 자동시작 --child).
        #
        # 잠그는 기준은 모드가 아니라 background 다. 모드로 잠그면 'watch' 와
        # 옛 합본이 다른 이름이 돼 둘 다 통과했다 — 실제로 겹친 조합이 그것이다.
        # 수동검색(background=False)은 별도 이름이라 매물감시와 **함께 뜬다**.
        _sikey = None
        try:
            from daangn_ext import single_instance
            _sikey = single_instance.key_for(_mode, MainWindow.MODES)
            _solo = single_instance.acquire(_sikey)
        except Exception:
            _solo = True                  # 가드가 앱을 못 켜게 만들지는 않는다
        if not _solo:
            # 클라는 바탕화면 아이콘으로 켠다. 경고를 띄우고 죽으면 '아이콘을
            # 눌렀는데 안 켜진다'로 보인다. 먼저 뜬 창을 앞으로 올려 준다.
            if not single_instance.summon(_sikey):
                QtWidgets.QMessageBox.information(
                    None, "이미 실행 중",
                    "이미 실행 중입니다 — 작업표시줄에서 그 창을 여세요.")
            raise SystemExit(0)           # 0 = 정상 종료(워치독이 되살리지 않는다)
        window = MainWindow(_mode)
        # closeEvent 를 안 거치는 종료(세션 로그아웃 등)에서도 임베드 창은 되돌린다.
        app.aboutToQuit.connect(window._emul_shutdown)
        window.show()
        if _sikey:
            # 뒤에 오는 실행이 이 창을 불러낼 수 있게 문을 열어둔다.
            try:
                single_instance.serve(_sikey, lambda: _raise_window(window))
            except Exception:
                pass
        raise SystemExit(app.exec())     # 종료코드 전달(워치독이 정상/크래시 구분)
    except SystemExit:
        raise
    except Exception:
        import traceback
        traceback.print_exc()            # _Tee 로 로그파일에도 기록됨
        raise SystemExit(1)              # 예외 탈출=크래시 → 워치독 재시작 신호


def _run_headless():
    """서버(화면 없음) 무인 런타임 — GUI 없이 등록·폴링·수확·알림. PyQt 불필요.
    설정: data/alert_settings.json(core_only/night), notify.json(텔레그램/시트).
    옵션 인자: --once(1회만), --no-harvest(수확 스킵), --interval=N(주기초)."""
    import sys as _sys, json as _json, os as _os, time as _time
    _chdir_app_dir()
    from daangn_ext.keyword_alert_api import MultiAccountAlerts, token_remaining
    from daangn_ext.supervisor import SupervisorPolicy
    from daangn.feed_sweep import FeedSweep as _FeedSweepOnce

    def log(m):
        print(f"[{_time.strftime('%H:%M:%S')}] {m}", flush=True)

    # 앱API→웹크롤 폴백 경고를 무인 런타임 로그로 끌어온다. 안 걸면 stderr 로만
    # 나가고, 스윕을 안 거치는 경로(수동/보정 수집)의 폴백은 시각 표시도 없이
    # 섞인다. 웹크롤은 명품을 억제하므로 그 경고가 곧 '0건의 진짜 이유'다.
    # GUI 쪽은 일부러 등록하지 않는다 — 경고를 내는 쪽이 레인 스레드라
    # 위젯에 직접 append 하면 Qt 스레드 규칙을 어긴다. GUI 는 스윕이
    # sweep_engine.run() 에서 self._log 로 등록하는 경로와, stderr→
    # karrot_monitor.log 티잉(_setup_logging)으로 덮인다.
    try:
        from daangn_ext.adaptive import set_app_fallback_logger
        set_app_fallback_logger(log)
    except Exception as e:
        log(f"[경고] 앱API 폴백 로거 등록 실패: {str(e)[:80]}")

    argv = _sys.argv
    once = "--once" in argv
    do_harvest = "--no-harvest" not in argv
    interval = 120
    for a in argv:
        if a.startswith("--interval="):
            try: interval = max(30, int(a.split("=", 1)[1]))
            except Exception: pass

    def _settings():
        try:
            with open("./data/alert_settings.json", encoding="utf-8") as f:
                return _json.load(f)
        except Exception:
            return {}

    def _notify_cfg():
        try:
            with open("./notify.json", encoding="utf-8") as f:
                return _json.load(f)
        except Exception:
            return {}

    def _night_factor_live():
        if not bool(_settings().get("night")):
            return 1
        h = _time.localtime().tm_hour
        if 0 <= h < 7: return 3
        if 22 <= h < 24 or 7 <= h < 9: return 2
        return 1

    # GUI 와 같은 간격 정책 공유 — 야간 감속 배수가 폴링·워치 스윕 양쪽에
    # 같은 곳(SupervisorPolicy)에서 계산되게 한다(따로 계산하면 둘이 갈라진다).
    policy = SupervisorPolicy(lambda: interval, _night_factor_live,
                              sweep_interval=WATCH_SWEEP_INTERVAL)

    def _notify(items, nt):
        if not items:
            return
        tok, chat = nt.get("tg_token"), nt.get("tg_chat")
        if tok and chat:
            try:
                from daangn.notify import TelegramSender, item_block
                tg = TelegramSender(tok, chat, log=log)
                for m in items:
                    tg.enqueue_item(item_block(
                        "신규 매물", m.get("region"), m.get("title"), m.get("price"),
                        m.get("url"), stamp=m.get("time"), stamp_label="등록"))
                tg.flush(); log(f"[텔레그램] {len(items)}건 전송")
            except Exception as e:
                log(f"[텔레그램] 실패: {str(e)[:60]}")
        if nt.get("sheet_url"):
            try:
                from daangn.notify import SheetWriter
                sw = SheetWriter(nt.get("sheet_url"), nt.get("sheet_cred") or "./credentials.json", log=log)
                for m in items:
                    ts = _time.strftime("%Y-%m-%d %H:%M", _time.localtime(int(m.get("time") or 0))) if m.get("time") else ""
                    sw.enqueue_row([ts, m.get("keyword") or "", m.get("title") or "", m.get("price") or "",
                                    m.get("region") or "", m.get("_account") or "", m.get("url") or ""])
                wrote, _ = sw.flush()
                if wrote: log(f"[구글시트] {wrote}행 기록")
            except Exception as e:
                log(f"[구글시트] 실패: {str(e)[:60]}")

    def _notify_lines(lines, nt, events=None):
        """워치리스트 변동 전송 — 텔레그램(매물 블록) + 구글시트(한 줄 요약)."""
        if not lines:
            return
        tok, chat = nt.get("tg_token"), nt.get("tg_chat")
        if tok and chat:
            try:
                from daangn.notify import TelegramSender
                tg = TelegramSender(tok, chat, log=log)
                _enqueue_watch_blocks(tg, events, lines)
                tg.flush()
                log(f"[텔레그램] 변동 {len(lines)}건 전송")
            except Exception as e:
                log(f"[텔레그램] 실패: {str(e)[:60]}")
        if nt.get("sheet_url"):
            try:
                from daangn.notify import SheetWriter
                sw = SheetWriter(nt.get("sheet_url"),
                                 nt.get("sheet_cred") or "./credentials.json", log=log)
                ts = _time.strftime("%Y-%m-%d %H:%M")
                for ln in lines:
                    sw.enqueue_row([ts, "가격변동", ln])
                wrote, _ = sw.flush()
                if wrote:
                    log(f"[구글시트] {wrote}행 기록")
            except Exception as e:
                log(f"[구글시트] 실패: {str(e)[:60]}")

    log("=== 헤드리스 무인 모니터 시작 ===")
    last_harvest = 0.0
    next_harvest = 0.0          # 첫 루프에서 곧바로 한 번 수확한다
    hstats = {}
    app_sweep_off_logged = [False]   # 틱마다 안 찍고 프로세스당 한 번만
    watch_store = watch_tracker = watch_budget = None
    last_watch_sweep = 0.0
    try:
        from daangn_ext import article_watch
        watch_store = article_watch.WatchStore("./data/watch.db")
        watch_tracker = article_watch.WatchTracker(watch_store)
        # GUI(_watch_budget)와 같은 풀·같은 재조회 규칙.
        watch_budget = article_watch.ProxyBudget(
            feed_proxies(_settings()), provider=lambda: feed_proxies(_settings()))
    except Exception as e:
        log(f"[가격추적] 초기화 실패 — 가격추적 없이 계속: {str(e)[:120]}")
    # 중복 판정은 watch 테이블이 한다(dead 행을 남기므로 '본 매물'의 진실이다).
    # article_id 없는 매치는 같은 DB 의 seen_key 테이블이 받아 재시작해도 남는다.
    # 이 집합은 저장소를 아예 못 열었을 때만 쓰는 마지막 방어선이다.
    fallback_seen = set()
    m = MultiAccountAlerts("./accounts.json", "./data/config.json")

    # ── 라우터 · 검색 스윕 — GUI(MainWindow.__init__)와 같은 생성 경로 ──
    # 이게 없으면 앱 슬롯 상한을 넘긴 키워드는 서버에서 아무도 안 본다.
    router = sweep_queue = sweep_runner = feed_runner = None
    alert_rules = AlertRulesCache()
    seed_state = {}
    # 스윕 스레드가 찾은 매물은 여기로만 건너온다 — sqlite 는 폴링 스레드 소유다.
    import queue as _queue
    sweep_found_q = _queue.Queue(maxsize=SWEEP_FIND_QUEUE_MAX)
    sweep_found_dropped = [0]

    def _server_keyword_list(core_only=False):
        """서버에 이미 등록된 키워드 목록(첫 유효 계정 기준).

        등록은 전 계정에 같은 키워드를 쓰므로 한 계정만 봐도 함대 상태를 안다.
        첫 실행 씨딩에만 쓰이며, routes 가 차면 호출조차 되지 않는다."""
        from daangn_ext.keyword_alert_api import KeywordAlertAPI
        valid = m._valid(core_only)
        if not valid:
            return {}
        _code, access, proxy = valid[0]
        api = KeywordAlertAPI(access, "./data/config.json", proxy=proxy)
        try:
            return api.list()
        finally:
            try:
                api.close()
            except Exception:
                pass

    try:
        from daangn_ext.sweep_queue import SweepQueue
        from daangn_ext.keyword_router import KeywordRouter, DEFAULT_SLOT_CAP
        sweep_queue = SweepQueue("./data/sweep_queue.json")
        _cap = int(_settings().get(SLOT_CAP_KEY) or DEFAULT_SLOT_CAP)
        router = KeywordRouter(m, sweep_queue, slot_cap=_cap)

        def _sweep_found(payload):
            """**스윕 스레드**에서 불린다 — 큐에 넣기만 하고 아무것도 안 만진다.

            여기서 WatchStore 를 건드리면 폴링 루프와 sqlite 커넥션 하나를
            두 스레드가 나눠 쓴다(drain_sweep_finds 참고). put_nowait 는
            절대 막히지 않으므로 스윕 스레드가 폴링에 붙잡히지 않는다."""
            try:
                sweep_found_q.put_nowait(payload)
            except _queue.Full:
                sweep_found_dropped[0] += 1

        def _sweep_cfg_builder():
            return headless_sweep_cfg(
                _settings(), sweep_queue.entries(), _notify_cfg(),
                proxies=headless_proxies(), proxy_provider=headless_proxies,
                # 헤드리스에서 수확을 소유하는 건 아래 폴링 루프(20분 주기)뿐이다.
                # 스윕까지 harvest_all 을 부르면 두 스레드가 스케줄에 맞춰 같은
                # 함대를 동시에 깨운다 — adb 폭주이자 accounts.json 동시쓰기이며,
                # LDPlayer 순차기동 규칙도 어긴다. 스윕이 필요한 건 신선한 토큰이라
                # 폴링 루프가 갱신해 둔 파일을 읽는 것으로 충분하다.
                # --no-harvest 면 아무도 갱신하지 않으므로 provider 자체를 뺀다
                # (= stabilize off, 기존 동작 그대로).
                token_provider=(read_token_quiet if do_harvest else None),
                # 지역 미지정이면 기본 지역으로 좁힌다는 사실을 서버 로그에
                # 남긴다 — 안 남기면 커버리지가 6537동에서 조용히 줄어든다.
                log=log,
                # 앱 알림이 이미 알린 매물은 스윕이 다시 안 알린다.
                already_notified=lambda aid: bool(
                    watch_store.get(str(aid))) if watch_store else False)

        sweep_runner = HeadlessSweepRunner(sweep_queue, _sweep_cfg_builder,
                                           log, _sweep_found)
    except Exception as e:
        log("[검색스윕] 초기화 실패 — 앱 슬롯 밖 키워드는 커버되지 않습니다: "
            f"{str(e)[:120]}")

    # ── 동 피드 — 앱 슬롯·조건표 무관, 웹 동 피드로 넓게 훑는다 ──
    def _feed_already_notified(href) -> bool:
        """피드가 낸 href 를 이미 봤는가 — GUI _already_notified 와 같은 규칙.

        href 행은 첫 상세 조회에서 숫자 id 로 옮겨 가지만 url 은 href 그대로라,
        커서를 잃고 같은 href 가 다시 올라와도 url 로 잡힌다."""
        if watch_store is None:
            return False
        try:
            key = str(href)
            return bool(watch_store.get(key) or watch_store.get_by_url(key))
        except Exception:
            return False

    feed_runner = HeadlessFeedRunner(
        lambda: feed_cfg(_settings(), _notify_cfg(), log=log,
                         already_notified=_feed_already_notified),
        log, _sweep_found)

    # 관측 상한 되돌리기 — 등록 엔드포인트의 일시적 오류 하나로 유효 상한이
    # 잘못 내려앉았을 때의 탈출구. 이 플래그가 생기기 전에는 서버에서
    # data/keyword_routes.json 을 손으로 고치는 것 말고 방법이 없었다.
    if "--reset-cap" in argv:
        if router is None:
            log("[라우터] 초기화 실패 — 상한 관측치를 되돌릴 수 없습니다")
        elif router.reset_observed_cap():
            log("[라우터] 앱 슬롯 상한 관측치 초기화 —"
                " 다음 등록부터 설정 상한을 다시 씁니다")
        else:
            log("[라우터] 상한 관측치가 이미 비어 있습니다 — 되돌릴 것이 없습니다")

    # 서버 부트스트랩: 조건표의 브랜드 등록 (--register) 후 --once면 종료
    if "--register" in argv:
        st = _settings(); co = bool(st.get("core_only"))
        # 등록 경로는 조건표 하나다. 브랜드만 등록해 놓고 조건표가 없으면
        # 모델·가격대와 무관하게 브랜드 전 매물이 알림으로 쏟아진다.
        from daangn_ext.alert_rules import brands as _rule_brands
        _rules = load_alert_rules().rules
        _brands = _rule_brands(_rules)
        if router is None:
            log("[등록] 라우터가 없어 등록을 건너뜁니다 — 초기화 실패 로그를 확인하세요")
        elif not _brands:
            log(f"[등록] 조건표가 비어 있습니다({ALERT_RULES_FILE}) —"
                " GUI 조건 탭 표에 적고 [조건 적용]을 먼저 누르세요")
        else:
            log(f"[등록] 조건표 {len(_rules)}줄 · 브랜드 {len(_brands)}개 등록"
                f" (커버 {'핵심' if co else '전국'})")
            # 등록 **전에** 씨딩한다 — 이 브랜치 이전 경로로 이미 서버가 꽉 차
            # 있으면, 인정하지 않고 등록하면 전부 실패해 스윕으로 밀린다.
            _allowed = rule_brand_keys(load_alert_rules())
            seed_router_from_server(
                router, lambda: _server_keyword_list(co), log, seed_state,
                allowed=_allowed,
                prune_fn=lambda extras: m.delete_keywords(extras, log=log, core_only=co))
            try:
                # 가격은 넘기지 않는다 — 당근 서버가 먼저 자르면 조건표가 볼
                # 매물 자체가 없어진다. 거르는 일은 조건표가 한다.
                for ks, days in brand_register_groups(_brands, _rules):
                    for r in router.add_many(ks, None, None, None, core_only=co,
                                             log=log, days=days,
                                             replace_cond=True):
                        log(f"[등록] {r.get('keyword')} → {r.get('route')}"
                            f" ({r.get('reason') or ''})")
                prune_to_rules(router, m, _allowed, log, co)
            except Exception as e:
                log(f"[등록] 실패: {str(e)[:80]}")
        if once:
            log("--register --once 완료")
            if sweep_runner is not None:
                sweep_runner.stop(join=8)
            if feed_runner is not None:
                feed_runner.stop(join=8)
            return
    try:
        while True:
            st = _settings()
            core_only = bool(st.get("core_only"))
            now = _time.time()
            # 자동수확(20분 주기)
            # 다음 수확 시각은 만료 직후로 잡힌다(GUI 와 같은 규칙).
            if do_harvest and now - last_harvest > next_harvest:
                try:
                    import ld_autoharvest
                    guest_proxy_sync("./accounts.json", log=log)
                    u, i, t, h = ld_autoharvest.harvest_all(
                        "./accounts.json", nudge=True, log=log, stats=hstats)
                    log(f"[수확] 갱신 {u} 신규 {i} 총 {t} (수확 {h})")
                except Exception as e:
                    log(f"[수확] 실패: {str(e)[:60]}")
                last_harvest = now
                import ld_autoharvest as _LA
                next_harvest = _LA.next_harvest_delay(hstats.get("min_remaining"),
                                                      ceil=harvest_interval())
            # ── 라우터 · 검색 스윕 — GUI _auto_poll_tick 과 같은 순서 ──
            # 씨딩 → 승격 → 스윕 재동기화. 재동기화는 루프 1회에 한 번뿐이라
            # 대기열이 요동쳐도 재시작은 폴링 주기당 한 번을 넘지 않는다.
            _allowed = rule_brand_keys(load_alert_rules())
            seed_router_from_server(
                router, lambda: _server_keyword_list(core_only), log, seed_state,
                allowed=_allowed,
                prune_fn=lambda extras: m.delete_keywords(extras, log=log,
                                                          core_only=core_only))
            if router is not None:
                try:
                    for p in router.rebalance(core_only=core_only, log=log):
                        log(f"[라우터] {p['keyword']} → 앱 알림 승격")
                except Exception as e:
                    log(f"[라우터] 승격 실패: {str(e)[:80]}")
                mirror_app_keywords_to_sweep(
                    router, sweep_queue, log=log,
                    enabled=sweep_mirror_enabled(
                        st, len(load_alert_rules())))
            if sweep_runner is not None:
                if sweep_app_enabled(st):
                    app_sweep_off_logged[0] = False
                    sweep_runner.resync()
                else:
                    if sweep_runner.running():
                        sweep_runner.stop()
                    if not app_sweep_off_logged[0]:
                        log("[검색스윕] 앱 스윕 꺼짐(설정) — 피드가 발굴합니다")
                        app_sweep_off_logged[0] = True
            if feed_runner is not None:
                if once:
                    # --once 는 스레드를 띄우지 않고 한 회차만 동기로 돈다 —
                    # 안 그러면 스레드가 뜨자마자 프로세스가 끝나 아무것도 못 본다.
                    try:
                        _fcfg = feed_once_cfg(feed_runner.cfg_builder(), log)
                        _fresult = _FeedSweepOnce(
                            _fcfg, on_log=log, on_found=_sweep_found).cycle_once()
                        log(f"[피드] 1회 결과: {_fresult}")
                    except Exception as e:
                        log(f"[피드] 1회 실패: {str(e)[:120]}")
                elif not feed_runner.running() and \
                        _settings().get("feed_enabled", FEED_DEFAULTS["feed_enabled"]) and \
                        len(load_alert_rules().rules):
                    feed_runner.start()
            # 스윕 스레드가 넘긴 매물을 여기(폴링 스레드)서 워치리스트에 넣는다.
            if sweep_found_dropped[0]:
                log(f"[검색스윕] 인계 큐가 차서 {sweep_found_dropped[0]}건 버림")
                sweep_found_dropped[0] = 0
            drain_sweep_finds(sweep_found_q, watch_tracker,
                              (sweep_queue.keywords if sweep_queue is not None
                               else list), log)
            # 유효 토큰 현황
            try:
                valid = len(m._valid(core_only))
            except Exception:
                valid = 0
            # 폴링
            try:
                matches = m.poll_all(core_only=core_only, log=log)
            except Exception as e:
                log(f"[폴링] 실패: {str(e)[:60]}"); matches = []
            matches, watch_only, _ = filter_by_conditions(
                matches, router, log, rules=alert_rules.get())
            fresh, dropped = dedupe_new_matches(matches, watch_store, fallback_seen)
            if dropped:
                log(f"[매칭] id 없는 payload {dropped}건 건너뜀")
            # 알림 여부와 무관하게 추적한다(GUI 경로와 같은 이유).
            try:
                added = watch_tracker.add_from_matches(fresh + watch_only) \
                    if watch_tracker else 0
                if added:
                    log(f"[가격추적] {added}건 추적 시작")
            except Exception as e:
                log(f"[가격추적] 등록 실패: {str(e)[:80]}")
            if fresh:
                log(f"[매칭] 신규 {len(fresh)}건 (유효계정 {valid})")
                _notify(fresh, _notify_cfg())
            else:
                log(f"[매칭] 신규 0 (유효계정 {valid}, 커버 {'핵심' if core_only else '전국'})")
            # 워치리스트 스윕 — 폴링과 같은 스레드에서 정책이 정한 간격으로만.
            if watch_tracker and watch_budget and \
                    headless_watch_due(last_watch_sweep, now, policy.sweep_ms() // 1000):
                last_watch_sweep = now
                try:
                    dropped = watch_tracker.enforce_cap()
                    if dropped:
                        log(f"[가격추적] 상한 초과 {dropped}건 추적 중단")
                    # 예산은 GUI(_WatchSweepThread)와 같이 생짜 WATCH_SWEEP_INTERVAL 로 계산한다.
                    # 스윕 "빈도"만 야간에 느려지고, 한 번 돌 때의 회차 예산 크기는 안 커진다 —
                    # GUI 와 다르게 계산하면 두 런타임이 다시 갈라진다.
                    budget = watch_sweep_budget(watch_store.active_count(),
                                                WATCH_SWEEP_INTERVAL)
                    if budget:
                        watch_budget.reload()
                        events = mark_range_entries(
                            watch_tracker.sweep(watch_budget.next, budget),
                            watch_store, router, rules=alert_rules.get())
                        lines = watch_event_lines(events)
                        if getattr(watch_tracker, "last_sweep_exhausted", False):
                            log("[가격추적] 계정 예산 소진 — 남은 대상은 다음 회차로")
                        if lines:
                            log("[가격추적] " + " / ".join(lines))
                            _notify_lines(lines, _notify_cfg(), events)
                except Exception as e:
                    log(f"[가격추적] 스윕 실패: {str(e)[:120]}")
            if once:
                log("--once 완료"); break
            eff = policy.poll_ms() // 1000
            log(f"다음 폴링 {eff}초 후")
            try:
                _time.sleep(eff)
            except KeyboardInterrupt:
                log("중단됨"); break
    finally:
        # 스윕 스레드를 남겨두면 프로세스가 안 끝난다(서버 재시작이 걸린다).
        # --once·KeyboardInterrupt·예외 어느 경로로 나가도 여기를 지난다.
        if sweep_runner is not None:
            sweep_runner.stop(join=8)
        if feed_runner is not None:
            feed_runner.stop(join=8)
        # 멈춘 뒤 큐에 남은 것까지 넣는다. 안 그러면 --once 는 스윕이 찾은 걸
        # 하나도 기록하지 못하고(시작하자마자 비어 있는 큐를 훑고 끝난다),
        # 상시 운영에서도 종료 때 마지막 주기 분이 통째로 사라진다.
        drain_sweep_finds(sweep_found_q, watch_tracker,
                          (sweep_queue.keywords if sweep_queue is not None
                           else list), log)


if __name__ == "__main__":
    import sys as _sys
    # Windows 콘솔(cp1252/cp949)서 한글 로그 UnicodeEncodeError 방지
    for _st in (_sys.stdout, _sys.stderr):
        try:
            _st.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass
    if "--headless" in _sys.argv:
        _run_headless()
    elif "--watchdog" in _sys.argv:
        _run_watchdog()
    else:
        _run_app()
