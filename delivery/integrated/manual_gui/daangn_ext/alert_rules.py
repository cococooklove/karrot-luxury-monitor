"""
알림 룰 테이블 — "브랜드로 넓게 수집하고, 엑셀 조건에 맞는 것만 알린다".

앱 알림 경로는 당근 서버가 판정하고 키워드 등록에 상한이 있다(수십 개).
그래서 모델별 가격대 수백 줄을 키워드로 등록할 수 없다. 대신 브랜드
("루이비통") 하나로 넓게 받아 놓고, 엑셀 수백 줄을 알림 직전 필터로 태운다.

  엑셀 한 줄 = 룰 하나 = "키워드 + 최소가격 ~ 최대가격"
  매물은 룰 중 하나라도 맞으면 알린다(OR).

가격이 상한만 넘긴 매물은 버리지 않고 추적한다 — 값이 내려와 범위 안에
들어오면 그때 알린다(mark_range_entries 와 같은 정책).

키워드 매칭은 search_filters 의 어절 AND 규칙을 그대로 쓴다(띄어쓰기 무시).
"""
from __future__ import annotations

import json
import time
from dataclasses import dataclass, field

from .price import parse_price_text
from .search_filters import (FIELD_SEP, contains_keyword, looks_wanted_ad,
                             normalize_text)

# verdict
HIT = "hit"       # 알린다
WATCH = "watch"   # 키워드는 맞지만 상한 초과 — 인하 대기로 추적
CUT = "cut"       # 어떤 룰에도 안 맞음
PASS = "pass"     # 룰 테이블이 비었음 — 필터하지 않는다


@dataclass(frozen=True)
class AlertRule:
    keyword: str                      # 매칭에 쓰는 말 전체(브랜드 + 제품명)
    brand: str = ""                   # 엑셀 '브랜드' 열. 등록은 이 값으로 한다
    product: str = ""                 # 엑셀 '제품명' 열. 화면 표시용
    min_price: int | None = None
    max_price: int | None = None
    exclude: tuple[str, ...] = ()
    days: int | None = None           # 끌올일수 — 검색 스윕 경로에서만 쓴다
    row: int = 0                      # 엑셀 행번호(로그·디버깅용)

    def brand_name(self) -> str:
        """등록에 쓸 브랜드.

        엑셀에 브랜드 열이 있으면 그 값이 진실이다. 없을 때만 키워드 첫
        어절로 짐작한다 — 그 짐작은 '보테가 베네타 카세트백'을 '보테가'로,
        '생 로랑 루루백'을 '생'으로 만든다. 두 어절 브랜드가 실제로 있다."""
        if str(self.brand or "").strip():
            return str(self.brand).strip()
        parts = str(self.keyword or "").split()
        return parts[0] if parts else ""

    def label(self) -> str:
        lo = f"{self.min_price:,}" if self.min_price else ""
        hi = f"{self.max_price:,}" if self.max_price else ""
        return f"{self.keyword} ({lo}~{hi})" if (lo or hi) else self.keyword

    def to_dict(self) -> dict:
        return {"keyword": self.keyword, "brand": self.brand,
                "product": self.product, "min": self.min_price,
                "max": self.max_price, "exclude": list(self.exclude),
                "days": self.days, "row": self.row}

    @classmethod
    def from_dict(cls, d: dict) -> "AlertRule":
        return cls(keyword=str(d.get("keyword") or ""),
                   brand=str(d.get("brand") or ""),
                   product=str(d.get("product") or ""),
                   min_price=_as_int(d.get("min")),
                   max_price=_as_int(d.get("max")),
                   exclude=tuple(str(x) for x in (d.get("exclude") or []) if x),
                   days=_as_int(d.get("days")),
                   row=int(d.get("row") or 0))


def _as_int(v) -> int | None:
    """'1,300,000' · '900,000원' · '285만원' · 1300000.0 · None 을 모두 받는다.

    알림 payload 의 가격은 '285만원'처럼 줄여 쓴 문자열이라 숫자만 뽑으면
    백만 배 어긋난다. 파싱은 price.parse_price_text 한 곳에만 둔다
    (article_watch 는 manual_gui 트리에만 있어 여기서 부르면 안 된다).
    """
    if v is None or v == "" or isinstance(v, bool):
        return None
    return parse_price_text(v) or None


@dataclass
class RuleTable:
    rules: list[AlertRule] = field(default_factory=list)
    drop_wanted: bool = True          # 삽니다/구합니다 글 컷
    applied_at: int = 0               # 적용 시각(epoch). 화면이 "언제 넣은 것"을 말한다
    source: str = ""                  # 불러온 엑셀 파일 이름

    def __len__(self) -> int:
        return len(self.rules)

    def detail(self) -> str:
        """브랜드 목록과 적용 시각. 줄 수는 섹션 제목이 이미 말한다."""
        if not self.rules:
            return "조건 없음 — 엑셀을 넣기 전까지 알리지 않습니다"
        bs = brands(self.rules)
        head = " · ".join(bs[:6]) + (f" 외 {len(bs) - 6}개" if len(bs) > 6 else "")
        when = (time.strftime("%m/%d %H:%M", time.localtime(self.applied_at))
                if self.applied_at else "")
        return head + (f"    ·    {when} 적용" if when else "")

    def summary(self) -> str:
        """화면 한 줄 요약. 조건이 없으면 그 사실을 분명히 말한다."""
        if not self.rules:
            return "조건 없음 — 엑셀을 넣기 전까지 알리지 않습니다"
        bs = brands(self.rules)
        head = " · ".join(bs[:3]) + (f" 외 {len(bs) - 3}개" if len(bs) > 3 else "")
        when = (time.strftime("%m/%d %H:%M", time.localtime(self.applied_at))
                if self.applied_at else "")
        return (f"조건 {len(self.rules)}개 · 브랜드 {len(bs)}개 — {head}"
                + (f"   ({when} 적용)" if when else ""))

    def verdict(self, title, price=None, body="") -> tuple[str, AlertRule | None]:
        """(판정, 맞은 룰). 룰이 없으면 (PASS, None) — 호출측(filter_by_conditions)이
        '조건 없음 = 알리지 않음'으로 처리한다."""
        if not self.rules:
            return PASS, None
        if self.drop_wanted and looks_wanted_ad(title):
            return CUT, None
        # 제목과 본문은 경계 표식으로 잇는다 — 그냥 붙이면 제목 끝 글자와
        # 본문 첫 글자가 한 어절로 뭉쳐 없던 단어가 생긴다.
        text = FIELD_SEP.join(normalize_text(x) for x in (title, body) if x)
        price = _as_int(price)
        watched: AlertRule | None = None
        for r in self.rules:
            if not contains_keyword(text, r.keyword):
                continue
            if any(contains_keyword(text, x) for x in r.exclude):
                continue
            # 가격을 못 읽으면 가격 조건은 건너뛴다 — 못 읽었다는 이유로 버리지 않는다.
            if price is None:
                return HIT, r
            if r.min_price is not None and price < r.min_price:
                continue
            if r.max_price is not None and price > r.max_price:
                watched = watched or r
                continue
            return HIT, r
        return (WATCH, watched) if watched else (CUT, None)

    # ── 저장/복원 ──
    def to_json(self) -> str:
        return json.dumps({"rules": [r.to_dict() for r in self.rules],
                           "drop_wanted": self.drop_wanted,
                           "applied_at": self.applied_at,
                           "source": self.source},
                          ensure_ascii=False, indent=1)

    def save(self, path) -> None:
        # 적용 시각은 저장하는 순간이 진실이다 — 화면이 "언제 넣은 조건인지"를
        # 말해야 클라가 자기가 넣은 그 파일인지 알아본다.
        if self.rules and not self.applied_at:
            self.applied_at = int(time.time())
        with open(path, "w", encoding="utf-8") as f:
            f.write(self.to_json())

    @classmethod
    def load(cls, path) -> "RuleTable":
        try:
            with open(path, encoding="utf-8") as f:
                d = json.load(f)
        except (OSError, ValueError):
            return cls()
        return cls(rules=[AlertRule.from_dict(x) for x in (d.get("rules") or [])],
                   drop_wanted=bool(d.get("drop_wanted", True)),
                   applied_at=int(d.get("applied_at") or 0),
                   source=str(d.get("source") or ""))


# ── 엑셀 로더 ──
# 클라 시트: 키워드 | 최소가격 | 최대가격 (+ 선택: 제외).
# 머리글은 부분일치로 잡는다 — "최소가격"·"최소금액"·"최소 가격" 다 받는다.
def load_rules_from_excel(path) -> tuple[list[AlertRule], list[str]]:
    """엑셀 → (룰, 오류·안내). **모든 시트**를 읽어 합친다.

    시트 하나만 읽던 동안, 브랜드별로 탭을 나눈 파일은 첫 탭만 들어가고
    나머지가 조용히 사라졌다. 오류도 안 났으니 알아챌 방법이 없었다.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = [(ws.title, list(ws.iter_rows(values_only=True)))
                  for ws in wb.worksheets]
    finally:
        wb.close()
    rules: list[AlertRule] = []
    errors: list[str] = []
    seen: set = set()
    read_any = False
    for title, rows in sheets:
        got, errs = parse_rule_rows(rows)
        if got:
            read_any = True
        for r in got:
            key = (normalize_text(r.keyword), r.min_price, r.max_price, r.exclude)
            if key in seen:
                continue
            seen.add(key)
            rules.append(r)
        # 시트가 여럿이면 어느 시트 몇 행인지까지 말해야 찾아갈 수 있다.
        for m in errs:
            errors.append(f"[{title}] {m}" if len(sheets) > 1 else m)
    if len(sheets) > 1 and read_any:
        counts = ", ".join(f"{t} {len(parse_rule_rows(r)[0])}줄" for t, r in sheets)
        errors.insert(0, f"시트 {len(sheets)}개를 모두 읽었습니다 — {counts}")
    return rules, errors


def parse_rule_rows(rows) -> tuple[list[AlertRule], list[str]]:
    """엑셀 행 튜플 목록 → (룰, 오류메시지). 테스트가 엑셀 없이 쓴다."""
    if not rows:
        return [], ["엑셀에 데이터가 없습니다"]

    # 머리글이 1행이라는 보장이 없다 — 제목 한 줄, 빈 줄, 안내문이 위에 붙은
    # 파일이 흔하다. 위에서부터 '키워드' 가 있는 행을 찾아 거기서 시작한다.
    head_at = None
    for n, row in enumerate(rows[:10]):
        cells = [normalize_text(c) for c in row]
        if any(c and ("키워드" in c or "keyword" in c or "브랜드" in c)
               for c in cells):
            head_at = n
            break
    if head_at is None:
        return [], ["엑셀 머리글에 '브랜드' 열이 없습니다"
                    " — [샘플 엑셀 저장]으로 형식을 받아 쓰세요"]
    head_row = head_at
    head = [normalize_text(c) for c in rows[head_at]]
    rows = rows[head_at:]

    def col(*names) -> int | None:
        for i, h in enumerate(head):
            if h and any(normalize_text(n) in h for n in names):
                return i
        return None

    i_brand, i_prod = col("브랜드"), col("제품", "모델", "상품")
    i_kw, i_min = col("키워드", "keyword"), col("최소")
    i_max, i_exc = col("최대"), col("제외")
    # 옛 시트도 그대로 읽는다 — 헤더가 다를 뿐 같은 뜻이다. 제외 컬럼은
    # "제외키워드"도 잡히고, 추가키워드는 키워드에 붙여 한 줄로 만든다
    # (어절 AND 라 "루이비통 오버 더 문 정품"과 뜻이 같다). 끌올일수는 이제
    # 안내하지 않지만, 적어 둔 옛 파일을 버리지는 않는다.
    i_add, i_days = col("추가"), col("끌올")
    if i_brand is None and i_kw is None:
        return [], ["엑셀 첫 행에 '브랜드' 열이 필요합니다"]

    rules: list[AlertRule] = []
    errors: list[str] = []
    seen: set[tuple] = set()
    # 브랜드를 첫 줄에만 적고 아래 줄은 비워 두는(또는 셀을 병합하는) 시트가
    # 흔하다 — 병합 셀은 첫 칸만 값이 읽히고 나머지는 빈 칸으로 온다. 그 빈
    # 칸을 제품명 첫 어절로 짐작하면 '트위스트'가 브랜드로 등록되고, 그 줄은
    # '트위스트'만으로 매칭돼 남의 브랜드 매물까지 잡는다. 엑셀 관례대로
    # 바로 위 브랜드를 이어받는다.
    last_brand = ""
    # 행 번호는 엑셀에서 보이는 그대로여야 찾아갈 수 있다.
    for n, row in enumerate(rows[1:], start=head_row + 2):
        def cell(i):
            return row[i] if (i is not None and i < len(row)) else None

        brand = str(cell(i_brand) or "").strip()
        prod = str(cell(i_prod) or "").strip()
        if i_brand is not None:
            if brand:
                last_brand = brand
            elif prod:
                if not last_brand:
                    errors.append(f"{n}행 '{prod}': 위에 브랜드가 없어 이어받을 수"
                                  " 없습니다 — 브랜드 칸을 채우세요 — 건너뜀")
                    continue
                brand = last_brand
        # '키워드' 한 열짜리 옛 시트와, '브랜드+제품명' 두 열짜리 새 시트를
        # 같은 규칙으로 읽는다. 매칭은 둘을 이어 붙인 말 전체로 한다.
        kw = str(cell(i_kw) or "").strip()
        if not kw:
            kw = " ".join(x for x in (brand, prod) if x)
        if not kw:
            continue
        add = str(cell(i_add) or "").replace(",", " ").split()
        if add:
            kw = " ".join([kw] + add)
        lo, hi = _as_int(cell(i_min)), _as_int(cell(i_max))
        if lo is not None and hi is not None and lo > hi:
            errors.append(f"{n}행 '{kw}': 최소가격이 최대가격보다 큽니다 — 건너뜀")
            continue
        # 제외는 쉼표로만 나눈다 — 공백으로 나누면 "A급 레플리카" 가 "A급"
        # 하나로 쪼개져 멀쩡한 매물까지 날아간다. 한 칸이 한 구절이다.
        exc = tuple(x.strip() for x in str(cell(i_exc) or "").split(",") if x.strip())
        key = (normalize_text(kw), lo, hi, exc)
        if key in seen:
            continue
        seen.add(key)
        rules.append(AlertRule(keyword=kw, brand=brand, product=prod,
                               min_price=lo, max_price=hi, exclude=exc,
                               days=_as_int(cell(i_days)), row=n))
    if not rules and not errors:
        errors.append("엑셀에서 읽은 키워드가 없습니다")
    return rules, errors


def brands(rules) -> list[str]:
    """등록할 브랜드 목록 — 키워드 첫 어절, 처음 나온 순서로 중복 제거.

    모델 수백 줄을 키워드로 올릴 수는 없다(앱 알림 등록에 상한이 있다).
    브랜드로 넓게 받아 놓고 거르는 것이 이 구조의 전부다."""
    out, seen = [], set()
    for r in rules or []:
        b = r.brand_name()
        k = normalize_text(b)
        if b and k not in seen:
            seen.add(k)
            out.append(b)
    return out


def brand_days(rules) -> dict[str, int]:
    """브랜드별 끌올일수. 줄마다 다르면 가장 느슨한(큰) 값을 쓴다 —
    좁게 잡으면 조건에 맞는 매물을 놓친다."""
    out: dict[str, int] = {}
    for r in rules or []:
        if r.days:
            b = r.brand_name()
            out[b] = max(out.get(b, 0), int(r.days))
    return out
